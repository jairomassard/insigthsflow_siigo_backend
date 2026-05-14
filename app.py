from sqlalchemy import func
from flask import Flask, jsonify, request, current_app
from flask import send_file
from flask_jwt_extended import (
    JWTManager, create_access_token, jwt_required,
    get_jwt, get_jwt_identity, decode_token,
    verify_jwt_in_request  # 👈 agrega esto
)
from flask_cors import cross_origin
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import date, datetime, timezone, timedelta
from licenciamiento import obtener_codigos_permitidos_cliente, cliente_tiene_permiso_en_paquete
 
from config import Config
from models import db, Usuario, Cliente, Perfil, SesionActiva, SiigoCredencial, SiigoFactura, SiigoFacturaItem, SiigoVendedor, SiigoCentroCosto, SiigoCustomer, SiigoNotaCredito, SiigoPagoProveedor, SiigoProveedor, SiigoCompra, SiigoCompraItem, SiigoCuentasPorCobrar, SiigoNomina, SiigoProducto, BalancePrueba, Permiso, PerfilPermiso, SiigoSyncConfig, SiigoSyncLog, SiigoSyncMetric, SystemNotification, PaqueteInsightflow, PaquetePermiso, ClientePaquete
from flask_cors import CORS
import os
from cryptography.fernet import Fernet, InvalidToken
import base64, json, requests

from flask_jwt_extended import jwt_required, get_jwt
from flask_jwt_extended import get_jwt_identity, verify_jwt_in_request


from siigo_api import auth as siigo_auth, SiigoError
from siigo.siigo_sync_refactor import sync_facturas_desde_siigo
from datetime import datetime
import requests
from siigo.siigo_sync_refactor import _request_with_retries, _str, siigo_auth_json, _headers_bearer, dec_local, contar_facturas_pendientes
from siigo.siigo_sync_pagos import sync_pagos_egresos_desde_siigo
from siigo.siigo_sync_compras import sync_compras_desde_siigo
from sqlalchemy import text, func
from sqlalchemy.exc import IntegrityError
from utils import _siigo_headers_bearer, _siigo_auth_json_for_client

from io import BytesIO
import pandas as pd
import zipfile

import requests
import time

import pandas as pd

from sqlalchemy.sql import text
import math

import unicodedata
from sqlalchemy.dialects.postgresql import insert

from openpyxl import load_workbook
# 👇 Importa también los helpers de seguridad
from decoradores_seguridad import (
    permiso_requerido,
    _is_superadmin,
    _perfil_tiene_permiso
)

from utils import local_to_utc, utc_to_local


import threading
from threading import Thread
import traceback

import re
import unicodedata
from decimal import Decimal, InvalidOperation

from balance import (
    regenerar_snapshot_saldos_corte,
    regenerar_snapshots_balance,
    construir_balance_general,
)

from siigo.siigo_sync_documentos_soporte_staging import sync_documentos_soporte_staging_desde_siigo
from siigo.siigo_insert_documentos_soporte_desde_staging import insertar_documentos_soporte_desde_staging

from urllib.parse import urlencode


HEADER_MAP = {
    "nombre": [
        "nombre",
        "empleado",
        "trabajador",
        "nombre empleado",
        "nombre trabajador",
    ],

    "identificacion": [
        "identificación",
        "identificacion",
        "documento",
        "cedula",
        "cédula",
        "numero documento",
        "número documento",
    ],

    "no_contrato": [
        "no contrato",
        "número de contrato",
        "numero de contrato",
        "contrato",
        "id contrato",
    ],

    "sueldo": [
        "sueldo",
        "salario",
        "salario básico",
        "salario basico",
        "basico",
        "básico",
    ],

    "aux_transporte": [
        "auxilio transporte/conectividad",
        "auxilio transporte",
        "auxilio de transporte",
        "auxilio de transporte/conectividad",
        "aux. de transporte/aux. de conectividad digital",
        "auxilio conectividad",
        "auxilio de conectividad",
        "aux conectividad",
        "transporte",
        "aux transporte",
        "Aux. de transporte/Aux. de conectividad digital",
    ],

    "auxilio_extralegal": [
        "auxilio extralegal",
        "aux extralegal",
        "bonificacion extralegal",
        "bonificación extralegal",
    ],

    "prima": [
        "prima",
        "prima servicios",
        "prima de servicios",
        "prima semestral",
        "prima mitad de año",
        "prima fin de año",
    ],

    "intereses_cesantias": [
        "intereses cesantias",
        "intereses de cesantias",
        "intereses a las cesantias",
        "intereses de las cesantias",
        "intereses cesantías",
        "intereses de cesantías",
        "intereses a las cesantías",
        "intereses de las cesantías",
    ],

    "total_ingresos": [
        "total ingresos",
        "total ingresos devengados",
        "total devengado",
        "total devengados",
        "ingresos totales",
        "devengados",
    ],

    "fondo_salud": [
        "fondo de salud",
        "salud",
        "eps",
        "aporte salud",
    ],

    "fondo_pension": [
        "fondo de pensión",
        "fondo de pension",
        "pensión",
        "pension",
        "afp",
        "aporte pension",
        "aporte pensión",
    ],

    "fondo_solidaridad": [
        "fondo de solidaridad pensional",
        "solidaridad pensional",
        "fondo solidaridad",
    ],

    "retefuente": [
        "retefuente",
        "retencion en la fuente",
        "retención en la fuente",
    ],

    "prestamos": [
        "prestamos",
        "préstamos",
        "prestamo",
        "préstamo",
        "descuento prestamos",
        "descuento préstamo",
    ],

    "total_deducciones": [
        "total deducciones",
        "total deducciones empleado",
        "total descuentos",
        "deducciones totales",
    ],

    "neto_pagar": [
        "neto a pagar",
        "neto pagar",
        "valor neto",
        "neto",
    ],
}

REQUIRED_DB_FIELDS = [
    "nombre",
    "identificacion",
    "no_contrato",
    "sueldo",
    "aux_transporte",
    "auxilio_extralegal",
    "total_ingresos",
    "fondo_salud",
    "fondo_pension",
    "fondo_solidaridad",
    "total_deducciones",
    "neto_pagar",
]


FERNET_KEY = os.environ.get("APP_CRYPTO_KEY")  # genera una vez y guárdala en .env
fernet = Fernet(FERNET_KEY) if FERNET_KEY else None

#PARTNER_ID_DEFAULT = "ProjectManagerApp"
PARTNER_ID_DEFAULT = os.getenv("SIIGO_PARTNER_ID", "ProjectManagerApp")
AUTH_TIMEOUT_SEC = int(os.getenv("SIIGO_AUTH_TIMEOUT", "60"))
SANDBOX = os.getenv("SIIGO_SANDBOX", "0") == "1"

def enc(b: str | None) -> bytes | None:
    if not b: return None
    if not fernet: return b.encode()
    return fernet.encrypt(b.encode())

def dec(b: bytes | None) -> str | None:
    if b is None: return None
    if not fernet: return b.decode()
    try:
        return fernet.decrypt(b).decode()
    except InvalidToken:
        return None

def cleanup_expired_sessions():
    now = datetime.now(timezone.utc)
    # borra sesiones expiradas si el token ya venció
    SesionActiva.query.filter(
        SesionActiva.expira_en.isnot(None),
        SesionActiva.expira_en < now
    ).delete(synchronize_session=False)
    db.session.commit()


def _looks_like_b64(s: str) -> bool:
    try:
        base64.b64decode(s + "==", validate=True)
        return True
    except Exception:
        return False

def _decoded_has_colon(s: str) -> bool:
    try:
        return ":" in base64.b64decode(s + "==").decode("utf-8", "ignore")
    except Exception:
        return False


# --- Función utilitaria compartida por /reportes/analisis_clientes y por  /reportes/facturas_cliente ---
def enriquecer_facturas(rows):
    today = date.today()
    enriched = []

    for r in rows:
        dias = None
        estado = "sano"

        # Asegurar campo pendiente
        pendiente = r.get("pendiente", r.get("saldo", 0)) or 0
        try:
            pendiente = float(pendiente)
        except Exception:
            pendiente = 0

        if round(pendiente, 2) == 0:
            estado = "pagado"
        else:
            venc = r.get("vencimiento")
            if venc:
                venc_dt = venc if isinstance(venc, date) else date.fromisoformat(str(venc))
                dias = (venc_dt - today).days
                if dias < 0:
                    estado = "vencido"
                elif dias <= 5:
                    estado = "alerta"
                else:
                    estado = "sano"

        r["dias_vencimiento"] = dias
        r["estado_cartera"] = estado
        enriched.append(r)

    return enriched


def quitar_tildes(texto):
    if not texto:
        return texto
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )



def interpretar_indicador(k: str, v: float | None) -> str:
    if v is None:
        return "Sin datos suficientes para interpretar."

    match k:
        case "liquidez":
            if v < 1:
                return "Riesgo de iliquidez: activo corriente insuficiente."
            elif v <= 3:
                return "Saludable: puede cubrir obligaciones a corto plazo."
            else:
                return "Exceso de liquidez: posible ineficiencia de capital."
        case "apalancamiento":
            if v > 0.8:
                return "Endeudamiento alto: activos financiados en su mayoría por deuda."
            elif v > 0.6:
                return "Nivel de endeudamiento moderado."
            else:
                return "Apalancamiento controlado: estructura sana."
        case "rentabilidad":
            if v < 0:
                return "Rentabilidad negativa: pérdidas netas."
            elif v < 0.1:
                return "Rentabilidad positiva pero baja."
            else:
                return "Rentabilidad sólida sobre ingresos."
        case "solvencia":
            if v < 1:
                return "Solvencia deficiente: activos no cubren pasivos."
            elif v <= 2:
                return "Solvencia adecuada: activos cubren deudas."
            else:
                return "Solvencia muy alta: estructura sólida."
        case "autonomia":
            if v < 0:
                return "Patrimonio negativo: dependencia total de deuda."
            elif v < 0.3:
                return "Alta dependencia del financiamiento externo."
            else:
                return "Buena autonomía financiera."
        case "endeudamiento_largo_plazo":
            if v <= 0:
                return "No interpretable (patrimonio negativo o nulo)."
            elif v > 1:
                return "Presión financiera a largo plazo elevada."
            else:
                return "Estructura de deuda a largo plazo saludable."
        case "capital_trabajo":
            if v < 0:
                return "Capital de trabajo negativo: riesgo operativo."
            elif v == 0:
                return "Capital de trabajo neutro."
            else:
                return "Colchón operativo positivo."
        case "cobertura_activo_pasivo":
            if v < 1:
                return "Activos insuficientes para cubrir pasivos."
            elif v <= 2:
                return "Cobertura aceptable de pasivos."
            else:
                return "Alta cobertura de pasivos: buena estructura."
        case "porcentaje_activo_no_corriente":
            if v > 0.7:
                return "Alta proporción de activos no líquidos."
            elif v > 0.4:
                return "Balance adecuado entre activos líquidos y fijos."
            else:
                return "Predominio de activos líquidos."
        case "porcentaje_pasivo_corto":
            if v > 0.7:
                return "Alta carga de pasivos a corto plazo."
            elif v > 0.4:
                return "Distribución equilibrada de pasivos."
            else:
                return "Predominio de deuda a largo plazo."
        case _:
            return "Sin interpretación disponible."




def obtener_idcliente_desde_request():
    """
    Obtiene el ID del cliente necesario para todas las sincronizaciones.
    Primero intenta el header 'X-ID-CLIENTE'. Si no existe, intenta extraerlo del JWT.
    """

    # 1. Intentar header primero
    x_id = request.headers.get("X-ID-CLIENTE")
    if x_id:
        print("obtener_idcliente_desde_request: desde header →", x_id)
        try:
            return int(x_id)
        except ValueError:
            return None

    # 2. Si no hay header, intentar JWT opcional
    try:
        verify_jwt_in_request(optional=True)  # 👈 aquí el truco
        identity = get_jwt_identity()
        if identity and isinstance(identity, dict):
            idc = identity.get("idcliente")
            print("obtener_idcliente_desde_request: desde JWT →", idc)
            if idc is not None:
                return int(idc)
    except Exception as e:
        print("obtener_idcliente_desde_request: error al leer JWT:", e)

    print("obtener_idcliente_desde_request: no se encontró idcliente")
    return None

# funciones para cargue campos de nomina
def normalizar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = texto.replace("\xa0", " ")
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def normalizar_numero(valor):
    """
    Convierte valores de Excel a Decimal seguro.
    Soporta:
    - None
    - NaN
    - enteros / floats
    - strings como '1.234.567,89' o '1234567.89'
    """
    if valor is None:
        return Decimal("0")

    try:
        if hasattr(valor, "__class__") and str(valor).lower() == "nan":
            return Decimal("0")
    except Exception:
        pass

    if isinstance(valor, (int, float, Decimal)):
        try:
            return Decimal(str(valor))
        except Exception:
            return Decimal("0")

    texto = str(valor).strip()
    if not texto:
        return Decimal("0")

    texto = texto.replace("$", "").replace(" ", "")

    # Caso formato latino: 1.234.567,89
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    # Caso 1234,56
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return Decimal(texto)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def resolver_columnas(headers_originales):
    """
    Retorna un dict:
    {
        "nombre": "Nombre",
        "identificacion": "Identificación",
        ...
    }
    mapeando el campo DB al nombre real de la columna del Excel.
    """
    headers_norm = {normalizar_texto(h): h for h in headers_originales if h is not None}
    columnas_resueltas = {}

    for campo_db, aliases in HEADER_MAP.items():
        for alias in aliases:
            alias_norm = normalizar_texto(alias)
            if alias_norm in headers_norm:
                columnas_resueltas[campo_db] = headers_norm[alias_norm]
                break

    return columnas_resueltas

# --- LÓGICA EXCLUSIVA PARA PROCESAMIENTO DE AUXILIAR CONTABLE SIIGO (LIBRO AUXILIAR) ---

def normalizar_valor_auxiliar(valor):
    """Convierte celdas de Excel a Decimal de forma ultra segura."""
    if valor is None or str(valor).lower() == 'nan':
        return Decimal("0.00")
    try:
        if isinstance(valor, (int, float, Decimal)):
            return Decimal(str(valor))
        texto = str(valor).replace("$", "").replace(" ", "").replace(".", "").replace(",", ".")
        return Decimal(texto)
    except:
        return Decimal("0.00")

def procesar_excel_auxiliar_v2(file_storage, idcliente):
    """Parser universal para el reporte de Siigo Nube."""
    # Leemos el Excel
    df = pd.read_excel(file_storage, skiprows=4)
    
    registros = []
    
    for _, row in df.iterrows():
        # 1. SEGURIDAD: Evitar filas vacías o totales de Siigo
        # Si la primera o segunda columna están vacías, saltar.
        if pd.isna(row.iloc[0]) or pd.isna(row.iloc[1]):
            continue
            
        try:
            # 2. MANEJO DE FECHA: Asegurar que sea una fecha válida
            f_val = row.iloc[0]
            f_contable = pd.to_datetime(f_val) if not isinstance(f_val, datetime) else f_val

            # 3. PROCESAMIENTO DE CUENTA: Separar código y nombre
            cuenta_raw = str(row.iloc[1]).strip()
            # Siigo a veces pone "240805 IVA..." o solo "240805"
            partes_cta = cuenta_raw.split(' ', 1) 
            codigo_cta = partes_cta[0]
            nombre_cta = partes_cta[1] if len(partes_cta) > 1 else ""
            
            # 4. COMPROBANTE
            comp_raw = str(row.iloc[2])
            tipo_comp = comp_raw.split('-')[0] if '-' in comp_raw else comp_raw
            num_comp = comp_raw.split('-')[-1] if '-' in comp_raw else ""

            # 5. CONSTRUCCIÓN DEL DICCIONARIO (Captura TOTAL de cuentas)
            registros.append({
                "idcliente": idcliente,
                "fecha_contable": f_contable.date(),
                "comprobante_tipo": tipo_comp,
                "comprobante_numero": num_comp,
                "cuenta_codigo": codigo_cta,
                "cuenta_nombre": nombre_cta,
                "tercero_nit": str(row.iloc[3]) if not pd.isna(row.iloc[3]) else "",
                "tercero_nombre": str(row.iloc[4]) if not pd.isna(row.iloc[4]) else "SIN TERCERO",
                "detalle": str(row.iloc[5]) if not pd.isna(row.iloc[5]) else "",
                "debito": normalizar_valor_auxiliar(row.iloc[6]),
                "credito": normalizar_valor_auxiliar(row.iloc[7]),
                "base_gravable": normalizar_valor_auxiliar(row.iloc[8]),
                "periodo_anio": f_contable.year,
                "periodo_mes": f_contable.month,
                "archivo_origen": getattr(file_storage, 'filename', 'upload_excel')
            })
        except Exception as e:
            # Opcional: imprimir el error para debug
            # print(f"Error en fila: {e}")
            continue
            
    return registros


def _obtener_busqueda_inteligente_facturas_data(idcliente, perfilid):
    from sqlalchemy import text
    from datetime import datetime, timedelta
    from decimal import Decimal

    q_idcliente = request.args.get("idcliente", type=int)
    if perfilid == 0 and q_idcliente:
        idcliente = q_idcliente

    if not idcliente:
        raise ValueError("No autorizado")

    q = (request.args.get("q") or "").strip().lower()
    factura = (request.args.get("factura") or request.args.get("idfactura") or "").strip()
    cliente = (request.args.get("cliente") or "").strip()
    desde = request.args.get("desde") or "2025-01-01"
    hasta = request.args.get("hasta") or datetime.now().strftime("%Y-%m-%d")
    cost_center = request.args.get("cost_center", type=int)
    estado_pago = (request.args.get("estado_pago") or "").strip().lower()
    estado_factura = (request.args.get("estado_factura") or request.args.get("estado") or "").strip().lower()
    limit = request.args.get("limit", type=int) or 5000

    if limit <= 0:
        limit = 5000

    if limit > 20000:
        limit = 20000

    # Incluir completo el día final.
    hasta_dt = datetime.strptime(hasta, "%Y-%m-%d") + timedelta(days=1)
    hasta_sql = hasta_dt.strftime("%Y-%m-%d")

    aliases = {
        "zapier": ["zapier", "zappier"],
        "zappier": ["zapier", "zappier"],
    }

    terminos = aliases.get(q, [q] if q else [])

    params = {
        "idcliente": idcliente,
        "desde": desde,
        "hasta": hasta_sql,
        "limit": limit,
        "estado_pago": estado_pago,
    }

    # ==========================================================
    # FUENTE OFICIAL PARA ESTA PÁGINA:
    # ventas_movimientos_enriquecidos
    #
    # Esta vista trae:
    # - FACTURA
    # - NOTA_CREDITO
    #
    # No usamos facturas_enriquecidas para evitar distorsiones.
    # ==========================================================

    where_main = [
        "m.idcliente = :idcliente",
        "m.fecha >= :desde",
        "m.fecha < :hasta",
    ]

    if factura:
        where_main.append("""
            (
                LOWER(COALESCE(m.documento, '')) LIKE :factura
                OR LOWER(COALESCE(m.documento_afectado, '')) LIKE :factura
            )
        """)
        params["factura"] = f"%{factura.lower()}%"

    if cliente:
        where_main.append("LOWER(TRIM(m.cliente_nombre)) = LOWER(TRIM(:cliente))")
        params["cliente"] = cliente

    if cost_center:
        where_main.append("m.cost_center = :cost_center")
        params["cost_center"] = cost_center

    if estado_factura:
        # Para notas crédito, normalmente el estado puede venir nulo.
        # Este filtro aplica sobre el estado del movimiento.
        where_main.append("LOWER(COALESCE(m.estado, '')) = :estado_factura")
        params["estado_factura"] = estado_factura

    exists_filter = ""
    descripcion_filter = ""

    if terminos:
        term_clauses_exists = []
        term_clauses_desc = []
        term_clauses_header = []

        for i, term in enumerate(terminos):
            key = f"term_{i}"
            params[key] = f"%{term}%"

            # Búsqueda en ítems solo aplica para facturas.
            term_clauses_exists.append(
                f"LOWER(COALESCE(fi.descripcion, '')) LIKE LOWER(:{key})"
            )

            term_clauses_desc.append(
                f"LOWER(COALESCE(fi2.descripcion, '')) LIKE LOWER(:{key})"
            )

            # Búsqueda general en encabezado de factura / nota crédito.
            term_clauses_header.extend([
                f"LOWER(COALESCE(m.documento, '')) LIKE LOWER(:{key})",
                f"LOWER(COALESCE(m.documento_afectado, '')) LIKE LOWER(:{key})",
                f"LOWER(COALESCE(m.cliente_nombre, '')) LIKE LOWER(:{key})",
                f"LOWER(COALESCE(m.centro_costo_nombre, '')) LIKE LOWER(:{key})",
                f"LOWER(COALESCE(m.vendedor_nombre, '')) LIKE LOWER(:{key})",
            ])

        exists_filter = f"""
            AND (
                {' OR '.join(term_clauses_header)}
                OR (
                    m.tipo_movimiento = 'FACTURA'
                    AND EXISTS (
                        SELECT 1
                        FROM siigo_factura_items fi
                        WHERE fi.factura_id = m.movimiento_id
                          AND fi.idcliente = m.idcliente
                          AND ({' OR '.join(term_clauses_exists)})
                    )
                )
            )
        """

        descripcion_filter = f"""
            AND ({' OR '.join(term_clauses_desc)})
        """

    where_clause = " AND ".join(where_main)

    sql_rows = text(f"""
        WITH movimientos AS (
            SELECT
                m.movimiento_id,
                m.idcliente,
                m.documento,
                m.tipo_movimiento,
                m.fecha,
                m.vencimiento,
                m.cliente_nombre,
                COALESCE(m.estado, '') AS estado,
                COALESCE(m.estado_pago, '') AS estado_pago,

                COALESCE(m.subtotal, 0) AS subtotal,
                COALESCE(m.impuestos_total, 0) AS impuestos,
                COALESCE(m.total, 0) AS total,

                COALESCE(m.pagos_total, 0) AS pagos_total,
                COALESCE(m.saldo, 0) AS saldo,

                m.cost_center,
                COALESCE(m.centro_costo_nombre, 'Sin centro de costo') AS centro_costo_nombre,
                COALESCE(m.centro_costo_codigo, '') AS centro_costo_codigo,

                m.seller_id,
                COALESCE(m.vendedor_nombre, 'Sin vendedor') AS vendedor_nombre,

                m.public_url,
                COALESCE(m.retenciones, '[]'::jsonb) AS retenciones,
                m.documento_afectado

            FROM ventas_movimientos_enriquecidos m
            WHERE {where_clause}
            {exists_filter}
        ),
        base AS (
            SELECT
                mv.movimiento_id AS factura_id,
                mv.documento AS idfactura,
                mv.documento,
                mv.tipo_movimiento,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN 'Factura'
                    WHEN mv.tipo_movimiento = 'NOTA_CREDITO' THEN 'Nota crédito'
                    ELSE mv.tipo_movimiento
                END AS tipo_documento_label,

                mv.fecha,
                mv.vencimiento,
                mv.cliente_nombre,
                mv.estado,

                mv.subtotal,
                mv.impuestos,
                mv.total,
                ABS(mv.total) AS total_abs,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN mv.pagos_total
                    ELSE 0
                END AS pagos_total,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN mv.saldo
                    ELSE 0
                END AS saldo,

                mv.public_url,
                ''::text AS observaciones,
                ''::text AS medio_pago,

                mv.cost_center,
                mv.centro_costo_nombre,
                mv.centro_costo_codigo,
                mv.vendedor_nombre,
                mv.documento_afectado,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN
                        COALESCE((
                            SELECT SUM((r->>'value')::numeric)
                            FROM jsonb_array_elements(mv.retenciones) AS r
                            WHERE jsonb_typeof(mv.retenciones) = 'array'
                              AND LOWER(COALESCE(r->>'type', '')) = 'reteica'
                        ), 0)
                    ELSE 0
                END AS reteica,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN
                        COALESCE((
                            SELECT SUM((r->>'value')::numeric)
                            FROM jsonb_array_elements(mv.retenciones) AS r
                            WHERE jsonb_typeof(mv.retenciones) = 'array'
                              AND LOWER(COALESCE(r->>'type', '')) = 'reteiva'
                        ), 0)
                    ELSE 0
                END AS reteiva,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN
                        COALESCE((
                            SELECT SUM((r->>'value')::numeric)
                            FROM jsonb_array_elements(mv.retenciones) AS r
                            WHERE jsonb_typeof(mv.retenciones) = 'array'
                              AND LOWER(COALESCE(r->>'type', '')) LIKE '%autorretencion%'
                        ), 0)
                    ELSE 0
                END AS autorretencion,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN
                        COALESCE((
                            SELECT SUM((r->>'value')::numeric)
                            FROM jsonb_array_elements(mv.retenciones) AS r
                            WHERE jsonb_typeof(mv.retenciones) = 'array'
                        ), 0)
                    ELSE 0
                END AS total_retenciones,

                CASE
                    WHEN mv.tipo_movimiento = 'FACTURA' THEN
                        (
                            SELECT STRING_AGG(TRIM(fi2.descripcion), ' || ' ORDER BY TRIM(fi2.descripcion))
                            FROM siigo_factura_items fi2
                            WHERE fi2.factura_id = mv.movimiento_id
                              AND fi2.idcliente = mv.idcliente
                              {descripcion_filter}
                        )
                    ELSE
                        CONCAT('Nota crédito que afecta la factura ', COALESCE(mv.documento_afectado, 'sin referencia'))
                END AS descripcion,

                CASE
                    WHEN mv.tipo_movimiento = 'NOTA_CREDITO' THEN 'no_aplica'
                    WHEN ABS(COALESCE(mv.saldo, 0)) <= 1 THEN 'pagada'
                    WHEN ABS(COALESCE(mv.total, 0) - COALESCE(mv.saldo, 0)) <= 1 THEN 'pendiente'
                    ELSE 'parcial'
                END AS estado_pago_real

            FROM movimientos mv
        )
        SELECT *
        FROM base
        WHERE (
            :estado_pago = ''
            OR estado_pago_real = :estado_pago
        )
        ORDER BY fecha DESC, idfactura DESC
        LIMIT :limit
    """)

    rows_raw = db.session.execute(sql_rows, params).mappings().all()

    def norm(v):
        if isinstance(v, Decimal):
            return float(v)
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return v

    rows = [{k: norm(v) for k, v in row.items()} for row in rows_raw]

    # ==========================================================
    # KPIs coherentes:
    # - facturas_emitidas: solo FACTURA en positivo
    # - notas_credito: solo NOTA_CREDITO en positivo
    # - ventas_netas: suma real de movimientos
    # - saldo: solo saldos de facturas
    # ==========================================================

    facturas_emitidas = float(
        sum((r.get("total") or 0) for r in rows if r.get("tipo_movimiento") == "FACTURA")
    )

    notas_credito = float(
        abs(sum((r.get("total") or 0) for r in rows if r.get("tipo_movimiento") == "NOTA_CREDITO"))
    )

    ventas_netas = float(sum((r.get("total") or 0) for r in rows))
    subtotal_neto = float(sum((r.get("subtotal") or 0) for r in rows))
    iva_neto = float(sum((r.get("impuestos") or 0) for r in rows))

    saldo_cartera = float(
        sum((r.get("saldo") or 0) for r in rows if r.get("tipo_movimiento") == "FACTURA")
    )

    summary = {
        "total_registros": len(rows),

        "cantidad_facturas": int(sum(1 for r in rows if r.get("tipo_movimiento") == "FACTURA")),
        "cantidad_notas_credito": int(sum(1 for r in rows if r.get("tipo_movimiento") == "NOTA_CREDITO")),

        "facturas_emitidas": facturas_emitidas,
        "notas_credito": notas_credito,
        "ventas_netas": ventas_netas,

        # Compatibilidad con frontend actual:
        # total_facturado se deja como venta neta comercial para que la gráfica no vuelva a inflarse.
        "total_facturado": ventas_netas,

        "total_facturado_bruto": facturas_emitidas,

        "subtotal": subtotal_neto,
        "iva": iva_neto,

        "reteica_total": float(sum((r.get("reteica") or 0) for r in rows)),
        "reteiva_total": float(sum((r.get("reteiva") or 0) for r in rows)),
        "autorretencion_total": float(sum((r.get("autorretencion") or 0) for r in rows)),
        "retenciones": float(sum((r.get("total_retenciones") or 0) for r in rows)),

        "saldo": saldo_cartera,
    }

    sql_series = text(f"""
        WITH movimientos AS (
            SELECT
                m.fecha,
                m.tipo_movimiento,
                COALESCE(m.total, 0) AS total,

                CASE
                    WHEN m.tipo_movimiento = 'NOTA_CREDITO' THEN 'no_aplica'
                    WHEN ABS(COALESCE(m.saldo, 0)) <= 1 THEN 'pagada'
                    WHEN ABS(COALESCE(m.total, 0) - COALESCE(m.saldo, 0)) <= 1 THEN 'pendiente'
                    ELSE 'parcial'
                END AS estado_pago_real

            FROM ventas_movimientos_enriquecidos m
            WHERE {where_clause}
            {exists_filter}
        )
        SELECT
            TO_CHAR(date_trunc('month', fecha), 'YYYY-MM') AS mes,

            COALESCE(SUM(
                CASE
                    WHEN tipo_movimiento = 'FACTURA' THEN total
                    ELSE 0
                END
            ), 0) AS facturas_emitidas,

            ABS(COALESCE(SUM(
                CASE
                    WHEN tipo_movimiento = 'NOTA_CREDITO' THEN total
                    ELSE 0
                END
            ), 0)) AS notas_credito,

            COALESCE(SUM(total), 0) AS ventas_netas,

            -- Compatibilidad con frontend actual:
            COALESCE(SUM(total), 0) AS total_facturado,

            COUNT(*) AS cantidad

        FROM movimientos
        WHERE (
            :estado_pago = ''
            OR estado_pago_real = :estado_pago
        )
        GROUP BY 1
        ORDER BY 1
    """)

    series_raw = db.session.execute(sql_series, params).mappings().all()
    series = [{k: norm(v) for k, v in row.items()} for row in series_raw]

    return {
        "rows": rows,
        "count": len(rows),
        "kpis": summary,
        "series": series,
        "filters": {
            "q": q,
            "factura": factura,
            "cliente": cliente,
            "cost_center": cost_center,
            "estado_pago": estado_pago,
            "estado_factura": estado_factura,
            "desde": desde,
            "hasta": hasta,
            "terminos_aplicados": terminos,
            "fuente": "ventas_movimientos_enriquecidos",
            "logica": "FACTURA + NOTA_CREDITO separadas",
        },
    }
    

# Helper para pagina de PNL y para la de indicadores
def construir_pnl_auxiliares(idcliente, desde, hasta):
    from sqlalchemy import text

    sql_evo = text("""
        SELECT
            periodo_anio,
            periodo_mes,

            SUM(CASE
                WHEN cuenta_codigo LIKE '41%' THEN (credito - debito)
                ELSE 0
            END) AS ingresos_operacionales,

            SUM(CASE
                WHEN cuenta_codigo LIKE '42%' THEN (credito - debito)
                ELSE 0
            END) AS ingresos_no_operacionales,

            SUM(CASE
                WHEN cuenta_codigo LIKE '6%' OR cuenta_codigo LIKE '7%' THEN (debito - credito)
                ELSE 0
            END) AS costos_venta,

            SUM(CASE
                WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN (debito - credito)
                ELSE 0
            END) AS gastos_operacionales,

            SUM(CASE
                WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN (debito - credito)
                ELSE 0
            END) AS gastos_no_operacionales,

            SUM(CASE
                WHEN cuenta_codigo LIKE '5160%'
                OR cuenta_codigo LIKE '5165%'
                OR cuenta_codigo LIKE '5260%'
                OR cuenta_codigo LIKE '5265%'
                THEN (debito - credito)
                ELSE 0
            END) AS dep_amort
        FROM auxiliar_contable
        WHERE idcliente = :idc
        AND fecha_contable BETWEEN :d AND :h
        GROUP BY periodo_anio, periodo_mes
        ORDER BY periodo_anio, periodo_mes
    """)

    sql_comp = text("""
        SELECT
            periodo_anio,
            periodo_mes,
            cuenta_codigo,
            LEFT(cuenta_codigo, 4) AS cuenta_padre,
            MAX(cuenta_nombre) AS nombre_cuenta,

            CASE
                WHEN cuenta_codigo LIKE '41%' THEN 'INGRESOS_OPERACIONALES'
                WHEN cuenta_codigo LIKE '42%' THEN 'INGRESOS_NO_OPERACIONALES'
                WHEN cuenta_codigo LIKE '6%'  OR cuenta_codigo LIKE '7%' THEN 'COSTOS_VENTA'
                WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN 'GASTOS_OPERACIONALES'
                WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN 'GASTOS_NO_OPERACIONALES'
                ELSE 'OTROS'
            END AS seccion,

            CASE
                WHEN LEFT(cuenta_codigo, 1) = '4' THEN 'CREDITO_MENOS_DEBITO'
                ELSE 'DEBITO_MENOS_CREDITO'
            END AS naturaleza,

            SUM(
                CASE
                    WHEN LEFT(cuenta_codigo, 1) = '4' THEN (credito - debito)
                    ELSE (debito - credito)
                END
            ) AS saldo
        FROM auxiliar_contable
        WHERE idcliente = :idc
        AND fecha_contable BETWEEN :d AND :h
        AND LEFT(cuenta_codigo, 1) IN ('4', '5', '6', '7')
        GROUP BY
            periodo_anio,
            periodo_mes,
            cuenta_codigo,
            LEFT(cuenta_codigo, 4),
            CASE
                WHEN cuenta_codigo LIKE '41%' THEN 'INGRESOS_OPERACIONALES'
                WHEN cuenta_codigo LIKE '42%' THEN 'INGRESOS_NO_OPERACIONALES'
                WHEN cuenta_codigo LIKE '6%'  OR cuenta_codigo LIKE '7%' THEN 'COSTOS_VENTA'
                WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN 'GASTOS_OPERACIONALES'
                WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN 'GASTOS_NO_OPERACIONALES'
                ELSE 'OTROS'
            END,
            CASE
                WHEN LEFT(cuenta_codigo, 1) = '4' THEN 'CREDITO_MENOS_DEBITO'
                ELSE 'DEBITO_MENOS_CREDITO'
            END
        HAVING SUM(
            CASE
                WHEN LEFT(cuenta_codigo, 1) = '4' THEN (credito - debito)
                ELSE (debito - credito)
            END
        ) <> 0
        ORDER BY periodo_anio, periodo_mes, cuenta_codigo
    """)

    res_evo = db.session.execute(sql_evo, {
        "idc": idcliente,
        "d": desde,
        "h": hasta
    }).mappings().all()

    res_comp = db.session.execute(sql_comp, {
        "idc": idcliente,
        "d": desde,
        "h": hasta
    }).mappings().all()

    evolucion = []
    totales = {
        "ingresos_operacionales": 0.0,
        "ingresos_no_operacionales": 0.0,
        "costos_venta": 0.0,
        "gastos_operacionales": 0.0,
        "gastos_no_operacionales": 0.0,
        "dep_amort": 0.0,
    }

    for r in res_evo:
        ing_op = float(r["ingresos_operacionales"] or 0)
        ing_no_op = float(r["ingresos_no_operacionales"] or 0)
        costos = float(r["costos_venta"] or 0)
        gastos_op = float(r["gastos_operacionales"] or 0)
        gastos_no_op = float(r["gastos_no_operacionales"] or 0)
        dep_amort = float(r["dep_amort"] or 0)

        ingresos_totales = ing_op + ing_no_op
        utilidad_bruta = ing_op - costos
        utilidad_operativa = utilidad_bruta - gastos_op
        ebitda = utilidad_operativa + dep_amort
        utilidad_antes_impuestos = utilidad_operativa + ing_no_op - gastos_no_op
        utilidad_neta = utilidad_antes_impuestos

        totales["ingresos_operacionales"] += ing_op
        totales["ingresos_no_operacionales"] += ing_no_op
        totales["costos_venta"] += costos
        totales["gastos_operacionales"] += gastos_op
        totales["gastos_no_operacionales"] += gastos_no_op
        totales["dep_amort"] += dep_amort

        base_margen = ingresos_totales if ingresos_totales != 0 else 0

        evolucion.append({
            "label": f"{r['periodo_anio']}-{int(r['periodo_mes']):02d}",
            "ingresos_operacionales": ing_op,
            "ingresos_no_operacionales": ing_no_op,
            "ingresos_totales": ingresos_totales,
            "costos_venta": costos,
            "gastos_operacionales": gastos_op,
            "gastos_no_operacionales": gastos_no_op,
            "utilidad_bruta": utilidad_bruta,
            "utilidad_operativa": utilidad_operativa,
            "ebitda": ebitda,
            "utilidad_antes_impuestos": utilidad_antes_impuestos,
            "utilidad_neta": utilidad_neta,
            "costos_gastos": costos + gastos_op + gastos_no_op,
            "margen_bruto": round((utilidad_bruta / base_margen) * 100, 2) if base_margen else 0,
            "margen_operativo": round((utilidad_operativa / base_margen) * 100, 2) if base_margen else 0,
            "margen_ebitda": round((ebitda / base_margen) * 100, 2) if base_margen else 0,
            "margen_neto": round((utilidad_neta / base_margen) * 100, 2) if base_margen else 0,
        })

    ingresos_operacionales = totales["ingresos_operacionales"]
    ingresos_no_operacionales = totales["ingresos_no_operacionales"]
    ingresos_totales = ingresos_operacionales + ingresos_no_operacionales
    costos_venta = totales["costos_venta"]
    gastos_operacionales = totales["gastos_operacionales"]
    gastos_no_operacionales = totales["gastos_no_operacionales"]
    dep_amort = totales["dep_amort"]

    utilidad_bruta = ingresos_operacionales - costos_venta
    utilidad_operativa = utilidad_bruta - gastos_operacionales
    ebitda = utilidad_operativa + dep_amort
    utilidad_antes_impuestos = utilidad_operativa + ingresos_no_operacionales - gastos_no_operacionales
    utilidad_neta = utilidad_antes_impuestos

    base_margen = ingresos_totales if ingresos_totales != 0 else 0

    cuentas_dict = {}
    for c in res_comp:
        cuenta_codigo = str(c["cuenta_codigo"])
        periodo = f"{c['periodo_anio']}-{int(c['periodo_mes']):02d}"

        if cuenta_codigo not in cuentas_dict:
            cuentas_dict[cuenta_codigo] = {
                "cuenta": cuenta_codigo,
                "cuenta_padre": str(c["cuenta_padre"]),
                "nombre": str(c["nombre_cuenta"]).strip().title(),
                "seccion": str(c["seccion"]),
                "naturaleza": str(c["naturaleza"]),
                "valores_mes": {},
                "total": 0.0,
            }

        val = float(c["saldo"] or 0)
        cuentas_dict[cuenta_codigo]["valores_mes"][periodo] = val
        cuentas_dict[cuenta_codigo]["total"] += val

    composicion = list(cuentas_dict.values())
    composicion.sort(key=lambda x: x["cuenta"])

    return {
        "kpis": {
            "ingresos_operacionales": ingresos_operacionales,
            "ingresos_no_operacionales": ingresos_no_operacionales,
            "ingresos_totales": ingresos_totales,
            "costos_venta": costos_venta,
            "utilidad_bruta": utilidad_bruta,
            "gastos_operacionales": gastos_operacionales,
            "utilidad_operativa": utilidad_operativa,
            "ebitda": ebitda,
            "gastos_no_operacionales": gastos_no_operacionales,
            "utilidad_antes_impuestos": utilidad_antes_impuestos,
            "utilidad_neta": utilidad_neta,
            "margen_bruto": round((utilidad_bruta / base_margen) * 100, 2) if base_margen else 0,
            "margen_operativo": round((utilidad_operativa / base_margen) * 100, 2) if base_margen else 0,
            "margen_ebitda": round((ebitda / base_margen) * 100, 2) if base_margen else 0,
            "margen_neto": round((utilidad_neta / base_margen) * 100, 2) if base_margen else 0,
        },
        "evolucion": evolucion,
        "composicion": composicion
    }



# =========================================================
# HELPERS DASHBOARD / RESUMEN EJECUTIVO INTELIGENTE
# =========================================================

from datetime import datetime, date, timedelta
from sqlalchemy import text


def _safe_float(val):
    try:
        return float(val or 0)
    except Exception:
        return 0.0


def _round2(val):
    return round(_safe_float(val), 2)


def _variacion(actual, anterior):
    actual = _safe_float(actual)
    anterior = _safe_float(anterior)
    diff = actual - anterior
    pct = (diff / anterior * 100) if anterior not in (0, None) else 0
    return {
        "actual": _round2(actual),
        "anterior": _round2(anterior),
        "diff": _round2(diff),
        "pct": _round2(pct),
    }


def _shift_months(dt, months):
    year = dt.year + ((dt.month - 1 + months) // 12)
    month = ((dt.month - 1 + months) % 12) + 1
    day = min(dt.day, 28)
    return dt.replace(year=year, month=month, day=day)


def _first_day_of_month(dt):
    return dt.replace(day=1)


def _last_day_of_month(dt):
    next_month = _shift_months(dt.replace(day=1), 1)
    return next_month - timedelta(days=1)


def _first_day_of_year(dt):
    return dt.replace(month=1, day=1)


def _ultimo_periodo_auxiliar_con_datos(idcliente):
    sql = text("""
        SELECT MAX(fecha_contable) AS ultima_fecha
        FROM auxiliar_contable
        WHERE idcliente = :idc
    """)
    row = db.session.execute(sql, {"idc": idcliente}).mappings().first()
    return row["ultima_fecha"] if row and row["ultima_fecha"] else None


def _resolver_corte_confiable_auxiliar(idcliente):
    """
    Define el corte confiable para lectura ejecutiva.

    Reglas:
    - Si la última fecha cargada cae exactamente en el último día de su mes,
      se considera mes cerrado y ese mismo día es el corte confiable.
    - Si la última fecha cargada es parcial dentro del mes,
      el corte confiable será el último día del mes anterior.
    - Si no existe data, retorna estructura vacía.
    """
    ultima_fecha = _ultimo_periodo_auxiliar_con_datos(idcliente)

    if not ultima_fecha:
        return {
            "ultima_fecha_auxiliar": None,
            "fecha_corte_confiable": None,
            "desde_ytd": None,
            "hasta_ytd": None,
            "mes_actual_parcial": False,
            "anio_corte": None,
            "mes_corte": None,
            "modo_periodo": "sin_datos",
        }

    ultimo_dia_del_mes = _last_day_of_month(ultima_fecha)

    if ultima_fecha == ultimo_dia_del_mes:
        fecha_corte_confiable = ultima_fecha
        mes_actual_parcial = False
    else:
        fecha_corte_confiable = _last_day_of_month(_shift_months(ultima_fecha, -1))
        mes_actual_parcial = True

    desde_ytd = date(fecha_corte_confiable.year, 1, 1)
    hasta_ytd = fecha_corte_confiable

    return {
        "ultima_fecha_auxiliar": ultima_fecha,
        "fecha_corte_confiable": fecha_corte_confiable,
        "desde_ytd": desde_ytd,
        "hasta_ytd": hasta_ytd,
        "mes_actual_parcial": mes_actual_parcial,
        "anio_corte": fecha_corte_confiable.year,
        "mes_corte": fecha_corte_confiable.month,
        "modo_periodo": "ytd_cerrado",
    }


def _obtener_config_dashboard(idcliente):
    sql = text("""
        SELECT
            idcliente,
            activo,
            mostrar_caja,
            mostrar_runway,
            modo_caja,
            cuentas_incluidas,
            cuentas_excluidas,
            modo_runway,
            meses_promedio_runway,
            meta_eficiencia_operativa,
            meta_ebitda,
            meta_margen_ebitda,
            meses_grafica,
            top_clientes,
            top_proveedores,
            top_gastos,
            indicador_estrella,
            modo_periodo_default
        FROM dashboard_resumen_config
        WHERE idcliente = :idc
        LIMIT 1
    """)
    row = db.session.execute(sql, {"idc": idcliente}).mappings().first()
    return dict(row) if row else None


def _normalizar_lista_cuentas_config(lista_raw):
    cuentas = []

    if not lista_raw:
        return cuentas

    if isinstance(lista_raw, list):
        for item in lista_raw:
            if isinstance(item, dict):
                codigo = str(item.get("codigo", "")).strip()
                nombre = str(item.get("nombre", "")).strip()
                if codigo:
                    cuentas.append({
                        "codigo": codigo,
                        "nombre": nombre
                    })
            else:
                codigo = str(item).strip()
                if codigo:
                    cuentas.append({
                        "codigo": codigo,
                        "nombre": ""
                    })
    return cuentas


def _resolver_meta_eficiencia(config):
    if not config:
        return 20.0
    return _round2(config.get("meta_eficiencia_operativa", 20.0))


def _calcular_caja_disponible_parametrizada(idcliente, hasta, config):
    if not config:
        return {
            "actual": None,
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_caja": "sin_configurar",
            "mensaje": "La caja disponible requiere parametrización para este cliente.",
            "cuentas_usadas": []
        }

    mostrar_caja = bool(config.get("mostrar_caja", False))
    modo_caja = str(config.get("modo_caja") or "sin_configurar").strip()

    cuentas_incluidas = _normalizar_lista_cuentas_config(config.get("cuentas_incluidas"))
    cuentas_excluidas = _normalizar_lista_cuentas_config(config.get("cuentas_excluidas"))

    if not mostrar_caja:
        return {
            "actual": None,
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_caja": modo_caja,
            "mensaje": "La visualización de caja disponible está desactivada para este cliente.",
            "cuentas_usadas": []
        }

    if modo_caja == "sin_configurar":
        return {
            "actual": None,
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_caja": modo_caja,
            "mensaje": "La caja disponible requiere parametrización de cuentas para este cliente.",
            "cuentas_usadas": []
        }

    if modo_caja == "inclusion":
        codigos = [c["codigo"] for c in cuentas_incluidas if c.get("codigo")]
        if not codigos:
            return {
                "actual": None,
                "requiere_parametrizacion": True,
                "visible": False,
                "modo_caja": modo_caja,
                "mensaje": "Modo inclusión activo, pero no hay cuentas incluidas configuradas.",
                "cuentas_usadas": []
            }

        condiciones = []
        params = {"idc": idcliente, "hasta": hasta}

        for idx, codigo in enumerate(codigos):
            key = f"c{idx}"
            condiciones.append(f"cuenta_codigo LIKE :{key}")
            params[key] = f"{codigo}%"

        where_cuentas = " OR ".join(condiciones)

        sql = text(f"""
            SELECT COALESCE(SUM(debito - credito), 0) AS caja_actual
            FROM auxiliar_contable
            WHERE idcliente = :idc
              AND fecha_contable <= :hasta
              AND ({where_cuentas})
        """)
        row = db.session.execute(sql, params).mappings().first()
        valor = _safe_float(row["caja_actual"] if row else 0)

        return {
            "actual": _round2(valor),
            "requiere_parametrizacion": False,
            "visible": True,
            "modo_caja": modo_caja,
            "mensaje": "Caja disponible calculada con cuentas incluidas parametrizadas.",
            "cuentas_usadas": cuentas_incluidas
        }

    if modo_caja == "exclusion":
        codigos = [c["codigo"] for c in cuentas_excluidas if c.get("codigo")]
        params = {"idc": idcliente, "hasta": hasta}

        filtro_exclusion = ""
        if codigos:
            condiciones = []
            for idx, codigo in enumerate(codigos):
                key = f"e{idx}"
                condiciones.append(f"cuenta_codigo NOT LIKE :{key}")
                params[key] = f"{codigo}%"
            filtro_exclusion = " AND " + " AND ".join(condiciones)

        sql = text(f"""
            SELECT COALESCE(SUM(debito - credito), 0) AS caja_actual
            FROM auxiliar_contable
            WHERE idcliente = :idc
              AND fecha_contable <= :hasta
              AND cuenta_codigo LIKE '11%'
              {filtro_exclusion}
        """)
        row = db.session.execute(sql, params).mappings().first()
        valor = _safe_float(row["caja_actual"] if row else 0)

        return {
            "actual": _round2(valor),
            "requiere_parametrizacion": False,
            "visible": True,
            "modo_caja": modo_caja,
            "mensaje": "Caja disponible calculada con clase 11 excluyendo cuentas parametrizadas.",
            "cuentas_usadas": cuentas_excluidas
        }

    return {
        "actual": None,
        "requiere_parametrizacion": True,
        "visible": False,
        "modo_caja": modo_caja,
        "mensaje": "El modo de caja configurado no es válido.",
        "cuentas_usadas": []
    }


def _calcular_cash_runway_parametrizado(idcliente, fecha_hasta, config, caja_info):
    if not config:
        return {
            "actual": None,
            "burn_promedio": None,
            "unidad": "meses",
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_runway": "sin_configurar",
            "mensaje": "El cash runway requiere parametrización para este cliente."
        }

    mostrar_runway = bool(config.get("mostrar_runway", False))
    modo_runway = str(config.get("modo_runway") or "sin_configurar").strip()
    meses_promedio = int(config.get("meses_promedio_runway") or 3)

    if not mostrar_runway:
        return {
            "actual": None,
            "burn_promedio": None,
            "unidad": "meses",
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_runway": modo_runway,
            "mensaje": "La visualización de cash runway está desactivada para este cliente."
        }

    if modo_runway == "sin_configurar":
        return {
            "actual": None,
            "burn_promedio": None,
            "unidad": "meses",
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_runway": modo_runway,
            "mensaje": "El cash runway requiere parametrización de fórmula para este cliente."
        }

    if caja_info.get("actual") in (None, 0) or caja_info.get("requiere_parametrizacion"):
        return {
            "actual": None,
            "burn_promedio": None,
            "unidad": "meses",
            "requiere_parametrizacion": True,
            "visible": False,
            "modo_runway": modo_runway,
            "mensaje": "No es posible calcular cash runway sin una caja disponible válida."
        }

    if modo_runway == "egresos_promedio":
        desde_ref = _first_day_of_month(_shift_months(fecha_hasta, -(meses_promedio - 1)))

        sql = text("""
            SELECT
                TO_CHAR(DATE_TRUNC('month', fecha), 'YYYY-MM') AS mes,
                COALESCE(SUM(total), 0) AS egresos_mes
            FROM siigo_compras
            WHERE idcliente = :idc
              AND fecha >= :desde
              AND fecha <= :hasta
            GROUP BY TO_CHAR(DATE_TRUNC('month', fecha), 'YYYY-MM')
            ORDER BY mes
        """)

        rows = db.session.execute(sql, {
            "idc": idcliente,
            "desde": desde_ref,
            "hasta": fecha_hasta
        }).mappings().all()

        egresos = [_safe_float(r["egresos_mes"]) for r in rows if _safe_float(r["egresos_mes"]) > 0]
        burn_promedio = _round2(sum(egresos) / len(egresos)) if egresos else 0

        runway = _round2(_safe_float(caja_info["actual"]) / burn_promedio) if burn_promedio > 0 else None

        return {
            "actual": runway,
            "burn_promedio": burn_promedio if burn_promedio > 0 else None,
            "unidad": "meses",
            "requiere_parametrizacion": False,
            "visible": runway is not None,
            "modo_runway": modo_runway,
            "mensaje": "Cash runway calculado con egresos promedio de compras/gastos."
        }

    return {
        "actual": None,
        "burn_promedio": None,
        "unidad": "meses",
        "requiere_parametrizacion": True,
        "visible": False,
        "modo_runway": modo_runway,
        "mensaje": "El modo de runway configurado no está implementado todavía."
    }


def _calcular_top_gastos(composicion, limite=5):
    gastos = []

    for item in composicion or []:
        seccion = str(item.get("seccion", "")).upper()
        cuenta = str(item.get("cuenta", ""))
        total = _safe_float(item.get("total", 0))
        nombre = str(item.get("nombre", "")).strip()

        if seccion != "GASTOS_OPERACIONALES":
            continue

        if cuenta.startswith(("5160", "5165", "5260", "5265")):
            continue

        gastos.append({
            "cuenta": cuenta,
            "nombre": nombre,
            "valor": _round2(total),
        })

    gastos.sort(key=lambda x: x["valor"], reverse=True)
    return gastos[:limite]


def _construir_explicaciones_y_acciones(
    eficiencia_actual,
    eficiencia_anterior,
    promedio_6m,
    runway_meses,
    top_gastos,
    ventas_actual,
    ventas_anterior,
    ebitda_actual,
    ebitda_anterior
):
    explicaciones = []
    acciones = []
    alertas = []

    diff_ef = eficiencia_actual - eficiencia_anterior
    diff_ventas = ventas_actual - ventas_anterior
    diff_ebitda = ebitda_actual - ebitda_anterior

    if diff_ef > 0:
        explicaciones.append(
            f"La eficiencia operativa mejoró {abs(diff_ef):.2f} puntos frente al período anterior."
        )
    elif diff_ef < 0:
        explicaciones.append(
            f"La eficiencia operativa cayó {abs(diff_ef):.2f} puntos frente al período anterior."
        )
    else:
        explicaciones.append(
            "La eficiencia operativa se mantuvo estable frente al período anterior."
        )

    if diff_ventas > 0 and diff_ebitda > 0:
        explicaciones.append(
            "El EBITDA y las ventas crecieron simultáneamente, señal de mejor tracción operativa."
        )
    elif diff_ventas > 0 and diff_ebitda < 0:
        explicaciones.append(
            "Las ventas crecieron, pero el EBITDA cayó; probablemente hubo presión en costos o gastos."
        )
    elif diff_ventas < 0 and diff_ebitda < 0:
        explicaciones.append(
            "Se observa contracción conjunta en ventas y EBITDA, lo que amerita revisión inmediata."
        )

    if promedio_6m > 0:
        if eficiencia_actual > promedio_6m:
            explicaciones.append(
                f"El indicador actual está {abs(eficiencia_actual - promedio_6m):.2f} puntos por encima del promedio de 6 meses."
            )
        elif eficiencia_actual < promedio_6m:
            explicaciones.append(
                f"El indicador actual está {abs(eficiencia_actual - promedio_6m):.2f} puntos por debajo del promedio de 6 meses."
            )

    if runway_meses > 0 and runway_meses < 1.5:
        alertas.append({
            "nivel": "alta",
            "titulo": "Caja ajustada",
            "descripcion": f"La caja actual cubre aproximadamente {runway_meses:.2f} meses de operación."
        })
    elif runway_meses >= 1.5 and runway_meses < 3:
        alertas.append({
            "nivel": "media",
            "titulo": "Runway moderado",
            "descripcion": f"La caja actual cubre aproximadamente {runway_meses:.2f} meses de operación."
        })
    else:
        alertas.append({
            "nivel": "baja",
            "titulo": "Caja con holgura",
            "descripcion": f"La caja actual cubre aproximadamente {runway_meses:.2f} meses de operación."
        })

    if top_gastos:
        top1 = top_gastos[0]
        acciones.append(
            f"Revisar la categoría '{top1['nombre']}' ({top1['cuenta']}), actualmente el gasto operacional más alto del período."
        )

    if eficiencia_actual < promedio_6m and promedio_6m > 0:
        acciones.append(
            "Analizar qué gastos crecieron más rápido que las ventas en el período actual."
        )

    if diff_ventas > 0 and diff_ebitda < 0:
        acciones.append(
            "Validar si el crecimiento comercial está generando presión excesiva en costos o gasto operativo."
        )

    if not acciones:
        acciones.append(
            "Mantener seguimiento mensual del EBITDA, ventas y caja para sostener la tendencia positiva."
        )

    return explicaciones[:3], acciones[:3], alertas[:3]


# Helpers Pagina de Cnfiguraciones varias Dashboar y otras

def _serializar_dashboard_resumen_config(row):
    if not row:
        return {
            "activo": True,
            "mostrar_caja": False,
            "mostrar_runway": False,
            "modo_caja": "sin_configurar",
            "cuentas_incluidas": [],
            "cuentas_excluidas": [],
            "modo_runway": "sin_configurar",
            "meses_promedio_runway": 3,
            "meta_eficiencia_operativa": 20.0,
            "meta_ebitda": None,
            "meta_margen_ebitda": None,
            "meses_grafica": 6,
            "top_clientes": 5,
            "top_proveedores": 5,
            "top_gastos": 5,
            "indicador_estrella": "eficiencia_operativa",
            "modo_periodo_default": "ytd_cerrado",
        }

    return {
        "id": row["id"],
        "idcliente": row["idcliente"],
        "activo": bool(row["activo"]),
        "mostrar_caja": bool(row["mostrar_caja"]),
        "mostrar_runway": bool(row["mostrar_runway"]),
        "modo_caja": row["modo_caja"] or "sin_configurar",
        "cuentas_incluidas": row["cuentas_incluidas"] or [],
        "cuentas_excluidas": row["cuentas_excluidas"] or [],
        "modo_runway": row["modo_runway"] or "sin_configurar",
        "meses_promedio_runway": int(row["meses_promedio_runway"] or 3),
        "meta_eficiencia_operativa": float(row["meta_eficiencia_operativa"] or 20),
        "meta_ebitda": float(row["meta_ebitda"]) if row["meta_ebitda"] is not None else None,
        "meta_margen_ebitda": float(row["meta_margen_ebitda"]) if row["meta_margen_ebitda"] is not None else None,
        "meses_grafica": int(row["meses_grafica"] or 6),
        "top_clientes": int(row["top_clientes"] or 5),
        "top_proveedores": int(row["top_proveedores"] or 5),
        "top_gastos": int(row["top_gastos"] or 5),
        "indicador_estrella": row["indicador_estrella"] or "eficiencia_operativa",
        "modo_periodo_default": row["modo_periodo_default"] or "ytd_cerrado",
    }


def _limpiar_cuentas_config(lista_raw):
    """
    Espera lista tipo:
    [
      {"codigo": "111005", "nombre": "Banco Davivienda Ahorros Ppal"}
    ]
    """
    salida = []

    if not isinstance(lista_raw, list):
        return salida

    for item in lista_raw:
        if not isinstance(item, dict):
            continue

        codigo = str(item.get("codigo", "")).strip()
        nombre = str(item.get("nombre", "")).strip()

        if not codigo:
            continue

        salida.append({
            "codigo": codigo,
            "nombre": nombre
        })

    return salida


def _buscar_cuentas_auxiliar_para_config(idcliente, q=None, limite=20):
    """
    Busca cuentas contables existentes en auxiliar_contable para autocompletar
    parametrizaciones del dashboard.

    Retorna cuentas únicas por código con su nombre más representativo.
    """
    q = (q or "").strip()
    limite = max(1, min(int(limite or 20), 50))

    params = {
        "idc": idcliente,
        "limite": limite,
    }

    filtros = ["idcliente = :idc"]

    if q:
        filtros.append("""
            (
                cuenta_codigo ILIKE :q
                OR cuenta_nombre ILIKE :q
            )
        """)
        params["q"] = f"%{q}%"

    where_sql = " AND ".join(filtros)

    sql = text(f"""
        SELECT
            cuenta_codigo AS codigo,
            MAX(cuenta_nombre) AS nombre,
            COUNT(*) AS apariciones,
            MAX(fecha_contable) AS ultima_fecha
        FROM auxiliar_contable
        WHERE {where_sql}
          AND cuenta_codigo IS NOT NULL
          AND TRIM(cuenta_codigo) <> ''
        GROUP BY cuenta_codigo
        ORDER BY
            CASE
                WHEN cuenta_codigo = :q_exact THEN 0
                WHEN cuenta_codigo ILIKE :q_prefix THEN 1
                ELSE 2
            END,
            cuenta_codigo ASC
        LIMIT :limite
    """)

    params["q_exact"] = q if q else ""
    params["q_prefix"] = f"{q}%" if q else ""

    rows = db.session.execute(sql, params).mappings().all()

    return [
        {
            "codigo": str(r["codigo"]).strip(),
            "nombre": str(r["nombre"] or "").strip(),
            "apariciones": int(r["apariciones"] or 0),
            "ultima_fecha": r["ultima_fecha"].strftime("%Y-%m-%d") if r["ultima_fecha"] else None,
        }
        for r in rows
    ]


#Helpers de control de Usuarios:
def _bool_from_payload(value, default=False):
    """
    Convierte valores enviados desde frontend a boolean real.
    Evita que strings como 'false' terminen evaluando como True.
    """
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes", "si", "sí")

    return bool(value)


def _validar_cupo_usuarios_cliente(idcliente, excluir_idusuario=None):
    """
    Valida si un cliente todavía tiene cupo para tener usuarios activos.

    - Cuenta solo usuarios activos.
    - Si excluir_idusuario viene, no cuenta ese usuario.
      Esto sirve para edición/reactivación/cambio de cliente.
    - Si limite_usuarios es NULL o <= 0, se interpreta como sin límite configurado
      para no romper clientes antiguos.
    """

    cliente = Cliente.query.get(idcliente)

    if not cliente:
        return False, {
            "error": "Cliente no encontrado",
            "detalle": f"No existe un cliente con idcliente={idcliente}."
        }, 404

    if not cliente.activo:
        return False, {
            "error": "Cliente inactivo",
            "detalle": "No se pueden crear o activar usuarios para un cliente inactivo.",
            "idcliente": cliente.idcliente,
            "cliente": cliente.nombre
        }, 400

    limite = cliente.limite_usuarios

    # Compatibilidad con clientes antiguos sin límite configurado.
    # Si quieres hacerlo estricto, cambia esta condición para bloquear cuando sea None.
    if limite is None or limite <= 0:
        return True, {
            "idcliente": cliente.idcliente,
            "cliente": cliente.nombre,
            "limite_usuarios": limite,
            "usuarios_activos": None,
            "cupos_disponibles": None,
            "sin_limite_configurado": True
        }, 200

    query = Usuario.query.filter(
        Usuario.idcliente == idcliente,
        Usuario.activo.is_(True)
    )

    if excluir_idusuario:
        query = query.filter(Usuario.idusuario != excluir_idusuario)

    usuarios_activos = query.count()
    cupos_disponibles = max(limite - usuarios_activos, 0)

    if usuarios_activos >= limite:
        return False, {
            "error": "Límite de usuarios alcanzado",
            "detalle": (
                f"El cliente tiene contratado un máximo de {limite} usuarios activos. "
                f"Actualmente ya tiene {usuarios_activos} usuarios activos."
            ),
            "idcliente": cliente.idcliente,
            "cliente": cliente.nombre,
            "limite_usuarios": limite,
            "usuarios_activos": usuarios_activos,
            "cupos_disponibles": 0
        }, 409

    return True, {
        "idcliente": cliente.idcliente,
        "cliente": cliente.nombre,
        "limite_usuarios": limite,
        "usuarios_activos": usuarios_activos,
        "cupos_disponibles": cupos_disponibles
    }, 200


def _validar_perfil_cliente(idperfil, idcliente):
    """
    Valida que el perfil exista y pertenezca al cliente indicado.
    """
    perfil = Perfil.query.filter_by(
        idperfil=idperfil,
        idcliente=idcliente
    ).first()

    if not perfil:
        return None, {
            "error": "Perfil no válido",
            "detalle": "El perfil enviado no existe o no pertenece al cliente seleccionado.",
            "idperfil": idperfil,
            "idcliente": idcliente
        }, 400

    return perfil, None, None


def _parse_date_yyyy_mm_dd(value):
    """
    Convierte un string YYYY-MM-DD a date.
    Si viene vacío o inválido, retorna None.
    Usa el datetime ya importado como clase.
    """
    if not value:
        return None

    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _parse_time_hh_mm(value, default_hour=2, default_minute=0):
    """
    Convierte HH:MM a time.
    Si viene inválido, retorna 02:00 por defecto.
    Usa datetime.strptime(...).time() para evitar problemas con datetime.time().
    """
    try:
        return datetime.strptime(str(value or "02:00")[:5], "%H:%M").time()
    except Exception:
        return datetime.strptime(f"{default_hour:02d}:{default_minute:02d}", "%H:%M").time()


def _extraer_espera_rate_limit(texto, default=6):
    """
    Extrae segundos desde mensajes de Siigo como:
    'Rate limit is exceeded. Try again in 2 seconds.'
    """
    try:
        match = re.search(r"Try again in\s+(\d+)\s+seconds", str(texto), re.IGNORECASE)
        if match:
            return max(int(match.group(1)) + 1, default)
    except Exception:
        pass

    return default




# Helpers de detalles de resultado sincronizaciones API siigo:
def _resumir_detalle_sync(detalle: str):
    """
    Lee el detalle técnico de sync-all y calcula:
    - total de pasos
    - pasos OK
    - pasos con error
    - endpoint fallido
    """
    if not detalle:
        return {
            "total_pasos": 0,
            "pasos_ok": 0,
            "pasos_error": 0,
            "endpoint_fallido": None,
        }

    lines = [l for l in detalle.split("\n") if l.strip()]
    total_pasos = 0
    pasos_ok = 0
    pasos_error = 0
    endpoint_fallido = None

    import re

    for line in lines:
        match = re.search(r"->\s+(\d{3})", line)
        if not match:
            continue

        total_pasos += 1
        status = int(match.group(1))

        if 200 <= status < 400:
            pasos_ok += 1
        else:
            pasos_error += 1
            if not endpoint_fallido:
                endpoint_fallido = line.split(" ")[0]

    return {
        "total_pasos": total_pasos,
        "pasos_ok": pasos_ok,
        "pasos_error": pasos_error,
        "endpoint_fallido": endpoint_fallido,
    }

# Helper para la Fecha de sincronizacion desde
def _resolver_sync_fecha_desde(data, config):
    """
    Resuelve la fecha global de sincronización de documentos Siigo.

    Prioridad:
    1. sync_fecha_desde enviado desde frontend.
    2. ds_fecha_desde enviado desde frontend, por compatibilidad.
    3. config.sync_fecha_desde.
    4. config.ds_fecha_desde, por compatibilidad.
    5. None: sin límite de fecha.
    """
    fecha = data.get("sync_fecha_desde")

    if not fecha:
        fecha = data.get("ds_fecha_desde")

    if not fecha and config and getattr(config, "sync_fecha_desde", None):
        fecha = config.sync_fecha_desde.isoformat()

    if not fecha and config and getattr(config, "ds_fecha_desde", None):
        fecha = config.ds_fecha_desde.isoformat()

    return fecha or None



#Helpers para ostrar notificaicones de proceso acabado en hitoria luego de usara boton manual de sincronizacion:
def _now_utc_aware():
    """
    Retorna fecha/hora UTC aware para guardar en columnas timezone=True.
    """
    import pytz
    return datetime.utcnow().replace(tzinfo=pytz.utc)


def _crear_log_sync_modulo_inicio(
    idcliente,
    endpoint,
    origen="manual_modulo",
    params=None,
    mensaje="Proceso iniciado.",
):
    """
    Crea un registro en siigo_sync_logs con estado EN_EJECUCION.
    Sirve para que el usuario vea en historial que el proceso manual arrancó.
    """
    try:
        now = _now_utc_aware()
        params = params or {}

        detalle = f"{endpoint} {params} -> EN_EJECUCION: {mensaje}"

        logrec = SiigoSyncLog(
            idcliente=idcliente,
            fecha_programada=now,
            ejecutado_en=now,
            resultado="EN_EJECUCION",
            detalle=detalle[:10000],
            origen=origen,
            total_pasos=1,
            pasos_ok=0,
            pasos_error=0,
            endpoint_fallido=None,
        )

        db.session.add(logrec)
        db.session.commit()

        return logrec.id

    except Exception as e:
        print(f"⚠️ Error creando log inicio módulo {endpoint}: {e}")
        db.session.rollback()
        return None


def _finalizar_log_sync_modulo(
    log_id,
    idcliente,
    endpoint,
    resultado,
    detalle,
    status_code=200,
    duracion_segundos=None,
):
    """
    Actualiza un log previamente creado como EN_EJECUCION.
    También registra métrica individual.
    """
    try:
        now = _now_utc_aware()

        resultado_final = "OK" if resultado == "OK" and status_code < 400 else "ERROR"
        pasos_ok = 1 if resultado_final == "OK" else 0
        pasos_error = 0 if resultado_final == "OK" else 1
        endpoint_fallido = None if resultado_final == "OK" else endpoint

        detalle_final = f"{endpoint} {{}} -> {status_code}: {detalle}"

        if log_id:
            logrec = SiigoSyncLog.query.get(log_id)
            if logrec:
                logrec.ejecutado_en = now
                logrec.resultado = resultado_final
                logrec.detalle = detalle_final[:10000]
                logrec.total_pasos = 1
                logrec.pasos_ok = pasos_ok
                logrec.pasos_error = pasos_error
                logrec.endpoint_fallido = endpoint_fallido
                db.session.add(logrec)

        metric = SiigoSyncMetric(
            idcliente=idcliente,
            endpoint=endpoint,
            duracion_segundos=duracion_segundos,
            status_code=status_code,
            resultado=resultado_final,
            detalle_resumen=detalle_final[:300],
        )
        db.session.add(metric)

        db.session.commit()

    except Exception as e:
        print(f"⚠️ Error finalizando log módulo {endpoint}: {e}")
        db.session.rollback()


# Helpers para cambio de pauete a clienets en la pagina del clienets del superadmin
def _paquete_base_actual_cliente(idcliente):
    """
    Retorna el paquete base activo actual del cliente.
    No considera módulos adicionales.
    """
    row = (
        db.session.query(ClientePaquete, PaqueteInsightflow)
        .join(
            PaqueteInsightflow,
            PaqueteInsightflow.idpaquete == ClientePaquete.idpaquete
        )
        .filter(
            ClientePaquete.idcliente == idcliente,
            ClientePaquete.activo.is_(True),
            PaqueteInsightflow.activo.is_(True),
            PaqueteInsightflow.es_modulo_adicional.is_(False),
        )
        .order_by(ClientePaquete.created_at.desc())
        .first()
    )

    if not row:
        return None, None

    cliente_paquete, paquete = row
    return cliente_paquete, paquete


def _cliente_as_dict_con_paquete(cliente):
    """
    Extiende Cliente.as_dict() agregando el paquete base activo actual.
    """
    data = cliente.as_dict()

    cliente_paquete, paquete = _paquete_base_actual_cliente(cliente.idcliente)

    data["paquete_actual"] = None
    data["paquete_codigo"] = None
    data["paquete_nombre"] = None
    data["idpaquete"] = None

    if paquete:
        data["paquete_actual"] = {
            "idcliente_paquete": cliente_paquete.id if cliente_paquete else None,
            "idpaquete": paquete.idpaquete,
            "codigo": paquete.codigo,
            "nombre": paquete.nombre,
            "descripcion": paquete.descripcion,
            "activo": paquete.activo,
        }
        data["paquete_codigo"] = paquete.codigo
        data["paquete_nombre"] = paquete.nombre
        data["idpaquete"] = paquete.idpaquete

    return data


def _obtener_codigos_paquete(idpaquete):
    """
    Obtiene los códigos de permisos activos configurados para un paquete.
    La fuente de verdad es paquete_permisos.
    """
    rows = (
        PaquetePermiso.query
        .filter(
            PaquetePermiso.idpaquete == idpaquete,
            PaquetePermiso.activo.is_(True),
        )
        .all()
    )

    return sorted({
        str(r.codigo_permiso).strip()
        for r in rows
        if r.codigo_permiso and str(r.codigo_permiso).strip()
    })


def _humanizar_codigo_permiso(codigo):
    """
    Convierte un código técnico en nombre legible.
    Ejemplo:
    ver_reporte_balance_general -> Ver Reporte Balance General
    """
    if not codigo:
        return ""

    partes = str(codigo).replace("-", "_").split("_")
    return " ".join(p.capitalize() for p in partes if p)


def _sincronizar_permisos_paquete_cliente(idcliente, paquete):
    """
    Sincroniza permisos del paquete hacia el cliente.

    Política:
    - Toma los códigos directamente desde paquete_permisos.
    - Crea en permisos los códigos que falten para el cliente.
    - Reactiva permisos existentes si estaban inactivos.
    - Asigna los permisos al perfil Administrador.
    - No depende de ningún cliente plantilla.
    - No elimina permisos antiguos.
    - No revoca permisos de otros perfiles.
    """
    from sqlalchemy import func

    codigos_permitidos = _obtener_codigos_paquete(paquete.idpaquete)

    if not codigos_permitidos:
        return {
            "ok": False,
            "error": "El paquete seleccionado no tiene permisos activos configurados.",
            "codigos_permitidos": [],
        }

    permisos_creados = 0
    permisos_existentes = 0
    permisos_reactivados = 0
    permisos_actualizados = 0

    permisos_cliente = []

    for codigo in codigos_permitidos:
        codigo = str(codigo).strip()

        if not codigo:
            continue

        permiso_cliente = (
            Permiso.query
            .filter_by(idcliente=idcliente, codigo=codigo)
            .first()
        )

        nombre_generado = _humanizar_codigo_permiso(codigo)
        descripcion_generada = f"Permiso generado automáticamente para {codigo}"

        if not permiso_cliente:
            permiso_cliente = Permiso(
                idcliente=idcliente,
                nombre=nombre_generado,
                codigo=codigo,
                descripcion=descripcion_generada,
                activo=True,
            )
            db.session.add(permiso_cliente)
            db.session.flush()
            permisos_creados += 1

        else:
            permisos_existentes += 1
            cambio = False

            if not permiso_cliente.activo:
                permiso_cliente.activo = True
                permisos_reactivados += 1
                cambio = True

            if not permiso_cliente.nombre:
                permiso_cliente.nombre = nombre_generado
                cambio = True

            if not permiso_cliente.descripcion:
                permiso_cliente.descripcion = descripcion_generada
                cambio = True

            if cambio:
                db.session.add(permiso_cliente)
                permisos_actualizados += 1

        permisos_cliente.append(permiso_cliente)

    perfil_admin = (
        Perfil.query
        .filter(
            Perfil.idcliente == idcliente,
            func.lower(Perfil.nombre) == "administrador",
        )
        .first()
    )

    if not perfil_admin:
        perfil_admin = Perfil(
            idcliente=idcliente,
            nombre="Administrador",
            descripcion="Perfil administrador del cliente",
        )
        db.session.add(perfil_admin)
        db.session.flush()

    asignaciones_creadas = 0
    asignaciones_existentes = 0
    asignaciones_reactivadas = 0

    for permiso in permisos_cliente:
        rel = (
            PerfilPermiso.query
            .filter_by(
                idcliente=idcliente,
                idperfil=perfil_admin.idperfil,
                idpermiso=permiso.idpermiso,
            )
            .first()
        )

        if not rel:
            rel = PerfilPermiso(
                idcliente=idcliente,
                idperfil=perfil_admin.idperfil,
                idpermiso=permiso.idpermiso,
                permitido=True,
            )
            db.session.add(rel)
            asignaciones_creadas += 1

        else:
            asignaciones_existentes += 1

            if not rel.permitido:
                rel.permitido = True
                db.session.add(rel)
                asignaciones_reactivadas += 1

    return {
        "ok": True,
        "codigos_permitidos": codigos_permitidos,
        "perfil_admin": {
            "idperfil": perfil_admin.idperfil,
            "nombre": perfil_admin.nombre,
        },
        "permisos_cliente": {
            "creados": permisos_creados,
            "existentes": permisos_existentes,
            "reactivados": permisos_reactivados,
            "actualizados": permisos_actualizados,
            "total_sincronizados": len(permisos_cliente),
        },
        "asignaciones_admin": {
            "creadas": asignaciones_creadas,
            "existentes": asignaciones_existentes,
            "reactivadas": asignaciones_reactivadas,
        },
    }



#---------------------------------------------------------------------------------------------------------
# ENDPOINTS DEL SISTEMA
#---------------------------------------------------------------------------------------------------------


def create_app():
    app = Flask(__name__, static_folder="static", static_url_path="")

    app.config.from_object(Config)

    # app.py (o donde configuras Flask/JWT)
    app.config["JWT_SECRET_KEY"] = os.getenv("JWT_SECRET_KEY", "pon-una-clave-larga-y-estable")
    app.config["JWT_ACCESS_TOKEN_EXPIRES"] = 60 * 60 * 24  # 24h opcional


    # ✅ CORS en la instancia CORRECTA
    # CORS aplicado globalmente con soporte completo
    CORS(
        app,
        resources={r"/*": {
            "origins": "https://insigthsflow.up.railway.app",
            "allow_headers": ["Content-Type", "Authorization", "X-ID-CLIENTE"],
            "supports_credentials": True
        }}
    )

    print("🔍 Usando esta URI de base de datos:", app.config["SQLALCHEMY_DATABASE_URI"])
    db.init_app(app)
    jwt = JWTManager(app)  # ← guarda la instancia

    # ---- revocación en memoria (en prod usa BD o Redis) ----
    BLOCKLIST = set()

    @jwt.token_in_blocklist_loader
    def is_token_revoked(jwt_header, jwt_payload):
        return jwt_payload.get("jti") in BLOCKLIST

    @jwt.invalid_token_loader
    def invalid_token(reason):
        return jsonify({"error": f"Invalid token: {reason}"}), 422

    @jwt.unauthorized_loader
    def missing_token(reason):
        return jsonify({"error": f"Missing token: {reason}"}), 401
    
    # (opcional) tokens expirados:
    @jwt.expired_token_loader
    def expired_token(jwt_header, jwt_payload):
        return jsonify({"error": "Token expirado"}), 401

    # Ruta de prueba
    @app.route("/")
    def index():
        return {"message": "Backend Siigo Insights funcionando ✅"}



    # Login con JWT
    @app.route("/auth/login", methods=["POST"])
    def login():
        data = request.get_json()

        if not data or "email" not in data or "password" not in data:
            return jsonify({"error": "Email y password requeridos"}), 400

        email = data["email"]
        password = data["password"]

        # Buscar usuario en la BD
        user = Usuario.query.filter_by(email=email).first()
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        # Verificar password
        if not check_password_hash(user.password_hash, password):
            return jsonify({"error": "Password incorrecto"}), 401

        # ========================
        # Detectar si es SuperAdmin
        # ========================
        if user.email == "superadmin@sistema.com":
            claims = {
                "idusuario": user.idusuario,
                "email": user.email,
                "perfilid": 0,       # 👑 SuperAdmin
                "idcliente": None    # 👑 No pertenece a un cliente
            }
        else:
            claims = {
                "idusuario": user.idusuario,
                "email": user.email,
                "perfilid": user.idperfil,
                "idcliente": user.idcliente
            }

        # Generar token con identity simple y claims adicionales
        is_superadmin = (user.email == "superadmin@sistema.com")

        extra_claims = {
            "idusuario": user.idusuario,
            "email": user.email,
            "perfilid": 0 if is_superadmin else user.idperfil,
            "idcliente": None if is_superadmin else user.idcliente,
        }

        # identity como string (recomendado por flask-jwt-extended)
        # (opcional) ajusta la expiración
        expires = timedelta(hours=4)
        access_token = create_access_token(
            identity=str(user.idusuario),
            additional_claims=extra_claims,
            expires_delta=expires
        )

        # Limpia sesiones expiradas
        cleanup_expired_sessions()

        # Obtén jti y exp del token
        decoded = decode_token(access_token)
        jti = decoded["jti"]
        exp_ts = decoded.get("exp")
        expira_en = datetime.fromtimestamp(exp_ts, tz=timezone.utc) if exp_ts else None

        # Enforzar límite de sesiones por cliente (superadmin no cuenta)
        idcliente = None if is_superadmin else user.idcliente
        if idcliente:
            cliente = Cliente.query.get(idcliente)
            limite = cliente.limite_sesiones or None
            if limite:
                activas = SesionActiva.query.filter_by(idcliente=idcliente).count()
                if activas >= limite:
                    return jsonify({"error": "Se alcanzó el límite de sesiones activas para este cliente."}), 429

        # Registrar sesión activa
        db.session.add(SesionActiva(
            jti=jti,
            idusuario=user.idusuario,
            idcliente=idcliente,
            expira_en=expira_en
        ))
        db.session.commit()

        return jsonify({"access_token": access_token, "usuario": user.email})



    @app.route("/auth/logout", methods=["POST"])
    @jwt_required()
    def logout():
        payload = get_jwt()
        jti = payload["jti"]

        # revoca en blocklist
        BLOCKLIST.add(jti)

        # elimina de sesiones activas (si existe)
        ses = SesionActiva.query.filter_by(jti=jti).first()
        if ses:
            db.session.delete(ses)
            db.session.commit()

        return jsonify({"message": "Sesión cerrada"}), 200



    # ==========================
    # CRUD Clientes
    # ==========================

    @app.route("/clientes", methods=["GET"])
    @jwt_required()
    def get_clientes():
        claims = get_jwt()

        if claims["perfilid"] == 0:
            clientes = (
                Cliente.query
                .order_by(Cliente.idcliente.asc())
                .all()
            )
        else:
            clientes = (
                Cliente.query
                .filter_by(idcliente=claims["idcliente"])
                .order_by(Cliente.idcliente.asc())
                .all()
            )

        return jsonify([_cliente_as_dict_con_paquete(c) for c in clientes]), 200


    @app.route("/clientes", methods=["POST"])
    @jwt_required()
    def create_cliente():
        claims = get_jwt()
        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}
        cliente = Cliente(
            nombre=data["nombre"],
            nit=data.get("nit"),
            email=data.get("email"),
            activo=bool(data.get("activo", True)),
            pais=data.get("pais"),
            ciudad=data.get("ciudad"),
            direccion=data.get("direccion"),
            telefono1=data.get("telefono1"),
            logo_url=data.get("logo_url"),
            limite_usuarios=(int(data["limite_usuarios"]) if data.get("limite_usuarios") not in (None, "",) else None),
            limite_sesiones=(int(data["limite_sesiones"]) if data.get("limite_sesiones") not in (None, "",) else None),
            timezone=data.get("timezone", "America/Bogota")  # 👈 nuevo campo con valor por defecto
        )
        db.session.add(cliente)
        db.session.commit()
        return jsonify(cliente.as_dict()), 201

    @app.route("/clientes/<int:idcliente>", methods=["PUT"])
    @jwt_required()
    def update_cliente(idcliente):
        claims = get_jwt()
        cliente = Cliente.query.get_or_404(idcliente)

        if claims["perfilid"] != 0 and cliente.idcliente != claims["idcliente"]:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}
        for field in [
            "nombre","nit","email","pais","ciudad","direccion","telefono1","logo_url", "timezone"
        ]:
            if field in data:
                setattr(cliente, field, data[field])

        if "activo" in data:
            cliente.activo = bool(data["activo"])

        if "limite_usuarios" in data:
            cliente.limite_usuarios = int(data["limite_usuarios"]) if data["limite_usuarios"] not in ("", None) else None
        if "limite_sesiones" in data:
            cliente.limite_sesiones = int(data["limite_sesiones"]) if data["limite_sesiones"] not in ("", None) else None

        db.session.commit()
        return jsonify(cliente.as_dict())


    @app.route("/clientes/<int:idcliente>/paquete", methods=["PUT", "OPTIONS"])
    @jwt_required()
    @cross_origin()
    def cambiar_paquete_cliente(idcliente):
        """
        Cambia el paquete base contratado de un cliente.

        Uso:
        PUT /clientes/9/paquete
        {
            "paquete": "completo"
        }

        Solo SuperAdmin.
        """
        from datetime import date
        from sqlalchemy import func

        claims = get_jwt()

        if claims.get("perfilid") != 0:
            return jsonify({"error": "No autorizado"}), 403

        cliente = Cliente.query.get(idcliente)

        if not cliente:
            return jsonify({
                "error": "Cliente no encontrado",
                "idcliente": idcliente,
            }), 404

        data = request.get_json() or {}

        paquete_codigo = (
            data.get("paquete")
            or data.get("codigo")
            or data.get("paquete_codigo")
        )

        idpaquete = data.get("idpaquete")

        if not paquete_codigo and not idpaquete:
            return jsonify({
                "error": "Debes enviar 'paquete' o 'idpaquete'.",
                "ejemplo": {"paquete": "completo"},
            }), 400

        q = PaqueteInsightflow.query.filter(
            PaqueteInsightflow.activo.is_(True),
            PaqueteInsightflow.es_modulo_adicional.is_(False),
        )

        if idpaquete:
            paquete = q.filter(PaqueteInsightflow.idpaquete == int(idpaquete)).first()
        else:
            paquete_codigo = str(paquete_codigo).strip().lower()
            paquete = q.filter(func.lower(PaqueteInsightflow.codigo) == paquete_codigo).first()

        if not paquete:
            return jsonify({
                "error": "Paquete no encontrado o inactivo.",
                "paquete": paquete_codigo,
                "idpaquete": idpaquete,
            }), 404

        try:
            paquete_anterior_rel, paquete_anterior = _paquete_base_actual_cliente(idcliente)

            paquetes_base_activos = (
                db.session.query(ClientePaquete)
                .join(PaqueteInsightflow, PaqueteInsightflow.idpaquete == ClientePaquete.idpaquete)
                .filter(
                    ClientePaquete.idcliente == idcliente,
                    ClientePaquete.activo.is_(True),
                    PaqueteInsightflow.es_modulo_adicional.is_(False),
                )
                .all()
            )

            for rel in paquetes_base_activos:
                if rel.idpaquete != paquete.idpaquete:
                    rel.activo = False
                    rel.fecha_fin = date.today()
                    db.session.add(rel)

            rel_nuevo = (
                ClientePaquete.query
                .filter_by(
                    idcliente=idcliente,
                    idpaquete=paquete.idpaquete,
                )
                .first()
            )

            if rel_nuevo:
                rel_nuevo.activo = True
                rel_nuevo.fecha_inicio = rel_nuevo.fecha_inicio or date.today()
                rel_nuevo.fecha_fin = None
            else:
                rel_nuevo = ClientePaquete(
                    idcliente=idcliente,
                    idpaquete=paquete.idpaquete,
                    activo=True,
                    fecha_inicio=date.today(),
                    fecha_fin=None,
                )

            db.session.add(rel_nuevo)
            db.session.flush()

            sync_result = _sincronizar_permisos_paquete_cliente(
                idcliente=idcliente,
                paquete=paquete,
            )

            if not sync_result.get("ok"):
                db.session.rollback()
                return jsonify(sync_result), 400

            db.session.commit()

            return jsonify({
                "message": "Paquete actualizado correctamente.",
                "cliente": _cliente_as_dict_con_paquete(cliente),
                "paquete_anterior": {
                    "idpaquete": paquete_anterior.idpaquete,
                    "codigo": paquete_anterior.codigo,
                    "nombre": paquete_anterior.nombre,
                } if paquete_anterior else None,
                "paquete_nuevo": {
                    "idpaquete": paquete.idpaquete,
                    "codigo": paquete.codigo,
                    "nombre": paquete.nombre,
                },
                "permisos": sync_result,
                "politica": (
                    "Cambio seguro: se agregaron/reactivaron permisos del nuevo paquete "
                    "y se asignaron al perfil Administrador. No se eliminaron permisos antiguos."
                ),
            }), 200

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo cambiar el paquete del cliente.",
                "detalle": str(e),
                "idcliente": idcliente,
                "paquete": paquete_codigo,
            }), 500



    @app.route("/clientes/<int:idcliente>", methods=["DELETE"])
    @jwt_required()
    def delete_cliente(idcliente):
        claims = get_jwt()
        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        cliente = Cliente.query.get_or_404(idcliente)
        db.session.delete(cliente)
        db.session.commit()
        return jsonify({"message": "Cliente eliminado"})




    @app.route("/clientes/<int:idcliente>/full_delete", methods=["DELETE"])
    @jwt_required()
    def delete_cliente_total(idcliente):
        """
        Elimina por completo un cliente y toda su data relacionada.

        Requiere doble confirmación:
        - 1ra: modal en frontend.
        - 2da: parámetro ?confirm=true en la URL.

        Solo SuperAdmin puede ejecutar esta acción.

        Importante:
        - Borra primero tablas hijas.
        - Luego tablas padre.
        - Al final elimina el registro del cliente.
        - No borra vistas directamente, como facturas_enriquecidas.
        """

        claims = get_jwt()

        if claims.get("perfilid") != 0:
            return jsonify({"error": "No autorizado"}), 403

        confirm = request.args.get("confirm", "false").lower()

        if confirm != "true":
            return jsonify({
                "warning": (
                    "⚠️ Falta confirmación final. "
                    "Para eliminar definitivamente, agrega '?confirm=true' al endpoint."
                ),
                "example": f"/clientes/{idcliente}/full_delete?confirm=true"
            }), 400

        cliente = Cliente.query.get(idcliente)

        if not cliente:
            return jsonify({
                "error": f"No existe un cliente con idcliente={idcliente}"
            }), 404

        nombre_cliente = cliente.nombre

        resumen = {}

        try:
            """
            Orden seguro basado en las FK reales confirmadas en Railway:

            - perfil_permisos depende de permisos y perfiles.
            - usuarios depende de perfiles.
            - siigo_factura_items depende de siigo_facturas.
            - siigo_cuentasporcobrar depende de siigo_compras.
            - clientes se elimina al final.
            - facturas_enriquecidas NO se elimina porque es una vista.
            """

            tablas_ordenadas = [
                # Seguridad / usuarios / permisos
                "sesiones_activas",
                "perfil_permisos",
                "usuarios",
                "permisos",
                "perfiles",

                # Licenciamiento / paquetes contratados
                "cliente_paquetes",

                # Facturación / cartera / ingresos
                "siigo_factura_items",
                "siigo_pagos_recibidos",
                "siigo_notas_credito",
                "siigo_facturas",

                # Compras / cuentas por pagar
                "siigo_cuentasporcobrar",
                "siigo_compras_items",
                "siigo_pagos_proveedores",
                "siigo_documentos_soporte_api_staging",
                "siigo_compras",

                # Catálogos Siigo
                "siigo_centros_costo",
                "siigo_customers",
                "siigo_proveedores",
                "siigo_productos",
                "siigo_vendedores",
                "siigo_nomina",

                # Configuración Siigo / sincronización
                "siigo_sync_logs",
                "siigo_sync_metrics",
                "siigo_sync_config",
                "siigo_credenciales",

                # Información financiera / contable
                "balance_prueba",
                "auxiliar_saldos_corte",
                "auxiliar_contable",

                # Configuración del dashboard / módulos / notificaciones
                "dashboard_resumen_config",
                "modulos_disponibles",
                "system_notifications",
            ]

            for tabla in tablas_ordenadas:
                result = db.session.execute(
                    text(f'DELETE FROM public."{tabla}" WHERE idcliente = :idcliente'),
                    {"idcliente": idcliente}
                )

                resumen[tabla] = result.rowcount if result.rowcount is not None else 0

            db.session.delete(cliente)
            db.session.commit()

            return jsonify({
                "message": f"✅ Cliente '{nombre_cliente}' y toda su información fueron eliminados correctamente.",
                "idcliente": idcliente,
                "detalles": resumen
            }), 200

        except Exception as e:
            db.session.rollback()

            return jsonify({
                "error": "No se pudo eliminar completamente el cliente.",
                "detalle": str(e),
                "idcliente": idcliente,
                "recomendacion": (
                    "La transacción fue reversada. Revisa si existe alguna nueva tabla relacionada "
                    "con idcliente o alguna llave foránea adicional no contemplada."
                )
            }), 500




    # ==========================
    # SuperAdmin CRUD Perfiles
    # ==========================
    @app.route("/admin/perfiles", methods=["GET"])
    @jwt_required()
    def admin_get_perfiles():
        claims = get_jwt()
        if claims["perfilid"] == 0:  # SuperAdmin → puede ver todos
            perfiles = Perfil.query.all()
        else:
            perfiles = Perfil.query.filter_by(idcliente=claims["idcliente"]).all()
        return jsonify([p.as_dict() for p in perfiles])

    @app.route("/admin/perfiles", methods=["POST"])
    @jwt_required()
    def admin_crear_perfil():
        claims = get_jwt()
        if claims["perfilid"] != 0:  
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json()
        perfil = Perfil(
            idcliente=data["idcliente"],   # 👑 SuperAdmin puede decidir a qué cliente asignarlo
            nombre=data["nombre"],
            descripcion=data.get("descripcion", "")
        )
        db.session.add(perfil)
        db.session.commit()
        return jsonify(perfil.as_dict()), 201

    @app.route("/admin/perfiles/<int:idperfil>", methods=["PUT"])
    @jwt_required()
    def admin_update_perfil(idperfil):
        claims = get_jwt()
        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        perfil = Perfil.query.get_or_404(idperfil)
        data = request.get_json()
        perfil.nombre = data.get("nombre", perfil.nombre)
        perfil.descripcion = data.get("descripcion", perfil.descripcion)
        db.session.commit()
        return jsonify(perfil.as_dict())

    @app.route("/admin/perfiles/<int:idperfil>", methods=["DELETE"])
    @jwt_required()
    def admin_delete_perfil(idperfil):
        claims = get_jwt()
        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        perfil = Perfil.query.get_or_404(idperfil)
        db.session.delete(perfil)
        db.session.commit()
        return jsonify({"message": "Perfil eliminado"})



    # ==========================
    # Superadmin CRUD Usuarios
    # ==========================

    @app.route("/admin/usuarios", methods=["GET"])
    @jwt_required()
    def admin_get_usuarios():
        claims = get_jwt()

        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        usuarios = Usuario.query.all()
        return jsonify([u.as_dict() for u in usuarios])


    @app.route("/admin/usuarios", methods=["POST"])
    @jwt_required()
    def admin_crear_usuario():
        claims = get_jwt()

        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        required = ["idcliente", "idperfil", "nombre", "email", "password"]
        faltantes = [field for field in required if not data.get(field)]

        if faltantes:
            return jsonify({
                "error": "Faltan campos obligatorios",
                "faltantes": faltantes
            }), 400

        try:
            idcliente = int(data["idcliente"])
            idperfil = int(data["idperfil"])
        except Exception:
            return jsonify({
                "error": "Datos inválidos",
                "detalle": "idcliente e idperfil deben ser valores numéricos."
            }), 400

        # Validar que el perfil pertenezca al cliente seleccionado
        perfil, error_perfil, status_perfil = _validar_perfil_cliente(idperfil, idcliente)
        if error_perfil:
            return jsonify(error_perfil), status_perfil

        # Validar límite de usuarios activos
        ok_cupo, info_cupo, status_cupo = _validar_cupo_usuarios_cliente(idcliente)
        if not ok_cupo:
            return jsonify(info_cupo), status_cupo

        try:
            user = Usuario(
                idcliente=idcliente,
                idperfil=perfil.idperfil,
                nombre=str(data["nombre"]).strip(),
                apellido=(str(data.get("apellido")).strip() if data.get("apellido") else None),
                email=str(data["email"]).strip().lower(),
                password_hash=generate_password_hash(
                    data["password"],
                    method="pbkdf2:sha256",
                    salt_length=16
                ),
                activo=True
            )

            db.session.add(user)
            db.session.commit()

            response = user.as_dict()
            response["control_usuarios"] = {
                **info_cupo,
                "usuarios_activos_despues": (
                    None
                    if info_cupo.get("usuarios_activos") is None
                    else info_cupo["usuarios_activos"] + 1
                ),
                "cupos_disponibles_despues": (
                    None
                    if info_cupo.get("cupos_disponibles") is None
                    else max(info_cupo["cupos_disponibles"] - 1, 0)
                )
            }

            return jsonify(response), 201

        except IntegrityError as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo crear el usuario por conflicto de datos únicos.",
                "detalle": "Probablemente ya existe un usuario con ese email.",
                "debug": str(e)
            }), 409

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo crear el usuario.",
                "detalle": str(e)
            }), 500


    @app.route("/admin/usuarios/<int:idusuario>", methods=["PUT"])
    @jwt_required()
    def admin_update_usuario(idusuario):
        claims = get_jwt()

        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        user = Usuario.query.get_or_404(idusuario)
        data = request.get_json() or {}

        nuevo_idcliente = user.idcliente
        nuevo_idperfil = user.idperfil
        nuevo_activo = user.activo

        if "idcliente" in data and data["idcliente"] not in (None, ""):
            try:
                nuevo_idcliente = int(data["idcliente"])
            except Exception:
                return jsonify({
                    "error": "idcliente inválido",
                    "detalle": "idcliente debe ser numérico."
                }), 400

        if "idperfil" in data and data["idperfil"] not in (None, ""):
            try:
                nuevo_idperfil = int(data["idperfil"])
            except Exception:
                return jsonify({
                    "error": "idperfil inválido",
                    "detalle": "idperfil debe ser numérico."
                }), 400

        if "activo" in data:
            nuevo_activo = _bool_from_payload(data.get("activo"), default=user.activo)

        # Si el usuario queda activo, validar cupo del cliente destino.
        # Esto cubre:
        # - Reactivar usuario.
        # - Mover usuario a otro cliente.
        # - Mantener activo en el mismo cliente, excluyéndose a sí mismo.
        if nuevo_activo:
            ok_cupo, info_cupo, status_cupo = _validar_cupo_usuarios_cliente(
                nuevo_idcliente,
                excluir_idusuario=user.idusuario
            )

            if not ok_cupo:
                return jsonify(info_cupo), status_cupo

        # Validar que el perfil pertenezca al cliente destino
        if nuevo_idperfil:
            perfil, error_perfil, status_perfil = _validar_perfil_cliente(
                nuevo_idperfil,
                nuevo_idcliente
            )

            if error_perfil:
                return jsonify(error_perfil), status_perfil

        try:
            if "nombre" in data:
                user.nombre = str(data["nombre"]).strip()

            if "apellido" in data:
                user.apellido = str(data["apellido"]).strip() if data["apellido"] else None

            if "email" in data:
                user.email = str(data["email"]).strip().lower()

            user.idcliente = nuevo_idcliente
            user.idperfil = nuevo_idperfil
            user.activo = nuevo_activo

            if "password" in data and data["password"]:
                user.password_hash = generate_password_hash(
                    data["password"],
                    method="pbkdf2:sha256",
                    salt_length=16
                )

            db.session.commit()
            return jsonify(user.as_dict())

        except IntegrityError as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo actualizar el usuario por conflicto de datos únicos.",
                "detalle": "Probablemente ya existe otro usuario con ese email.",
                "debug": str(e)
            }), 409

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo actualizar el usuario.",
                "detalle": str(e)
            }), 500


    @app.route("/admin/usuarios/<int:idusuario>", methods=["DELETE"])
    @jwt_required()
    def admin_delete_usuario(idusuario):
        claims = get_jwt()

        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        user = Usuario.query.get_or_404(idusuario)

        db.session.delete(user)
        db.session.commit()

        return jsonify({"message": "Usuario eliminado"})



    # ==========================
    # Registro Inicial de Cliente por el superadmin
    # ==========================

    @app.route("/clientes/registro_inicial", methods=["POST"])
    @jwt_required()
    def registro_inicial_cliente():
        claims = get_jwt()

        if claims["perfilid"] != 0:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        cliente_data = data.get("cliente") or {}
        usuario_data = data.get("usuario") or {}

        paquetes_codigos = data.get("paquetes") or []
        idpaquetes = data.get("idpaquetes") or []

        if not cliente_data.get("nombre"):
            return jsonify({"error": "Falta nombre del cliente"}), 400

        if not usuario_data.get("nombre") or not usuario_data.get("email") or not usuario_data.get("password"):
            return jsonify({
                "error": "Faltan datos del usuario administrador: nombre, email o password"
            }), 400

        if not paquetes_codigos and not idpaquetes:
            return jsonify({
                "error": "Debe indicar al menos un paquete contratado en 'paquetes' o 'idpaquetes'"
            }), 400

        # =====================================================
        # Validación contractual: límite mínimo de usuarios
        # =====================================================
        limite_usuarios_raw = cliente_data.get("limite_usuarios")

        if limite_usuarios_raw not in (None, ""):
            try:
                limite_usuarios_int = int(limite_usuarios_raw)
            except Exception:
                return jsonify({
                    "error": "Límite de usuarios inválido",
                    "detalle": "limite_usuarios debe ser un valor numérico."
                }), 400

            if limite_usuarios_int < 1:
                return jsonify({
                    "error": "Límite de usuarios inválido",
                    "detalle": (
                        "El cliente debe tener al menos 1 usuario permitido, "
                        "porque el registro inicial crea un usuario administrador."
                    )
                }), 400
        else:
            # Si no llega límite, por seguridad dejamos mínimo 1,
            # porque este endpoint crea el usuario administrador inicial.
            limite_usuarios_int = 1

        # =====================================================
        # Validación contractual: límite de sesiones
        # =====================================================
        limite_sesiones_raw = cliente_data.get("limite_sesiones")

        if limite_sesiones_raw not in (None, ""):
            try:
                limite_sesiones_int = int(limite_sesiones_raw)
            except Exception:
                return jsonify({
                    "error": "Límite de sesiones inválido",
                    "detalle": "limite_sesiones debe ser un valor numérico."
                }), 400

            if limite_sesiones_int < 1:
                return jsonify({
                    "error": "Límite de sesiones inválido",
                    "detalle": "El cliente debe tener al menos 1 sesión concurrente permitida."
                }), 400
        else:
            # Si no llega límite, por seguridad dejamos mínimo 1.
            limite_sesiones_int = 1

        try:
            # =====================================================
            # 1. Resolver paquetes contratados
            # =====================================================
            q_paquetes = PaqueteInsightflow.query.filter_by(activo=True)

            if paquetes_codigos:
                paquetes_codigos = [
                    str(c).strip().lower()
                    for c in paquetes_codigos
                    if str(c).strip()
                ]

                paquetes = (
                    q_paquetes
                    .filter(PaqueteInsightflow.codigo.in_(paquetes_codigos))
                    .all()
                )
            else:
                paquetes = (
                    q_paquetes
                    .filter(PaqueteInsightflow.idpaquete.in_(idpaquetes))
                    .all()
                )

            if not paquetes:
                return jsonify({
                    "error": "No se encontraron paquetes activos con los datos enviados."
                }), 400

            # =====================================================
            # 2. Crear Cliente
            # =====================================================
            cliente = Cliente(
                nombre=str(cliente_data["nombre"]).strip(),
                nit=str(cliente_data.get("nit")).strip() if cliente_data.get("nit") else None,
                email=str(cliente_data.get("email")).strip().lower() if cliente_data.get("email") else None,
                activo=bool(cliente_data.get("activo", True)),
                pais=str(cliente_data.get("pais")).strip() if cliente_data.get("pais") else None,
                ciudad=str(cliente_data.get("ciudad")).strip() if cliente_data.get("ciudad") else None,
                direccion=str(cliente_data.get("direccion")).strip() if cliente_data.get("direccion") else None,
                telefono1=str(cliente_data.get("telefono1")).strip() if cliente_data.get("telefono1") else None,
                logo_url=str(cliente_data.get("logo_url")).strip() if cliente_data.get("logo_url") else None,
                limite_usuarios=limite_usuarios_int,
                limite_sesiones=limite_sesiones_int,
                timezone=cliente_data.get("timezone") or "America/Bogota"
            )

            db.session.add(cliente)
            db.session.flush()

            # =====================================================
            # 3. Asociar paquetes al cliente
            # =====================================================
            for paquete in paquetes:
                db.session.add(ClientePaquete(
                    idcliente=cliente.idcliente,
                    idpaquete=paquete.idpaquete,
                    activo=True,
                    fecha_inicio=date.today(),
                    fecha_fin=None
                ))

            db.session.flush()

            # =====================================================
            # 4. Obtener códigos permitidos por paquetes contratados
            # =====================================================
            codigos_permitidos = obtener_codigos_permitidos_cliente(cliente.idcliente)

            if not codigos_permitidos:
                db.session.rollback()
                return jsonify({
                    "error": "Los paquetes seleccionados no tienen permisos activos configurados."
                }), 400

            # =====================================================
            # 5. Crear Perfil Administrador
            # =====================================================
            perfil_admin = Perfil(
                idcliente=cliente.idcliente,
                nombre="Administrador",
                descripcion="Perfil administrador del cliente"
            )

            db.session.add(perfil_admin)
            db.session.flush()

            # =====================================================
            # 6. Crear permisos del cliente desde los códigos del paquete
            #    IMPORTANTE:
            #    Antes esto clonaba desde Permiso.idcliente == 1.
            #    Eso hacía que el sistema dependiera de Binaria como cliente plantilla.
            #    Ahora los permisos se crean desde paquete_permisos / codigos_permitidos.
            # =====================================================

            def nombre_permiso_desde_codigo(codigo: str) -> str:
                """
                Convierte un código como 'ver_reporte_balance_general'
                en un nombre legible como 'Ver Reporte Balance General'.
                """
                return str(codigo or "").replace("_", " ").strip().title()

            permisos_creados = []

            for codigo in sorted(codigos_permitidos):
                codigo_limpio = str(codigo).strip()

                if not codigo_limpio:
                    continue

                permiso_existente = (
                    Permiso.query
                    .filter(
                        Permiso.idcliente == cliente.idcliente,
                        Permiso.codigo == codigo_limpio
                    )
                    .first()
                )

                if permiso_existente:
                    permisos_creados.append(permiso_existente)
                    continue

                nuevo = Permiso(
                    idcliente=cliente.idcliente,
                    nombre=nombre_permiso_desde_codigo(codigo_limpio),
                    codigo=codigo_limpio,
                    descripcion=f"Permiso generado automáticamente para el código '{codigo_limpio}'.",
                    activo=True
                )

                db.session.add(nuevo)
                db.session.flush()

                permisos_creados.append(nuevo)

            if not permisos_creados:
                db.session.rollback()
                return jsonify({
                    "error": "No se pudieron crear permisos para el cliente desde los códigos del paquete.",
                    "detalle": (
                        "Verifica que paquete_permisos tenga códigos activos para los paquetes seleccionados."
                    )
                }), 400
            

            # =====================================================
            # 7. Asignar todos los permisos contratados al perfil administrador
            # =====================================================
            for permiso in permisos_creados:
                db.session.add(PerfilPermiso(
                    idcliente=cliente.idcliente,
                    idperfil=perfil_admin.idperfil,
                    idpermiso=permiso.idpermiso,
                    permitido=True
                ))

            # =====================================================
            # 8. Crear Usuario Administrador
            # =====================================================
            password_hash = generate_password_hash(
                usuario_data["password"],
                method="pbkdf2:sha256",
                salt_length=16
            )

            usuario_admin = Usuario(
                idcliente=cliente.idcliente,
                idperfil=perfil_admin.idperfil,
                nombre=str(usuario_data["nombre"]).strip(),
                apellido=str(usuario_data.get("apellido")).strip() if usuario_data.get("apellido") else None,
                email=str(usuario_data["email"]).strip().lower(),
                password_hash=password_hash,
                activo=True
            )

            db.session.add(usuario_admin)
            db.session.commit()

            return jsonify({
                "message": "Cliente registrado correctamente con paquete contratado.",
                "cliente": cliente.as_dict(),
                "paquetes": [
                    {
                        "idpaquete": p.idpaquete,
                        "codigo": p.codigo,
                        "nombre": p.nombre
                    }
                    for p in paquetes
                ],
                "perfil_admin": {
                    "idperfil": perfil_admin.idperfil,
                    "nombre": perfil_admin.nombre
                },
                "usuario_admin": {
                    "idusuario": usuario_admin.idusuario,
                    "email": usuario_admin.email,
                    "nombre": usuario_admin.nombre,
                    "apellido": usuario_admin.apellido
                },
                "control_usuarios": {
                    "limite_usuarios": cliente.limite_usuarios,
                    "usuarios_activos": 1,
                    "cupos_disponibles": max(cliente.limite_usuarios - 1, 0)
                },
                "control_sesiones": {
                    "limite_sesiones": cliente.limite_sesiones
                },
                "permisos_asignados": len(permisos_creados)
            }), 201

        except IntegrityError as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo registrar el cliente por conflicto de datos únicos.",
                "detalle": "Probablemente ya existe un usuario con ese email o existe un conflicto de datos únicos.",
                "debug": str(e)
            }), 409

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo registrar el cliente.",
                "detalle": str(e)
            }), 500




    # ==========================
    # Cliente Admin CRUD Perfiles
    # ==========================

    @app.route("/perfiles", methods=["GET"])
    @jwt_required()
    def cliente_listar_perfiles():
        claims = get_jwt()

        if claims["perfilid"] == 0:
            return jsonify({"error": "SuperAdmin no debe usar este endpoint"}), 403

        idcliente = claims.get("idcliente")

        if not idcliente:
            return jsonify({"error": "Token inválido: falta idcliente"}), 403

        perfiles = (
            Perfil.query
            .filter_by(idcliente=idcliente)
            .order_by(Perfil.nombre.asc())
            .all()
        )

        return jsonify([p.as_dict() for p in perfiles]), 200





    @app.route("/perfiles", methods=["POST"])
    @jwt_required()
    def cliente_crear_perfil():
        claims = get_jwt()
        if claims["perfilid"] == 0:  
            return jsonify({"error": "SuperAdmin no puede crear perfiles de clientes"}), 403

        data = request.get_json()
        perfil = Perfil(
            idcliente=claims["idcliente"],  # 🔐 siempre el cliente del token
            nombre=data["nombre"],
            descripcion=data.get("descripcion", "")
        )
        db.session.add(perfil)
        db.session.commit()
        return jsonify(perfil.as_dict()), 201

    @app.route("/perfiles/<int:idperfil>", methods=["PUT"])
    @jwt_required()
    def cliente_update_perfil(idperfil):
        claims = get_jwt()
        if claims["perfilid"] == 0:  
            return jsonify({"error": "SuperAdmin no puede editar perfiles de clientes"}), 403

        perfil = Perfil.query.get_or_404(idperfil)
        if perfil.idcliente != claims["idcliente"]:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json()
        perfil.nombre = data.get("nombre", perfil.nombre)
        perfil.descripcion = data.get("descripcion", perfil.descripcion)
        db.session.commit()
        return jsonify(perfil.as_dict())

    @app.route("/perfiles/<int:idperfil>", methods=["DELETE"])
    @jwt_required()
    def cliente_delete_perfil(idperfil):
        claims = get_jwt()
        if claims["perfilid"] == 0:
            return jsonify({"error": "SuperAdmin no puede borrar perfiles de clientes"}), 403

        perfil = Perfil.query.get_or_404(idperfil)
        if perfil.idcliente != claims["idcliente"]:
            return jsonify({"error": "No autorizado"}), 403

        db.session.delete(perfil)
        db.session.commit()
        return jsonify({"message": "Perfil eliminado"})


    # ==========================
    # Cliente Admin CRUD Usuarios
    # ==========================

    @app.route("/usuarios", methods=["POST"])
    @jwt_required()
    def crear_usuario():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        perfilid = claims.get("perfilid")

        # Solo admins de cliente pueden crear usuarios.
        # El SuperAdmin debe usar /admin/usuarios.
        if perfilid == 0:
            return jsonify({
                "error": "SuperAdmin no puede crear usuarios de clientes desde este endpoint"
            }), 403

        data = request.get_json() or {}

        required = ["idperfil", "nombre", "email", "password"]
        faltantes = [field for field in required if not data.get(field)]

        if faltantes:
            return jsonify({
                "error": "Faltan campos obligatorios",
                "faltantes": faltantes
            }), 400

        # Validar límite de usuarios activos del cliente autenticado
        ok_cupo, info_cupo, status_cupo = _validar_cupo_usuarios_cliente(idcliente)
        if not ok_cupo:
            return jsonify(info_cupo), status_cupo

        try:
            idperfil_nuevo = int(data["idperfil"])
        except Exception:
            return jsonify({
                "error": "Perfil inválido",
                "detalle": "idperfil debe ser numérico."
            }), 400

        # Validar que el perfil pertenezca al mismo cliente del JWT
        perfil, error_perfil, status_perfil = _validar_perfil_cliente(
            idperfil_nuevo,
            idcliente
        )

        if error_perfil:
            return jsonify(error_perfil), status_perfil

        try:
            password_hash = generate_password_hash(
                data["password"],
                method="pbkdf2:sha256",
                salt_length=16
            )

            usuario = Usuario(
                idcliente=idcliente,
                idperfil=perfil.idperfil,
                nombre=str(data["nombre"]).strip(),
                apellido=(str(data.get("apellido")).strip() if data.get("apellido") else None),
                email=str(data["email"]).strip().lower(),
                password_hash=password_hash,
                activo=True
            )

            db.session.add(usuario)
            db.session.commit()

            return jsonify({
                "idusuario": usuario.idusuario,
                "idcliente": usuario.idcliente,
                "idperfil": usuario.idperfil,
                "nombre": usuario.nombre,
                "apellido": usuario.apellido,
                "email": usuario.email,
                "activo": usuario.activo,
                "perfil": perfil.nombre,
                "control_usuarios": {
                    **info_cupo,
                    "usuarios_activos_despues": (
                        None
                        if info_cupo.get("usuarios_activos") is None
                        else info_cupo["usuarios_activos"] + 1
                    ),
                    "cupos_disponibles_despues": (
                        None
                        if info_cupo.get("cupos_disponibles") is None
                        else max(info_cupo["cupos_disponibles"] - 1, 0)
                    )
                }
            }), 201

        except IntegrityError as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo crear el usuario por conflicto de datos únicos.",
                "detalle": "Probablemente ya existe un usuario con ese email.",
                "debug": str(e)
            }), 409

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo crear el usuario.",
                "detalle": str(e)
            }), 500


    @app.route("/usuarios/<int:idusuario>", methods=["PUT"])
    @jwt_required()
    def cliente_update_usuario(idusuario):
        claims = get_jwt()

        if claims["perfilid"] == 0:
            return jsonify({
                "error": "SuperAdmin no debe usar este endpoint"
            }), 403

        idcliente = claims.get("idcliente")

        usuario = Usuario.query.get_or_404(idusuario)

        if usuario.idcliente != idcliente:
            return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        try:
            if "nombre" in data:
                usuario.nombre = str(data["nombre"]).strip()

            if "apellido" in data:
                usuario.apellido = str(data["apellido"]).strip() if data["apellido"] else None

            if "email" in data:
                usuario.email = str(data["email"]).strip().lower()

            # Opcional: si tu frontend cliente permite cambiar perfil,
            # esta parte ya queda lista y segura.
            if "idperfil" in data and data["idperfil"] not in (None, ""):
                try:
                    nuevo_idperfil = int(data["idperfil"])
                except Exception:
                    return jsonify({
                        "error": "Perfil inválido",
                        "detalle": "idperfil debe ser numérico."
                    }), 400

                perfil, error_perfil, status_perfil = _validar_perfil_cliente(
                    nuevo_idperfil,
                    idcliente
                )

                if error_perfil:
                    return jsonify(error_perfil), status_perfil

                usuario.idperfil = perfil.idperfil

            # Opcional: si más adelante permites activar/inactivar desde esta página,
            # validamos cupo cuando se intente activar.
            if "activo" in data:
                nuevo_activo = _bool_from_payload(data.get("activo"), default=usuario.activo)

                if nuevo_activo and not usuario.activo:
                    ok_cupo, info_cupo, status_cupo = _validar_cupo_usuarios_cliente(
                        idcliente,
                        excluir_idusuario=usuario.idusuario
                    )

                    if not ok_cupo:
                        return jsonify(info_cupo), status_cupo

                usuario.activo = nuevo_activo

            if "password" in data and data["password"]:
                usuario.password_hash = generate_password_hash(
                    data["password"],
                    method="pbkdf2:sha256",
                    salt_length=16
                )

            db.session.commit()
            return jsonify(usuario.as_dict())

        except IntegrityError as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo actualizar el usuario por conflicto de datos únicos.",
                "detalle": "Probablemente ya existe otro usuario con ese email.",
                "debug": str(e)
            }), 409

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo actualizar el usuario.",
                "detalle": str(e)
            }), 500


    @app.route("/usuarios/<int:idusuario>", methods=["DELETE"])
    @jwt_required()
    def cliente_delete_usuario(idusuario):
        claims = get_jwt()

        if claims["perfilid"] == 0:
            return jsonify({
                "error": "SuperAdmin no debe usar este endpoint"
            }), 403

        usuario = Usuario.query.get_or_404(idusuario)

        if usuario.idcliente != claims["idcliente"]:
            return jsonify({"error": "No autorizado"}), 403

        db.session.delete(usuario)
        db.session.commit()

        return jsonify({"message": "Usuario eliminado"})


    @app.route("/usuarios", methods=["GET"])
    @jwt_required()
    def cliente_listar_usuarios():
        claims = get_jwt()

        if claims["perfilid"] == 0:
            return jsonify({
                "error": "SuperAdmin no debe usar este endpoint"
            }), 403

        idcliente = claims.get("idcliente")

        usuarios = Usuario.query.filter_by(idcliente=idcliente).all()

        return jsonify([u.as_dict() for u in usuarios])




    # para que el frontend consulte el rol y el cliente del token
    @app.route("/auth/whoami", methods=["GET"])
    @jwt_required()
    def whoami():
        user_id = get_jwt_identity()
        user = db.session.get(Usuario, user_id)
        if not user:
            return jsonify({"error": "Usuario no encontrado"}), 404

        data = {
            "idusuario": user.idusuario,
            "idcliente": user.idcliente,
            "idperfil": user.idperfil,
            "perfilid": 0 if user.idcliente is None else user.idperfil,
            "email": user.email,
        }

        if user.idcliente:
            cliente = db.session.get(Cliente, user.idcliente)
            if cliente:
                data["cliente"] = {
                    "id": cliente.idcliente,
                    "nombre": cliente.nombre,
                    "logo_url": cliente.logo_url,
                }

        return jsonify(data), 200


    @app.route("/config/siigo", methods=["GET"])
    @jwt_required()
    def get_siigo_config():
        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0:
            if not q_idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
            idcliente = q_idcliente
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        sync_cfg = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        def mask(s: str | None):
            if not s:
                return None
            return s[:2] + "•" * (max(0, len(s) - 4)) + s[-2:]

        if not cfg:
            return jsonify({
                "idcliente": idcliente,
                "base_url": None,
                "client_id": None,
                "client_secret_mask": None,
                "username": None,
                "password_mask": None,
                "partner_id": None,
                "updated_at": None,
                "ds_fecha_desde": sync_cfg.ds_fecha_desde.isoformat() if sync_cfg and sync_cfg.ds_fecha_desde else None,
            })

        return jsonify({
            "idcliente": idcliente,
            "base_url": cfg.base_url,
            "client_id": cfg.client_id,
            "client_secret_mask": mask(dec(cfg.client_secret)),
            "username": cfg.username,
            "password_mask": mask(dec(cfg.password)),
            "partner_id": cfg.partner_id,
            "updated_at": cfg.updated_at.isoformat() if cfg.updated_at else None,
            "ds_fecha_desde": sync_cfg.ds_fecha_desde.isoformat() if sync_cfg and sync_cfg.ds_fecha_desde else None,
        })


    @app.route("/config/sync", methods=["POST"])
    @jwt_required()
    def upsert_siigo_sync_config():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)

        if perfilid == 0:
            if q_idcliente:
                idcliente = q_idcliente
            elif not idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        hora_ejecucion_raw = data.get("hora_ejecucion") or "02:00"
        frecuencia_dias_raw = data.get("frecuencia_dias", 1)
        activo_raw = data.get("activo", True)

        hora_ejecucion = _parse_time_hh_mm(hora_ejecucion_raw, 2, 0)

        try:
            frecuencia_dias = int(frecuencia_dias_raw or 1)
            if frecuencia_dias < 1:
                frecuencia_dias = 1
        except Exception:
            frecuencia_dias = 1

        if isinstance(activo_raw, bool):
            activo = activo_raw
        elif isinstance(activo_raw, str):
            activo = activo_raw.strip().lower() in ("true", "1", "yes", "si", "sí", "on")
        else:
            activo = bool(activo_raw)

        sync_fecha_desde = _parse_date_yyyy_mm_dd(
            data.get("sync_fecha_desde") or data.get("ds_fecha_desde")
        )

        try:
            config = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

            if not config:
                config = SiigoSyncConfig(
                    idcliente=idcliente,
                    hora_ejecucion=hora_ejecucion,
                    frecuencia_dias=frecuencia_dias,
                    activo=activo,
                    sync_fecha_desde=sync_fecha_desde,
                    ds_fecha_desde=sync_fecha_desde,
                )
                db.session.add(config)
            else:
                config.hora_ejecucion = hora_ejecucion
                config.frecuencia_dias = frecuencia_dias
                config.activo = activo
                config.sync_fecha_desde = sync_fecha_desde
                config.ds_fecha_desde = sync_fecha_desde

            db.session.commit()

            return jsonify({
                "message": "Configuración de sincronización guardada",
                "config": config.as_dict() if hasattr(config, "as_dict") else {
                    "id": config.id,
                    "idcliente": config.idcliente,
                    "hora_ejecucion": str(config.hora_ejecucion) if config.hora_ejecucion else None,
                    "frecuencia_dias": config.frecuencia_dias,
                    "activo": config.activo,
                    "sync_fecha_desde": config.sync_fecha_desde.isoformat() if config.sync_fecha_desde else None,
                    "ds_fecha_desde": config.ds_fecha_desde.isoformat() if config.ds_fecha_desde else None,
                    "ultimo_ejecutado": config.ultimo_ejecutado.isoformat() if config.ultimo_ejecutado else None,
                    "ultimo_auto_ejecutado": config.ultimo_auto_ejecutado.isoformat() if config.ultimo_auto_ejecutado else None,
                    "resultado_ultima_sync": config.resultado_ultima_sync,
                    "detalle_ultima_sync": config.detalle_ultima_sync,
                }
            }), 200

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No se pudo guardar la configuración de sincronización.",
                "detalle": str(e),
                "idcliente": idcliente,
            }), 500

    @app.route("/config/siigo", methods=["PUT"])
    @jwt_required()
    def upsert_siigo_config():
        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0:
            if not q_idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
            idcliente = q_idcliente
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        base_url = data.get("base_url")
        client_id = data.get("client_id")
        client_secret = data.get("client_secret")
        username = data.get("username")
        password = data.get("password")
        partner_id = data.get("partner_id")
        ds_fecha_desde_raw = data.get("ds_fecha_desde")

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            cfg = SiigoCredencial(idcliente=idcliente)

        if base_url is not None:
            cfg.base_url = base_url
        if client_id is not None:
            cfg.client_id = client_id
        if client_secret:
            cfg.client_secret = enc(client_secret)
        if username is not None:
            cfg.username = username
        if password:
            cfg.password = enc(password)
        if partner_id is not None:
            cfg.partner_id = partner_id

        db.session.add(cfg)

        sync_cfg = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        # Solo tocamos ds_fecha_desde si viene explícitamente en el payload.
        # Esto evita borrar la configuración si el frontend viejo no manda el campo.
        if "sync_fecha_desde" in data or "ds_fecha_desde" in data:
            if not sync_cfg:
                sync_cfg = SiigoSyncConfig(
                    idcliente=idcliente,
                    hora_ejecucion=_parse_time_hh_mm("02:00"),
                    frecuencia_dias=1,
                    activo=True,
                )

            fecha_global = _parse_date_yyyy_mm_dd(data.get("sync_fecha_desde") or data.get("ds_fecha_desde"))

            sync_cfg.sync_fecha_desde = fecha_global
            sync_cfg.ds_fecha_desde = fecha_global

            db.session.add(sync_cfg)

        db.session.commit()

        return jsonify({
            "message": "Configuración guardada",
            "sync_fecha_desde": sync_cfg.sync_fecha_desde.isoformat() if sync_cfg and sync_cfg.sync_fecha_desde else None,
            "ds_fecha_desde": sync_cfg.ds_fecha_desde.isoformat() if sync_cfg and sync_cfg.ds_fecha_desde else None,
         }), 200


    @app.route("/siigo/test_auth", methods=["POST"])
    @jwt_required()
    def siigo_test_auth():
        """
        Autenticación contra Siigo con dos flujos (JSON y Basic).
        Si SIIGO_SANDBOX=1 (o ?force_sandbox=1), responde OK sin llamar a Siigo.
        """
        # -------- sandbox / simulación ----------
        force_sandbox = request.args.get("force_sandbox") == "1"
        if SANDBOX or force_sandbox:
            return jsonify({
                "ok": True,
                "flow": "sandbox",
                "endpoint": None,
                "token_type": "bearer",
                "expires_in": 3600
            }), 200
        # ----------------------------------------

        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0:
            if not q_idcliente:
                return jsonify({"ok": False, "error": "Falta idcliente"}), 400
            idcliente = q_idcliente
        else:
            if not idcliente:
                return jsonify({"ok": False, "error": "No autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg or not cfg.base_url or not cfg.client_id or not cfg.client_secret:
            return jsonify({"ok": False, "error": "Credenciales incompletas"}), 400

        base_url = (cfg.base_url or "").rstrip("/")
        username = cfg.client_id or ""                 # Usuario API (correo)
        access_key = dec(cfg.client_secret) or ""      # Access Key
        partner_id = os.getenv("SIIGO_PARTNER_ID", PARTNER_ID_DEFAULT).strip() or PARTNER_ID_DEFAULT

        base_headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
            "Partner-Id": partner_id,
        }

        timeout = AUTH_TIMEOUT_SEC

        # ---------- 1) Flujo JSON (oficial) ----------
        json_auth_urls = [
            f"{base_url}/auth",
            f"{base_url}/v1/auth",
            f"{base_url}/oauth2/token",
        ]
        for url in json_auth_urls:
            try:
                r = requests.post(
                    url,
                    headers=base_headers,
                    json={"username": username, "access_key": access_key},
                    timeout=timeout,
                )
                if r.status_code == 200:
                    try:
                        data = r.json() or {}
                        print("=== TOKEN DE SIIGO ===")
                        print(data)
                    except Exception:
                        data = {}
                    return jsonify({
                        "ok": True,
                        "flow": "json",
                        "endpoint": url,
                        "token_type": data.get("token_type"),
                        "expires_in": data.get("expires_in"),
                        "access_token": data.get("access_token")  # 👈 para que aparezca en respuesta
                    }), 200

                if r.status_code in (401, 403):
                    return jsonify({
                        "ok": False,
                        "flow": "json",
                        "endpoint": url,
                        "error": f"Credenciales inválidas ({r.status_code})",
                        "body": (r.text or "")[:800],
                    }), r.status_code

                if r.status_code == 404:
                    # probar siguiente variante
                    continue

                return jsonify({
                    "ok": False,
                    "flow": "json",
                    "endpoint": url,
                    "error": f"HTTP {r.status_code}",
                    "body": (r.text or "")[:800],
                }), 502
            except requests.RequestException as e:
                return jsonify({
                    "ok": False,
                    "flow": "json",
                    "endpoint": url,
                    "error": f"Conexión fallida: {str(e)}",
                }), 502

        # ---------- 2) Flujo Basic (fallback) ----------
        pair = f"{username}:{access_key}"
        basic_token = base64.b64encode(pair.encode("utf-8")).decode("utf-8")
        basic_headers = dict(base_headers)
        basic_headers["Authorization"] = f"Basic {basic_token}"

        basic_urls = [
            f"{base_url}/auth",
            f"{base_url}/v1/auth",
            f"{base_url}/oauth2/token",
        ]
        for url in basic_urls:
            try:
                r = requests.post(url, headers=basic_headers, timeout=timeout)
                if r.status_code == 200:
                    try:
                        data = r.json() or {}
                    except Exception:
                        data = {}
                    return jsonify({
                        "ok": True,
                        "flow": "basic",
                        "endpoint": url,
                        "token_type": data.get("token_type"),
                        "expires_in": data.get("expires_in"),
                    }), 200

                if r.status_code in (401, 403):
                    return jsonify({
                        "ok": False,
                        "flow": "basic",
                        "endpoint": url,
                        "error": f"Credenciales inválidas ({r.status_code})",
                        "body": (r.text or "")[:800],
                    }), r.status_code

                if r.status_code == 404:
                    continue

                return jsonify({
                    "ok": False,
                    "flow": "basic",
                    "endpoint": url,
                    "error": f"HTTP {r.status_code}",
                    "body": (r.text or "")[:800],
                }), 502
            except requests.RequestException as e:
                return jsonify({
                    "ok": False,
                    "flow": "basic",
                    "endpoint": url,
                    "error": f"Conexión fallida: {str(e)}",
                }), 502

        # ---------- ninguna ruta respondió ----------
        return jsonify({
            "ok": False,
            "error": "No se encontró un endpoint de auth que responda correctamente.",
            "tried_json": json_auth_urls,
            "tried_basic": basic_urls,
            "partner_id_used": partner_id,
            "base_url_used": base_url,
        }), 502



    # (Opcional) Endpoint de diagnóstico rápido
    @app.route("/siigo/_diag", methods=["GET"])
    @jwt_required()
    def siigo_diag():
        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        partner_id = os.getenv("SIIGO_PARTNER_ID", PARTNER_ID_DEFAULT).strip() or PARTNER_ID_DEFAULT

        if cfg and cfg.client_secret:
            try:
                access_key = dec(cfg.client_secret)
            except Exception as e:
                access_key = f"ERROR: {e}"
        else:
            access_key = None

        print("DEBUG auth credentials:")
        print("  Username:", cfg.client_id if cfg else "❌ no cfg")
        print("  Access Key:", access_key or "❌ vacía")

        return jsonify({
            "idcliente": idcliente,
            "base_url": (cfg.base_url if cfg else None),
            "client_id": (cfg.client_id if cfg else None),
            "secret_stored": bool(cfg and cfg.client_secret),
            "partner_id": partner_id,
        })
        



    @app.route("/siigo/sync-facturas", methods=["POST"])
    def siigo_sync_facturas():
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        deep = request.args.get("deep") in ("1", "true", "yes")
        batch = request.args.get("batch", default=None, type=int)
        only_missing = request.args.get("only_missing", default="1") in ("1", "true", "yes")
        since = request.args.get("since")
        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        kwargs = {
            "idcliente": idcliente,
            "deep": deep,
            "only_missing": only_missing,
            "since": since,
        }

        if batch:
            kwargs["batch_size"] = batch

        endpoint_log = "/siigo/sync-facturas"
        params_log = {
            "deep": deep,
            "batch": batch,
            "only_missing": only_missing,
            "since": since,
        }

        nombre_proceso = "Facturas detalle" if deep else "Facturas de venta"

        try:
            if modo_sync_all:
                # Ejecución sincrónica cuando viene desde sync-all.
                mensaje = sync_facturas_desde_siigo(**kwargs)
                return jsonify({"mensaje": mensaje}), 200

            # Modo UI manual: background con registro en historial.
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje=(
                    f"{nombre_proceso}: proceso iniciado"
                    + (f" con fecha desde {since}." if since else " sin límite de fecha.")
                ),
            )

            def trabajo_lento(local_kwargs, local_log_id, local_endpoint_log, local_nombre_proceso):
                with app.app_context():
                    inicio = time.time()

                    try:
                        mensaje = sync_facturas_desde_siigo(**local_kwargs)
                        duracion = round(time.time() - inicio, 2)

                        detalle = (
                            f"{local_nombre_proceso} finalizado correctamente. "
                            f"{mensaje}"
                        )

                        print(f"[siigo_sync_facturas] ✅ Terminado: {mensaje}")

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=local_kwargs["idcliente"],
                            endpoint=local_endpoint_log,
                            resultado="OK",
                            detalle=detalle,
                            status_code=200,
                            duracion_segundos=duracion,
                        )

                    except Exception as e:
                        duracion = round(time.time() - inicio, 2)
                        detalle_error = traceback.format_exc()

                        print(f"[siigo_sync_facturas] ❌ Error: {e}")
                        traceback.print_exc()

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=local_kwargs["idcliente"],
                            endpoint=local_endpoint_log,
                            resultado="ERROR",
                            detalle=detalle_error,
                            status_code=500,
                            duracion_segundos=duracion,
                        )

            t = Thread(
                target=trabajo_lento,
                args=(kwargs, log_id, endpoint_log, nombre_proceso),
                daemon=True,
            )
            t.start()

            return jsonify({
                "mensaje": (
                    f"{nombre_proceso}: proceso iniciado"
                    + (f" con fecha desde {since}." if since else " sin límite de fecha.")
                    + " Puedes revisar el resultado en el historial de sincronizaciones."
                ),
                "log_id": log_id,
                "origen": "manual_modulo",
                "estado": "EN_EJECUCION",
                "since": since,
            }), 202

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    @app.route("/siigo/sync-facturas-completar-detalle", methods=["POST"])
    def siigo_sync_facturas_completar_detalle():
        idcliente = obtener_idcliente_desde_request()

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        data = request.get_json(silent=True) or {}

        # Mantener estable: lotes de 100
        batch_size = data.get("batch_size") or request.args.get("batch", default=100, type=int)
        since = data.get("since") or request.args.get("since")
        max_lotes = data.get("max_lotes") or request.args.get("max_lotes", default=100, type=int)

        try:
            batch_size = int(batch_size or 100)
        except Exception:
            batch_size = 100

        try:
            max_lotes = int(max_lotes or 100)
        except Exception:
            max_lotes = 100

        # Protección para evitar cargas exageradas por error
        if batch_size < 1:
            batch_size = 100

        if batch_size > 100:
            batch_size = 100

        if max_lotes < 1:
            max_lotes = 100

        if max_lotes > 100:
            max_lotes = 100

        endpoint_log = "/siigo/sync-facturas-completar-detalle"

        params_log = {
            "batch_size": batch_size,
            "since": since,
            "max_lotes": max_lotes,
        }

        try:
            from siigo.siigo_sync_refactor import sync_facturas_desde_siigo, contar_facturas_pendientes

            pendientes_iniciales = contar_facturas_pendientes(
                idcliente=idcliente,
                since=since,
            )

            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje=(
                    f"Completar detalle de facturas: proceso iniciado. "
                    f"Pendientes iniciales: {pendientes_iniciales}. "
                    f"Lote máximo: {batch_size}. "
                    + (f"Fecha desde: {since}." if since else "Sin límite de fecha.")
                ),
            )

            def trabajo_completar_detalle(
                local_idcliente,
                local_log_id,
                local_batch_size,
                local_since,
                local_max_lotes,
            ):
                with app.app_context():
                    inicio = time.time()

                    try:
                        from siigo.siigo_sync_refactor import sync_facturas_desde_siigo, contar_facturas_pendientes

                        pendientes_antes_global = contar_facturas_pendientes(
                            idcliente=local_idcliente,
                            since=local_since,
                        )

                        detalle_lines = []
                        detalle_lines.append("Completar detalle de facturas iniciado.")
                        detalle_lines.append(f"Pendientes iniciales: {pendientes_antes_global}.")
                        detalle_lines.append(f"Tamaño de lote: {local_batch_size}.")
                        detalle_lines.append(f"Máximo de lotes permitidos: {local_max_lotes}.")
                        detalle_lines.append(f"Fecha desde: {local_since or 'SIN LIMITE'}.")
                        detalle_lines.append("")

                        lotes_ejecutados = 0
                        pendientes_actuales = pendientes_antes_global

                        while pendientes_actuales > 0 and lotes_ejecutados < local_max_lotes:
                            lotes_ejecutados += 1

                            pendientes_antes_lote = contar_facturas_pendientes(
                                idcliente=local_idcliente,
                                since=local_since,
                            )

                            if pendientes_antes_lote <= 0:
                                pendientes_actuales = 0
                                break

                            mensaje_lote = sync_facturas_desde_siigo(
                                idcliente=local_idcliente,
                                deep=True,
                                batch_size=local_batch_size,
                                only_missing=True,
                                since=local_since,
                            )

                            pendientes_despues_lote = contar_facturas_pendientes(
                                idcliente=local_idcliente,
                                since=local_since,
                            )

                            procesadas_lote = max(
                                pendientes_antes_lote - pendientes_despues_lote,
                                0,
                            )

                            detalle_lines.append(
                                f"Lote {lotes_ejecutados}: "
                                f"pendientes antes={pendientes_antes_lote}, "
                                f"procesadas={procesadas_lote}, "
                                f"pendientes después={pendientes_despues_lote}."
                            )
                            detalle_lines.append(str(mensaje_lote))
                            detalle_lines.append("")

                            pendientes_actuales = pendientes_despues_lote

                            # Evita ciclo infinito si Siigo no permite enriquecer alguna factura
                            if procesadas_lote == 0:
                                detalle_lines.append(
                                    "El proceso se detuvo porque el último lote no redujo los pendientes. "
                                    "Puede haber facturas con error al consultar detalle en Siigo."
                                )
                                break

                        pendientes_finales = contar_facturas_pendientes(
                            idcliente=local_idcliente,
                            since=local_since,
                        )

                        procesadas_total = max(
                            pendientes_antes_global - pendientes_finales,
                            0,
                        )

                        progreso = (
                            100
                            if pendientes_antes_global == 0
                            else round((procesadas_total / pendientes_antes_global) * 100, 1)
                        )

                        finalizado = pendientes_finales == 0

                        detalle_lines.append("Resumen final:")
                        detalle_lines.append(f"Pendientes iniciales: {pendientes_antes_global}.")
                        detalle_lines.append(f"Procesadas en esta ejecución: {procesadas_total}.")
                        detalle_lines.append(f"Pendientes finales: {pendientes_finales}.")
                        detalle_lines.append(f"Lotes ejecutados: {lotes_ejecutados}.")
                        detalle_lines.append(f"Progreso de esta ejecución: {progreso}%.")
                        detalle_lines.append(f"Finalizado: {'sí' if finalizado else 'no'}.")

                        if not finalizado and lotes_ejecutados >= local_max_lotes:
                            detalle_lines.append(
                                "El proceso se detuvo porque alcanzó el máximo de lotes configurado. "
                                "Puedes ejecutarlo nuevamente para continuar."
                            )

                        duracion = round(time.time() - inicio, 2)
                        detalle = "\n".join(detalle_lines)

                        resultado = "OK" if finalizado else "PARCIAL"

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=local_idcliente,
                            endpoint=endpoint_log,
                            resultado=resultado,
                            detalle=detalle,
                            status_code=200,
                            duracion_segundos=duracion,
                        )

                    except Exception as e:
                        duracion = round(time.time() - inicio, 2)
                        detalle_error = traceback.format_exc()

                        print(f"[siigo_sync_facturas_completar_detalle] ❌ Error: {e}")
                        traceback.print_exc()

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=local_idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle_error,
                            status_code=500,
                            duracion_segundos=duracion,
                        )

            t = Thread(
                target=trabajo_completar_detalle,
                args=(idcliente, log_id, batch_size, since, max_lotes),
                daemon=True,
            )
            t.start()

            return jsonify({
                "mensaje": (
                    f"Completar detalle de facturas iniciado. "
                    f"Pendientes iniciales: {pendientes_iniciales}. "
                    f"El sistema procesará lotes de hasta {batch_size} facturas automáticamente. "
                    f"Puedes revisar el avance en el historial de sincronizaciones."
                ),
                "log_id": log_id,
                "origen": "manual_modulo",
                "estado": "EN_EJECUCION",
                "batch_size": batch_size,
                "max_lotes": max_lotes,
                "pendientes_iniciales": pendientes_iniciales,
                "since": since,
            }), 202

        except Exception as e:
            return jsonify({"error": str(e)}), 500




    @app.route("/siigo/debug-invoice", methods=["GET"])
    @jwt_required()
    def siigo_debug_invoice():
        """
        Muestra:
        - detalle crudo de Siigo para una factura (por ?uuid=... o ?name=...)
        - lo guardado en siigo_facturas y siigo_factura_items
        Uso:
        /siigo/debug-invoice?uuid=<uuid_de_siigo>
        /siigo/debug-invoice?name=<name/FV-...>
        Nota: requiere JWT; usa el idcliente del token (o ?idcliente= en superadmin).
        """
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        # superadmin puede inspeccionar otro cliente
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado o sin cliente en el token"}), 403

        uuid = request.args.get("uuid", type=str)
        name = request.args.get("name", type=str)
        if not uuid and not name:
            return jsonify({"error": "Proporciona ?uuid= o ?name="}), 400

        # cargar credenciales
        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg or not cfg.base_url or not cfg.client_id or not cfg.client_secret:
            return jsonify({"error": "Credenciales de Siigo no configuradas para este cliente"}), 400

        # autenticar
        auth_data = _siigo_auth_json_for_client(cfg)
        if auth_data.get("_error"):
            return jsonify({"error": auth_data["_error"]}), 502
        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token desde Siigo"}), 502

        base_url = (cfg.base_url or "").rstrip("/")
        headers = _siigo_headers_bearer(token)

        # 1) Traer detalle crudo desde Siigo
        raw_detail = None
        tried = []
        try:
            if uuid:
                url = f"{base_url}/v1/invoices/{uuid}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    raw_detail = r.json()
                elif name:
                    # fallback por nombre si mandaste también name
                    url2 = f"{base_url}/v1/invoices/{name}"
                    tried.append(url2)
                    r2 = requests.get(url2, headers=headers, timeout=60)
                    if r2.status_code == 200:
                        raw_detail = r2.json()
                    else:
                        # búsqueda por query name
                        url3 = f"{base_url}/v1/invoices?name={name}"
                        tried.append(url3)
                        r3 = requests.get(url3, headers=headers, timeout=60)
                        if r3.status_code == 200:
                            payload = r3.json() or {}
                            res = payload.get("results") or []
                            raw_detail = res[0] if res else None
            else:
                # sólo name
                url = f"{base_url}/v1/invoices/{name}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    raw_detail = r.json()
                else:
                    url2 = f"{base_url}/v1/invoices?name={name}"
                    tried.append(url2)
                    r2 = requests.get(url2, headers=headers, timeout=60)
                    if r2.status_code == 200:
                        payload = r2.json() or {}
                        res = payload.get("results") or []
                        raw_detail = res[0] if res else None
        except requests.RequestException as e:
            return jsonify({"error": f"Error consultando Siigo: {str(e)}", "tried": tried}), 502

        # 2) Lo que tienes guardado en BD
        factura_db = None
        items_db = []
        if name:
            factura_db = SiigoFactura.query.filter_by(idcliente=idcliente, idfactura=name).first()
        elif uuid:
            factura_db = SiigoFactura.query.filter_by(idcliente=idcliente, siigo_uuid=uuid).first()
            # si no la encuentra por uuid, intenta por name si el detalle lo trae
            if not factura_db and raw_detail and isinstance(raw_detail, dict):
                nm = raw_detail.get("name")
                if nm:
                    factura_db = SiigoFactura.query.filter_by(idcliente=idcliente, idfactura=nm).first()

        if factura_db:
            items_db = SiigoFacturaItem.query.filter_by(factura_id=factura_db.id).all()

        # 3) Serializar para JSON (sin SQLAlchemy-automagic)
        def _dec(v):
            # helper simple: para Decimals/fechas
            from decimal import Decimal
            if isinstance(v, Decimal):
                return float(v)
            if isinstance(v, (datetime, )):
                return v.isoformat()
            return v

        def factura_to_dict(f):
            if not f:
                return None
            return {
                "id": f.id,
                "idcliente": f.idcliente,
                "idfactura": f.idfactura,
                "siigo_uuid": f.siigo_uuid,
                "fecha": f.fecha.isoformat() if f.fecha else None,
                "vencimiento": f.vencimiento.isoformat() if f.vencimiento else None,
                "cliente_nombre": f.cliente_nombre,
                "vendedor": f.vendedor,
                "seller_id": f.seller_id,
                "estado": f.estado,
                "total": _dec(f.total),
                "saldo": _dec(f.saldo),
                "subtotal": _dec(f.subtotal),
                "impuestos_total": _dec(f.impuestos_total),
                "descuentos_total": _dec(f.descuentos_total),
                "pagos_total": _dec(f.pagos_total),
                "saldo_calculado": _dec(f.saldo_calculado),
                "estado_pago": f.estado_pago,
                "moneda": f.moneda,
                "medio_pago": f.medio_pago,
                "observaciones": f.observaciones,
                "metadata_created": f.metadata_created.isoformat() if f.metadata_created else None,
                "metadata_updated": f.metadata_updated.isoformat() if f.metadata_updated else None,
                "created_at": f.created_at.isoformat() if f.created_at else None,
                "customer_id": f.customer_id,
                "customer_identificacion": f.customer_identificacion,
            }

        def item_to_dict(it):
            return {
                "id": it.id,
                "factura_id": it.factura_id,
                "descripcion": it.descripcion,
                "cantidad": _dec(it.cantidad),
                "precio": _dec(it.precio),
                "impuestos": _dec(it.impuestos),
                "producto_id": it.producto_id,
                "codigo": it.codigo,
                "sku": it.sku,
                "iva_porcentaje": _dec(it.iva_porcentaje),
                "iva_valor": _dec(it.iva_valor),
                "descuento_valor": _dec(it.descuento_valor),
                "total_item": _dec(it.total_item),
            }

        resp = {
            "query_params": {"uuid": uuid, "name": name, "idcliente": idcliente},
            "siigo_tried_urls": tried,
            "siigo_raw_detail": raw_detail,  # <- lo crudo que devuelve Siigo (tal cual)
            "db_factura": factura_to_dict(factura_db),
            "db_items": [item_to_dict(i) for i in items_db],
        }
        return jsonify(resp), 200



    # Sincronización de vendedores y de centros de costos
    # Sincronización de vendedores y centros de costos
    @app.route("/siigo/sync-catalogos", methods=["POST"])
    def siigo_sync_catalogos():
        idcliente = obtener_idcliente_desde_request()
        print(f"🔹 Sync catálogos iniciado para cliente {idcliente}")

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        # Cuando este endpoint es llamado desde /siigo/sync-all,
        # NO debe crear log individual para no duplicar el historial.
        # El log general lo crea /siigo/sync-all al final.
        es_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_actual = "/siigo/sync-catalogos"
        ejecutado_en = datetime.now(timezone.utc)

        detalle_lines = []
        pasos_ok = 0
        pasos_error = 0
        endpoint_fallido = None

        def guardar_log(resultado: str, detalle: str):
            """
            Guarda historial solo cuando la ejecución es manual por módulo.
            Si viene desde sync-all, no guarda log individual para evitar duplicados.
            """
            if es_sync_all:
                return None

            try:
                log = SiigoSyncLog(
                    idcliente=idcliente,
                    fecha_programada=ejecutado_en,
                    ejecutado_en=ejecutado_en,
                    origen="manual_modulo",
                    resultado=resultado,
                    total_pasos=2,
                    pasos_ok=pasos_ok,
                    pasos_error=pasos_error,
                    endpoint_fallido=endpoint_fallido,
                    detalle=detalle,
                )
                db.session.add(log)
                db.session.commit()
                return log.id

            except Exception as log_error:
                db.session.rollback()
                raise Exception(f"Error guardando log de sincronización catálogos: {str(log_error)}")

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()

        if not cred:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ Credenciales no encontradas para el cliente."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Credenciales no encontradas",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "Credenciales no encontradas",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        access_key = dec(cred.client_secret)

        if not access_key:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ No se pudo desencriptar access_key de Siigo."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "No se pudo desencriptar access_key",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "No se pudo desencriptar access_key",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        try:
            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data["access_token"]
            headers = _headers_bearer(token)

            # -------------------------
            # Sync vendedores
            # -------------------------
            vendedores_insertados = 0

            r_v = _request_with_retries(
                "GET",
                f"{cred.base_url.rstrip('/')}/v1/users",
                headers=headers
            )

            if r_v.status_code == 200:
                db.session.query(SiigoVendedor).filter_by(idcliente=idcliente).delete(
                    synchronize_session=False
                )

                payload = r_v.json()

                # Puede ser lista o dict con results
                if isinstance(payload, list):
                    results = payload
                else:
                    results = payload.get("results", [])

                for u in results:
                    vid = u.get("id")
                    nombre = _str(
                        u.get("name")
                        or u.get("full_name")
                        or u.get("display_name")
                        or f"{u.get('first_name', '')} {u.get('last_name', '')}".strip()
                    )

                    if vid:
                        db.session.add(SiigoVendedor(
                            id=vid,
                            idcliente=idcliente,
                            nombre=nombre,
                            activo=bool(u.get("active", True)),
                            metadata_json=u
                        ))
                        vendedores_insertados += 1

                pasos_ok += 1
                detalle_lines.append(
                    f"✅ GET /v1/users -> {r_v.status_code}. "
                    f"Vendedores sincronizados: {vendedores_insertados}."
                )

            else:
                pasos_error += 1
                endpoint_fallido = endpoint_fallido or "/v1/users"
                detalle_lines.append(
                    f"❌ GET /v1/users -> {r_v.status_code}. "
                    f"Respuesta: {r_v.text[:1000]}"
                )

            # -------------------------
            # Sync centros de costo
            # -------------------------
            centros_insertados = 0

            r_cc = _request_with_retries(
                "GET",
                f"{cred.base_url.rstrip('/')}/v1/cost-centers",
                headers=headers
            )

            if r_cc.status_code == 200:
                db.session.query(SiigoCentroCosto).filter_by(idcliente=idcliente).delete(
                    synchronize_session=False
                )

                payload = r_cc.json()

                # Puede ser lista o dict con results
                if isinstance(payload, list):
                    results = payload
                else:
                    results = payload.get("results", [])

                for c in results:
                    cid = c.get("id")
                    nombre = _str(c.get("name") or c.get("description") or c.get("code"))

                    if cid:
                        db.session.add(SiigoCentroCosto(
                            id=cid,
                            idcliente=idcliente,
                            nombre=nombre,
                            codigo=_str(c.get("code")),
                            activo=bool(c.get("active", True)),
                            metadata_json=c
                        ))
                        centros_insertados += 1

                pasos_ok += 1
                detalle_lines.append(
                    f"✅ GET /v1/cost-centers -> {r_cc.status_code}. "
                    f"Centros de costo sincronizados: {centros_insertados}."
                )

            else:
                pasos_error += 1
                endpoint_fallido = endpoint_fallido or "/v1/cost-centers"
                detalle_lines.append(
                    f"❌ GET /v1/cost-centers -> {r_cc.status_code}. "
                    f"Respuesta: {r_cc.text[:1000]}"
                )

            # IMPORTANTE:
            # Se conserva el comportamiento anterior: si Siigo devuelve error en un catálogo,
            # no frenamos necesariamente toda la operación con HTTP 500.
            # Guardamos detalle e historial, pero el endpoint puede responder 200 si no hubo excepción Python.
            db.session.commit()

            detalle = "\n".join(detalle_lines)
            resultado_log = "OK" if pasos_error == 0 else "ERROR"

            try:
                log_id = guardar_log(resultado_log, detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Los catálogos se procesaron, pero falló el registro del historial.",
                    "detalle": detalle,
                    "error_log": str(log_error),
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                }), 500

            if pasos_error > 0:
                return jsonify({
                    "mensaje": "Catálogos procesados con alertas.",
                    "estado": "ERROR",
                    "log_id": log_id,
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                    "endpoint_fallido": endpoint_fallido,
                    "detalle": detalle,
                    "resumen": {
                        "vendedores_insertados": vendedores_insertados,
                        "centros_costo_insertados": centros_insertados,
                    }
                }), 200

            return jsonify({
                "mensaje": "Catálogos sincronizados correctamente",
                "estado": "OK",
                "log_id": log_id,
                "pasos_ok": pasos_ok,
                "pasos_error": pasos_error,
                "detalle": detalle,
                "resumen": {
                    "vendedores_insertados": vendedores_insertados,
                    "centros_costo_insertados": centros_insertados,
                }
            }), 200

        except Exception as e:
            db.session.rollback()

            pasos_error = max(pasos_error, 1)
            endpoint_fallido = endpoint_fallido or endpoint_actual
            detalle_lines.append(f"❌ Error general en sincronización de catálogos: {str(e)}")
            detalle = "\n".join(detalle_lines)

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": str(e),
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": str(e),
                "detalle": detalle,
                "log_id": log_id,
            }), 500      

    # ------------------------------------------
    # Catálogo de Vendedores
    # ------------------------------------------
    @app.route("/catalogos/vendedores", methods=["GET"])
    @jwt_required()
    def get_catalogo_vendedores():
        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        try:
            vendedores = SiigoVendedor.query.all()
            data = [
                {
                    "id": v.id,
                    "nombre": v.nombre,
                    "activo": v.activo,
                }
                for v in vendedores
            ]
            return jsonify(data), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500


    # ------------------------------------------
    # Catálogo de Centros de Costo
    # ------------------------------------------
    @app.route("/catalogos/centros-costo/general", methods=["GET"])
    @jwt_required()
    def get_catalogo_centros_costo():
        claims = get_jwt()
        perfilid = claims["perfilid"]
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        try:
            centros = SiigoCentroCosto.query.all()
            data = [
                {
                    "id": c.id,
                    "nombre": c.nombre,
                    "codigo": c.codigo,
                    "activo": c.activo,
                }
                for c in centros
            ]
            return jsonify(data), 200
        except Exception as e:
            return jsonify({"error": str(e)}), 500


    # Sincronización de clientes (terceros) desde Siigo
    @app.route("/siigo/sync-customers", methods=["POST"])
    def siigo_sync_customers():
        idcliente = obtener_idcliente_desde_request()
        print(f"🔹 Sync clientes iniciado para cliente {idcliente}")

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        # Cuando este endpoint es llamado desde /siigo/sync-all,
        # NO debe crear log individual para no duplicar el historial.
        # El log general lo crea /siigo/sync-all al final.
        es_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_actual = "/siigo/sync-customers"
        ejecutado_en = datetime.now(timezone.utc)

        detalle_lines = []
        pasos_ok = 0
        pasos_error = 0
        endpoint_fallido = None

        def guardar_log(resultado: str, detalle: str):
            """
            Guarda historial solo cuando la ejecución es manual por módulo.
            Si viene desde sync-all, no guarda log individual para evitar duplicados.
            """
            if es_sync_all:
                return None

            try:
                log = SiigoSyncLog(
                    idcliente=idcliente,
                    fecha_programada=ejecutado_en,
                    ejecutado_en=ejecutado_en,
                    origen="manual_modulo",
                    resultado=resultado,
                    total_pasos=1,
                    pasos_ok=pasos_ok,
                    pasos_error=pasos_error,
                    endpoint_fallido=endpoint_fallido,
                    detalle=detalle,
                )
                db.session.add(log)
                db.session.commit()
                return log.id

            except Exception as log_error:
                db.session.rollback()
                raise Exception(f"Error guardando log de sincronización clientes: {str(log_error)}")

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()

        if not cred:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ Credenciales no encontradas para el cliente."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Credenciales no encontradas",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "Credenciales no encontradas",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        access_key = dec(cred.client_secret)

        if not access_key:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ No se pudo desencriptar access_key de Siigo."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "No se pudo desencriptar access_key",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "No se pudo desencriptar access_key",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        try:
            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data["access_token"]
            headers = _headers_bearer(token)

            base_url = cred.base_url.rstrip("/")
            page = 1
            page_size = 100
            total_insertados = 0
            total_paginas = 0

            # Limpiar tabla antes de insertar
            db.session.query(SiigoCustomer).filter_by(idcliente=idcliente).delete()

            while True:
                url = f"{base_url}/v1/customers?page={page}&page_size={page_size}"
                r = _request_with_retries("GET", url, headers=headers)

                if r.status_code != 200:
                    pasos_error += 1
                    endpoint_fallido = endpoint_fallido or "/v1/customers"
                    detalle_lines.append(
                        f"❌ GET /v1/customers page={page} page_size={page_size} -> {r.status_code}. "
                        f"Respuesta: {r.text[:1000]}"
                    )
                    break

                payload = r.json()

                if isinstance(payload, list):
                    results = payload
                else:
                    results = payload.get("results", [])

                if not results:
                    detalle_lines.append(
                        f"ℹ️ GET /v1/customers page={page} -> {r.status_code}. "
                        f"Sin más clientes para sincronizar."
                    )
                    break

                total_paginas += 1

                for c in results:
                    cid = c.get("id")

                    if not cid:
                        continue

                    sc = SiigoCustomer(
                        id=cid,
                        idcliente=idcliente,
                        identification=c.get("identification"),
                        name=c.get("name"),
                        first_name=c.get("first_name"),
                        last_name=c.get("last_name"),
                        branch_office=c.get("branch_office"),
                        email=(c.get("email") if isinstance(c.get("email"), str) else None),
                        phone=(c.get("phone") if isinstance(c.get("phone"), str) else None),
                        address=c.get("address"),
                        contacts=c.get("contacts"),
                        metadata_json=c,
                    )

                    db.session.add(sc)
                    total_insertados += 1

                detalle_lines.append(
                    f"✅ GET /v1/customers page={page} page_size={page_size} -> {r.status_code}. "
                    f"Clientes recibidos en página: {len(results)}. "
                    f"Acumulado insertado: {total_insertados}."
                )

                # ¿Hay más páginas?
                links = payload.get("_links", {}) if isinstance(payload, dict) else {}

                if not links.get("next") or not links["next"].get("href"):
                    break

                page += 1

            if pasos_error == 0:
                pasos_ok = 1

            db.session.commit()

            detalle_lines.append(
                f"📊 Resumen clientes: páginas procesadas: {total_paginas}, "
                f"clientes sincronizados: {total_insertados}, errores: {pasos_error}."
            )

            detalle = "\n".join(detalle_lines)
            resultado_log = "OK" if pasos_error == 0 else "ERROR"

            try:
                log_id = guardar_log(resultado_log, detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Los clientes se procesaron, pero falló el registro del historial.",
                    "detalle": detalle,
                    "error_log": str(log_error),
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                }), 500

            if pasos_error > 0:
                return jsonify({
                    "mensaje": f"Clientes procesados con alertas. Sincronizados: {total_insertados}",
                    "estado": "ERROR",
                    "log_id": log_id,
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                    "endpoint_fallido": endpoint_fallido,
                    "detalle": detalle,
                    "resumen": {
                        "clientes_insertados": total_insertados,
                        "paginas_procesadas": total_paginas,
                    }
                }), 200

            return jsonify({
                "mensaje": f"Clientes sincronizados: {total_insertados}",
                "estado": "OK",
                "log_id": log_id,
                "pasos_ok": pasos_ok,
                "pasos_error": pasos_error,
                "detalle": detalle,
                "resumen": {
                    "clientes_insertados": total_insertados,
                    "paginas_procesadas": total_paginas,
                }
            }), 200

        except Exception as e:
            db.session.rollback()

            pasos_error = max(pasos_error, 1)
            endpoint_fallido = endpoint_fallido or endpoint_actual
            detalle_lines.append(f"❌ Error general en sincronización de clientes: {str(e)}")
            detalle = "\n".join(detalle_lines)

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": str(e),
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": str(e),
                "detalle": detalle,
                "log_id": log_id,
            }), 500


    # ==========================
    # Reportes y Dashboards 
    # ==========================


    # ============================================================
    # Endpoint reporte Financiero Ventas
    # Fuente comercial: ventas_movimientos_enriquecidos
    # ============================================================

    @app.route("/reportes/facturas_enriquecidas", methods=["GET"])
    @jwt_required()
    def get_facturas_enriquecidas():
        from sqlalchemy.sql import text

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        seller_id   = request.args.get("seller_id", type=int)
        cost_center = request.args.get("cost_center", type=int)
        cliente     = request.args.get("cliente")
        limit       = request.args.get("limit", type=int) or 5000

        incluye_impuesto = str(request.args.get("incluye_impuesto", "1")).lower() in [
            "1", "true", "si", "sí", "yes"
        ]

        incluye_nota_credito = str(request.args.get("incluye_nota_credito", "1")).lower() in [
            "1", "true", "si", "sí", "yes"
        ]

        try:
            # ============================================================
            # 1) Detalle de facturas reales
            #    Se mantiene desde siigo_facturas para no romper modales antiguos.
            # ============================================================
            wh_rows = ["f.idcliente = :idcliente"]
            params = {
                "idcliente": idcliente,
                "limit": limit,
            }

            if desde:
                wh_rows.append("f.fecha >= :desde")
                params["desde"] = desde

            if hasta:
                wh_rows.append("f.fecha <= :hasta")
                params["hasta"] = hasta

            if seller_id:
                wh_rows.append("f.seller_id = :seller_id")
                params["seller_id"] = seller_id

            if cost_center:
                wh_rows.append("f.cost_center = :cost_center")
                params["cost_center"] = cost_center

            if cliente:
                wh_rows.append("f.cliente_nombre = :cliente")
                params["cliente"] = cliente

            where_rows = " AND ".join(wh_rows)

            sql_rows = text(f"""
                SELECT DISTINCT ON (f.id)
                    f.id              AS factura_id,
                    f.idcliente,
                    f.idfactura,
                    f.fecha,
                    f.vencimiento,
                    f.cliente_nombre  AS cliente_nombre,
                    f.estado,
                    f.estado_pago,
                    COALESCE(f.subtotal, 0) AS subtotal,
                    COALESCE(f.impuestos_total, 0) AS impuestos_total,
                    COALESCE(f.total, 0) AS total,
                    COALESCE(f.pagos_total, 0) AS pagos_total,
                    COALESCE(f.saldo, 0) AS saldo,
                    COALESCE(f.saldo_calculado, f.saldo, 0) AS saldo_calculado,
                    f.medio_pago,
                    f.observaciones,
                    f.public_url,
                    f.cost_center,
                    COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre,
                    cc.codigo AS centro_costo_codigo,
                    f.seller_id,
                    v.nombre AS vendedor_nombre
                FROM siigo_facturas f
                LEFT JOIN siigo_vendedores v
                    ON v.id = f.seller_id
                AND v.idcliente = f.idcliente
                LEFT JOIN siigo_centros_costo cc
                    ON cc.id = f.cost_center
                AND cc.idcliente = f.idcliente
                LEFT JOIN siigo_customers c
                    ON c.id::text = f.customer_id::text
                AND c.idcliente = f.idcliente
                WHERE {where_rows}
                ORDER BY f.id DESC
                LIMIT :limit
            """)

            rows = [dict(r) for r in db.session.execute(sql_rows, params).mappings().all()]

            # ============================================================
            # 2) KPIs / series / top clientes
            #    Fuente oficial para reporte comercial: ventas_movimientos_enriquecidos
            # ============================================================
            wh_mov = ["m.idcliente = :idcliente"]

            if desde:
                wh_mov.append("m.fecha >= :desde")

            if hasta:
                wh_mov.append("m.fecha <= :hasta")

            if seller_id:
                wh_mov.append("m.seller_id = :seller_id")

            if cost_center:
                wh_mov.append("m.cost_center = :cost_center")

            if cliente:
                wh_mov.append("m.cliente_nombre = :cliente")

            if not incluye_nota_credito:
                wh_mov.append("m.tipo_movimiento = 'FACTURA'")

            where_mov = " AND ".join(wh_mov)

            # Campo principal igual a Siigo:
            # - incluye_impuesto = true  => total
            # - incluye_impuesto = false => subtotal
            campo_valor = "total" if incluye_impuesto else "subtotal"

            cte_common = f"""
                WITH comp AS (
                    SELECT
                        m.*,

                        COALESCE((
                            SELECT SUM((elem->>'value')::numeric)
                            FROM jsonb_array_elements(m.retenciones) elem
                            WHERE jsonb_typeof(m.retenciones) = 'array'
                            AND LOWER(elem->>'type') LIKE '%autorretencion%'
                        ), 0) AS autorretencion,

                        COALESCE((
                            SELECT SUM((elem->>'value')::numeric)
                            FROM jsonb_array_elements(m.retenciones) elem
                            WHERE jsonb_typeof(m.retenciones) = 'array'
                            AND (elem->>'type') IS NOT NULL
                            AND LOWER(elem->>'type') NOT LIKE '%autorretencion%'
                        ), 0) AS retenciones_sin_auto,

                        COALESCE(m.subtotal, 0) AS subtotal_b,
                        COALESCE(m.impuestos_total, 0) AS impuestos_b,
                        COALESCE(m.total, 0) AS total_b,
                        COALESCE(m.{campo_valor}, 0) AS valor_siigo_b,
                        COALESCE(m.saldo, 0) AS saldo_b

                    FROM ventas_movimientos_enriquecidos m
                    WHERE {where_mov}
                ),
                ajuste AS (
                    SELECT
                        date_trunc('month', fecha)::date AS mes,
                        tipo_movimiento,
                        subtotal_b,
                        impuestos_b,
                        total_b,
                        valor_siigo_b,
                        autorretencion,
                        retenciones_sin_auto,

                        CASE
                            WHEN tipo_movimiento = 'FACTURA' THEN saldo_b
                            ELSE 0
                        END AS saldo_b,

                        CASE
                            WHEN tipo_movimiento = 'FACTURA' THEN GREATEST(total_b - saldo_b, 0)
                            ELSE 0
                        END AS pagado_b,

                        CASE
                            WHEN tipo_movimiento = 'FACTURA' THEN GREATEST(saldo_b, 0)
                            ELSE 0
                        END AS pendiente_b

                    FROM comp
                )
            """

            # ============================================================
            # KPIs principales
            # ============================================================
            sql_kpis = text(cte_common + """
                SELECT
                    -- Valor principal de ventas netas según modo Siigo
                    COALESCE(SUM(valor_siigo_b), 0) AS subtotal,

                    -- Impuesto neto comercial:
                    -- ventas con impuesto - ventas sin impuesto
                    COALESCE(SUM(total_b) - SUM(subtotal_b), 0) AS impuestos,

                    COALESCE(SUM(autorretencion), 0) AS autorretencion,

                    -- Facturas emitidas según modo actual:
                    -- con impuesto => total
                    -- sin impuesto => subtotal
                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN valor_siigo_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_facturado,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN subtotal_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS facturas_emitidas_sin_impuesto,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN total_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS facturas_emitidas_con_impuesto,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'NOTA_CREDITO'
                                THEN subtotal_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS notas_credito_sin_impuesto,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'NOTA_CREDITO'
                                THEN total_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS notas_credito_con_impuesto,

                    -- Nota crédito según modo actual, se devuelve negativa para auditoría
                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'NOTA_CREDITO'
                                THEN valor_siigo_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS notas_credito,

                    -- Nota crédito positiva para mostrar en frontend
                    ABS(
                        COALESCE(
                            SUM(
                                CASE
                                    WHEN tipo_movimiento = 'NOTA_CREDITO'
                                    THEN valor_siigo_b
                                    ELSE 0
                                END
                            ),
                            0
                        )
                    ) AS notas_credito_abs,

                    COALESCE(SUM(retenciones_sin_auto), 0) AS retenciones,

                    COALESCE(
                        SUM(total_b - (autorretencion + retenciones_sin_auto)),
                        0
                    ) AS total_utilizable,

                    COALESCE(SUM(pagado_b), 0) AS pagado,
                    COALESCE(SUM(pendiente_b), 0) AS pendiente,

                    COALESCE(SUM(subtotal_b), 0) AS ventas_sin_impuesto,
                    COALESCE(SUM(total_b), 0) AS ventas_con_impuesto
                FROM ajuste
            """)

            kpis = dict(db.session.execute(sql_kpis, params).mappings().first() or {})

            # ============================================================
            # Series mensuales
            # ============================================================
            sql_series = text(cte_common + """
                SELECT
                    mes::text AS periodo,
                    TO_CHAR(mes, 'Mon YYYY') AS label,

                    -- Ventas netas según modo Siigo
                    COALESCE(SUM(valor_siigo_b), 0) AS subtotal,

                    -- Impuesto neto real del mes
                    COALESCE(SUM(total_b) - SUM(subtotal_b), 0) AS impuestos,

                    COALESCE(SUM(autorretencion), 0) AS autorretencion,

                    -- Facturas emitidas según modo actual
                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN valor_siigo_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS total_facturado,

                    -- Notas crédito según modo actual, positiva para tooltip/gráfica
                    ABS(
                        COALESCE(
                            SUM(
                                CASE
                                    WHEN tipo_movimiento = 'NOTA_CREDITO'
                                    THEN valor_siigo_b
                                    ELSE 0
                                END
                            ),
                            0
                        )
                    ) AS notas_credito,

                    -- Notas crédito negativa para auditoría
                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'NOTA_CREDITO'
                                THEN valor_siigo_b
                                ELSE 0
                            END
                        ),
                        0
                    ) AS notas_credito_neto,

                    COALESCE(SUM(retenciones_sin_auto), 0) AS retenciones,

                    COALESCE(
                        SUM(total_b - (autorretencion + retenciones_sin_auto)),
                        0
                    ) AS total_utilizable,

                    COALESCE(SUM(pagado_b), 0) AS pagado,
                    COALESCE(SUM(pendiente_b), 0) AS pendiente,

                    COALESCE(SUM(subtotal_b), 0) AS ventas_sin_impuesto,
                    COALESCE(SUM(total_b), 0) AS ventas_con_impuesto
                FROM ajuste
                GROUP BY mes
                ORDER BY mes
            """)

            series = [dict(r) for r in db.session.execute(sql_series, params).mappings().all()]

            # ============================================================
            # Estados pagado / pendiente
            # ============================================================
            sql_estados = text(cte_common + """
                SELECT 'Pagado' AS estado, COALESCE(SUM(pagado_b), 0) AS valor FROM ajuste
                UNION ALL
                SELECT 'Pendiente', COALESCE(SUM(pendiente_b), 0) AS valor FROM ajuste
            """)

            estados = [dict(r) for r in db.session.execute(sql_estados, params).mappings().all()]

            # ============================================================
            # Top clientes según modo Siigo actual
            # ============================================================
            sql_top_clientes = text(f"""
                SELECT
                    m.cliente_nombre AS cliente,
                    COALESCE(SUM(m.{campo_valor}), 0) AS total,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN m.tipo_movimiento = 'FACTURA'
                                THEN m.{campo_valor}
                                ELSE 0
                            END
                        ),
                        0
                    ) AS facturas_emitidas,

                    ABS(
                        COALESCE(
                            SUM(
                                CASE
                                    WHEN m.tipo_movimiento = 'NOTA_CREDITO'
                                    THEN m.{campo_valor}
                                    ELSE 0
                                END
                            ),
                            0
                        )
                    ) AS notas_credito

                FROM ventas_movimientos_enriquecidos m
                WHERE {where_mov}
                GROUP BY m.cliente_nombre
                ORDER BY total DESC
                LIMIT 5
            """)

            top_clientes = [dict(r) for r in db.session.execute(sql_top_clientes, params).mappings().all()]

            return jsonify({
                "rows": rows,
                "kpis": kpis,
                "series": series,
                "estados": estados,
                "top_clientes": top_clientes,
                "count": len(rows),
                "config": {
                    "incluye_impuesto": incluye_impuesto,
                    "incluye_nota_credito": incluye_nota_credito,
                    "campo_valor": campo_valor,
                    "fuente_kpis": "ventas_movimientos_enriquecidos",
                    "fuente_rows": "siigo_facturas",
                    "logica": "ventas_netas = facturas_emitidas - notas_credito"
                }
            })

        except Exception as e:
            return jsonify({"error": str(e)}), 500


    # ============================================================
    # Endpoint detalle de movimientos comerciales
    # Para modales de barras, clientes y conciliación contra Siigo
    # Fuente: ventas_movimientos_enriquecidos
    # ============================================================

    @app.route("/reportes/ventas_movimientos_detalle", methods=["GET"])
    @jwt_required()
    def ventas_movimientos_detalle():
        from sqlalchemy.sql import text

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        seller_id   = request.args.get("seller_id", type=int)
        cost_center = request.args.get("cost_center", type=int)
        cliente     = request.args.get("cliente")
        tipo        = request.args.get("tipo")  # FACTURA / NOTA_CREDITO opcional
        limit       = request.args.get("limit", type=int) or 10000

        incluye_impuesto = str(request.args.get("incluye_impuesto", "1")).lower() in [
            "1", "true", "si", "sí", "yes"
        ]

        incluye_nota_credito = str(request.args.get("incluye_nota_credito", "1")).lower() in [
            "1", "true", "si", "sí", "yes"
        ]

        campo_valor = "total" if incluye_impuesto else "subtotal"

        try:
            wh = ["m.idcliente = :idcliente"]
            params = {
                "idcliente": idcliente,
                "limit": limit,
            }

            if desde:
                wh.append("m.fecha >= :desde")
                params["desde"] = desde

            if hasta:
                wh.append("m.fecha <= :hasta")
                params["hasta"] = hasta

            if seller_id:
                wh.append("m.seller_id = :seller_id")
                params["seller_id"] = seller_id

            if cost_center:
                wh.append("m.cost_center = :cost_center")
                params["cost_center"] = cost_center

            if cliente:
                wh.append("m.cliente_nombre = :cliente")
                params["cliente"] = cliente

            if tipo:
                wh.append("m.tipo_movimiento = :tipo")
                params["tipo"] = tipo

            if not incluye_nota_credito:
                wh.append("m.tipo_movimiento = 'FACTURA'")

            where_clause = " AND ".join(wh)

            sql = text(f"""
                WITH base AS (
                    SELECT
                        m.movimiento_id,
                        m.idcliente,
                        m.documento,
                        m.tipo_movimiento,
                        m.fecha,
                        m.vencimiento,
                        m.cliente_nombre,
                        m.estado,
                        m.estado_pago,
                        COALESCE(m.subtotal, 0) AS subtotal,
                        COALESCE(m.impuestos_total, 0) AS impuestos_total,
                        COALESCE(m.total, 0) AS total,
                        COALESCE(m.{campo_valor}, 0) AS valor,
                        COALESCE(m.pagos_total, 0) AS pagos_total,
                        COALESCE(m.saldo, 0) AS saldo,
                        m.cost_center,
                        m.centro_costo_nombre,
                        m.centro_costo_codigo,
                        m.seller_id,
                        m.vendedor_nombre,
                        m.public_url,
                        m.documento_afectado,

                        CASE
                            WHEN m.tipo_movimiento = 'FACTURA'
                            THEN GREATEST(COALESCE(m.total, 0) - COALESCE(m.saldo, 0), 0)
                            ELSE 0
                        END AS pagado,

                        CASE
                            WHEN m.tipo_movimiento = 'FACTURA'
                            THEN GREATEST(COALESCE(m.saldo, 0), 0)
                            ELSE 0
                        END AS pendiente

                    FROM ventas_movimientos_enriquecidos m
                    WHERE {where_clause}
                )
                SELECT
                    *
                FROM base
                ORDER BY fecha DESC, tipo_movimiento ASC, documento DESC
                LIMIT :limit
            """)

            rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]

            sql_resumen = text(f"""
                WITH base AS (
                    SELECT
                        m.tipo_movimiento,
                        COALESCE(m.subtotal, 0) AS subtotal,
                        COALESCE(m.total, 0) AS total,
                        COALESCE(m.{campo_valor}, 0) AS valor,
                        COALESCE(m.saldo, 0) AS saldo
                    FROM ventas_movimientos_enriquecidos m
                    WHERE {where_clause}
                )
                SELECT
                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN valor
                                ELSE 0
                            END
                        ),
                        0
                    ) AS facturas_emitidas,

                    ABS(
                        COALESCE(
                            SUM(
                                CASE
                                    WHEN tipo_movimiento = 'NOTA_CREDITO'
                                    THEN valor
                                    ELSE 0
                                END
                            ),
                            0
                        )
                    ) AS notas_credito,

                    COALESCE(SUM(valor), 0) AS ventas_netas,

                    COALESCE(SUM(total), 0) AS ventas_con_impuesto,
                    COALESCE(SUM(subtotal), 0) AS ventas_sin_impuesto,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN GREATEST(total - saldo, 0)
                                ELSE 0
                            END
                        ),
                        0
                    ) AS pagado,

                    COALESCE(
                        SUM(
                            CASE
                                WHEN tipo_movimiento = 'FACTURA'
                                THEN GREATEST(saldo, 0)
                                ELSE 0
                            END
                        ),
                        0
                    ) AS pendiente,

                    COUNT(*) AS total_movimientos,

                    COUNT(*) FILTER (WHERE tipo_movimiento = 'FACTURA') AS total_facturas,
                    COUNT(*) FILTER (WHERE tipo_movimiento = 'NOTA_CREDITO') AS total_notas_credito

                FROM base
            """)

            resumen = dict(db.session.execute(sql_resumen, params).mappings().first() or {})

            return jsonify({
                "rows": rows,
                "resumen": resumen,
                "count": len(rows),
                "config": {
                    "incluye_impuesto": incluye_impuesto,
                    "incluye_nota_credito": incluye_nota_credito,
                    "campo_valor": campo_valor,
                    "fuente": "ventas_movimientos_enriquecidos",
                    "logica": "ventas_netas = facturas_emitidas - notas_credito"
                }
            })

        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # ================================
    # ENDPOINT: Facturas por mes (para modal del dashboard)
    # ================================
    @app.route("/reportes/facturas_detalle_mes", methods=["GET"])
    @jwt_required()
    def facturas_detalle_mes():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        # Superadmin permite ?idcliente=
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        seller_id   = request.args.get("seller_id", type=int)
        cost_center = request.args.get("cost_center", type=int)
        cliente     = request.args.get("cliente")

        wh = ["f.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde:
            wh.append("f.fecha >= :desde")
            params["desde"] = desde

        if hasta:
            wh.append("f.fecha <= :hasta")
            params["hasta"] = hasta

        if seller_id:
            wh.append("f.seller_id = :seller_id")
            params["seller_id"] = seller_id

        if cost_center:
            wh.append("f.cost_center = :cost_center")
            params["cost_center"] = cost_center

        if cliente:
            wh.append("f.cliente_nombre = :cliente")
            params["cliente"] = cliente

        where_clause = " AND ".join(wh)

        sql = text(f"""
            SELECT
                f.factura_id AS id,
                f.idfactura,
                f.fecha,
                f.vencimiento,
                f.cliente_nombre,
                f.subtotal,
                f.impuestos_total AS impuestos,
                f.total,
                f.pagos_total AS pagado,
                f.saldo,
                f.cost_center,
                COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre,
                v.nombre AS vendedor_nombre,
                f.public_url,
                f.retenciones,
                f.medio_pago,
                f.estado,
                f.estado_pago
            FROM facturas_enriquecidas f
            LEFT JOIN siigo_centros_costo cc
                ON cc.id = f.cost_center
                AND cc.idcliente = f.idcliente
            LEFT JOIN siigo_vendedores v
                ON v.id = f.seller_id
                AND v.idcliente = f.idcliente
            WHERE {where_clause}
            ORDER BY f.fecha DESC
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify({"rows": rows})

    
    # Modal para ver facturas por cliente en pagina Finaciero/ventas
    @app.route("/reportes/facturas_por_cliente", methods=["GET"])
    @jwt_required()
    def facturas_por_cliente_financieroventas():
        from sqlalchemy.sql import text

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        cliente_nombre = request.args.get("cliente")
        if not cliente_nombre:
            return jsonify({"rows": []})

        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        seller_id   = request.args.get("seller_id", type=int)
        cost_center = request.args.get("cost_center", type=int)

        wh = ["f.idcliente = :idcliente", "f.cliente_nombre = :cliente"]
        params = {"idcliente": idcliente, "cliente": cliente_nombre}

        if desde:
            wh.append("f.fecha >= :desde")
            params["desde"] = desde

        if hasta:
            wh.append("f.fecha <= :hasta")
            params["hasta"] = hasta

        if seller_id:
            wh.append("f.seller_id = :seller_id")
            params["seller_id"] = seller_id

        if cost_center:
            wh.append("f.cost_center = :cost_center")
            params["cost_center"] = cost_center

        sql = text(f"""
            SELECT
                f.idfactura,
                f.fecha,
                f.vencimiento,
                f.cliente_nombre,
                f.subtotal,
                f.impuestos_total AS impuestos,
                f.total,
                f.pagos_total AS pagado,
                f.saldo,
                COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre,
                v.nombre AS vendedor_nombre,
                f.public_url
            FROM facturas_enriquecidas f
            LEFT JOIN siigo_centros_costo cc
                ON cc.id = f.cost_center
                AND cc.idcliente = f.idcliente
            LEFT JOIN siigo_vendedores v
                ON v.id = f.seller_id
                AND v.idcliente = f.idcliente       
            WHERE { " AND ".join(wh) }
            ORDER BY f.fecha DESC
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify({"rows": rows})



    # Modal facturas pagadas y pendientes en pagina financiero/ventas
    # Modal facturas por estado (Pagado / Pendiente) en página financiero/ventas
    @app.route("/reportes/facturas_por_estado", methods=["GET"])
    @jwt_required()
    def facturas_por_estado():
        from sqlalchemy.sql import text

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        # Superadmin puede forzar ?idcliente=
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        estado = request.args.get("estado")  # "Pagado" o "Pendiente"
        if not estado:
            return jsonify({"rows": []})

        # ---- Filtros compartidos (igual que en facturas_enriquecidas) ----
        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        seller_id   = request.args.get("seller_id", type=int)
        cost_center = request.args.get("cost_center", type=int)
        cliente     = request.args.get("cliente")

        wh = ["f.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde:
            wh.append("f.fecha >= :desde")
            params["desde"] = desde

        if hasta:
            wh.append("f.fecha <= :hasta")
            params["hasta"] = hasta

        if seller_id:
            wh.append("f.seller_id = :seller_id")
            params["seller_id"] = seller_id

        if cost_center:
            wh.append("f.cost_center = :cost_center")
            params["cost_center"] = cost_center

        if cliente:
            wh.append("f.cliente_nombre = :cliente")
            params["cliente"] = cliente

        where_clause = " AND ".join(wh)

        # ---- CTE basado en FACTURAS_ENRIQUECIDAS (MISMA BASE QUE EL REPORTE) ----
        cte = f"""
            WITH base AS (
                SELECT
                    f.idcliente,
                    f.idfactura,
                    f.fecha,
                    f.vencimiento,
                    f.cliente_nombre,
                    COALESCE(f.subtotal, 0)          AS subtotal,
                    COALESCE(f.impuestos_total, 0)   AS impuestos,
                    COALESCE(f.total, 0)             AS total,
                    COALESCE(f.saldo, 0)             AS saldo,
                    COALESCE(f.total, 0) - COALESCE(f.saldo, 0) AS pagado,
                    f.public_url,
                    f.cost_center,
                    f.seller_id
                FROM facturas_enriquecidas f
                WHERE {where_clause}
            ),
            calc AS (
                SELECT
                    *,
                    saldo AS pendiente
                FROM base
            )
        """

        # ---- Filtro por estado de pago real (coherente con el pie) ----
        # El pie hace:
        #  - Pagado   = SUM(total - saldo)
        #  - Pendiente = SUM(saldo)
        #
        # Para el modal:
        #  - "Pagado"   => facturas donde pagado > 0
        #  - "Pendiente" => facturas donde pendiente > 0
        if estado.lower() == "pagado":
            filtro_estado = "pagado > 0"
        else:
            filtro_estado = "pendiente > 0"

        sql = text(cte + f"""
            SELECT
                c.idfactura,
                c.fecha,
                c.vencimiento,
                c.cliente_nombre,
                c.subtotal,
                c.impuestos,      -- alias correcto para que el frontend use f.impuestos
                c.total,
                c.pagado,
                c.pendiente,
                c.saldo,  -- <-- NECESARIO PARA EL FRONTEND
                c.public_url,
                COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre,
                v.nombre AS vendedor_nombre
            FROM calc c
            LEFT JOIN siigo_centros_costo cc
                ON cc.id = c.cost_center
            AND cc.idcliente = c.idcliente
            LEFT JOIN siigo_vendedores v
                ON v.id = c.seller_id
            AND v.idcliente = c.idcliente
            WHERE {filtro_estado}
            ORDER BY c.fecha DESC
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify({"rows": rows})




    @app.route("/catalogos/clientes-facturas", methods=["GET"])
    @jwt_required()
    def catalogo_clientes_facturas():
        claims = get_jwt()
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if claims.get("perfilid") == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        try:
            wh = ["f.idcliente = :idcliente"]
            params = {"idcliente": idcliente}
            if desde:
                wh.append("f.fecha >= :desde")
                params["desde"] = desde
            if hasta:
                wh.append("f.fecha <= :hasta")
                params["hasta"] = hasta
            where_clause = " AND ".join(wh)

            sql = text(f"""
                SELECT DISTINCT
                    f.cliente_nombre AS id,
                    f.cliente_nombre AS nombre
                FROM facturas_enriquecidas f
                WHERE {where_clause}
                AND TRIM(COALESCE(f.cliente_nombre, '')) <> ''
                ORDER BY nombre
            """)
            rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
            return jsonify(rows)


        except Exception as e:
            return jsonify({"error": str(e)}), 500


    @app.route("/siigo/sync-pendientes", methods=["GET"])
    @jwt_required()
    def contar_pendientes():
        idcliente = get_jwt_identity()
        cantidad = contar_facturas_pendientes(idcliente)
        return jsonify({"pendientes": cantidad})



    # --- ENDPOINT: Clientes Insights (optimizado, enriquecido y alineado con Ingresos por Ventas) ---
    # ============================================================
    # ENDPOINT: Clientes Insights
    # Alineado con Ingresos por Ventas:
    # - Ventas comerciales: ventas_movimientos_enriquecidos
    # - Cartera / pagos / vencimientos: siigo_facturas
    # ============================================================

    @app.route("/reportes/analisis_clientes", methods=["GET"])
    @jwt_required()
    def get_clientes_insights():
        from sqlalchemy.sql import text
        from datetime import datetime
        from collections import defaultdict
        import re

        def normalizar_cliente(nombre: str) -> str:
            nombre = (nombre or "").strip().lower()
            nombre = nombre.replace(".", "").replace(",", "")
            nombre = re.sub(r"\s+", " ", nombre)
            return nombre

        def money_float(valor):
            try:
                return float(valor or 0)
            except Exception:
                return 0.0

        def money_fmt(valor, decimales=0):
            valor = money_float(valor)
            return f"$ {valor:,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        cliente = request.args.get("cliente")
        cost_center = request.args.get("cost_center", type=int)
        filtro_estado = request.args.get("estado")
        limit_facturas_raw = request.args.get("limit_facturas", default="8")

        mostrar_todas_facturas = False

        if str(limit_facturas_raw).lower().strip() in ["all", "todos", "todas", "0"]:
            mostrar_todas_facturas = True
            limit_facturas = 999999
        else:
            try:
                limit_facturas = int(limit_facturas_raw)
            except Exception:
                limit_facturas = 8

            if limit_facturas <= 0:
                limit_facturas = 8

            # Dejamos un límite de seguridad solo cuando NO se pide all.
            if limit_facturas > 500:
                limit_facturas = 500

        fecha_desde_val = validar_fecha(desde) if desde else None
        fecha_hasta_val = validar_fecha(hasta) if hasta else None

        try:
            # ============================================================
            # 1) Filtros comerciales
            # Fuente: ventas_movimientos_enriquecidos
            # Aquí viven ventas netas, facturas emitidas y notas crédito.
            # ============================================================
            wh_mov = ["m.idcliente = :idcliente"]
            params_mov = {"idcliente": idcliente}

            if fecha_desde_val:
                wh_mov.append("m.fecha >= :desde")
                params_mov["desde"] = fecha_desde_val

            if fecha_hasta_val:
                wh_mov.append("m.fecha < (CAST(:hasta AS date) + INTERVAL '1 day')")
                params_mov["hasta"] = fecha_hasta_val

            if cliente:
                wh_mov.append("LOWER(TRIM(m.cliente_nombre)) = LOWER(TRIM(:cliente))")
                params_mov["cliente"] = cliente

            if cost_center:
                wh_mov.append("m.cost_center = :cost_center")
                params_mov["cost_center"] = cost_center

            where_mov = " AND ".join(wh_mov)

            # ============================================================
            # 2) Filtros de cartera
            # Fuente: siigo_facturas, pero enriqueciendo cliente y centro.
            # Aquí viven saldos, vencimientos, pagos y facturas recientes.
            # ============================================================
            estado_sql = """
                CASE
                    WHEN ROUND(COALESCE(fb.saldo, 0)::numeric, 2) = 0 THEN 'pagado'
                    WHEN fb.vencimiento IS NOT NULL
                        AND fb.vencimiento < CURRENT_DATE
                        AND COALESCE(fb.saldo, 0) > 0 THEN 'vencido'
                    WHEN fb.vencimiento IS NOT NULL
                        AND fb.vencimiento >= CURRENT_DATE
                        AND fb.vencimiento <= CURRENT_DATE + INTERVAL '5 days'
                        AND COALESCE(fb.saldo, 0) > 0 THEN 'alerta'
                    ELSE 'sano'
                END
            """

            wh_fac = ["fb.idcliente = :idcliente"]
            params_fac = {
                "idcliente": idcliente,
                "limit_facturas": limit_facturas,
                "mostrar_todas_facturas": mostrar_todas_facturas,
            }

            if fecha_desde_val:
                wh_fac.append("fb.fecha >= :desde")
                params_fac["desde"] = fecha_desde_val

            if fecha_hasta_val:
                wh_fac.append("fb.fecha < (CAST(:hasta AS date) + INTERVAL '1 day')")
                params_fac["hasta"] = fecha_hasta_val

            if cliente:
                wh_fac.append("LOWER(TRIM(fb.cliente_nombre_ok)) = LOWER(TRIM(:cliente))")
                params_fac["cliente"] = cliente

            if cost_center:
                wh_fac.append("fb.cost_center = :cost_center")
                params_fac["cost_center"] = cost_center

            filtro_estado_normalizado = None
            if filtro_estado:
                filtro_estado_normalizado = filtro_estado.lower().strip()
                if filtro_estado_normalizado in ["pagado", "vencido", "alerta", "sano"]:
                    wh_fac.append(f"({estado_sql}) = :estado")
                    params_fac["estado"] = filtro_estado_normalizado
                else:
                    filtro_estado_normalizado = None

            where_fac = " AND ".join(wh_fac)

            facturas_base_cte = """
                WITH facturas_base AS (
                    SELECT
                        f.*,
                        regexp_replace(
                            COALESCE(
                                NULLIF(TRIM(BOTH '"' FROM f.cliente_nombre), ''),
                                NULLIF(TRIM(BOTH '"' FROM c.name), ''),
                                'Desconocido'
                            ),
                            '[\\{{\\}}\\[\\]\\"]',
                            '',
                            'g'
                        ) AS cliente_nombre_ok,
                        COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre_ok
                    FROM siigo_facturas f
                    LEFT JOIN siigo_customers c
                        ON c.id::text = f.customer_id::text
                    AND c.idcliente = f.idcliente
                    LEFT JOIN siigo_centros_costo cc
                        ON cc.id = f.cost_center
                    AND cc.idcliente = f.idcliente
                )
            """

            # ============================================================
            # 3) Resumen comercial global
            # ============================================================
            sql_resumen_comercial = text(f"""
                SELECT
                    COUNT(DISTINCT m.cliente_nombre) AS clientes_facturados,

                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'FACTURA') AS cantidad_facturas,
                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'NOTA_CREDITO') AS cantidad_notas_credito,

                    COALESCE(SUM(m.total), 0) AS ventas_netas,
                    COALESCE(SUM(m.subtotal), 0) AS ventas_sin_impuesto,
                    COALESCE(SUM(m.total) - SUM(m.subtotal), 0) AS impuestos,

                    COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'FACTURA' THEN m.total
                            ELSE 0
                        END
                    ), 0) AS total_facturado_siigo,

                    ABS(COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'NOTA_CREDITO' THEN m.total
                            ELSE 0
                        END
                    ), 0)) AS notas_credito

                FROM ventas_movimientos_enriquecidos m
                WHERE {where_mov}
            """)

            resumen_comercial = dict(
                db.session.execute(sql_resumen_comercial, params_mov).mappings().first() or {}
            )

            # ============================================================
            # 4) Resumen cartera global
            # ============================================================
            sql_resumen_cartera = text(facturas_base_cte + f"""
                SELECT
                    COALESCE(SUM(COALESCE(fb.total, 0) - COALESCE(fb.saldo, 0)), 0) AS total_pagado,
                    COALESCE(SUM(COALESCE(fb.saldo, 0)), 0) AS saldo_pendiente,

                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(fb.saldo, 0) > 0
                                AND fb.vencimiento < CURRENT_DATE
                            THEN fb.saldo
                            ELSE 0
                        END
                    ), 0) AS saldo_vencido,

                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(fb.saldo, 0) > 0
                                AND (fb.vencimiento IS NULL OR fb.vencimiento >= CURRENT_DATE)
                            THEN fb.saldo
                            ELSE 0
                        END
                    ), 0) AS saldo_por_vencer,

                    COUNT(*) AS cantidad_facturas_cartera

                FROM facturas_base fb
                WHERE {where_fac}
            """)

            resumen_cartera = dict(
                db.session.execute(sql_resumen_cartera, params_fac).mappings().first() or {}
            )

            ventas_netas = money_float(resumen_comercial.get("ventas_netas"))
            ventas_sin_impuesto = money_float(resumen_comercial.get("ventas_sin_impuesto"))
            impuestos = money_float(resumen_comercial.get("impuestos"))
            total_facturado_siigo = money_float(resumen_comercial.get("total_facturado_siigo"))
            notas_credito = money_float(resumen_comercial.get("notas_credito"))

            total_pagado = money_float(resumen_cartera.get("total_pagado"))
            saldo_pendiente = money_float(resumen_cartera.get("saldo_pendiente"))
            saldo_vencido = money_float(resumen_cartera.get("saldo_vencido"))
            saldo_por_vencer = money_float(resumen_cartera.get("saldo_por_vencer"))

            pct_pagado = (total_pagado / total_facturado_siigo * 100) if total_facturado_siigo else 0
            pct_vencido = (saldo_vencido / saldo_pendiente * 100) if saldo_pendiente else 0

            resumen = {
                "clientes_facturados": int(resumen_comercial.get("clientes_facturados") or 0),
                "cantidad_facturas": int(resumen_comercial.get("cantidad_facturas") or 0),
                "cantidad_notas_credito": int(resumen_comercial.get("cantidad_notas_credito") or 0),

                "ventas_netas": ventas_netas,
                "ventas_sin_impuesto": ventas_sin_impuesto,
                "impuestos": impuestos,
                "total_facturado_siigo": total_facturado_siigo,
                "notas_credito": notas_credito,

                # Compatibilidad con frontend anterior.
                "total_facturado": ventas_netas,

                "total_pagado": total_pagado,
                "saldo_pendiente": saldo_pendiente,
                "saldo_vencido": saldo_vencido,
                "saldo_por_vencer": saldo_por_vencer,

                "pct_pagado": round(pct_pagado, 2),
                "pct_vencido": round(pct_vencido, 2),

                "ventas_netas_str": money_fmt(ventas_netas),
                "ventas_sin_impuesto_str": money_fmt(ventas_sin_impuesto),
                "impuestos_str": money_fmt(impuestos),
                "total_facturado_siigo_str": money_fmt(total_facturado_siigo),
                "notas_credito_str": money_fmt(notas_credito),

                # Compatibilidad con frontend anterior.
                "total_facturado_str": money_fmt(ventas_netas),

                "total_pagado_str": money_fmt(total_pagado),
                "saldo_pendiente_str": money_fmt(saldo_pendiente),
                "saldo_vencido_str": money_fmt(saldo_vencido),
                "saldo_por_vencer_str": money_fmt(saldo_por_vencer),
            }

            # ============================================================
            # 5) Clientes comerciales
            # ============================================================
            sql_clientes_comercial = text(f"""
                SELECT
                    m.cliente_nombre AS cliente,

                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'FACTURA') AS cantidad_facturas,
                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'NOTA_CREDITO') AS cantidad_notas_credito,
                    COUNT(DISTINCT m.cost_center) AS cantidad_centros_costo,

                    COALESCE(SUM(m.total), 0) AS ventas_netas,
                    COALESCE(SUM(m.subtotal), 0) AS ventas_sin_impuesto,
                    COALESCE(SUM(m.total) - SUM(m.subtotal), 0) AS impuestos,

                    COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'FACTURA' THEN m.total
                            ELSE 0
                        END
                    ), 0) AS total_facturado_siigo,

                    ABS(COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'NOTA_CREDITO' THEN m.total
                            ELSE 0
                        END
                    ), 0)) AS notas_credito,

                    MAX(m.fecha) AS ultima_factura

                FROM ventas_movimientos_enriquecidos m
                WHERE {where_mov}
                GROUP BY m.cliente_nombre
            """)

            clientes_comercial_rows = [
                dict(r) for r in db.session.execute(sql_clientes_comercial, params_mov).mappings().all()
            ]

            clientes_comercial = {}
            for r in clientes_comercial_rows:
                key = normalizar_cliente(r.get("cliente"))
                clientes_comercial[key] = r

            # ============================================================
            # 6) Clientes cartera
            # ============================================================
            sql_clientes_cartera = text(facturas_base_cte + f"""
                SELECT
                    fb.cliente_nombre_ok AS cliente,

                    COALESCE(SUM(COALESCE(fb.total, 0) - COALESCE(fb.saldo, 0)), 0) AS total_pagado,
                    COALESCE(SUM(COALESCE(fb.saldo, 0)), 0) AS saldo_pendiente,

                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(fb.saldo, 0) > 0
                                AND fb.vencimiento < CURRENT_DATE
                            THEN fb.saldo
                            ELSE 0
                        END
                    ), 0) AS saldo_vencido,

                    COALESCE(SUM(
                        CASE
                            WHEN COALESCE(fb.saldo, 0) > 0
                                AND (fb.vencimiento IS NULL OR fb.vencimiento >= CURRENT_DATE)
                            THEN fb.saldo
                            ELSE 0
                        END
                    ), 0) AS saldo_por_vencer

                FROM facturas_base fb
                WHERE {where_fac}
                GROUP BY fb.cliente_nombre_ok
            """)

            clientes_cartera_rows = [
                dict(r) for r in db.session.execute(sql_clientes_cartera, params_fac).mappings().all()
            ]

            clientes_cartera = {}
            for r in clientes_cartera_rows:
                key = normalizar_cliente(r.get("cliente"))
                clientes_cartera[key] = r

            # ============================================================
            # 7) Centros de costo comerciales por cliente
            # ============================================================
            sql_cc_comercial = text(f"""
                SELECT
                    m.cliente_nombre,
                    COALESCE(m.centro_costo_nombre, 'Sin centro de costo') AS centro_costo_nombre,
                    m.cost_center,

                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'FACTURA') AS cantidad_facturas,
                    COUNT(*) FILTER (WHERE m.tipo_movimiento = 'NOTA_CREDITO') AS cantidad_notas_credito,

                    COALESCE(SUM(m.total), 0) AS ventas_netas,
                    COALESCE(SUM(m.subtotal), 0) AS ventas_sin_impuesto,
                    COALESCE(SUM(m.total) - SUM(m.subtotal), 0) AS impuestos,

                    COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'FACTURA' THEN m.total
                            ELSE 0
                        END
                    ), 0) AS total_facturado_siigo,

                    ABS(COALESCE(SUM(
                        CASE
                            WHEN m.tipo_movimiento = 'NOTA_CREDITO' THEN m.total
                            ELSE 0
                        END
                    ), 0)) AS notas_credito

                FROM ventas_movimientos_enriquecidos m
                WHERE {where_mov}
                GROUP BY m.cliente_nombre, m.cost_center, m.centro_costo_nombre
                ORDER BY ventas_netas DESC
            """)

            centros_comercial_rows = [
                dict(r) for r in db.session.execute(sql_cc_comercial, params_mov).mappings().all()
            ]

            # ============================================================
            # 8) Centros de costo cartera por cliente
            # ============================================================
            sql_cc_cartera = text(facturas_base_cte + f"""
                SELECT
                    fb.cliente_nombre_ok AS cliente_nombre,
                    fb.centro_costo_nombre_ok AS centro_costo_nombre,
                    fb.cost_center,

                    COALESCE(SUM(COALESCE(fb.total, 0) - COALESCE(fb.saldo, 0)), 0) AS total_pagado,
                    COALESCE(SUM(COALESCE(fb.saldo, 0)), 0) AS saldo_pendiente

                FROM facturas_base fb
                WHERE {where_fac}
                GROUP BY fb.cliente_nombre_ok, fb.cost_center, fb.centro_costo_nombre_ok
            """)

            centros_cartera_rows = [
                dict(r) for r in db.session.execute(sql_cc_cartera, params_fac).mappings().all()
            ]

            centros_merge = {}

            for r in centros_comercial_rows:
                cliente_key = normalizar_cliente(r.get("cliente_nombre"))
                cc_key = f"{cliente_key}::{r.get('cost_center')}::{r.get('centro_costo_nombre')}"

                centros_merge[cc_key] = {
                    "cliente_key": cliente_key,
                    "centro_costo_nombre": r.get("centro_costo_nombre") or "Sin centro de costo",
                    "cost_center": r.get("cost_center"),
                    "cantidad_facturas": int(r.get("cantidad_facturas") or 0),
                    "cantidad_notas_credito": int(r.get("cantidad_notas_credito") or 0),

                    "ventas_netas": money_float(r.get("ventas_netas")),
                    "ventas_sin_impuesto": money_float(r.get("ventas_sin_impuesto")),
                    "impuestos": money_float(r.get("impuestos")),
                    "total_facturado_siigo": money_float(r.get("total_facturado_siigo")),
                    "notas_credito": money_float(r.get("notas_credito")),

                    "total_pagado": 0,
                    "saldo_pendiente": 0,
                }

            for r in centros_cartera_rows:
                cliente_key = normalizar_cliente(r.get("cliente_nombre"))
                cc_key = f"{cliente_key}::{r.get('cost_center')}::{r.get('centro_costo_nombre')}"

                if cc_key not in centros_merge:
                    centros_merge[cc_key] = {
                        "cliente_key": cliente_key,
                        "centro_costo_nombre": r.get("centro_costo_nombre") or "Sin centro de costo",
                        "cost_center": r.get("cost_center"),
                        "cantidad_facturas": 0,
                        "cantidad_notas_credito": 0,

                        "ventas_netas": 0,
                        "ventas_sin_impuesto": 0,
                        "impuestos": 0,
                        "total_facturado_siigo": 0,
                        "notas_credito": 0,

                        "total_pagado": 0,
                        "saldo_pendiente": 0,
                    }

                centros_merge[cc_key]["total_pagado"] = money_float(r.get("total_pagado"))
                centros_merge[cc_key]["saldo_pendiente"] = money_float(r.get("saldo_pendiente"))

            centros_por_cliente = defaultdict(list)

            for item in centros_merge.values():
                item["total_facturado"] = item["ventas_netas"]

                item["ventas_netas_str"] = money_fmt(item["ventas_netas"])
                item["ventas_sin_impuesto_str"] = money_fmt(item["ventas_sin_impuesto"])
                item["impuestos_str"] = money_fmt(item["impuestos"])
                item["total_facturado_siigo_str"] = money_fmt(item["total_facturado_siigo"])
                item["notas_credito_str"] = money_fmt(item["notas_credito"])
                item["total_facturado_str"] = money_fmt(item["ventas_netas"])
                item["saldo_pendiente_str"] = money_fmt(item["saldo_pendiente"])

                centros_por_cliente[item["cliente_key"]].append(item)

            # ============================================================
            # 9) Estados por cliente desde facturas_base
            # ============================================================
            sql_estados = text(facturas_base_cte + f"""
                SELECT
                    fb.cliente_nombre_ok AS cliente_nombre,
                    ({estado_sql}) AS estado_cartera,
                    COUNT(*) AS cantidad,
                    COALESCE(SUM(fb.saldo), 0) AS saldo_pendiente
                FROM facturas_base fb
                WHERE {where_fac}
                GROUP BY fb.cliente_nombre_ok, estado_cartera
            """)

            estados_rows = [
                dict(r) for r in db.session.execute(sql_estados, params_fac).mappings().all()
            ]

            estados_por_cliente = defaultdict(lambda: {
                "pagado": 0,
                "sano": 0,
                "alerta": 0,
                "vencido": 0,
            })

            estados_saldo_por_cliente = defaultdict(lambda: {
                "pagado": 0.0,
                "sano": 0.0,
                "alerta": 0.0,
                "vencido": 0.0,
            })

            for r in estados_rows:
                cliente_key = normalizar_cliente(r.get("cliente_nombre"))
                estado = r.get("estado_cartera") or "sano"
                cantidad = int(r.get("cantidad") or 0)
                saldo_estado = money_float(r.get("saldo_pendiente"))

                estados_por_cliente[cliente_key][estado] = cantidad
                estados_saldo_por_cliente[cliente_key][estado] = saldo_estado

            # ============================================================
            # 10) Facturas recientes reales desde facturas_base
            # ============================================================
            sql_facturas = text(f"""
                WITH facturas_base AS (
                    SELECT
                        f.*,
                        regexp_replace(
                            COALESCE(
                                NULLIF(TRIM(BOTH '"' FROM f.cliente_nombre), ''),
                                NULLIF(TRIM(BOTH '"' FROM c.name), ''),
                                'Desconocido'
                            ),
                            '[\\{{\\}}\\[\\]\\"]',
                            '',
                            'g'
                        ) AS cliente_nombre_ok,
                        COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre_ok
                    FROM siigo_facturas f
                    LEFT JOIN siigo_customers c
                        ON c.id::text = f.customer_id::text
                    AND c.idcliente = f.idcliente
                    LEFT JOIN siigo_centros_costo cc
                        ON cc.id = f.cost_center
                    AND cc.idcliente = f.idcliente
                ),
                ranked AS (
                    SELECT
                        fb.idfactura,
                        fb.fecha,
                        fb.vencimiento,

                        COALESCE(fb.subtotal, 0) AS ventas_netas,
                        COALESCE(fb.impuestos_total, 0) AS impuestos,
                        COALESCE(fb.total, 0) AS total_facturado_siigo,

                        COALESCE(fb.total, 0) AS total,
                        COALESCE(fb.total, 0) - COALESCE(fb.saldo, 0) AS pagado,
                        COALESCE(fb.saldo, 0) AS pendiente,

                        fb.public_url,
                        fb.cliente_nombre_ok AS cliente_nombre,
                        fb.cost_center,
                        fb.centro_costo_nombre_ok AS centro_costo_nombre,

                        ({estado_sql}) AS estado_cartera,

                        CASE
                            WHEN fb.vencimiento IS NULL THEN NULL
                            ELSE (fb.vencimiento - CURRENT_DATE)
                        END AS dias_vencimiento,

                        ROW_NUMBER() OVER (
                            PARTITION BY fb.cliente_nombre_ok
                            ORDER BY fb.fecha DESC, fb.idfactura DESC
                        ) AS rn

                    FROM facturas_base fb
                    WHERE {where_fac}
                )
                SELECT *
                FROM ranked
                WHERE (:mostrar_todas_facturas = true OR rn <= :limit_facturas)
                ORDER BY cliente_nombre, fecha DESC
            """)

            facturas_rows = [
                dict(r) for r in db.session.execute(sql_facturas, params_fac).mappings().all()
            ]

            facturas_por_cliente = defaultdict(list)

            for r in facturas_rows:
                cliente_key = normalizar_cliente(r.get("cliente_nombre"))
                fecha = r.get("fecha")
                vencimiento = r.get("vencimiento")

                ventas_netas_f = money_float(r.get("ventas_netas"))
                impuestos_f = money_float(r.get("impuestos"))
                total_siigo_f = money_float(r.get("total_facturado_siigo"))
                total = money_float(r.get("total"))
                pagado = money_float(r.get("pagado"))
                pendiente = money_float(r.get("pendiente"))
                dias_vencimiento = r.get("dias_vencimiento")

                fecha_str = (
                    fecha.strftime("%d/%m/%Y")
                    if fecha and hasattr(fecha, "strftime")
                    else str(fecha) if fecha else None
                )

                vencimiento_str = (
                    vencimiento.strftime("%d/%m/%Y")
                    if vencimiento and hasattr(vencimiento, "strftime")
                    else str(vencimiento) if vencimiento else None
                )

                facturas_por_cliente[cliente_key].append({
                    "idfactura": r.get("idfactura"),
                    "fecha": fecha_str,
                    "vencimiento": vencimiento_str,

                    "ventas_netas": ventas_netas_f,
                    "impuestos": impuestos_f,
                    "total_facturado_siigo": total_siigo_f,

                    "ventas_netas_str": money_fmt(ventas_netas_f),
                    "impuestos_str": money_fmt(impuestos_f),
                    "total_facturado_siigo_str": money_fmt(total_siigo_f),

                    "total": total,
                    "total_str": money_fmt(total),

                    "pagado": pagado,
                    "pendiente": pendiente,
                    "pagado_str": money_fmt(pagado),
                    "pendiente_str": money_fmt(pendiente),

                    "public_url": r.get("public_url"),
                    "cliente_nombre": r.get("cliente_nombre"),
                    "cliente_key": cliente_key,
                    "cost_center": r.get("cost_center"),
                    "centro_costo_nombre": r.get("centro_costo_nombre") or "Sin centro de costo",
                    "estado_cartera": r.get("estado_cartera") or "sano",
                    "dias_vencimiento": int(dias_vencimiento) if dias_vencimiento is not None else None,
                })

            # ============================================================
            # 11) Armar clientes enriquecidos
            # ============================================================
            if filtro_estado_normalizado:
                # Cuando se filtra por estado, mostramos clientes que tienen facturas en ese estado.
                all_keys = set(clientes_cartera.keys())
            else:
                all_keys = set(clientes_comercial.keys()) | set(clientes_cartera.keys())

            clientes = []

            for cliente_key in sorted(all_keys):
                comercial = clientes_comercial.get(cliente_key, {})
                cartera = clientes_cartera.get(cliente_key, {})

                cliente_nombre = (
                    comercial.get("cliente")
                    or cartera.get("cliente")
                    or "Sin cliente"
                )

                ventas_netas = money_float(comercial.get("ventas_netas"))
                ventas_sin_impuesto = money_float(comercial.get("ventas_sin_impuesto"))
                impuestos = money_float(comercial.get("impuestos"))
                total_facturado_siigo = money_float(comercial.get("total_facturado_siigo"))
                notas_credito = money_float(comercial.get("notas_credito"))

                cantidad_facturas = int(comercial.get("cantidad_facturas") or 0)
                cantidad_notas_credito = int(comercial.get("cantidad_notas_credito") or 0)

                cantidad_centros_costo = len(centros_por_cliente.get(cliente_key, []))

                pagado = money_float(cartera.get("total_pagado"))
                pendiente = money_float(cartera.get("saldo_pendiente"))
                vencido = money_float(cartera.get("saldo_vencido"))
                saldo_por_vencer_cliente = money_float(cartera.get("saldo_por_vencer"))

                pct_cliente_pagado = (pagado / total_facturado_siigo * 100) if total_facturado_siigo else 0
                pct_cliente_pendiente = (pendiente / total_facturado_siigo * 100) if total_facturado_siigo else 0
                pct_cliente_vencido = (vencido / pendiente * 100) if pendiente else 0

                ultima_factura = comercial.get("ultima_factura")
                ultima_factura_str = (
                    ultima_factura.strftime("%d/%m/%Y")
                    if ultima_factura and hasattr(ultima_factura, "strftime")
                    else str(ultima_factura) if ultima_factura else None
                )

                clientes.append({
                    "cliente": cliente_nombre,
                    "cliente_key": cliente_key,

                    "cantidad_facturas": cantidad_facturas,
                    "cantidad_notas_credito": cantidad_notas_credito,
                    "cantidad_centros_costo": cantidad_centros_costo,

                    "ventas_netas": ventas_netas,
                    "ventas_sin_impuesto": ventas_sin_impuesto,
                    "impuestos": impuestos,
                    "total_facturado_siigo": total_facturado_siigo,
                    "notas_credito": notas_credito,

                    # Compatibilidad.
                    "total_facturado": ventas_netas,

                    "total_pagado": pagado,
                    "saldo_pendiente": pendiente,
                    "saldo_vencido": vencido,
                    "saldo_por_vencer": saldo_por_vencer_cliente,

                    "ventas_netas_str": money_fmt(ventas_netas),
                    "ventas_sin_impuesto_str": money_fmt(ventas_sin_impuesto),
                    "impuestos_str": money_fmt(impuestos),
                    "total_facturado_siigo_str": money_fmt(total_facturado_siigo),
                    "notas_credito_str": money_fmt(notas_credito),

                    # Compatibilidad.
                    "total_facturado_str": money_fmt(ventas_netas),

                    "total_pagado_str": money_fmt(pagado),
                    "saldo_pendiente_str": money_fmt(pendiente),
                    "saldo_vencido_str": money_fmt(vencido),
                    "saldo_por_vencer_str": money_fmt(saldo_por_vencer_cliente),

                    "pct_pagado": round(pct_cliente_pagado, 2),
                    "pct_pendiente": round(pct_cliente_pendiente, 2),
                    "pct_vencido": round(pct_cliente_vencido, 2),

                    "ultima_factura": ultima_factura_str,
                    "centros_costo": centros_por_cliente.get(cliente_key, []),
                    "facturas_recientes": facturas_por_cliente.get(cliente_key, []),
                    "estados": estados_por_cliente.get(cliente_key, {
                        "pagado": 0,
                        "sano": 0,
                        "alerta": 0,
                        "vencido": 0,
                    }),
                    "estados_saldo": estados_saldo_por_cliente.get(cliente_key, {
                        "pagado": 0,
                        "sano": 0,
                        "alerta": 0,
                        "vencido": 0,
                    }),
                })

            # ============================================================
            # 12) Catálogos para filtros
            # ============================================================
            wh_cat_mov = ["m.idcliente = :idcliente"]
            wh_cat_fac = ["fb.idcliente = :idcliente"]
            params_cat = {"idcliente": idcliente}

            if fecha_desde_val:
                wh_cat_mov.append("m.fecha >= :desde")
                wh_cat_fac.append("fb.fecha >= :desde")
                params_cat["desde"] = fecha_desde_val

            if fecha_hasta_val:
                wh_cat_mov.append("m.fecha < (CAST(:hasta AS date) + INTERVAL '1 day')")
                wh_cat_fac.append("fb.fecha < (CAST(:hasta AS date) + INTERVAL '1 day')")
                params_cat["hasta"] = fecha_hasta_val

            where_cat_mov = " AND ".join(wh_cat_mov)
            where_cat_fac = " AND ".join(wh_cat_fac)

            sql_catalogo_clientes = text(f"""
                WITH facturas_base AS (
                    SELECT
                        f.*,
                        regexp_replace(
                            COALESCE(
                                NULLIF(TRIM(BOTH '"' FROM f.cliente_nombre), ''),
                                NULLIF(TRIM(BOTH '"' FROM c.name), ''),
                                'Desconocido'
                            ),
                            '[\\{{\\}}\\[\\]\\"]',
                            '',
                            'g'
                        ) AS cliente_nombre_ok
                    FROM siigo_facturas f
                    LEFT JOIN siigo_customers c
                        ON c.id::text = f.customer_id::text
                    AND c.idcliente = f.idcliente
                ),
                clientes AS (
                    SELECT DISTINCT
                        m.cliente_nombre AS id,
                        m.cliente_nombre AS nombre
                    FROM ventas_movimientos_enriquecidos m
                    WHERE {where_cat_mov}
                    AND TRIM(COALESCE(m.cliente_nombre, '')) <> ''

                    UNION

                    SELECT DISTINCT
                        fb.cliente_nombre_ok AS id,
                        fb.cliente_nombre_ok AS nombre
                    FROM facturas_base fb
                    WHERE {where_cat_fac}
                    AND TRIM(COALESCE(fb.cliente_nombre_ok, '')) <> ''
                )
                SELECT id, nombre
                FROM clientes
                ORDER BY nombre
            """)

            catalogo_clientes = [
                dict(r) for r in db.session.execute(sql_catalogo_clientes, params_cat).mappings().all()
            ]

            sql_catalogo_cc = text(f"""
                WITH facturas_base AS (
                    SELECT
                        f.*,
                        COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre_ok
                    FROM siigo_facturas f
                    LEFT JOIN siigo_centros_costo cc
                        ON cc.id = f.cost_center
                    AND cc.idcliente = f.idcliente
                ),
                centros AS (
                    SELECT DISTINCT
                        m.cost_center AS id,
                        COALESCE(m.centro_costo_nombre, 'Sin centro de costo') AS nombre
                    FROM ventas_movimientos_enriquecidos m
                    WHERE {where_cat_mov}
                    AND m.cost_center IS NOT NULL

                    UNION

                    SELECT DISTINCT
                        fb.cost_center AS id,
                        fb.centro_costo_nombre_ok AS nombre
                    FROM facturas_base fb
                    WHERE {where_cat_fac}
                    AND fb.cost_center IS NOT NULL
                )
                SELECT id, nombre
                FROM centros
                ORDER BY nombre
            """)

            catalogo_centros_costo = [
                dict(r) for r in db.session.execute(sql_catalogo_cc, params_cat).mappings().all()
            ]

            # ============================================================
            # 13) Top charts
            # ============================================================
            top_facturacion = sorted(
                clientes,
                key=lambda x: x["ventas_netas"],
                reverse=True
            )[:10]

            top_saldo = sorted(
                clientes,
                key=lambda x: x["saldo_pendiente"],
                reverse=True
            )[:10]

            return jsonify({
                "resumen": resumen,
                "clientes": clientes,
                "catalogos": {
                    "clientes": catalogo_clientes,
                    "centros_costo": catalogo_centros_costo,
                    "estados": [
                        {"id": "pagado", "nombre": "Pagado"},
                        {"id": "sano", "nombre": "Sano"},
                        {"id": "alerta", "nombre": "Por vencer pronto"},
                        {"id": "vencido", "nombre": "Vencido"},
                    ],
                },
                "charts": {
                    "top_facturacion": [
                        {
                            "cliente": c["cliente"],
                            "ventas_netas": c["ventas_netas"],
                            "ventas_sin_impuesto": c["ventas_sin_impuesto"],
                            "total_facturado": c["ventas_netas"],
                            "total_facturado_siigo": c["total_facturado_siigo"],
                            "notas_credito": c["notas_credito"],
                            "saldo_pendiente": c["saldo_pendiente"],
                        }
                        for c in top_facturacion
                    ],
                    "top_saldo": [
                        {
                            "cliente": c["cliente"],
                            "saldo_pendiente": c["saldo_pendiente"],
                            "saldo_vencido": c["saldo_vencido"],
                        }
                        for c in top_saldo
                    ],
                },
                "params": {
                    "idcliente": idcliente,
                    "desde": desde,
                    "hasta": hasta,
                    "cliente": cliente,
                    "cost_center": cost_center,
                    "estado": filtro_estado,
                    "limit_facturas": "all" if mostrar_todas_facturas else limit_facturas,
                    "mostrar_todas_facturas": mostrar_todas_facturas,
                },
                "config": {
                    "fuente_ventas": "ventas_movimientos_enriquecidos",
                    "fuente_cartera": "siigo_facturas_enriquecida_con_cliente",
                    "logica_ventas": "ventas_netas = facturas_emitidas - notas_credito",
                    "ventas_netas": "con impuesto",
                }
            })

        except Exception as e:
            return jsonify({"error": str(e)}), 500




    # --- ENDPOINT: Facturas por cliente/centro de costo (paginadas) ---
    @app.route("/reportes/facturas_cliente_cartera", methods=["GET"])
    @jwt_required()
    def get_facturas_cliente_cartera():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde       = request.args.get("desde")
        hasta       = request.args.get("hasta")
        cliente     = request.args.get("cliente")
        cost_center = request.args.get("cost_center", type=int)
        limit       = request.args.get("limit", type=int)
        offset      = request.args.get("offset", type=int) or 0
        filtro_estado = request.args.get("estado")  # 👈 'sano' | 'alerta' | 'vencido' | 'pagado'

        try:
            wh = ["f.idcliente = :idcliente"]
            params = {"idcliente": idcliente, "limit": limit, "offset": offset}

            if desde:
                wh.append("f.fecha >= :desde")
                params["desde"] = desde
            if hasta:
                wh.append("f.fecha <= :hasta")
                params["hasta"] = hasta
            if cliente:
                wh.append("LOWER(TRIM(f.cliente_nombre)) = LOWER(TRIM(:cliente))")
                params["cliente"] = cliente
            if cost_center:
                wh.append("f.cost_center = :cost_center")
                params["cost_center"] = cost_center

            where_clause = " AND ".join(wh)

            sql = text(f"""
                SELECT
                    f.idfactura,
                    f.fecha,
                    f.vencimiento,
                    f.total,
                    (f.total - f.saldo) AS pagado,
                    f.saldo AS pendiente,
                    f.public_url,
                    f.cliente_nombre,
                    COALESCE(f.centro_costo_nombre, 'Sin centro de costo') AS centro_costo_nombre
                FROM facturas_enriquecidas f
                WHERE {where_clause}
                ORDER BY f.fecha DESC
                LIMIT COALESCE(:limit, 999999)
            """)
            rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
            enriched = enriquecer_facturas(rows)

            # aplicar filtro de estado si corresponde
            if filtro_estado:
                enriched = [r for r in enriched if r["estado_cartera"] == filtro_estado.lower()]

            return jsonify({"rows": enriched})

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    # --- ENDPOINT: Catálogo de centros de costo ---
    @app.route("/catalogos/centros-costo", methods=["GET"])
    @jwt_required()
    def catalogo_centros_costo():
        claims = get_jwt()
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if claims.get("perfilid") == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        try:
            wh = ["f.idcliente = :idcliente"]
            params = {"idcliente": idcliente}

            if desde:
                wh.append("f.fecha >= :desde")
                params["desde"] = desde
            if hasta:
                wh.append("f.fecha <= :hasta")
                params["hasta"] = hasta

            where_clause = " AND ".join(wh)

            sql = text(f"""
                SELECT DISTINCT
                    f.cost_center AS id,
                    COALESCE(cc.nombre, 'Sin centro de costo') AS nombre
                FROM facturas_enriquecidas f
                LEFT JOIN siigo_centros_costo cc ON f.cost_center = cc.id
                WHERE {where_clause}
                AND f.cost_center IS NOT NULL
                ORDER BY nombre
            """)
            rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
            return jsonify(rows)

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    # --- ENDPOINT: Sincronizar notas crédito desde Siigo ---
    @app.route("/siigo/sync-notas-credito", methods=["POST"])
    def siigo_sync_notas_credito():
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        since = request.args.get("since") or request.args.get("fecha_desde")
        since_date = _parse_date_yyyy_mm_dd(since)
        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_log = "/siigo/sync-notas-credito"
        params_log = {
            "since": since,
        }

        log_id = None
        inicio = time.time()

        if not modo_sync_all:
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje=(
                    "Notas crédito: proceso iniciado"
                    + (f" con fecha desde {since}." if since else " sin límite de fecha.")
                ),
            )

        try:
            cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
            if not cred or not cred.client_id or not cred.client_secret or not cred.base_url:
                detalle = "Credenciales de Siigo no configuradas."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=400,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({"error": detalle}), 400

            access_key = dec(cred.client_secret)
            if not access_key:
                detalle = "No se pudo desencriptar el Access Key."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=400,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({"error": detalle}), 400

            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data.get("access_token")
            if not token:
                detalle = "Error al obtener token Siigo."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=500,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({"error": detalle}), 500

            page = 1
            nuevas = 0
            actualizadas = 0
            omitidas_por_fecha = 0
            total_leidas = 0

            while True:
                url = f"{cred.base_url.rstrip('/')}/v1/credit-notes?page={page}&page_size=100"
                r = _request_with_retries("GET", url, headers=_headers_bearer(token))

                if r.status_code != 200:
                    detalle = f"Siigo error {r.status_code}: {r.text}"

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=r.status_code,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({"error": detalle}), r.status_code

                data = r.json() or {}
                notas = data.get("results") or []

                if not notas:
                    break

                for n in notas:
                    total_leidas += 1

                    nota_id = _str(n.get("name"))
                    if not nota_id:
                        continue

                    fecha_str = n.get("date")
                    try:
                        fecha = datetime.fromisoformat(str(fecha_str)).date() if fecha_str else None
                    except Exception:
                        fecha = None

                    # Filtro global de fecha inicial de datos Siigo.
                    if since_date:
                        if not fecha or fecha < since_date:
                            omitidas_por_fecha += 1
                            continue

                    total = float(n.get("total") or 0)
                    estado = _str(n.get("status"))
                    observaciones = _str(n.get("observations"))
                    motivo = _str(n.get("reason"))
                    uuid = _str(n.get("id"))
                    customer = n.get("customer") or {}
                    cliente_nombre = _str(customer.get("name"))
                    customer_id = _str(customer.get("id"))

                    factura_afectada_id = _str((n.get("invoice") or {}).get("name"))
                    factura_afectada_uuid = _str((n.get("invoice") or {}).get("id"))
                    metadata_json = n

                    nota = SiigoNotaCredito.query.filter_by(
                        idcliente=idcliente,
                        nota_id=nota_id
                    ).first()

                    if nota:
                        changes = 0

                        if nota.fecha != fecha:
                            nota.fecha = fecha
                            changes += 1
                        if nota.total != total:
                            nota.total = total
                            changes += 1
                        if nota.estado != estado:
                            nota.estado = estado
                            changes += 1
                        if nota.observaciones != observaciones:
                            nota.observaciones = observaciones
                            changes += 1
                        if nota.motivo != motivo:
                            nota.motivo = motivo
                            changes += 1
                        if nota.uuid != uuid:
                            nota.uuid = uuid
                            changes += 1
                        if nota.cliente_nombre != cliente_nombre:
                            nota.cliente_nombre = cliente_nombre
                            changes += 1
                        if nota.customer_id != customer_id:
                            nota.customer_id = customer_id
                            changes += 1
                        if nota.factura_afectada_id != factura_afectada_id:
                            nota.factura_afectada_id = factura_afectada_id
                            changes += 1
                        if nota.factura_afectada_uuid != factura_afectada_uuid:
                            nota.factura_afectada_uuid = factura_afectada_uuid
                            changes += 1
                        if nota.metadata_json != metadata_json:
                            nota.metadata_json = metadata_json
                            changes += 1

                        if changes > 0:
                            actualizadas += 1

                    else:
                        nota = SiigoNotaCredito(
                            idcliente=idcliente,
                            nota_id=nota_id,
                            fecha=fecha,
                            total=total,
                            estado=estado,
                            observaciones=observaciones,
                            motivo=motivo,
                            uuid=uuid,
                            cliente_nombre=cliente_nombre,
                            customer_id=customer_id,
                            factura_afectada_id=factura_afectada_id,
                            factura_afectada_uuid=factura_afectada_uuid,
                            metadata_json=metadata_json,
                        )
                        db.session.add(nota)
                        nuevas += 1

                db.session.commit()

                next_href = ((data.get("_links") or {}).get("next") or {}).get("href")
                if not next_href or len(notas) < 100:
                    break

                page += 1

            mensaje = (
                f"Sincronización de notas crédito completa: "
                f"{nuevas} nuevas, {actualizadas} actualizadas, "
                f"{omitidas_por_fecha} omitidas por fecha."
            )

            respuesta = {
                "mensaje": mensaje,
                "since": since_date.isoformat() if since_date else None,
                "total_leidas_siigo": total_leidas,
                "nuevas": nuevas,
                "actualizadas": actualizadas,
                "omitidas_por_fecha": omitidas_por_fecha,
                "total_procesadas": nuevas + actualizadas,
            }

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="OK",
                    detalle=str(respuesta),
                    status_code=200,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify(respuesta), 200

        except Exception as e:
            detalle_error = traceback.format_exc()

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="ERROR",
                    detalle=detalle_error,
                    status_code=500,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify({"error": str(e)}), 500


    # --- ENDPOINT: Debug de una Nota Crédito ---
    @app.route("/siigo/debug-nota-credito", methods=["GET"])
    @jwt_required()
    def siigo_debug_nota_credito():
        """
        Muestra:
        - detalle crudo de Siigo para una nota crédito (por ?uuid=... o ?name=...)
        - lo guardado en siigo_notas_credito
        Uso:
        /siigo/debug-nota-credito?uuid=<uuid_de_siigo>
        /siigo/debug-nota-credito?name=<name/NC-...>
        Nota: requiere JWT; usa el idcliente del token (o ?idcliente= en superadmin).
        """
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        # superadmin puede inspeccionar otro cliente
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado o sin cliente en el token"}), 403

        uuid = request.args.get("uuid", type=str)
        name = request.args.get("name", type=str)
        if not uuid and not name:
            return jsonify({"error": "Proporciona ?uuid= o ?name="}), 400

        # cargar credenciales
        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg or not cfg.base_url or not cfg.client_id or not cfg.client_secret:
            return jsonify({"error": "Credenciales de Siigo no configuradas para este cliente"}), 400

        # autenticar
        auth_data = _siigo_auth_json_for_client(cfg)
        if auth_data.get("_error"):
            return jsonify({"error": auth_data["_error"]}), 502
        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token desde Siigo"}), 502

        base_url = (cfg.base_url or "").rstrip("/")
        headers = _siigo_headers_bearer(token)

        # 1) Traer detalle crudo desde Siigo
        raw_detail = None
        tried = []
        try:
            if uuid:
                url = f"{base_url}/v1/credit-notes/{uuid}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    raw_detail = r.json()
                elif name:
                    url2 = f"{base_url}/v1/credit-notes/{name}"
                    tried.append(url2)
                    r2 = requests.get(url2, headers=headers, timeout=60)
                    if r2.status_code == 200:
                        raw_detail = r2.json()
                    else:
                        url3 = f"{base_url}/v1/credit-notes?name={name}"
                        tried.append(url3)
                        r3 = requests.get(url3, headers=headers, timeout=60)
                        if r3.status_code == 200:
                            payload = r3.json() or {}
                            res = payload.get("results") or []
                            raw_detail = res[0] if res else None
            else:
                url = f"{base_url}/v1/credit-notes/{name}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    raw_detail = r.json()
                else:
                    url2 = f"{base_url}/v1/credit-notes?name={name}"
                    tried.append(url2)
                    r2 = requests.get(url2, headers=headers, timeout=60)
                    if r2.status_code == 200:
                        payload = r2.json() or {}
                        res = payload.get("results") or []
                        raw_detail = res[0] if res else None
        except requests.RequestException as e:
            return jsonify({"error": f"Error consultando Siigo: {str(e)}", "tried": tried}), 502

        # 2) Lo que tienes guardado en BD
        nota_db = None
        if name:
            nota_db = SiigoNotaCredito.query.filter_by(idcliente=idcliente, nota_id=name).first()
        elif uuid:
            nota_db = SiigoNotaCredito.query.filter_by(idcliente=idcliente, uuid=uuid).first()
            if not nota_db and raw_detail and isinstance(raw_detail, dict):
                nm = raw_detail.get("name")
                if nm:
                    nota_db = SiigoNotaCredito.query.filter_by(idcliente=idcliente, nota_id=nm).first()

        # 3) Serializar para JSON
        def _dec(v):
            from decimal import Decimal
            if isinstance(v, Decimal):
                return float(v)
            if isinstance(v, (datetime, )):
                return v.isoformat()
            return v

        def nota_to_dict(n):
            if not n:
                return None
            return {
                "id": n.id,
                "idcliente": n.idcliente,
                "nota_id": n.nota_id,
                "uuid": n.uuid,
                "fecha": n.fecha.isoformat() if n.fecha else None,
                "total": _dec(n.total),
                "estado": n.estado,
                "motivo": n.motivo,
                "observaciones": n.observaciones,
                "cliente_nombre": n.cliente_nombre,
                "customer_id": n.customer_id,
                "factura_afectada_id": n.factura_afectada_id,
                "metadata_json": n.metadata_json,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            }

        resp = {
            "query_params": {"uuid": uuid, "name": name, "idcliente": idcliente},
            "siigo_tried_urls": tried,
            "siigo_raw_detail": raw_detail,   # <- lo crudo que devuelve Siigo
            "db_nota_credito": nota_to_dict(nota_db),
        }
        return jsonify(resp), 200


    #----ENDPOINT DEBUG consulta documentos soporte
    @app.route("/siigo/debug-documentos-soporte", methods=["GET"])
    def siigo_debug_documentos_soporte():
        """
        Endpoint temporal SOLO LECTURA.
        Consulta documentos soporte desde Siigo API y devuelve el JSON crudo.

        No guarda.
        No actualiza.
        No borra.
        No toca siigo_compras.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        page_size = request.args.get("page_size", default=5, type=int)
        page = request.args.get("page", default=1, type=int)

        url = f"{base_url}/v1/purchase-support-documents?page_size={page_size}&page={page}"

        try:
            r = requests.get(url, headers=headers, timeout=90)

            try:
                data = r.json()
            except ValueError:
                return jsonify({
                    "error": "Siigo respondió algo que no es JSON",
                    "status_code": r.status_code,
                    "url": url,
                    "text": r.text
                }), 500

            results = []

            if isinstance(data, dict):
                if isinstance(data.get("results"), list):
                    results = data.get("results")
                elif isinstance(data.get("data"), dict) and isinstance(data["data"].get("results"), list):
                    results = data["data"].get("results")

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Consulta realizada sin guardar información en InsightFlow.",
                "status_code": r.status_code,
                "url": url,
                "cantidad_results": len(results),
                "data": data
            }), r.status_code

        except Exception as e:
            return jsonify({
                "error": "Error consultando documentos soporte en Siigo",
                "detalle": str(e)
            }), 500


    @app.route("/siigo/debug-documentos-soporte/<string:documento_id>", methods=["GET"])
    def siigo_debug_documento_soporte_detalle(documento_id):
        """
        Endpoint temporal SOLO LECTURA.
        Consulta un documento soporte específico por ID Siigo.

        No guarda.
        No actualiza.
        No borra.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        url = f"{base_url}/v1/purchase-support-documents/{documento_id}"

        try:
            r = requests.get(url, headers=headers, timeout=90)

            try:
                data = r.json()
            except ValueError:
                return jsonify({
                    "error": "Siigo respondió algo que no es JSON",
                    "status_code": r.status_code,
                    "url": url,
                    "text": r.text
                }), 500

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Consulta de detalle realizada sin guardar información en InsightFlow.",
                "status_code": r.status_code,
                "url": url,
                "data": data
            }), r.status_code

        except Exception as e:
            return jsonify({
                "error": "Error consultando detalle del documento soporte en Siigo",
                "detalle": str(e)
            }), 500



    @app.route("/siigo/debug-document-types-ds", methods=["GET"])
    def siigo_debug_document_types_ds():
        """
        Endpoint temporal SOLO LECTURA.
        Consulta tipos de comprobante Documento Soporte en Siigo.

        No guarda nada.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        url = f"{base_url}/v1/document-types?type=DS"

        try:
            r = requests.get(url, headers=headers, timeout=90)

            try:
                data = r.json()
            except ValueError:
                return jsonify({
                    "error": "Siigo respondió algo que no es JSON",
                    "status_code": r.status_code,
                    "url": url,
                    "text": r.text
                }), 500

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Consulta de tipos DS realizada sin guardar información.",
                "status_code": r.status_code,
                "url": url,
                "data": data
            }), r.status_code

        except Exception as e:
            return jsonify({
                "error": "Error consultando tipos DS en Siigo",
                "detalle": str(e)
            }), 500




    # --- ENDPOINT: Reporte de Cuentas por Cobrar (Aging Report) ---
    @app.route("/reportes/cuentas-por-cobrar", methods=["GET"])
    @jwt_required()
    def cuentas_por_cobrar():
        """
        Reporte Aging / Cuentas por Cobrar.

        Fuente principal:
        - siigo_facturas

        Regla:
        - Solo se consideran facturas con saldo > 0.
        - No usa facturas_enriquecidas para evitar efectos de notas crédito comerciales
        sobre la cartera.
        """

        from sqlalchemy.sql import text
        from collections import defaultdict
        from datetime import datetime
        import re

        def normalizar_cliente(nombre: str) -> str:
            nombre = (nombre or "").strip().lower()
            nombre = nombre.replace(".", "").replace(",", "")
            nombre = re.sub(r"\s+", " ", nombre)
            return nombre

        def money_float(valor):
            try:
                return float(valor or 0)
            except Exception:
                return 0.0

        def money_fmt(valor, decimales=2):
            valor = money_float(valor)
            return f"$ {valor:,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        def calcular_bucket(dias_vencidos):
            dias = int(dias_vencidos or 0)

            if dias <= 0:
                return "Por vencer"
            if dias <= 30:
                return "1-30"
            if dias <= 60:
                return "31-60"
            if dias <= 90:
                return "61-90"
            return "91+"

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado o sin cliente en el token"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        incluir_detalle = request.args.get("detalle", "0") == "1"

        fecha_desde_val = validar_fecha(desde) if desde else None
        fecha_hasta_val = validar_fecha(hasta) if hasta else None

        condiciones = [
            "fb.idcliente = :idcliente",
            "COALESCE(fb.saldo, 0) > 0",
        ]

        params = {
            "idcliente": idcliente,
        }

        if fecha_desde_val:
            condiciones.append("fb.fecha >= :desde")
            params["desde"] = fecha_desde_val

        if fecha_hasta_val:
            condiciones.append("fb.fecha <= :hasta")
            params["hasta"] = fecha_hasta_val

        where_sql = " AND ".join(condiciones)

        facturas_base_cte = """
            WITH facturas_base AS (
                SELECT
                    f.id,
                    f.idcliente,
                    f.idfactura,
                    f.fecha,
                    f.vencimiento,
                    f.total,
                    f.subtotal,
                    f.impuestos_total,
                    f.pagos_total,
                    f.saldo,
                    f.public_url,
                    f.customer_id,
                    f.cost_center,
                    f.seller_id,

                    REPLACE(
                        REPLACE(
                            REPLACE(
                                REPLACE(
                                    COALESCE(
                                        NULLIF(TRIM(BOTH '"' FROM f.cliente_nombre), ''),
                                        NULLIF(TRIM(BOTH '"' FROM c.name), ''),
                                        'Sin cliente'
                                    ),
                                    '{', ''
                                ),
                                '}', ''
                            ),
                            '[', ''
                        ),
                        ']', ''
                    ) AS cliente_nombre_ok,

                    COALESCE(cc.nombre, 'Sin centro de costo') AS centro_costo_nombre_ok,

                    COALESCE(v.nombre, 'Sin vendedor') AS vendedor_nombre_ok

                FROM siigo_facturas f
                LEFT JOIN siigo_customers c
                    ON c.id::text = f.customer_id::text
                AND c.idcliente = f.idcliente
                LEFT JOIN siigo_centros_costo cc
                    ON cc.id = f.cost_center
                AND cc.idcliente = f.idcliente
                LEFT JOIN siigo_vendedores v
                    ON v.id::text = f.seller_id::text
                AND v.idcliente = f.idcliente
            )
        """

        try:
            # ------------------------------------------------------------
            # 1) Base consolidada por cliente / centro / vendedor
            # ------------------------------------------------------------
            query_base = text(facturas_base_cte + f"""
                SELECT
                    fb.cliente_nombre_ok AS cliente_nombre,
                    fb.centro_costo_nombre_ok AS centro_costo_nombre,
                    fb.vendedor_nombre_ok AS vendedor_nombre,

                    COUNT(*) AS num_facturas,
                    COALESCE(SUM(fb.saldo), 0) AS saldo_total,

                    COALESCE(SUM(
                        CASE
                            WHEN fb.vencimiento IS NULL OR CURRENT_DATE <= fb.vencimiento
                            THEN fb.saldo ELSE 0
                        END
                    ), 0) AS saldo_sano,

                    COALESCE(SUM(
                        CASE
                            WHEN fb.vencimiento IS NOT NULL
                                AND CURRENT_DATE > fb.vencimiento
                                AND CURRENT_DATE - fb.vencimiento <= 30
                            THEN fb.saldo ELSE 0
                        END
                    ), 0) AS saldo_1_30,

                    COALESCE(SUM(
                        CASE
                            WHEN fb.vencimiento IS NOT NULL
                                AND CURRENT_DATE - fb.vencimiento BETWEEN 31 AND 60
                            THEN fb.saldo ELSE 0
                        END
                    ), 0) AS saldo_31_60,

                    COALESCE(SUM(
                        CASE
                            WHEN fb.vencimiento IS NOT NULL
                                AND CURRENT_DATE - fb.vencimiento BETWEEN 61 AND 90
                            THEN fb.saldo ELSE 0
                        END
                    ), 0) AS saldo_61_90,

                    COALESCE(SUM(
                        CASE
                            WHEN fb.vencimiento IS NOT NULL
                                AND CURRENT_DATE - fb.vencimiento > 90
                            THEN fb.saldo ELSE 0
                        END
                    ), 0) AS saldo_mas_90

                FROM facturas_base fb
                WHERE {where_sql}
                GROUP BY fb.cliente_nombre_ok, fb.centro_costo_nombre_ok, fb.vendedor_nombre_ok
                ORDER BY saldo_total DESC
            """)

            result = db.session.execute(query_base, params).mappings().all()
            rows_raw = [dict(r) for r in result]

            # ------------------------------------------------------------
            # 2) Agrupar por nombre normalizado
            # ------------------------------------------------------------
            agrupado = defaultdict(list)
            nombre_visible_por_key = {}

            for r in rows_raw:
                clave = normalizar_cliente(r.get("cliente_nombre"))
                agrupado[clave].append(r)

                if clave not in nombre_visible_por_key:
                    nombre_visible_por_key[clave] = r.get("cliente_nombre") or "Sin cliente"

            # ------------------------------------------------------------
            # 3) Consolidado por cliente
            # ------------------------------------------------------------
            consolidado = []

            for cliente_key, grupo in agrupado.items():
                base = {
                    "cliente_key": cliente_key,
                    "cliente_nombre": nombre_visible_por_key.get(cliente_key, "Sin cliente"),
                    "centro_costo_nombre": grupo[0].get("centro_costo_nombre"),
                    "vendedor_nombre": grupo[0].get("vendedor_nombre"),
                    "num_facturas": 0,
                    "aging": {
                        "por_vencer": 0.0,
                        "1_30": 0.0,
                        "31_60": 0.0,
                        "61_90": 0.0,
                        "91_mas": 0.0,
                    },
                    "total": 0.0,
                }

                for r in grupo:
                    base["total"] += money_float(r.get("saldo_total"))
                    base["aging"]["por_vencer"] += money_float(r.get("saldo_sano"))
                    base["aging"]["1_30"] += money_float(r.get("saldo_1_30"))
                    base["aging"]["31_60"] += money_float(r.get("saldo_31_60"))
                    base["aging"]["61_90"] += money_float(r.get("saldo_61_90"))
                    base["aging"]["91_mas"] += money_float(r.get("saldo_mas_90"))
                    base["num_facturas"] += int(r.get("num_facturas") or 0)

                base["total_str"] = money_fmt(base["total"])
                consolidado.append(base)

            consolidado.sort(key=lambda x: x["total"], reverse=True)

            # ------------------------------------------------------------
            # 4) Resumen global
            # ------------------------------------------------------------
            total_global = sum(r["total"] for r in consolidado)
            facturas_vivas = sum(r["num_facturas"] for r in consolidado)

            total_por_vencer = sum(r["aging"]["por_vencer"] for r in consolidado)
            total_1_30 = sum(r["aging"]["1_30"] for r in consolidado)
            total_31_60 = sum(r["aging"]["31_60"] for r in consolidado)
            total_61_90 = sum(r["aging"]["61_90"] for r in consolidado)
            total_91_mas = sum(r["aging"]["91_mas"] for r in consolidado)

            total_vencido = total_1_30 + total_31_60 + total_61_90 + total_91_mas
            pct_vencido = (total_vencido / total_global * 100) if total_global else 0

            resumen_global = {
                "facturas_vivas": facturas_vivas,

                "total_global": money_fmt(total_global, 1),
                "total_por_vencer": money_fmt(total_por_vencer),
                "total_vencido": money_fmt(total_vencido),

                "pct_vencido": round(pct_vencido, 2),

                "total_1_30": money_fmt(total_1_30),
                "total_31_60": money_fmt(total_31_60),
                "total_61_90": money_fmt(total_61_90),
                "total_91_mas": money_fmt(total_91_mas),

                # Valores numéricos adicionales por si el frontend los necesita.
                "total_global_num": total_global,
                "total_por_vencer_num": total_por_vencer,
                "total_vencido_num": total_vencido,
                "total_1_30_num": total_1_30,
                "total_31_60_num": total_31_60,
                "total_61_90_num": total_61_90,
                "total_91_mas_num": total_91_mas,
            }

            # ------------------------------------------------------------
            # 5) Detalle de facturas
            # ------------------------------------------------------------
            detalle = []

            if incluir_detalle:
                query_detalle = text(facturas_base_cte + f"""
                    SELECT
                        fb.idfactura,
                        fb.cliente_nombre_ok AS cliente_nombre,
                        fb.centro_costo_nombre_ok AS centro_costo_nombre,
                        fb.vendedor_nombre_ok AS vendedor_nombre,

                        TO_CHAR(fb.fecha, 'DD/MM/YYYY') AS fecha,
                        TO_CHAR(fb.vencimiento, 'DD/MM/YYYY') AS vencimiento,

                        CASE
                            WHEN fb.vencimiento IS NULL THEN 0
                            ELSE CURRENT_DATE - fb.vencimiento
                        END AS dias_vencidos,

                        CASE
                            WHEN fb.fecha IS NULL THEN 0
                            ELSE CURRENT_DATE - fb.fecha
                        END AS dias_transcurridos,

                        COALESCE(fb.total, 0) AS total,
                        COALESCE(fb.pagos_total, COALESCE(fb.total, 0) - COALESCE(fb.saldo, 0)) AS pagos_total,
                        COALESCE(fb.saldo, 0) AS saldo,
                        fb.public_url

                    FROM facturas_base fb
                    WHERE {where_sql}
                    ORDER BY fb.cliente_nombre_ok, fb.vencimiento, fb.fecha, fb.idfactura
                """)

                result_detalle = db.session.execute(query_detalle, params).mappings().all()

                for row in result_detalle:
                    r = dict(row)

                    dias_vencidos = int(r.get("dias_vencidos") or 0)
                    dias_transcurridos = int(r.get("dias_transcurridos") or 0)
                    cliente_key = normalizar_cliente(r.get("cliente_nombre"))

                    saldo = money_float(r.get("saldo"))
                    total = money_float(r.get("total"))
                    pagos_total = money_float(r.get("pagos_total"))

                    r["cliente_key"] = cliente_key
                    r["aging_bucket"] = calcular_bucket(dias_vencidos)
                    r["dias_vencidos"] = dias_vencidos
                    r["dias_transcurridos"] = dias_transcurridos

                    r["saldo"] = saldo
                    r["total"] = total
                    r["pagos_total"] = pagos_total

                    r["saldo_str"] = money_fmt(saldo)
                    r["total_str"] = money_fmt(total)
                    r["pagos_total_str"] = money_fmt(pagos_total)

                    detalle.append(r)

            # ------------------------------------------------------------
            # 6) Proyección por fecha
            # ------------------------------------------------------------
            query_proyeccion = text(facturas_base_cte + f"""
                SELECT
                    fb.vencimiento::date AS fecha,
                    COALESCE(SUM(fb.saldo), 0) AS total,
                    CASE
                        WHEN fb.vencimiento::date < CURRENT_DATE THEN true
                        ELSE false
                    END AS vencido,

                    json_agg(
                        json_build_object(
                            'idfactura', fb.idfactura,
                            'cliente_nombre', fb.cliente_nombre_ok,
                            'fecha', TO_CHAR(fb.fecha, 'DD/MM/YYYY'),
                            'vencimiento', TO_CHAR(fb.vencimiento, 'DD/MM/YYYY'),
                            'saldo', fb.saldo,
                            'public_url', fb.public_url,
                            'dias_vencidos',
                                CASE
                                    WHEN fb.vencimiento IS NULL THEN 0
                                    ELSE CURRENT_DATE - fb.vencimiento
                                END,
                            'dias_transcurridos',
                                CASE
                                    WHEN fb.fecha IS NULL THEN 0
                                    ELSE CURRENT_DATE - fb.fecha
                                END
                        )
                        ORDER BY fb.saldo DESC
                    ) AS facturas

                FROM facturas_base fb
                WHERE {where_sql}
                AND fb.vencimiento IS NOT NULL
                GROUP BY fb.vencimiento::date
                ORDER BY fecha
            """)

            result_proyeccion = db.session.execute(query_proyeccion, params).mappings().all()
            proyeccion_por_fecha = []

            for row in result_proyeccion:
                r = dict(row)

                facturas_json = []
                for f in r.get("facturas") or []:
                    cliente_key = normalizar_cliente(f.get("cliente_nombre"))

                    saldo = money_float(f.get("saldo"))
                    dias_vencidos = int(f.get("dias_vencidos") or 0)
                    dias_transcurridos = int(f.get("dias_transcurridos") or 0)

                    facturas_json.append({
                        "idfactura": f.get("idfactura"),
                        "cliente_nombre": f.get("cliente_nombre"),
                        "cliente_key": cliente_key,
                        "fecha": f.get("fecha"),
                        "vencimiento": f.get("vencimiento"),
                        "saldo": saldo,
                        "saldo_str": money_fmt(saldo),
                        "public_url": f.get("public_url"),
                        "dias_vencidos": dias_vencidos,
                        "dias_transcurridos": dias_transcurridos,
                        "aging_bucket": calcular_bucket(dias_vencidos),
                    })

                fecha = r.get("fecha")
                total_fecha = money_float(r.get("total"))

                proyeccion_por_fecha.append({
                    "fecha": fecha.strftime("%d/%m/%Y") if fecha and hasattr(fecha, "strftime") else str(fecha),
                    "total": total_fecha,
                    "total_str": money_fmt(total_fecha),
                    "vencido": bool(r.get("vencido")),
                    "facturas": facturas_json,
                })

            return jsonify({
                "resumen_global": resumen_global,
                "consolidado": consolidado,
                "detalle": detalle if incluir_detalle else None,
                "proyeccion_por_fecha": proyeccion_por_fecha,
                "params": {
                    "idcliente": idcliente,
                    "desde": desde,
                    "hasta": hasta,
                    "detalle": incluir_detalle,
                    "fuente": "siigo_facturas",
                    "solo_saldo_pendiente": True,
                },
            })

        except Exception as e:
            return jsonify({"error": str(e)}), 500


    @app.route("/siigo/sync-pagos-egresos", methods=["POST"])
    @jwt_required()
    def siigo_sync_pagos_egresos():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        deep = request.args.get("deep") in ("1","true","yes")
        batch = request.args.get("batch", default=None, type=int)
        only_missing = request.args.get("only_missing", default="1") in ("1","true","yes")
        since = request.args.get("since")  # opcional filtro por fecha

        try:
            mensaje = sync_pagos_egresos_desde_siigo(
                idcliente=idcliente,
                deep=deep,
                only_missing=only_missing,
                batch_size=batch,
                since=since
            )
            return jsonify({"mensaje": mensaje})
        except Exception as e:
            return jsonify({"error": str(e)}), 500


    @app.route("/siigo/debug-pago", methods=["GET"])
    @jwt_required()
    def siigo_debug_pago():
        """
        Devuelve detalle crudo de un recibo de egreso de Siigo y lo guardado en la BD (si existe).
        Uso:
        /siigo/debug-pago?idpago=12345
        Nota: requiere JWT. Superadmin puede pasar ?idcliente=.
        """
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        if perfilid == 0:
            idcliente = request.args.get("idcliente", type=int) or idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado o sin cliente en el token"}), 403

        idpago = request.args.get("idpago")
        if not idpago:
            return jsonify({"error": "Falta ?idpago="}), 400

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        token_data = _siigo_auth_json_for_client(cred)
        if token_data.get("_error"):
            return jsonify({"error": token_data["_error"]}), 502

        token = token_data.get("access_token")
        headers = _siigo_headers_bearer(token)
        base_url = cred.base_url.rstrip("/")
        tried_urls = []

        # intentar traer el JSON desde la API
        raw_data = None
        try:
            url = f"{base_url}/v1/payment-receipts/{idpago}"
            tried_urls.append(url)
            r = requests.get(url, headers=headers)
            if r.status_code == 200:
                raw_data = r.json()
            else:
                return jsonify({"error": f"HTTP {r.status_code}", "tried": tried_urls, "siigo_response": r.text}), 502
        except Exception as e:
            return jsonify({"error": str(e), "tried": tried_urls}), 502

        # buscar en la BD local
        pago = SiigoPagoProveedor.query.filter_by(idcliente=idcliente, idpago=idpago).first()
        pago_db = None
        if pago:
            pago_db = {
                "id": pago.id,
                "fecha": pago.fecha.isoformat() if pago.fecha else None,
                "proveedor_nombre": pago.proveedor_nombre,
                "metodo_pago": pago.metodo_pago,
                "valor": float(pago.valor or 0),
                "factura_aplicada": pago.factura_aplicada,
            }

        return jsonify({
            "idcliente": idcliente,
            "idpago": idpago,
            "tried": tried_urls,
            "siigo_raw": raw_data,
            "db_pago": pago_db
        })



    # Endpoint de depuración
    @app.route("/siigo/debug-cruce-pago", methods=["GET"])
    @jwt_required()
    def siigo_debug_cruce_pago():
        """
        Endpoint para depurar el cruce entre pagos y documentos soporte (idcompra).
        Uso: /siigo/debug-cruce-pago?idcompra=DS-1-1548
        Requiere autenticación JWT.
        """
        from models import SiigoCompra, SiigoPagoProveedor


        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 403


        idcompra = request.args.get("idcompra")
        if not idcompra:
            return jsonify({"error": "Falta parámetro idcompra"}), 400


        compra = SiigoCompra.query.filter_by(idcliente=idcliente, idcompra=idcompra).first()
        if not compra:
            return jsonify({"error": f"Compra {idcompra} no encontrada"}), 404


        pagos = SiigoPagoProveedor.query.filter_by(idcliente=idcliente, factura_aplicada=idcompra).all()


        pagos_info = [
            {
                "idpago": p.idpago,
                "fecha": p.fecha.isoformat() if p.fecha else None,
                "valor": float(p.valor or 0),
                "metodo_pago": p.metodo_pago,
                "proveedor": p.proveedor_nombre
            } for p in pagos
        ]


        total_pagado = sum(float(p.valor or 0) for p in pagos)
        total_compra = float(compra.total or 0)
        saldo = round(total_compra - total_pagado, 2)


        return jsonify({
            "idcompra": idcompra,
            "total_compra": total_compra,
            "total_pagado": total_pagado,
            "saldo": saldo,
            "pagos": pagos_info,
            "estado": "PAGADO" if saldo <= 0 else ("PARCIAL" if total_pagado > 0 else "PENDIENTE")
        })




    # Permite actualizar pagos y estados de facturas/documentos soporte
    @app.route("/siigo/actualizar-estado-pagos", methods=["POST"])
    @jwt_required()
    def actualizar_estado_pagos():
        """
        Recalcula el estado de pago de cada factura/documento soporte
        cruzando con los pagos cargados desde Siigo.
        Guarda el estado ("SI", "NO", "PARCIAL") en cada línea de siigo_pagos_proveedores.
        """
        from models import SiigoPagoProveedor, SiigoCompra
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 403

        # 1. Obtener compras del cliente con sus totales (clave: factura_proveedor + proveedor_identificacion)
        compras = SiigoCompra.query.filter_by(idcliente=idcliente).all()
        compras_dict = {}
        for c in compras:
            clave = (c.factura_proveedor, c.proveedor_identificacion)
            compras_dict[clave] = float(c.total or 0)

        print(f"📦 Compras cargadas en memoria: {len(compras_dict)}")

        # 2. Obtener todos los pagos del cliente
        pagos = SiigoPagoProveedor.query.filter_by(idcliente=idcliente).all()
        pagos_por_factura = {}

        # 3. Agrupar pagos por (factura_aplicada, proveedor_identificacion)
        for pago in pagos:
            if not pago.factura_aplicada or not pago.proveedor_identificacion:
                print(f"⚠️ Pago sin factura/proveedor -> idpago={pago.idpago}")
                continue
            clave = (pago.factura_aplicada, pago.proveedor_identificacion)
            pagos_por_factura.setdefault(clave, []).append(pago)

        print(f"💰 Pagos agrupados por factura: {len(pagos_por_factura)}")

        # 4. Calcular estado para cada grupo de pagos de una factura
        for clave, pagos_factura in pagos_por_factura.items():
            factura, prov_id = clave
            total_pagado = round(sum(float(p.valor or 0) for p in pagos_factura), 2)
            total_factura = compras_dict.get(clave, 0)

            if total_factura == 0:
                estado = "NO"
            elif total_pagado >= total_factura:
                estado = "SI"
            elif total_pagado > 0:
                estado = "PARCIAL"
            else:
                estado = "NO"

            print(f"🔎 Cruce factura={factura}, proveedor={prov_id} -> total_factura={total_factura}, total_pagado={total_pagado}, estado={estado}")

            # 5. Guardar el estado en todos los pagos vinculados a esa factura
            for pago in pagos_factura:
                pago.factura_pagada = estado

        db.session.commit()

        print("✅ Cruce finalizado y estados actualizados")

        # 6. Resumen de estados
        resumen = {
            "SI": SiigoPagoProveedor.query.filter_by(idcliente=idcliente, factura_pagada="SI").count(),
            "NO": SiigoPagoProveedor.query.filter_by(idcliente=idcliente, factura_pagada="NO").count(),
            "PARCIAL": SiigoPagoProveedor.query.filter_by(idcliente=idcliente, factura_pagada="PARCIAL").count(),
            "NULL": SiigoPagoProveedor.query.filter_by(idcliente=idcliente, factura_pagada=None).count(),
        }

        return jsonify({
            "status": "ok",
            "msg": "Estados de pago actualizados correctamente",
            "facturas_actualizadas": len(pagos_por_factura),
            "pagos_actualizados": len(pagos),
            "resumen_estados": resumen
        })




    @app.route("/siigo/sync-compras", methods=["POST"])
    def siigo_sync_compras():
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        deep = request.args.get("deep") in ("1", "true", "yes")
        batch = request.args.get("batch", default=None, type=int)
        only_missing = request.args.get("only_missing", default="1") in ("1", "true", "yes")
        since = request.args.get("since")
        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_log = "/siigo/sync-compras"
        params_log = {
            "deep": deep,
            "batch": batch,
            "only_missing": only_missing,
            "since": since,
        }

        try:
            if modo_sync_all:
                # Ejecutar directamente cuando viene desde sync-all.
                resultado = sync_compras_desde_siigo(
                    idcliente=idcliente,
                    deep=deep,
                    batch_size=batch if batch else 50,
                    only_missing=only_missing,
                    since=since,
                )

                # Si la función interna devuelve error, no responder 200.
                if isinstance(resultado, dict) and resultado.get("error"):
                    detalle = str(resultado.get("detalle") or "")
                    status = 429 if "429" in str(resultado.get("error")) or "requests_limit" in detalle else 500

                    return jsonify({
                        "error": resultado.get("error"),
                        "detalle": resultado.get("detalle"),
                        "mensaje": "Error sincronizando compras desde Siigo."
                    }), status

                return jsonify({"mensaje": f"Compras sincronizadas: {resultado}"}), 200

            # Modo UI manual: background con historial.
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje=(
                    "Compras: proceso iniciado"
                    + (f" con fecha desde {since}." if since else " sin límite de fecha.")
                ),
            )

            def run_background(local_log_id):
                with app.app_context():
                    inicio = time.time()

                    try:
                        print(f"[sync-compras] 🔁 Iniciando para cliente {idcliente}")

                        resultado = sync_compras_desde_siigo(
                            idcliente=idcliente,
                            deep=deep,
                            batch_size=batch if batch else 50,
                            only_missing=only_missing,
                            since=since,
                        )

                        duracion = round(time.time() - inicio, 2)

                        if isinstance(resultado, dict) and resultado.get("error"):
                            detalle = str(resultado.get("detalle") or resultado)
                            status = 429 if "429" in str(resultado.get("error")) or "requests_limit" in detalle else 500

                            print(f"[sync-compras] ❌ Error: {resultado}")

                            _finalizar_log_sync_modulo(
                                log_id=local_log_id,
                                idcliente=idcliente,
                                endpoint=endpoint_log,
                                resultado="ERROR",
                                detalle=f"Error sincronizando compras: {resultado}",
                                status_code=status,
                                duracion_segundos=duracion,
                            )
                            return

                        print(f"[sync-compras] ✅ Finalizado para cliente {idcliente}: {resultado}")

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="OK",
                            detalle=f"Compras sincronizadas correctamente: {resultado}",
                            status_code=200,
                            duracion_segundos=duracion,
                        )

                    except Exception as e:
                        duracion = round(time.time() - inicio, 2)
                        detalle_error = traceback.format_exc()

                        print(f"[sync-compras] ❌ Error en background: {e}")
                        traceback.print_exc()

                        _finalizar_log_sync_modulo(
                            log_id=local_log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle_error,
                            status_code=500,
                            duracion_segundos=duracion,
                        )

            threading.Thread(target=run_background, args=(log_id,), daemon=True).start()

            return jsonify({
                "mensaje": (
                    "Compras: proceso iniciado"
                    + (f" con fecha desde {since}." if since else " sin límite de fecha.")
                    + " Puedes revisar el resultado en el historial de sincronizaciones."
                ),
                "log_id": log_id,
                "origen": "manual_modulo",
                "estado": "EN_EJECUCION",
                "since": since,
            }), 202

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    @app.route("/siigo/debug-compra", methods=["GET"])
    @jwt_required()
    def siigo_debug_compra():
        from models import SiigoCompra, SiigoCompraItem

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        # superadmin puede ver otro cliente
        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado o sin cliente en el token"}), 403

        compra_id = request.args.get("idcompra")
        uuid = request.args.get("uuid")

        if not compra_id and not uuid:
            return jsonify({"error": "Proporciona ?idcompra= o ?uuid="}), 400

        # obtener credenciales y autenticar
        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales de Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if auth_data.get("_error"):
            return jsonify({"error": auth_data["_error"]}), 502

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo token de Siigo"}), 502

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        tried = []
        raw_detail = None

        try:
            if uuid:
                url = f"{base_url}/v1/support-documents/{uuid}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    raw_detail = r.json()
            elif compra_id:
                url = f"{base_url}/v1/support-documents?document={compra_id}"
                tried.append(url)
                r = requests.get(url, headers=headers, timeout=60)
                if r.status_code == 200:
                    results = r.json().get("results")
                    if results:
                        raw_detail = results[0]
        except Exception as e:
            return jsonify({"error": str(e), "tried": tried}), 502

        # buscar en la base de datos local
        compra_db = SiigoCompra.query.filter_by(idcliente=idcliente, idcompra=compra_id).first()
        items_db = []
        if compra_db:
            items_db = SiigoCompraItem.query.filter_by(compra_id=compra_db.id).all()

        def _dec(v):
            from decimal import Decimal
            if isinstance(v, Decimal): return float(v)
            if isinstance(v, datetime): return v.isoformat()
            return v

        def compra_to_dict(c):
            if not c: return None
            return {
                "id": c.id,
                "idcliente": c.idcliente,
                "idcompra": c.idcompra,
                "fecha": c.fecha.isoformat() if c.fecha else None,
                "vencimiento": c.vencimiento.isoformat() if c.vencimiento else None,
                "proveedor_nombre": c.proveedor_nombre,
                "estado": c.estado,
                "total": _dec(c.total),
                "saldo": _dec(c.saldo),
                "created_at": c.created_at.isoformat() if c.created_at else None,
            }

        def item_to_dict(i):
            return {
                "id": i.id,
                "compra_id": i.compra_id,
                "descripcion": i.descripcion,
                "cantidad": _dec(i.cantidad),
                "precio": _dec(i.precio),
                "impuestos": _dec(i.impuestos),
            }

        return jsonify({
            "query_params": {"uuid": uuid, "idcompra": compra_id, "idcliente": idcliente},
            "siigo_tried_urls": tried,
            "siigo_raw_detail": raw_detail,
            "db_compra": compra_to_dict(compra_db),
            "db_items": [item_to_dict(x) for x in items_db]
        })




    @app.route("/siigo/debug-purchases-raw", methods=["GET"])
    def siigo_debug_purchases_raw():
        # NO jwt_required, para pruebas
        # Aquí puedes pasar idcliente por query param si quieres
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/debug-purchases-raw?idcliente=1)"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        url = f"{base_url}/v1/purchases?page_size=20"  # número pequeño para ver ejemplo
        try:
            r = requests.get(url, headers=headers, timeout=60)
        except Exception as e:
            return jsonify({"error": "Error al hacer la petición HTTP", "detalle": str(e)}), 500

        status = r.status_code
        # Intenta parsear JSON, si no, devuelve texto
        try:
            body = r.json()
        except Exception:
            body = r.text

        return jsonify({
            "status": status,
            "url": url,
            "body": body
        }), status

    # Consultar detalle de un Recibo de Pago en la api de Siigo para ver detalle de lo que entrega - raw
    # se consulta asi: http://localhost:5000/siigo/debug-pago-raw?idcliente=1&idpago=713   (RP-1-713)
    @app.route("/siigo/debug-todos-pagos-raw", methods=["GET"])
    def siigo_debug_todos_pagos_raw():
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente"}), 400

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        token_data = _siigo_auth_json_for_client(cred)
        token = token_data.get("access_token")
        headers = _siigo_headers_bearer(token)
        base_url = cred.base_url.rstrip("/")

        # recorrer muchas páginas
        all_pagos = []
        page = 1
        page_size = 100
        while True:
            url = f"{base_url}/v1/payment-receipts?page={page}&page_size={page_size}"
            r = requests.get(url, headers=headers, timeout=60)
            if r.status_code != 200:
                break
            data = r.json()
            resultados = data.get("results") or []
            all_pagos.extend(resultados)
            if len(resultados) < page_size:
                break
            page += 1

        return jsonify({"pagos": all_pagos})




    # Buscar el ID real (UUID) del recibo de pago por nombre (ej: RP-1-713)
    @app.route("/siigo/debug-buscar-pago", methods=["GET"])
    def siigo_debug_buscar_pago():
        """
        Busca el ID real (UUID) del recibo de pago por nombre (ej: RP-1-713)
        Uso: /siigo/debug-buscar-pago?idcliente=1&nombre_pago=RP-1-713
        """
        idcliente = request.args.get("idcliente", type=int)
        nombre_pago = request.args.get("nombre_pago")

        if not idcliente or not nombre_pago:
            return jsonify({
                "error": "Faltan parámetros. Uso: /siigo/debug-buscar-pago?idcliente=1&nombre_pago=RP-1-713"
            }), 400

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        # Autenticación
        token_data = _siigo_auth_json_for_client(cred)
        if token_data.get("_error"):
            return jsonify({"error": token_data["_error"]}), 502

        token = token_data.get("access_token")
        headers = _siigo_headers_bearer(token)
        base_url = cred.base_url.rstrip("/")
        url = f"{base_url}/v1/payment-receipts?page_size=100"

        try:
            r = requests.get(url, headers=headers, timeout=60)
        except Exception as e:
            return jsonify({"error": "Error HTTP", "detalle": str(e)}), 500

        try:
            pagos = r.json().get("results", [])
        except Exception:
            return jsonify({"error": "Respuesta no es JSON", "detalle": r.text}), 500

        encontrados = []
        for pago in pagos:
            doc = pago.get("document", {})
            if doc.get("name") == nombre_pago:
                encontrados.append({
                    "uuid": pago.get("id"),
                    "documento": doc,
                    "total": pago.get("total"),
                    "fecha": pago.get("date"),
                    "proveedor": pago.get("provider", {}).get("name")
                })

        return jsonify({
            "busqueda": nombre_pago,
            "coincidencias": encontrados
        })





    @app.route("/siigo/debug-documentos-soporte-tipo", methods=["GET"])
    def siigo_debug_documentos_soporte_tipo():
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/debug-documentos-soporte-tipo?idcliente=1)"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        # 1. Obtener tipos de documento de compra
        doc_types_url = f"{base_url}/v1/document-types?type=FC"
        try:
            r_doc = requests.get(doc_types_url, headers=headers, timeout=60)
            doc_types = r_doc.json()
        except Exception as e:
            return jsonify({"error": "Error al consultar tipos de documento", "detalle": str(e)}), 500

        # 2. Filtrar solo los tipos que son documento soporte
        soporte_ids = [
            d["id"] for d in doc_types if d.get("document_support") is True
        ]

        # 3. Consultar /purchases
        purchases_url = f"{base_url}/v1/purchases?page_size=50"
        try:
            r = requests.get(purchases_url, headers=headers, timeout=60)
            compras_data = r.json()
        except Exception as e:
            return jsonify({"error": "Error al consultar compras", "detalle": str(e)}), 500

        # 4. Filtrar solo los que tienen document.id en soporte_ids
        resultados = compras_data.get("results", [])
        docs_soporte = [
            compra for compra in resultados
            if compra.get("document", {}).get("id") in soporte_ids
        ]

        return jsonify({
            "status": r.status_code,
            "document_support_ids": soporte_ids,
            "total_documentos_soporte": len(docs_soporte),
            "documentos_soporte": docs_soporte
        })


    @app.route("/siigo/debug-compras-json-completo", methods=["GET"])
    def siigo_debug_compras_json_completo():
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        url = f"{base_url}/v1/purchases?page_size=300"
        try:
            r = requests.get(url, headers=headers, timeout=60)
            return jsonify({
                "status": r.status_code,
                "url": url,
                "body": r.json()
            }), r.status_code
        except Exception as e:
            return jsonify({"error": "Fallo en la consulta", "detalle": str(e)}), 500




    @app.route("/siigo/debug-proveedores-raw", methods=["GET"])
    def siigo_debug_proveedores_raw():
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/debug-proveedores-raw?idcliente=1)"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        posibles_rutas = [
            "contacts",
        ]

        resultados = []
        for ruta in posibles_rutas:
            url = f"{base_url}/v1/{ruta}?page_size=20"
            try:
                r = requests.get(url, headers=headers, timeout=60)
            except Exception as e:
                resultados.append({
                    "ruta": ruta,
                    "status": "error",
                    "error": str(e),
                })
                continue

            status = r.status_code
            body = None
            try:
                body = r.json()
            except Exception:
                body = r.text

            resultados.append({
                "ruta": ruta,
                "status": status,
                "body": body,
            })

            if status == 200:
                # encontramos una ruta funcional, la devolvemos ya
                return jsonify({
                    "ruta_valida": ruta,
                    "url": url,
                    "status": status,
                    "body": body
                }), 200

        # Si ninguna funcionó, devolvemos todos los intentos para debugging
        return jsonify({
            "error": "No se encontró ruta válida para proveedores",
            "intentos": resultados
        }), 404


    # Endpoint consulta y prueba extracion proveedores en el Navegador (probar asi:  http://localhost:5000/siigo/debug-proveedores?idcliente=1)
    @app.route("/siigo/debug-proveedores", methods=["GET"])
    def siigo_debug_proveedores2():
        from datetime import datetime

        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/debug-proveedores?idcliente=1)"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        page = 1
        page_size = 100
        proveedores = []

        while True:
            url = f"{base_url}/v1/customers?page={page}&page_size={page_size}"
            try:
                r = requests.get(url, headers=headers, timeout=60)
            except Exception as e:
                return jsonify({"error": f"Error al conectar con Siigo: {str(e)}"}), 500

            if r.status_code != 200:
                return jsonify({
                    "error": "Error en respuesta de Siigo",
                    "status": r.status_code,
                    "body": r.text,
                    "url": url
                }), 500

            data = r.json()
            if not isinstance(data, list):
                data = data.get("results", [])

            solo_proveedores = [c for c in data if c.get("type") == "Supplier"]
            proveedores.extend(solo_proveedores)

            # ¿hay más páginas?
            if isinstance(r.json(), dict):
                links = r.json().get("_links", {})
                if not links.get("next") or not links["next"].get("href"):
                    break
            else:
                # No paginación si es lista directa
                break

            page += 1

        return jsonify({
            "total_proveedores": len(proveedores),
            "proveedores": proveedores[:20],  # solo los primeros 20 para visualización
            "timestamp": datetime.utcnow().isoformat() + "Z"
        }), 200



    # realiza la sincronización de proveedores trayendo proveedores, clientes y otros
    @app.route("/siigo/sync-proveedores", methods=["POST"])
    def siigo_sync_proveedores():
        idcliente = obtener_idcliente_desde_request()
        print(f"🔹 Sync proveedores iniciado para cliente {idcliente}")

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        # Cuando este endpoint es llamado desde /siigo/sync-all,
        # NO debe crear log individual para no duplicar el historial.
        # El log general lo crea /siigo/sync-all al final.
        es_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_actual = "/siigo/sync-proveedores"
        ejecutado_en = datetime.now(timezone.utc)

        detalle_lines = []
        pasos_ok = 0
        pasos_error = 0
        endpoint_fallido = None

        def guardar_log(resultado: str, detalle: str):
            """
            Guarda historial solo cuando la ejecución es manual por módulo.
            Si viene desde sync-all, no guarda log individual para evitar duplicados.
            """
            if es_sync_all:
                return None

            try:
                log = SiigoSyncLog(
                    idcliente=idcliente,
                    fecha_programada=ejecutado_en,
                    ejecutado_en=ejecutado_en,
                    origen="manual_modulo",
                    resultado=resultado,
                    total_pasos=1,
                    pasos_ok=pasos_ok,
                    pasos_error=pasos_error,
                    endpoint_fallido=endpoint_fallido,
                    detalle=detalle,
                )
                db.session.add(log)
                db.session.commit()
                return log.id

            except Exception as log_error:
                db.session.rollback()
                raise Exception(f"Error guardando log de sincronización proveedores: {str(log_error)}")

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()

        if not cred:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ Credenciales Siigo no configuradas para el cliente."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Credenciales Siigo no configuradas",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "Credenciales Siigo no configuradas",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        access_key = dec(cred.client_secret)

        if not access_key:
            pasos_error = 1
            endpoint_fallido = endpoint_actual
            detalle = "❌ No se pudo desencriptar access_key de Siigo."

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "No se pudo desencriptar access_key",
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": "No se pudo desencriptar access_key",
                "log_id": log_id,
                "detalle": detalle,
            }), 400

        try:
            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data["access_token"]
            headers = _headers_bearer(token)

            base_url = cred.base_url.rstrip("/")
            page = 1
            page_size = 100
            total_insertados = 0
            total_actualizados = 0
            total_paginas = 0

            while True:
                url = f"{base_url}/v1/customers?page={page}&page_size={page_size}"
                r = requests.get(url, headers=headers)

                if r.status_code != 200:
                    pasos_error += 1
                    endpoint_fallido = endpoint_fallido or "/v1/customers"
                    detalle_lines.append(
                        f"❌ GET /v1/customers page={page} page_size={page_size} -> {r.status_code}. "
                        f"Respuesta: {r.text[:1000]}"
                    )
                    break

                payload = r.json()
                results = payload if isinstance(payload, list) else payload.get("results", [])

                if not results:
                    detalle_lines.append(
                        f"ℹ️ GET /v1/customers page={page} -> {r.status_code}. "
                        f"Sin más proveedores/terceros para sincronizar."
                    )
                    break

                total_paginas += 1
                procesados_pagina = 0

                for c in results:
                    # Si no tiene identificación, ignorar
                    identificacion = c.get("identification")
                    if not identificacion:
                        continue

                    nombre_raw = c.get("name", [])

                    if isinstance(nombre_raw, list):
                        nombre = " ".join([str(n) for n in nombre_raw if n is not None]).strip()
                    else:
                        nombre = str(nombre_raw or "").strip()

                    tipo_ident = quitar_tildes(c.get("id_type", {}).get("name", ""))
                    dv = c.get("check_digit", "")
                    direccion = c.get("address", {}).get("address", "")
                    ciudad = quitar_tildes(c.get("address", {}).get("city", {}).get("city_name", ""))
                    telefonos = c.get("phones", [])
                    telefono = ""

                    if telefonos:
                        tel = telefonos[0]
                        telefono = f"{tel.get('indicative', '')} {tel.get('number', '')}".strip()

                    estado = "Activo" if c.get("active") else "Inactivo"

                    stmt = insert(SiigoProveedor).values(
                        idcliente=idcliente,
                        nombre=nombre,
                        tipo_identificacion=tipo_ident,
                        identificacion=identificacion,
                        digito_verificacion=dv,
                        direccion=direccion,
                        ciudad=ciudad,
                        telefono=telefono,
                        estado=estado
                    ).on_conflict_do_update(
                        index_elements=["idcliente", "identificacion"],
                        set_={
                            "nombre": nombre,
                            "tipo_identificacion": tipo_ident,
                            "digito_verificacion": dv,
                            "direccion": direccion,
                            "ciudad": ciudad,
                            "telefono": telefono,
                            "estado": estado
                        }
                    )

                    res = db.session.execute(stmt)

                    if res.rowcount == 1:
                        total_insertados += 1
                    else:
                        total_actualizados += 1

                    procesados_pagina += 1

                detalle_lines.append(
                    f"✅ GET /v1/customers page={page} page_size={page_size} -> {r.status_code}. "
                    f"Terceros procesados en página: {procesados_pagina}. "
                    f"Acumulado insertados: {total_insertados}, actualizados: {total_actualizados}."
                )

                if "_links" in payload and payload["_links"].get("next"):
                    page += 1
                else:
                    break

            if pasos_error == 0:
                pasos_ok = 1

            db.session.commit()

            detalle_lines.append(
                f"📊 Resumen proveedores: páginas procesadas: {total_paginas}, "
                f"insertados: {total_insertados}, actualizados: {total_actualizados}, "
                f"errores: {pasos_error}."
            )

            detalle = "\n".join(detalle_lines)
            resultado_log = "OK" if pasos_error == 0 else "ERROR"

            try:
                log_id = guardar_log(resultado_log, detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Los proveedores se procesaron, pero falló el registro del historial.",
                    "detalle": detalle,
                    "error_log": str(log_error),
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                }), 500

            if pasos_error > 0:
                return jsonify({
                    "mensaje": (
                        f"Proveedores procesados con alertas: "
                        f"{total_insertados}, actualizados: {total_actualizados}"
                    ),
                    "estado": "ERROR",
                    "log_id": log_id,
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                    "endpoint_fallido": endpoint_fallido,
                    "detalle": detalle,
                    "resumen": {
                        "proveedores_insertados": total_insertados,
                        "proveedores_actualizados": total_actualizados,
                        "paginas_procesadas": total_paginas,
                    }
                }), 200

            return jsonify({
                "mensaje": f"Proveedores sincronizados: {total_insertados}, actualizados: {total_actualizados}",
                "estado": "OK",
                "log_id": log_id,
                "pasos_ok": pasos_ok,
                "pasos_error": pasos_error,
                "detalle": detalle,
                "resumen": {
                    "proveedores_insertados": total_insertados,
                    "proveedores_actualizados": total_actualizados,
                    "paginas_procesadas": total_paginas,
                }
            }), 200

        except Exception as e:
            db.session.rollback()

            pasos_error = max(pasos_error, 1)
            endpoint_fallido = endpoint_fallido or endpoint_actual
            detalle_lines.append(f"❌ Error general en sincronización de proveedores: {str(e)}")
            detalle = "\n".join(detalle_lines)

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": str(e),
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": str(e),
                "detalle": detalle,
                "log_id": log_id,
            }), 500


    # PReview en pagina de los proveedores que se cargaran
    @app.route("/siigo/preview-proveedores", methods=["GET"])
    def siigo_preview_proveedores():
        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/preview-proveedores?idcliente=1)"}), 400

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        access_key = dec(cred.client_secret)
        if not access_key:
            return jsonify({"error": "No se pudo desencriptar access_key"}), 400

        try:
            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data["access_token"]
            headers = _headers_bearer(token)

            base_url = cred.base_url.rstrip("/")
            page = 1
            page_size = 100
            proveedores = []

            while True:
                url = f"{base_url}/v1/customers?page={page}&page_size={page_size}"
                r = requests.get(url, headers=headers)
                if r.status_code != 200:
                    break

                payload = r.json()
                results = payload if isinstance(payload, list) else payload.get("results", [])
                if not results:
                    break

                for c in results:
                    if c.get("type") != "Supplier":
                        continue

                    nombre = " ".join(c.get("name", [])).strip()
                    tipo_ident = quitar_tildes(c.get("id_type", {}).get("name", ""))
                    identificacion = c.get("identification", "")
                    dv = c.get("check_digit", "")
                    direccion = c.get("address", {}).get("address", "")
                    ciudad = quitar_tildes(c.get("address", {}).get("city", {}).get("city_name", ""))
                    telefonos = c.get("phones", [])
                    telefono = ""
                    if telefonos:
                        tel = telefonos[0]
                        telefono = f"{tel.get('indicative', '')} {tel.get('number', '')}".strip()

                    estado = "Activo" if c.get("active") else "Inactivo"

                    proveedor = {
                        "nombre": nombre,
                        "tipo_identificacion": tipo_ident,
                        "identificacion": identificacion,
                        "digito_verificacion": dv,
                        "direccion": direccion,
                        "ciudad": ciudad,
                        "telefono": telefono,
                        "estado": estado
                    }

                    proveedores.append(proveedor)

                if "_links" in payload and payload["_links"].get("next"):
                    page += 1
                else:
                    break

            return jsonify({
                "proveedores": proveedores,
                "total": len(proveedores)
            }), 200

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    @app.route("/siigo/cargar-proveedores", methods=["POST"])
    @jwt_required()
    def cargar_proveedores_excel():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        if "archivo" not in request.files:
            return jsonify({"error": "Archivo no proporcionado"}), 400

        file = request.files["archivo"]

        try:
            # Usa la fila 6 como encabezado
            df = pd.read_excel(file, header=6)
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        requeridos = [
            "Nombre tercero", "Tipo de identificación", "Identificación",
            "Digito verificación", "Dirección", "Ciudad", "Teléfono.", "Estado"
        ]
        faltantes = [col for col in requeridos if col not in df.columns]
        if faltantes:
            return jsonify({"error": f"Faltan columnas requeridas en el archivo: {', '.join(faltantes)}"}), 400

        nuevos = 0
        actualizados = 0
        for _, row in df.iterrows():
            identificacion = str(row["Identificación"]).strip()
            proveedor = SiigoProveedor.query.filter_by(idcliente=idcliente, identificacion=identificacion).first()

            datos = dict(
                nombre=str(row["Nombre tercero"]).strip(),
                tipo_identificacion=str(row["Tipo de identificación"]).strip(),
                digito_verificacion=str(row.get("Digito verificación", "")).strip(),
                direccion=str(row.get("Dirección", "")).strip(),
                ciudad=str(row.get("Ciudad", "")).strip(),
                telefono=str(row.get("Teléfono.", "")).strip(),
                estado=str(row.get("Estado", "")).strip(),
            )

            if proveedor:
                for k, v in datos.items():
                    setattr(proveedor, k, v)
                actualizados += 1
            else:
                db.session.add(SiigoProveedor(idcliente=idcliente, identificacion=identificacion, **datos))
                nuevos += 1

        db.session.commit()
        return jsonify({"mensaje": f"Proveedores cargados correctamente. Nuevos: {nuevos}, Actualizados: {actualizados}"})



    #Reporte de Compras a Proveedores
    # --- ENDPOINT: Reporte General de Compras por Proveedor ---
    @app.route("/reportes/compras/proveedores", methods=["GET"])
    @jwt_required()
    def reporte_compras_proveedores():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        incluir_detalle = request.args.get("detalle", "0") == "1"
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        estado = request.args.get("estado")  # "pagado" | "pendiente"

        condiciones = ["idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        if desde and validar_fecha(desde):
            condiciones.append("fecha >= :desde")
            params["desde"] = desde
        if hasta and validar_fecha(hasta):
            condiciones.append("fecha <= :hasta")
            params["hasta"] = hasta

        # --- Filtro por estado real de siigo_compras ---
        if estado:
            estado_norm = estado.strip().lower()
            if estado_norm in ["pagado", "pendiente"]:
                condiciones.append("LOWER(estado) = :estado")
                params["estado"] = estado_norm

        where_sql = " AND ".join(condiciones)

        # --- Resumen por proveedor ---
        query = f"""
            SELECT
                COALESCE(proveedor_identificacion, '') AS proveedor_identificacion,
                COALESCE(proveedor_nombre, '') AS proveedor_nombre,
                COUNT(*) AS num_compras,
                SUM(COALESCE(total, 0)) AS total_compras,
                SUM(CASE WHEN estado = 'pendiente' THEN saldo ELSE 0 END) AS total_saldo,
                MAX(fecha) AS ultima_fecha
            FROM siigo_compras
            WHERE {where_sql}
            GROUP BY proveedor_identificacion, proveedor_nombre
            HAVING SUM(COALESCE(total, 0)) > 0
            ORDER BY total_compras DESC
        """
        rows = db.session.execute(text(query), params).mappings().all()
        resultado = [dict(r) for r in rows]

        for r in resultado:
            r["total_pagado"] = float(r["total_compras"] or 0) - float(r["total_saldo"] or 0)

        detalle = []
        if incluir_detalle:
            query_detalle = f"""
                SELECT
                    idcompra,
                    factura_proveedor,
                    proveedor_identificacion,
                    proveedor_nombre,
                    fecha,
                    vencimiento,
                    total,
                    saldo,
                    estado
                FROM siigo_compras
                WHERE {where_sql}
                ORDER BY proveedor_nombre, fecha DESC
            """
            rows_detalle = db.session.execute(text(query_detalle), params).mappings().all()
            detalle = [dict(r) for r in rows_detalle]

        return jsonify({
            "resumen": resultado,
            "detalle": detalle if incluir_detalle else None
        })




    # --- ENDPOINT: Reporte de Ítems de Compra ---
    @app.route("/reportes/compras/items", methods=["GET"])
    @jwt_required()
    def reporte_compras_items():
        from sqlalchemy.sql import text

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        proveedor = request.args.get("proveedor")
        codigo = request.args.get("codigo")

        condiciones = ["sc.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde:
            condiciones.append("sc.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            condiciones.append("sc.fecha <= :hasta")
            params["hasta"] = hasta
        if proveedor:
            condiciones.append("sc.proveedor_identificacion = :proveedor")
            params["proveedor"] = proveedor
        if codigo:
            condiciones.append("sci.codigo = :codigo")
            params["codigo"] = codigo

        where_sql = " AND ".join(condiciones)

        query = f"""
            SELECT
                sci.descripcion,
                sci.codigo,
                SUM(sci.cantidad) AS total_cantidad,
                SUM(sci.precio * sci.cantidad) AS total_gastado,
                CASE WHEN SUM(sci.cantidad) > 0 THEN
                    SUM(sci.precio * sci.cantidad) / SUM(sci.cantidad)
                ELSE 0 END AS precio_promedio
            FROM siigo_compras_items sci
            JOIN siigo_compras sc ON sci.compra_id = sc.id
            WHERE {where_sql}
            GROUP BY sci.descripcion, sci.codigo
            ORDER BY total_gastado DESC
        """

        rows = db.session.execute(text(query), params).mappings().all()
        resultado = [dict(r) for r in rows]

        return jsonify({"items": resultado})



    # --- ENDPOINT: Reporte Financiero Compras y Gastos ---
    # --- ENDPOINT: Reporte Financiero Compras y Gastos ---
    # --- ENDPOINT 1: Reporte Financiero Compras y Gastos (KPIs + Evolución Mensual) ---
    @app.route("/reportes/financiero/compras-gastos", methods=["GET"])
    @jwt_required()
    def reporte_financiero_compras_gastos():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos")

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        desde_valida = validar_fecha(desde) if desde else None
        hasta_valida = validar_fecha(hasta) if hasta else None

        condiciones = ["c.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde_valida:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = desde_valida
        if hasta_valida:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = hasta_valida
        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        # --- Evolución mensual (contable) ---
        query_evolucion = f"""
            SELECT
                date_trunc('month', c.fecha) AS mes,
                SUM(COALESCE(c.total,0)) AS total_compras,
                SUM(COALESCE(c.total,0) - COALESCE(c.saldo,0)) AS total_pagadas,
                SUM(COALESCE(c.saldo,0)) AS total_pendientes
            FROM siigo_compras c
            WHERE {where_sql}
            GROUP BY mes
            ORDER BY mes
        """
        rows_evol = db.session.execute(text(query_evolucion), params).mappings().all()

        # --- KPIs generales (contables + parciales) ---
        query_kpis = f"""
            SELECT
                COUNT(*) AS total_facturas,
                SUM(COALESCE(c.total,0)) AS total_compras,

                -- Contable (incluye parciales)
                SUM(COALESCE(c.total,0) - COALESCE(c.saldo,0)) AS total_pagado,
                SUM(COALESCE(c.saldo,0)) AS total_saldo,

                -- Conteos contables por saldo
                SUM(CASE WHEN COALESCE(c.saldo,0) = 0 THEN 1 ELSE 0 END) AS facturas_pagadas,
                SUM(CASE WHEN COALESCE(c.total,0) > 0 AND COALESCE(c.saldo,0) >= COALESCE(c.total,0) THEN 1 ELSE 0 END) AS facturas_pendientes,
                SUM(CASE WHEN COALESCE(c.saldo,0) > 0 AND COALESCE(c.saldo,0) < COALESCE(c.total,0) THEN 1 ELSE 0 END) AS facturas_parciales,

                -- Valor parcial (saldo de parciales)
                SUM(CASE WHEN COALESCE(c.saldo,0) > 0 AND COALESCE(c.saldo,0) < COALESCE(c.total,0) THEN COALESCE(c.saldo,0) ELSE 0 END) AS saldo_parcial,

                -- KPIs por tipo documento (siguen por total, no por estado)
                SUM(CASE WHEN c.idcompra LIKE 'FC-%' THEN 1 ELSE 0 END) AS compras_x_factura,
                SUM(CASE WHEN c.idcompra LIKE 'FC-%' THEN COALESCE(c.total,0) ELSE 0 END) AS valor_compras_x_factura,
                SUM(CASE WHEN c.idcompra LIKE 'DS-%' THEN 1 ELSE 0 END) AS compras_x_cta_cobro,
                SUM(CASE WHEN c.idcompra LIKE 'DS-%' THEN COALESCE(c.total,0) ELSE 0 END) AS valor_compras_x_cta_cobro
            FROM siigo_compras c
            WHERE {where_sql}
        """
        row_kpis = db.session.execute(text(query_kpis), params).mappings().first()

        return jsonify({
            "kpis": dict(row_kpis or {}),
            "evolucion": [dict(r) for r in rows_evol]
        })


    # --- ENDPOINT 2: Top 15 Proveedores por valor de compras ---
    @app.route("/reportes/financiero/compras-gastos/top-proveedores", methods=["GET"])
    @jwt_required()
    def top_proveedores_compras():
        from sqlalchemy.sql import text
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos")

        condiciones = ["c.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = hasta
        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT
                c.proveedor_nombre,
                COUNT(*) AS num_facturas,
                SUM(c.total) AS total_compras
            FROM siigo_compras c
            WHERE {where_sql}
            GROUP BY c.proveedor_nombre
            ORDER BY total_compras DESC
            LIMIT 15
        """)
        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify(rows)


    # --- ENDPOINT 3: Top 15 Proveedores por número de facturas ---
    @app.route("/reportes/financiero/compras-gastos/top-proveedores-facturas", methods=["GET"])
    @jwt_required()
    def top_proveedores_facturas():
        from sqlalchemy.sql import text
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos")

        condiciones = ["c.idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        if desde:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = hasta
        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT
                c.proveedor_nombre,
                COUNT(*) AS num_facturas
            FROM siigo_compras c
            WHERE {where_sql}
            GROUP BY c.proveedor_nombre
            ORDER BY num_facturas DESC
            LIMIT 15
        """)
        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify(rows)


    # --- ENDPOINT 4: Detalle de facturas por mes ---
    # --- ENDPOINT 4: Detalle de facturas por mes ---
    @app.route("/reportes/financiero/compras-gastos/detalle", methods=["GET"])
    @jwt_required()
    def detalle_facturas_mes():
        from sqlalchemy.sql import text

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        mes = request.args.get("mes")
        estado = request.args.get("estado")
        centro_costos = request.args.get("centro_costos")
        tipo_documento = request.args.get("tipo_documento", "todos")

        if not mes:
            return jsonify({"error": "Mes requerido"}), 400

        condiciones = [
            "c.idcliente = :idcliente",
            "TO_CHAR(c.fecha, 'YYYY-MM') = :mes"
        ]
        params = {"idcliente": idcliente, "mes": mes}

        if estado == "pagado":
            condiciones.append("COALESCE(c.saldo,0) = 0")
        elif estado == "pendiente":
            condiciones.append("COALESCE(c.total,0) > 0 AND COALESCE(c.saldo,0) >= COALESCE(c.total,0)")
        elif estado == "parcial":
            condiciones.append("COALESCE(c.saldo,0) > 0 AND COALESCE(c.saldo,0) < COALESCE(c.total,0)")

        if tipo_documento == "factura":
            condiciones.append("c.idcompra ILIKE 'FC-%'")
        elif tipo_documento == "documento_soporte":
            condiciones.append("c.idcompra ILIKE 'DS-%'")

        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT
                c.id,
                c.proveedor_nombre,
                c.idcompra AS factura,
                c.factura_proveedor,
                TO_CHAR(c.fecha::date, 'YYYY-MM-DD') AS fecha,
                TO_CHAR(c.vencimiento::date, 'YYYY-MM-DD') AS vencimiento,
                COALESCE(c.estado, '') AS estado_raw,
                COALESCE(c.total,0) AS total,
                COALESCE(c.saldo,0) AS saldo,
                (COALESCE(c.total,0) - COALESCE(c.saldo,0)) AS pagado_calc,

                CASE
                    WHEN COALESCE(c.saldo,0) = 0 THEN 'pagado'
                    WHEN COALESCE(c.total,0) > 0 AND COALESCE(c.saldo,0) >= COALESCE(c.total,0) THEN 'pendiente'
                    WHEN COALESCE(c.saldo,0) > 0 AND COALESCE(c.saldo,0) < COALESCE(c.total,0) THEN 'parcial'
                    ELSE 'pendiente'
                END AS estado_calc,

                CASE
                    WHEN COALESCE(c.total,0) > 0 AND COALESCE(c.saldo,0) > COALESCE(c.total,0) THEN true
                    ELSE false
                END AS anomalia_saldo_mayor_total,

                CASE
                    WHEN c.idcompra ILIKE 'FC-%' THEN 'factura'
                    WHEN c.idcompra ILIKE 'DS-%' THEN 'documento_soporte'
                    ELSE 'otro'
                END AS tipo_documento,

                sc.nombre AS centro_costo_nombre
            FROM siigo_compras c
            LEFT JOIN siigo_centros_costo sc
                ON c.cost_center = sc.id
            AND sc.idcliente = c.idcliente
            WHERE {where_sql}
            ORDER BY c.fecha DESC, c.proveedor_nombre ASC, c.idcompra ASC, c.id ASC
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify(rows)


    # --- ENDPOINT 5: Detalle de facturas por proveedor ---
    @app.route("/reportes/financiero/compras-gastos/detalle-proveedor", methods=["GET"])
    @jwt_required()
    def detalle_proveedor():
        from sqlalchemy.sql import text

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        proveedor = request.args.get("proveedor")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos")
        tipo_documento = request.args.get("tipo_documento", "todos")

        if not proveedor:
            return jsonify({"error": "Proveedor requerido"}), 400

        condiciones = [
            "c.idcliente = :idcliente",
            "c.proveedor_nombre = :proveedor"
        ]
        params = {"idcliente": idcliente, "proveedor": proveedor}

        if desde:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = desde

        if hasta:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = hasta

        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        if tipo_documento == "factura":
            condiciones.append("c.idcompra ILIKE 'FC-%'")
        elif tipo_documento == "documento_soporte":
            condiciones.append("c.idcompra ILIKE 'DS-%'")

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT 
                c.id,
                c.idcompra AS idcompra,
                c.proveedor_nombre,
                c.idcompra AS factura,
                c.factura_proveedor,
                TO_CHAR(c.fecha::date, 'YYYY-MM-DD') AS fecha,
                TO_CHAR(c.vencimiento::date, 'YYYY-MM-DD') AS vencimiento,
                COALESCE(c.total, 0) AS total,
                COALESCE(c.saldo, 0) AS saldo,
                c.estado AS estado_raw,
                (COALESCE(c.total,0) - COALESCE(c.saldo,0)) AS pagado_calc,

                CASE
                    WHEN COALESCE(c.saldo,0) <= 0 THEN 'pagado'
                    WHEN COALESCE(c.saldo,0) >= COALESCE(c.total,0) THEN 'pendiente'
                    ELSE 'parcial'
                END AS estado_calc,

                CASE
                    WHEN COALESCE(c.saldo,0) > COALESCE(c.total,0) THEN true
                    ELSE false
                END AS anomalia_saldo_mayor_total,

                CASE
                    WHEN c.idcompra ILIKE 'FC-%' THEN 'factura'
                    WHEN c.idcompra ILIKE 'DS-%' THEN 'documento_soporte'
                    ELSE 'otro'
                END AS tipo_documento,

                sc.nombre AS centro_costo_nombre
            FROM siigo_compras c
            LEFT JOIN siigo_centros_costo sc
                ON c.cost_center = sc.id
                AND sc.idcliente = c.idcliente
            WHERE {where_sql}
            ORDER BY c.fecha DESC, c.proveedor_nombre ASC, c.idcompra ASC, c.id ASC
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify(rows)


    # Endpoint para llamar Centros de costos enunciados en la BD de Siigo Compras
    @app.route("/catalogos/centros-costo-reales", methods=["GET"])
    @jwt_required()
    def catalogo_centros_costo_reales():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception as e:
                app.logger.warning(f"⚠️ Fecha inválida: {fecha_str} — {e}")
                return None

        desde_valida = validar_fecha(desde) if desde else None
        hasta_valida = validar_fecha(hasta) if hasta else None

        condiciones = ["c.idcliente = :idcliente", "c.cost_center IS NOT NULL"]
        params = {"idcliente": idcliente}

        if desde_valida:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = desde_valida
        if hasta_valida:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = hasta_valida

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT DISTINCT
                c.cost_center AS id,
                COALESCE(cc.nombre, 'Sin centro de costo') AS nombre
            FROM siigo_compras c
            LEFT JOIN siigo_centros_costo cc ON c.cost_center = cc.id
            WHERE {where_sql}
            ORDER BY nombre
        """)

        try:
            app.logger.info(f"🔍 Ejecutando consulta de centros de costo reales — SQL: {sql} — Params: {params}")
            rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
            return jsonify(rows)
        except Exception as e:
            app.logger.error(f"❌ Error al ejecutar consulta de centros-costo-reales — {e}", exc_info=True)
            return jsonify({
                "error": "Error interno al consultar centros de costo reales.",
                "detalle": str(e)
            }), 500




    # --- ENDPOINT: Reporte Financiero Consolidado ---
    # ============================================================
    # ENDPOINT: Reporte Financiero Consolidado
    # Ingresos corregidos con ventas_movimientos_enriquecidos
    # ============================================================

    @app.route("/reportes/financiero/consolidado", methods=["GET"])
    @jwt_required()
    def reporte_financiero_consolidado():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        # -------- Filtros --------
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos", type=int)

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        fecha_desde_val = validar_fecha(desde) if desde else None
        fecha_hasta_val = validar_fecha(hasta) if hasta else None

        # ============================================================
        # 1) Ingresos comerciales tipo Siigo
        # Fuente: ventas_movimientos_enriquecidos
        # Regla: ingresos = facturas emitidas - notas crédito, CON impuesto
        # ============================================================

        condiciones_ing = ["m.idcliente = :idcliente"]
        params_ing = {"idcliente": idcliente}

        if fecha_desde_val:
            condiciones_ing.append("m.fecha >= :desde")
            params_ing["desde"] = fecha_desde_val

        if fecha_hasta_val:
            condiciones_ing.append("m.fecha <= :hasta")
            params_ing["hasta"] = fecha_hasta_val

        if centro_costos:
            condiciones_ing.append("m.cost_center = :centro_costos")
            params_ing["centro_costos"] = centro_costos

        where_ing = " AND ".join(condiciones_ing)

        sql_ingresos = text(f"""
            SELECT
                date_trunc('month', m.fecha)::date AS mes,

                -- Ingresos netos con impuesto, igual a Siigo con "Incluye impuesto"
                COALESCE(SUM(m.total), 0) AS ingresos,

                -- Base sin impuesto, útil para auditoría
                COALESCE(SUM(m.subtotal), 0) AS ingresos_sin_impuesto,

                -- Impuesto neto comercial
                COALESCE(SUM(m.total) - SUM(m.subtotal), 0) AS impuestos_netos,

                -- Facturas emitidas antes de notas crédito
                COALESCE(SUM(
                    CASE WHEN m.tipo_movimiento = 'FACTURA'
                    THEN m.total ELSE 0 END
                ), 0) AS facturas_emitidas,

                -- Notas crédito del periodo en positivo
                ABS(COALESCE(SUM(
                    CASE WHEN m.tipo_movimiento = 'NOTA_CREDITO'
                    THEN m.total ELSE 0 END
                ), 0)) AS notas_credito,

                COUNT(*) FILTER (WHERE m.tipo_movimiento = 'FACTURA') AS facturas_venta,
                COUNT(*) FILTER (WHERE m.tipo_movimiento = 'NOTA_CREDITO') AS notas_credito_count

            FROM ventas_movimientos_enriquecidos m
            WHERE {where_ing}
            GROUP BY date_trunc('month', m.fecha)::date
        """)

        ingresos_rows = db.session.execute(sql_ingresos, params_ing).mappings().all()

        # ============================================================
        # 2) Egresos compras/gastos
        # Fuente: siigo_compras, como viene actualmente
        # ============================================================

        condiciones_egr = ["c.idcliente = :idcliente"]
        params_egr = {"idcliente": idcliente}

        if fecha_desde_val:
            condiciones_egr.append("c.fecha >= :desde")
            params_egr["desde"] = fecha_desde_val

        if fecha_hasta_val:
            condiciones_egr.append("c.fecha <= :hasta")
            params_egr["hasta"] = fecha_hasta_val

        if centro_costos:
            condiciones_egr.append("c.cost_center = :centro_costos")
            params_egr["centro_costos"] = centro_costos

        where_egr = " AND ".join(condiciones_egr)

        sql_egresos = text(f"""
            SELECT
                date_trunc('month', c.fecha)::date AS mes,
                COALESCE(SUM(c.total), 0) AS egresos,
                COUNT(*) AS facturas_compra
            FROM siigo_compras c
            WHERE {where_egr}
            GROUP BY date_trunc('month', c.fecha)::date
        """)

        egresos_rows = db.session.execute(sql_egresos, params_egr).mappings().all()

        # ============================================================
        # 3) Nómina
        # ============================================================

        condiciones_nomina = ["idcliente = :idcliente"]
        params_nomina = {"idcliente": idcliente}

        if fecha_desde_val:
            condiciones_nomina.append("periodo >= :desde_nomina")
            params_nomina["desde_nomina"] = fecha_desde_val

        if fecha_hasta_val:
            condiciones_nomina.append("periodo <= :hasta_nomina")
            params_nomina["hasta_nomina"] = fecha_hasta_val

        where_nomina = " AND ".join(condiciones_nomina)

        sql_nomina = text(f"""
            SELECT
                date_trunc('month', periodo)::date AS mes,
                COALESCE(SUM(total_ingresos), 0) AS nomina
            FROM siigo_nomina
            WHERE {where_nomina}
            GROUP BY date_trunc('month', periodo)::date
        """)

        nomina_rows = db.session.execute(sql_nomina, params_nomina).mappings().all()

        # ============================================================
        # 4) Merge mensual
        # ============================================================

        ingresos_dict = {str(r["mes"]): dict(r) for r in ingresos_rows}
        egresos_dict = {str(r["mes"]): dict(r) for r in egresos_rows}
        nomina_dict = {str(r["mes"]): dict(r) for r in nomina_rows}

        meses = sorted(
            set(ingresos_dict.keys())
            | set(egresos_dict.keys())
            | set(nomina_dict.keys())
        )

        evolucion = []

        total_ingresos = 0
        total_ingresos_sin_impuesto = 0
        total_impuestos_netos = 0
        total_facturas_emitidas = 0
        total_notas_credito = 0
        total_egresos = 0
        total_nomina = 0
        facturas_venta = 0
        facturas_compra = 0
        notas_credito_count = 0
        utilidad_acumulada = 0

        for mes in meses:
            ing = ingresos_dict.get(mes, {
                "ingresos": 0,
                "ingresos_sin_impuesto": 0,
                "impuestos_netos": 0,
                "facturas_emitidas": 0,
                "notas_credito": 0,
                "facturas_venta": 0,
                "notas_credito_count": 0,
            })

            egr = egresos_dict.get(mes, {
                "egresos": 0,
                "facturas_compra": 0,
            })

            nom = nomina_dict.get(mes, {
                "nomina": 0,
            })

            ingresos = ing["ingresos"] or 0
            ingresos_sin_impuesto = ing["ingresos_sin_impuesto"] or 0
            impuestos_netos = ing["impuestos_netos"] or 0
            facturas_emitidas = ing["facturas_emitidas"] or 0
            notas_credito = ing["notas_credito"] or 0

            egresos_base = egr["egresos"] or 0
            nomina_mes = nom["nomina"] or 0
            egresos = egresos_base + nomina_mes

            utilidad = ingresos - egresos
            margen = (utilidad / ingresos * 100) if ingresos > 0 else 0

            utilidad_acumulada += utilidad

            evolucion.append({
                "mes": mes,

                # Para mantener compatibilidad con la página:
                "ingresos": ingresos,
                "ingresos_netos": ingresos,

                # Nuevos campos claros:
                "ingresos_con_impuesto": ingresos,
                "ingresos_sin_impuesto": ingresos_sin_impuesto,
                "impuestos_netos": impuestos_netos,
                "facturas_emitidas": facturas_emitidas,
                "notas_credito": notas_credito,

                "egresos": egresos,
                "egresos_base": egresos_base,
                "nomina": nomina_mes,

                "utilidad": utilidad,
                "margen": round(margen, 2),
                "utilidad_acumulada": utilidad_acumulada,

                "facturas_venta": ing["facturas_venta"] or 0,
                "notas_credito_count": ing["notas_credito_count"] or 0,
                "facturas_compra": egr["facturas_compra"] or 0,
            })

            total_ingresos += ingresos
            total_ingresos_sin_impuesto += ingresos_sin_impuesto
            total_impuestos_netos += impuestos_netos
            total_facturas_emitidas += facturas_emitidas
            total_notas_credito += notas_credito
            total_egresos += egresos
            total_nomina += nomina_mes
            facturas_venta += ing["facturas_venta"] or 0
            notas_credito_count += ing["notas_credito_count"] or 0
            facturas_compra += egr["facturas_compra"] or 0

        # ============================================================
        # 5) KPIs globales
        # ============================================================

        utilidad_total = total_ingresos - total_egresos
        margen_total = (utilidad_total / total_ingresos * 100) if total_ingresos > 0 else 0

        kpis = {
            # Compatibilidad con la página actual:
            "ingresos": total_ingresos,
            "ingresos_netos": total_ingresos,

            # Nuevos campos claros:
            "ingresos_con_impuesto": total_ingresos,
            "ingresos_sin_impuesto": total_ingresos_sin_impuesto,
            "impuestos_netos": total_impuestos_netos,
            "facturas_emitidas": total_facturas_emitidas,
            "notas_credito": total_notas_credito,

            "egresos": total_egresos,
            "nomina": total_nomina,
            "utilidad": utilidad_total,
            "margen": round(margen_total, 2),
            "facturas_venta": facturas_venta,
            "notas_credito_count": notas_credito_count,
            "facturas_compra": facturas_compra,
        }

        # ============================================================
        # 6) Top clientes
        # Fuente: movimientos comerciales con impuesto
        # ============================================================

        sql_top_clientes = text(f"""
            SELECT
                m.cliente_nombre AS nombre,

                -- Venta neta con impuesto
                COALESCE(SUM(m.total), 0) AS total,

                COALESCE(SUM(
                    CASE WHEN m.tipo_movimiento = 'FACTURA'
                    THEN m.total ELSE 0 END
                ), 0) AS facturas_emitidas,

                ABS(COALESCE(SUM(
                    CASE WHEN m.tipo_movimiento = 'NOTA_CREDITO'
                    THEN m.total ELSE 0 END
                ), 0)) AS notas_credito

            FROM ventas_movimientos_enriquecidos m
            WHERE {where_ing}
            GROUP BY m.cliente_nombre
            ORDER BY total DESC
            LIMIT 10
        """)

        top_clientes = [
            dict(r) for r in db.session.execute(sql_top_clientes, params_ing).mappings().all()
        ]

        # ============================================================
        # 7) Top proveedores
        # ============================================================

        sql_top_proveedores = text(f"""
            SELECT
                COALESCE(c.proveedor_nombre, 'Sin proveedor') AS nombre,
                COALESCE(SUM(c.total), 0) AS total
            FROM siigo_compras c
            WHERE {where_egr}
            GROUP BY COALESCE(c.proveedor_nombre, 'Sin proveedor')
            ORDER BY total DESC
            LIMIT 10
        """)

        top_proveedores = [
            dict(r) for r in db.session.execute(sql_top_proveedores, params_egr).mappings().all()
        ]

        return jsonify({
            "kpis": kpis,
            "evolucion": evolucion,
            "top_clientes": top_clientes,
            "top_proveedores": top_proveedores,
            "config": {
                "fuente_ingresos": "ventas_movimientos_enriquecidos",
                "fuente_egresos": "siigo_compras",
                "ingresos": "ventas_netas_con_impuesto",
                "logica": "ingresos = facturas_emitidas - notas_credito"
            }
        })


    # ============================================================
    # ENDPOINT: Detalle de ingresos / clientes para consolidado
    # Ahora devuelve movimientos comerciales: FACTURA y NOTA_CREDITO
    # ============================================================

    @app.route("/reportes/facturas_cliente", methods=["GET"])
    @jwt_required()
    def facturas_por_cliente():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        cliente = request.args.get("cliente")
        centro_costos = request.args.get("centro_costos", type=int)
        if not centro_costos:
            centro_costos = request.args.get("cost_center", type=int)
        limit = request.args.get("limit", type=int) or 10000

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        condiciones = ["m.idcliente = :idcliente"]
        params = {
            "idcliente": idcliente,
            "limit": limit,
        }

        if desde and validar_fecha(desde):
            condiciones.append("m.fecha >= :desde")
            params["desde"] = desde

        if hasta and validar_fecha(hasta):
            condiciones.append("m.fecha <= :hasta")
            params["hasta"] = hasta

        if cliente:
            condiciones.append("m.cliente_nombre = :cliente")
            params["cliente"] = cliente

        if centro_costos:
            condiciones.append("m.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT
                m.movimiento_id,
                m.documento AS idfactura,
                m.documento,
                m.tipo_movimiento,
                m.fecha,
                m.vencimiento,
                m.cliente_nombre,
                m.estado,
                m.estado_pago,
                m.subtotal,
                m.impuestos_total,
                m.total,
                m.pagos_total AS pagado,
                m.saldo,
                m.cost_center,
                m.centro_costo_nombre,
                m.centro_costo_codigo,
                m.seller_id,
                m.vendedor_nombre,
                m.public_url,
                m.documento_afectado,

                CASE
                    WHEN m.tipo_movimiento = 'FACTURA'
                    THEN GREATEST(COALESCE(m.total, 0) - COALESCE(m.saldo, 0), 0)
                    ELSE 0
                END AS valor_pagado,

                CASE
                    WHEN m.tipo_movimiento = 'FACTURA'
                    THEN GREATEST(COALESCE(m.saldo, 0), 0)
                    ELSE 0
                END AS valor_pendiente

            FROM ventas_movimientos_enriquecidos m
            WHERE {where_sql}
            ORDER BY m.fecha DESC, m.tipo_movimiento ASC, m.documento DESC
            LIMIT :limit
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]

        sql_resumen = text(f"""
            SELECT
                COALESCE(SUM(CASE WHEN m.tipo_movimiento = 'FACTURA' THEN m.total ELSE 0 END), 0) AS facturas_emitidas,
                ABS(COALESCE(SUM(CASE WHEN m.tipo_movimiento = 'NOTA_CREDITO' THEN m.total ELSE 0 END), 0)) AS notas_credito,
                COALESCE(SUM(m.total), 0) AS ventas_netas,
                COUNT(*) FILTER (WHERE m.tipo_movimiento = 'FACTURA') AS total_facturas,
                COUNT(*) FILTER (WHERE m.tipo_movimiento = 'NOTA_CREDITO') AS total_notas_credito,
                COUNT(*) AS total_movimientos
            FROM ventas_movimientos_enriquecidos m
            WHERE {where_sql}
        """)

        resumen = dict(db.session.execute(sql_resumen, params).mappings().first() or {})

        return jsonify({
            "rows": rows,
            "resumen": resumen,
            "total": len(rows),
            "idcliente": idcliente,
            "fuente": "ventas_movimientos_enriquecidos"
        })


    # ============================================================
    # ENDPOINT: Facturas de proveedor para modal consolidado
    # Se mantiene sobre siigo_compras
    # ============================================================

    @app.route("/reportes/facturas_proveedor", methods=["GET"])
    @jwt_required()
    def facturas_por_proveedor():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)

        if perfilid == 0:
            if q_idcliente:
                idcliente = q_idcliente
            elif not idcliente:
                return jsonify({"error": "Falta idcliente para consulta SuperAdmin"}), 400
        else:
            if not idcliente:
                return jsonify({"error": "Token sin cliente"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        proveedor = request.args.get("proveedor")
        centro_costos = request.args.get("centro_costos", type=int)
        limit = request.args.get("limit", type=int) or 10000

        condiciones = ["c.idcliente = :idcliente"]
        params = {
            "idcliente": idcliente,
            "limit": limit,
        }

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        fecha_desde_val = validar_fecha(desde) if desde else None
        fecha_hasta_val = validar_fecha(hasta) if hasta else None

        if fecha_desde_val:
            condiciones.append("c.fecha >= :desde")
            params["desde"] = fecha_desde_val

        if fecha_hasta_val:
            condiciones.append("c.fecha <= :hasta")
            params["hasta"] = fecha_hasta_val

        if proveedor:
            condiciones.append(
                "LOWER(COALESCE(c.proveedor_nombre, 'sin proveedor')) = LOWER(:proveedor)"
            )
            params["proveedor"] = proveedor.strip()

        if centro_costos:
            condiciones.append("c.cost_center = :centro_costos")
            params["centro_costos"] = centro_costos

        where_sql = " AND ".join(condiciones)

        sql = text(f"""
            SELECT
                c.id,
                c.idcompra,
                c.proveedor_nombre,
                c.factura_proveedor,
                c.fecha,
                c.vencimiento,
                c.total,
                c.saldo,
                c.cost_center,
                cc.nombre AS centro_costo_nombre,
                CASE
                    WHEN LOWER(COALESCE(c.estado, '')) = 'pagado' THEN 'Pagada'
                    ELSE 'No Pagada'
                END AS estado
            FROM siigo_compras c
            LEFT JOIN siigo_centros_costo cc
                ON cc.id = c.cost_center
            AND cc.idcliente = c.idcliente
            WHERE {where_sql}
            ORDER BY c.fecha DESC, c.proveedor_nombre ASC, c.idcompra ASC, c.id ASC
            LIMIT :limit
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]

        return jsonify({
            "rows": rows,
            "total": len(rows),
            "idcliente": idcliente
        }), 200


    # ============================================================
    # ENDPOINT: Catálogo de centros de costo para consolidado
    # Toma centros usados en ingresos y egresos
    # ============================================================

    @app.route("/catalogos/centros-costo-consolidado", methods=["GET"])
    @jwt_required()
    def catalogo_centros_costo_consolidado():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        def validar_fecha(fecha_str):
            try:
                return datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except Exception:
                return None

        fecha_desde_val = validar_fecha(desde) if desde else None
        fecha_hasta_val = validar_fecha(hasta) if hasta else None

        params = {"idcliente": idcliente}

        wh_mov = ["m.idcliente = :idcliente", "m.cost_center IS NOT NULL"]
        wh_comp = ["c.idcliente = :idcliente", "c.cost_center IS NOT NULL"]

        if fecha_desde_val:
            wh_mov.append("m.fecha >= :desde")
            wh_comp.append("c.fecha >= :desde")
            params["desde"] = fecha_desde_val

        if fecha_hasta_val:
            wh_mov.append("m.fecha <= :hasta")
            wh_comp.append("c.fecha <= :hasta")
            params["hasta"] = fecha_hasta_val

        where_mov = " AND ".join(wh_mov)
        where_comp = " AND ".join(wh_comp)

        sql = text(f"""
            WITH centros AS (
                SELECT DISTINCT
                    m.cost_center AS id,
                    COALESCE(m.centro_costo_nombre, 'Sin centro de costo') AS nombre
                FROM ventas_movimientos_enriquecidos m
                WHERE {where_mov}

                UNION

                SELECT DISTINCT
                    c.cost_center AS id,
                    COALESCE(cc.nombre, 'Sin centro de costo') AS nombre
                FROM siigo_compras c
                LEFT JOIN siigo_centros_costo cc
                    ON cc.id = c.cost_center
                AND cc.idcliente = c.idcliente
                WHERE {where_comp}
            )
            SELECT id, nombre
            FROM centros
            WHERE id IS NOT NULL
            ORDER BY nombre
        """)

        rows = [dict(r) for r in db.session.execute(sql, params).mappings().all()]
        return jsonify(rows)



    # ------------------------------------------
    # Cargar Documentos soporte desde archivo Excel
    # ------------------------------------------
    @app.route("/importar/soporte-excel", methods=["POST"])
    @jwt_required()
    def importar_documentos_soporte_desde_excel():
        from decimal import Decimal, InvalidOperation

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        if "archivo" not in request.files:
            return jsonify({"error": "Archivo no proporcionado"}), 400

        file = request.files["archivo"]

        def _norm_text(v) -> str:
            if pd.isna(v):
                return ""
            return str(v).replace("\xa0", " ").replace("\ufeff", "").strip()

        def _norm_lower(v) -> str:
            return _norm_text(v).lower()

        def _norm_upper(v) -> str:
            return _norm_text(v).upper()

        def _clean_number_text(v) -> str:
            """
            Convierte valores tipo:
            1896.0 -> 1896
            ' DS-1 ' -> DS-1
            nan -> ''
            """
            if pd.isna(v):
                return ""
            s = str(v).strip()
            if s.endswith(".0"):
                s = s[:-2]
            return s

        def _to_decimal(v) -> Decimal:
            """
            Convierte valores de Excel a Decimal de forma segura.
            Acepta números, strings, valores vacíos y NaN.
            """
            if v is None or pd.isna(v):
                return Decimal("0")

            if isinstance(v, Decimal):
                return v

            s = str(v).strip()
            if not s:
                return Decimal("0")

            # Limpieza básica por si viene con símbolos
            s = (
                s.replace("$", "")
                .replace(" ", "")
                .replace("\xa0", "")
            )

            # Si viene formato colombiano: 3.873.219,74
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s and "." not in s:
                s = s.replace(",", ".")

            try:
                return Decimal(s)
            except InvalidOperation:
                return Decimal("0")

        try:
            df = pd.read_excel(file, header=7)
            df.columns = df.columns.str.strip().str.replace("\xa0", " ")
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        requeridos = [
            "Número comprobante",
            "Consecutivo",
            "Nombre tercero",
            "Centro costo",
            "Total",
            "Fecha elaboración",
            "Factura proveedor",
            "Tipo de registro",
        ]

        faltantes = [col for col in requeridos if col not in df.columns]
        if faltantes:
            return jsonify({
                "error": f"Faltan columnas requeridas: {', '.join(faltantes)}"
            }), 400

        # ---------------------------
        # BLINDAJE / VALIDACIÓN PREVIA
        # ---------------------------
        possible_tipo_cols = [
            "Tipo de Transacción",
            "Tipo de transacción",
            "Tipo transacción",
            "Tipo de Transaccion",
            "Tipo transaccion",
            "Tipo transacción",
        ]

        tipo_col = next((c for c in possible_tipo_cols if c in df.columns), None)

        df["__tipo_reg__"] = df["Tipo de registro"].apply(_norm_lower)
        df["__num__"] = df["Número comprobante"].apply(
            lambda x: _norm_upper(_clean_number_text(x).split(".")[0])
        )
        df["__pref__"] = df["__num__"].apply(lambda s: s.split("-")[0] if s else "")
        df["__consecutivo__"] = df["Consecutivo"].apply(
            lambda x: _clean_number_text(x).split(".")[0]
        )
        df["__idcompra__"] = df.apply(
            lambda r: f"{r['__num__']}-{r['__consecutivo__']}"
            if r["__num__"] and r["__consecutivo__"]
            else "",
            axis=1
        )

        sample = df[df["__tipo_reg__"] == "secuencia"].copy()

        if sample.empty:
            return jsonify({
                "error": "El archivo no contiene filas 'secuencia'. Revisa que sea el reporte correcto."
            }), 400

        total_seq = len(sample)
        cnt_ds = int((sample["__pref__"] == "DS").sum())
        cnt_fc = int((sample["__pref__"] == "FC").sum())
        cnt_fv = int((sample["__pref__"] == "FV").sum())
        cnt_otro = total_seq - (cnt_ds + cnt_fc + cnt_fv)

        if cnt_fc > 0 or cnt_fv > 0:
            return jsonify({
                "error": (
                    "Archivo inválido: el reporte contiene transacciones distintas a Documento Soporte (FC/FV). "
                    "En Siigo debes exportar desde: Reportes > Compras y Gastos > Movimientos de documentos Compras > "
                    "Movimiento Factura de compra, seleccionando Tipo de Transacción = Documento Soporte."
                ),
                "diagnostico": {
                    "filas_secuencia_analizadas": total_seq,
                    "DS": cnt_ds,
                    "FC": cnt_fc,
                    "FV": cnt_fv,
                    "OTRO": cnt_otro,
                },
            }), 400

        if tipo_col:
            sample["__tipo_tx__"] = sample[tipo_col].apply(_norm_lower)
            invalid_tx = sample[~sample["__tipo_tx__"].str.contains("documento soporte", na=False)]

            if not invalid_tx.empty:
                return jsonify({
                    "error": (
                        f"Archivo inválido: la columna '{tipo_col}' contiene valores diferentes a 'Documento Soporte'. "
                        "Reexporta el reporte seleccionando Tipo de Transacción = Documento Soporte."
                    ),
                    "diagnostico": {
                        "columna_tipo_transaccion": tipo_col,
                        "filas_secuencia_analizadas": total_seq,
                        "ejemplos_invalidos": invalid_tx[tipo_col].head(10).tolist(),
                    },
                }), 400

        # Nos quedamos con todas las filas DS, no solo Secuencia.
        # Esto es clave porque el total real del documento viene en la fila "Formas de pago".
        df_ds = df[
            (df["__pref__"] == "DS") &
            (df["__idcompra__"] != "")
        ].copy()

        if df_ds.empty:
            return jsonify({
                "error": "No se encontraron Documentos Soporte DS válidos en el archivo."
            }), 400

        centros_costo = {
            c.nombre.strip().upper(): c.id
            for c in SiigoCentroCosto.query.filter_by(idcliente=idcliente).all()
        }

        registros_creados = 0
        registros_actualizados = 0
        items_creados = 0
        compras_omitidas = []
        errores_documentos = []

        documentos_procesados = 0

        try:
            for idcompra, grupo in df_ds.groupby("__idcompra__"):
                documentos_procesados += 1

                filas_secuencia = grupo[grupo["__tipo_reg__"] == "secuencia"].copy()
                filas_formas_pago = grupo[grupo["__tipo_reg__"] == "formas de pago"].copy()
                filas_impuesto_total = grupo[grupo["__tipo_reg__"] == "impuesto total"].copy()

                if filas_secuencia.empty:
                    compras_omitidas.append({
                        "idcompra": idcompra,
                        "motivo": "Documento sin filas Secuencia"
                    })
                    continue

                # Tomamos la primera fila Secuencia como fila maestra para datos generales
                row_base = filas_secuencia.iloc[0]

                fecha_elab = pd.to_datetime(
                    row_base.get("Fecha elaboración"),
                    dayfirst=True,
                    errors="coerce"
                )

                if pd.isna(fecha_elab):
                    errores_documentos.append({
                        "idcompra": idcompra,
                        "error": "Fecha elaboración inválida"
                    })
                    continue

                venc = pd.to_datetime(
                    row_base.get("Fecha vencimiento"),
                    dayfirst=True,
                    errors="coerce"
                )
                venc_date = None if pd.isna(venc) else venc.date()

                centro_costo_nombre = _norm_upper(row_base.get("Centro costo", ""))
                cost_center = centros_costo.get(centro_costo_nombre)

                factura_proveedor = _norm_upper(row_base.get("Factura proveedor", ""))
                if factura_proveedor.lower() in ("nan", "nat", ""):
                    factura_proveedor = None

                # ---------------------------
                # TOTAL REAL DEL DOCUMENTO
                # ---------------------------
                # 1) Valor principal: fila "Formas de pago"
                # 2) Fallback: suma de Secuencias si el Excel no trae Formas de pago
                total_formas_pago = sum(
                    _to_decimal(v) for v in filas_formas_pago["Total"].tolist()
                )

                total_secuencias = sum(
                    _to_decimal(v) for v in filas_secuencia["Total"].tolist()
                )

                total_impuesto_total = sum(
                    _to_decimal(v) for v in filas_impuesto_total["Total"].tolist()
                )

                if total_formas_pago > 0:
                    total_documento = total_formas_pago
                    fuente_total = "formas_de_pago"
                else:
                    total_documento = total_secuencias
                    fuente_total = "suma_secuencias_fallback"

                compra = db.session.query(SiigoCompra).filter_by(
                    idcliente=idcliente,
                    idcompra=idcompra
                ).first()

                if compra:
                    # Actualizamos el encabezado del documento.
                    # Esto corrige documentos ya cargados con total parcial.
                    compra.fecha = fecha_elab.date()
                    compra.vencimiento = venc_date
                    compra.proveedor_nombre = _norm_text(row_base.get("Nombre tercero", ""))
                    compra.proveedor_identificacion = _clean_number_text(row_base.get("Identificación", ""))
                    compra.total = total_documento
                    compra.cost_center = cost_center
                    compra.factura_proveedor = factura_proveedor

                    # Importante:
                    # No tocamos saldo aquí, porque el saldo debe venir del cruce con cuentas por pagar.
                    # Si aún no se ha cruzado, quedará como esté y luego se corrige al correr /siigo/cross-accounts-payable.

                    # Reconstruimos ítems para evitar duplicados y garantizar que queden todos.
                    db.session.query(SiigoCompraItem).filter_by(
                        idcliente=idcliente,
                        compra_id=compra.id
                    ).delete(synchronize_session=False)

                    registros_actualizados += 1

                else:
                    compra = SiigoCompra(
                        idcliente=idcliente,
                        idcompra=idcompra,
                        fecha=fecha_elab.date(),
                        vencimiento=venc_date,
                        proveedor_nombre=_norm_text(row_base.get("Nombre tercero", "")),
                        proveedor_identificacion=_clean_number_text(row_base.get("Identificación", "")),
                        estado=None,
                        total=total_documento,
                        saldo=0,
                        cost_center=cost_center,
                        creado=fecha_elab + pd.Timedelta(minutes=15),
                        factura_proveedor=factura_proveedor,
                    )

                    db.session.add(compra)
                    db.session.flush()
                    registros_creados += 1

                # ---------------------------
                # ÍTEMS DEL DOCUMENTO
                # ---------------------------
                for _, item_row in filas_secuencia.iterrows():
                    impuestos = (
                        _to_decimal(item_row.get("Valor Impuesto Cargo", 0)) +
                        _to_decimal(item_row.get("Valor Impuesto Cargo 2", 0))
                    )

                    item = SiigoCompraItem(
                        compra_id=compra.id,
                        idcliente=idcliente,
                        descripcion=_norm_text(item_row.get("Nombre", "")),
                        cantidad=_to_decimal(item_row.get("Cantidad", 0)),
                        precio=_to_decimal(item_row.get("Total", 0)),
                        impuestos=impuestos,
                        codigo="" if pd.isna(item_row.get("Código", "")) else _norm_text(item_row.get("Código", "")),
                    )

                    db.session.add(item)
                    items_creados += 1

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": f"Error importando Documentos Soporte: {str(e)}"
            }), 500

        return jsonify({
            "mensaje": (
                f"Importación completada. "
                f"Documentos creados: {registros_creados}. "
                f"Documentos actualizados: {registros_actualizados}. "
                f"Ítems creados: {items_creados}."
            ),
            "omitidas": compras_omitidas,
            "errores_documentos": errores_documentos,
            "debug": {
                "documentos_procesados": documentos_procesados,
                "filas_secuencia_analizadas": total_seq,
                "DS": cnt_ds,
                "FC": cnt_fc,
                "FV": cnt_fv,
                "OTRO": cnt_otro,
                "tipo_col": tipo_col,
                "nota": (
                    "El total del documento se toma de la fila 'Formas de pago'. "
                    "Si no existe, se usa como fallback la suma de las filas 'Secuencia'."
                )
            }
        })




    # Cargar Documentos soporte desde archivo Excel para exportarlo a un ZIP con dos archivos Excel (preview)
    @app.route("/importar/soporte-excel-preview", methods=["POST"])
    @jwt_required()
    def importar_documentos_soporte_desde_excel_preview():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        if "archivo" not in request.files:
            return jsonify({"error": "Archivo no proporcionado"}), 400

        file = request.files["archivo"]

        try:
            df = pd.read_excel(file, header=7)
            df.columns = df.columns.str.strip().str.replace('\xa0', ' ')
            print("🧩 Columnas detectadas:", df.columns.tolist())
        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        requeridos = ["Número comprobante", "Consecutivo", "Nombre tercero", "Centro costo", "Total", "Fecha creación"]
        faltantes = [col for col in requeridos if col not in df.columns]
        if faltantes:
            return jsonify({"error": f"Faltan columnas requeridas: {', '.join(faltantes)}"}), 400

        centros_costo = {
            c.nombre.strip().upper(): c.id for c in SiigoCentroCosto.query.filter_by(idcliente=idcliente).all()
        }

        compras_preview = []
        items_preview = []
        compras = {}

        for _, row in df.iterrows():
            tipo_registro_raw = row.get("Tipo de registro", "")
            tipo_registro = str(tipo_registro_raw).replace('\xa0', ' ').replace('\ufeff', '').strip().lower()
            print(f"🧪 Tipo de registro detectado: '{tipo_registro}' (original: {repr(tipo_registro_raw)})")

            if tipo_registro != "secuencia":
                continue

            idcompra = f"{str(row['Número comprobante']).strip()}-{str(row['Consecutivo']).strip()}"

            if idcompra not in compras:
                centro_costo_nombre = str(row.get("Centro costo", "")).strip().upper()
                cost_center = centros_costo.get(centro_costo_nombre)

                compras[idcompra] = len(compras) + 1  # Fake ID

                compras_preview.append({
                    "idcliente": idcliente,
                    "idcompra": idcompra,
                    "fecha": pd.to_datetime(row["Fecha creación"]).date(),
                    "vencimiento": pd.to_datetime(row.get("Fecha vencimiento"), errors="coerce").date() if row.get("Fecha vencimiento") else None,
                    "proveedor_nombre": str(row.get("Nombre tercero", "")).strip(),
                    "proveedor_identificacion": str(row.get("Identificación", "")).strip().split('.')[0],
                    "estado": None,
                    "total": float(row.get("Total", 0) or 0),
                    "saldo": 0,
                    "cost_center": cost_center,
                    "creado": pd.to_datetime(row["Fecha creación"]) + pd.Timedelta(minutes=15)
                })

            compra_id = compras[idcompra]

            impuestos = float(row.get("Valor Impuesto Cargo", 0) or 0) + float(row.get("Valor Impuesto", 0) or 0)

            items_preview.append({
                "compra_id": compra_id,
                "descripcion": str(row.get("Nombre", "")).strip(),
                "cantidad": float(row.get("Cantidad", 0) or 0),
                "precio": float(row.get("Total", 0) or 0),
                "impuestos": impuestos,
                "codigo": "" if pd.isna(row.get("Código", "")) else str(row.get("Código", "")).strip()
            })

        # Exportar a Excel en memoria
        compras_df = pd.DataFrame(compras_preview)
        items_df = pd.DataFrame(items_preview)

        in_memory_zip = BytesIO()
        with zipfile.ZipFile(in_memory_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            compras_buffer = BytesIO()
            items_buffer = BytesIO()

            compras_df.to_excel(compras_buffer, index=False)
            items_df.to_excel(items_buffer, index=False)

            zf.writestr("preview_siigo_compras.xlsx", compras_buffer.getvalue())
            zf.writestr("preview_siigo_compras_items.xlsx", items_buffer.getvalue())

        in_memory_zip.seek(0)
        return send_file(
            in_memory_zip,
            mimetype="application/zip",
            as_attachment=True,
            download_name="preview_siigo_compras.zip"
        )




    # Endpoint para validar cuentas por pagar vs Siigo
    @app.route("/debug/siigo/accounts-payable-check", methods=["GET"])
    @jwt_required()
    def siigo_accounts_payable_check():
        import requests
        from sqlalchemy.sql import text
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        # ⚠️ Tu token de Siigo (ajusta según lo tengas guardado)
        siigo_token = "Bearer TU_TOKEN_SIIGO_AQUI"

        try:
            # Paso 1: Llamar API de Siigo
            response = requests.get(
                "https://api.siigo.com/v1/reports/accounts-payable",
                headers={"Authorization": siigo_token}
            )
            response.raise_for_status()
            siigo_data = response.json()
        except Exception as e:
            return jsonify({"error": f"Error consultando Siigo: {str(e)}"}), 500

        # Paso 2: Limpiar y normalizar data del API Siigo
        siigo_map = []
        for item in siigo_data:
            siigo_map.append({
                "proveedor": item.get("provider", "").strip().upper(),
                "factura": item.get("document_number", "").strip().upper(),
                "saldo_api": float(item.get("balance", 0)),
                "total_api": float(item.get("total", 0)),
                "fecha": item.get("issue_date"),
                "vencimiento": item.get("due_date")
            })

        # Paso 3: Cargar todas tus facturas locales en un mapa rápido
        sql = text("""
            SELECT 
                proveedor_nombre, idcompra, saldo, total
            FROM siigo_compras
            WHERE idcliente = :idcliente
        """)
        local_facturas = {
            row["idcompra"].strip().upper(): {
                "proveedor": row["proveedor_nombre"].strip().upper(),
                "saldo_local": float(row["saldo"]),
                "total_local": float(row["total"])
            }
            for row in db.session.execute(sql, {"idcliente": idcliente}).mappings().all()
        }

        # Paso 4: Comparar y buscar discrepancias
        discrepancias = []
        for item in siigo_map:
            factura = item["factura"]
            match = local_facturas.get(factura)

            if match:
                delta_saldo = abs(match["saldo_local"] - item["saldo_api"])
                if delta_saldo > 10:  # margen mínimo de diferencia
                    discrepancias.append({
                        "factura": factura,
                        "proveedor_siigo": item["proveedor"],
                        "proveedor_local": match["proveedor"],
                        "saldo_api": item["saldo_api"],
                        "saldo_local": match["saldo_local"],
                        "diferencia": delta_saldo
                    })
            else:
                discrepancias.append({
                    "factura": factura,
                    "proveedor_siigo": item["proveedor"],
                    "proveedor_local": "No encontrada",
                    "saldo_api": item["saldo_api"],
                    "saldo_local": None,
                    "diferencia": None
                })

        return jsonify({
            "total_siigo": len(siigo_map),
            "coincidencias": len(siigo_map) - len(discrepancias),
            "discrepancias": discrepancias
        })


    # Endpoint para obtener cuentas por pagar raw desde Siigo (para debug)
    @app.route("/debug/siigo/accounts-payable-raw", methods=["GET"])
    @jwt_required()
    def siigo_accounts_payable_raw():
        from flask import jsonify
        import requests

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "ID cliente no encontrado en JWT"}), 401

        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales de Siigo no encontradas para este cliente"}), 404

        access_key = dec(cred.client_secret)
        if not access_key:
            return jsonify({"error": "No se pudo desencriptar client_secret"}), 500

        try:
            # Obtener token
            token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            token = token_data["access_token"]

            headers = {
                "Content-Type": "application/json",
                "Accept": "application/json",
                "Authorization": f"Bearer {token}",
            }

            if cred.partner_id:
                headers["Partner-Id"] = cred.partner_id

            all_items = []
            page = 1
            page_size = 100  # máximo permitido en la mayoría de APIs
            base_url = cred.base_url.rstrip('/')

            while True:
                url = f"{base_url}/v1/accounts-payable?page={page}&page_size={page_size}"
                res = requests.get(url, headers=headers)

                if not res.ok:
                    return jsonify({
                        "error": f"Error desde Siigo (página {page})",
                        "detalle": res.text,
                        "headers": dict(res.headers),
                        "url_usada": url
                    }), res.status_code

                data = res.json()
                results = data.get("results", [])
                all_items.extend(results)

                # Verificar si hay siguiente página
                next_link = data.get("_links", {}).get("next", {}).get("href")
                if not next_link:
                    break  # no hay más páginas

                page += 1

            # 🔁 Transformamos los resultados al formato esperado por el frontend
            transformed = []
            for item in all_items:
                transformed.append({
                    "document": f'{item["due"]["prefix"]}-{item["due"]["consecutive"]}',
                    "date": item["due"]["date"],
                    "due_date": item["due"]["date"],  # seguimos usando la misma
                    "supplier": {
                        "identification": item["provider"]["identification"],
                        "name": item["provider"]["name"],
                    },
                    "amount": item["due"]["balance"],  # usamos el mismo valor como monto
                    "balance": item["due"]["balance"],
                    "cost_center": item.get("cost_center", {}).get("name"),
                })

            return jsonify(transformed)

        except Exception as e:
            return jsonify({
                "error": "Error interno al consultar Siigo",
                "detalle": str(e)
            }), 500



    # 1. Sincronizar cuentas por pagar desde Siigo
    @app.route("/siigo/sync-accounts-payable", methods=["POST"])
    def sync_accounts_payable():
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"
        endpoint_log = "/siigo/sync-accounts-payable"
        log_id = None
        inicio = time.time()

        if not modo_sync_all:
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params={},
                mensaje="Cuentas por pagar: proceso iniciado."
            )

        try:
            cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
            if not cred:
                detalle = "Credenciales no encontradas."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=404,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({
                    "error": detalle,
                    "tipo": "credenciales_no_encontradas"
                }), 404

            if not cred.base_url or not cred.client_id or not cred.client_secret:
                detalle = "Credenciales de Siigo incompletas."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=400,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({
                    "error": detalle,
                    "tipo": "credenciales_incompletas"
                }), 400

            access_key = dec(cred.client_secret)
            if not access_key:
                detalle = "No se pudo desencriptar el Access Key de Siigo."

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=400,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({
                    "error": detalle,
                    "tipo": "access_key_invalida"
                }), 400

            try:
                token_data = siigo_auth_json(cred.base_url, cred.client_id, access_key)
            except Exception as e:
                detalle = f"Error autenticando contra Siigo: {str(e)}"

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=502,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({
                    "error": "Error autenticando contra Siigo",
                    "detalle": str(e),
                    "tipo": "siigo_auth_error"
                }), 502

            token = token_data.get("access_token")
            if not token:
                detalle = f"Siigo no devolvió access_token. Respuesta: {token_data}"

                if not modo_sync_all:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=detalle,
                        status_code=502,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

                return jsonify({
                    "error": "Siigo no devolvió access_token",
                    "detalle": token_data,
                    "tipo": "siigo_token_error"
                }), 502

            headers = {
                "Authorization": f"Bearer {token}",
                "Accept": "application/json",
            }

            if cred.partner_id:
                headers["Partner-Id"] = cred.partner_id

            base_url = cred.base_url.rstrip("/")
            url = f"{base_url}/v1/accounts-payable?page=1&page_size=100"

            all_results = []
            page_count = 0
            max_pages = 200

            while url:
                page_count += 1

                if page_count > max_pages:
                    detalle = (
                        f"Se superó el máximo de páginas consultando cuentas por pagar. "
                        f"max_pages={max_pages}, total_parcial={len(all_results)}"
                    )

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=500,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": "Se superó el máximo de páginas consultando cuentas por pagar",
                        "tipo": "accounts_payable_max_pages",
                        "max_pages": max_pages,
                        "total_parcial": len(all_results)
                    }), 500

                try:
                    res = requests.get(url, headers=headers, timeout=90)
                except requests.Timeout:
                    detalle = f"Timeout consultando cuentas por pagar en Siigo. url={url}"

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=504,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": "Timeout consultando cuentas por pagar en Siigo",
                        "tipo": "siigo_timeout",
                        "url": url,
                        "total_parcial": len(all_results)
                    }), 504

                except requests.RequestException as e:
                    detalle = f"Error de conexión consultando cuentas por pagar en Siigo: {str(e)}"

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=502,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": "Error de conexión consultando cuentas por pagar en Siigo",
                        "detalle": str(e),
                        "tipo": "siigo_request_error",
                        "url": url,
                        "total_parcial": len(all_results)
                    }), 502

                if res.status_code != 200:
                    try:
                        detalle_siigo = res.json()
                    except Exception:
                        detalle_siigo = res.text[:1000]

                    status = 429 if res.status_code == 429 else 502
                    detalle = (
                        f"Siigo respondió HTTP {res.status_code} al consultar cuentas por pagar. "
                        f"Detalle: {detalle_siigo}"
                    )

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=status,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": f"Siigo respondió error HTTP {res.status_code} al consultar cuentas por pagar",
                        "detalle": detalle_siigo,
                        "tipo": "siigo_accounts_payable_error",
                        "url": url,
                        "total_parcial": len(all_results)
                    }), status

                try:
                    data = res.json() or {}
                except Exception:
                    detalle = f"Siigo devolvió una respuesta no JSON. Respuesta: {res.text[:1000]}"

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=502,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": "Siigo devolvió una respuesta no JSON al consultar cuentas por pagar",
                        "detalle": res.text[:1000],
                        "tipo": "siigo_non_json_response",
                        "url": url,
                        "total_parcial": len(all_results)
                    }), 502

                results = data.get("results") or []
                if not isinstance(results, list):
                    detalle = f"Formato inesperado en respuesta de cuentas por pagar: {data}"

                    if not modo_sync_all:
                        _finalizar_log_sync_modulo(
                            log_id=log_id,
                            idcliente=idcliente,
                            endpoint=endpoint_log,
                            resultado="ERROR",
                            detalle=detalle,
                            status_code=502,
                            duracion_segundos=round(time.time() - inicio, 2),
                        )

                    return jsonify({
                        "error": "Formato inesperado en respuesta de cuentas por pagar",
                        "detalle": data,
                        "tipo": "siigo_unexpected_format",
                        "url": url,
                        "total_parcial": len(all_results)
                    }), 502

                all_results.extend(results)

                next_href = ((data.get("_links") or {}).get("next") or {}).get("href")

                if next_href:
                    if str(next_href).startswith("http"):
                        url = next_href
                    else:
                        url = f"{base_url}{next_href}"
                else:
                    url = None

            SiigoCuentasPorCobrar.query.filter_by(idcliente=idcliente).delete()

            insertadas = 0
            omitidas = 0

            for item in all_results:
                try:
                    due = item.get("due") or {}
                    provider = item.get("provider") or {}

                    prefix = due.get("prefix")
                    consecutive = due.get("consecutive")
                    due_date = due.get("date")
                    balance = due.get("balance")

                    provider_identification = provider.get("identification")
                    provider_name = provider.get("name")

                    if not prefix or not consecutive:
                        omitidas += 1
                        continue

                    documento = f"{prefix}-{consecutive}"

                    centro_costo = None
                    cost_center = item.get("cost_center")
                    if isinstance(cost_center, dict):
                        centro_costo = cost_center.get("name")
                    elif cost_center:
                        centro_costo = str(cost_center)

                    row = SiigoCuentasPorCobrar(
                        idcliente=idcliente,
                        documento=documento,
                        fecha=None,
                        fecha_vencimiento=due_date,
                        proveedor_identificacion=provider_identification,
                        proveedor_nombre=provider_name,
                        valor=balance,
                        saldo=balance,
                        centro_costo=centro_costo
                    )
                    db.session.add(row)
                    insertadas += 1

                except Exception as e:
                    print(f"Error procesando cuenta por pagar: {e} | item={item}")
                    omitidas += 1
                    continue

            db.session.commit()

            respuesta = {
                "mensaje": f"{insertadas} registros de cuentas por pagar sincronizados.",
                "total_resultados_siigo": len(all_results),
                "insertadas": insertadas,
                "omitidas": omitidas,
                "paginas_consultadas": page_count
            }

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="OK",
                    detalle=str(respuesta),
                    status_code=200,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify(respuesta), 200

        except Exception as e:
            db.session.rollback()
            traceback.print_exc()

            detalle = f"Error interno sincronizando cuentas por pagar: {str(e)}"

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="ERROR",
                    detalle=detalle,
                    status_code=500,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify({
                "error": "Error interno sincronizando cuentas por pagar",
                "detalle": str(e),
                "tipo": "accounts_payable_internal_error"
            }), 500
        

    # 2. Cruce con compras locales
    @app.route("/siigo/cross-accounts-payable", methods=["POST"])
    def cross_accounts_payable():
        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"
        endpoint_log = "/siigo/cross-accounts-payable"
        log_id = None
        inicio = time.time()

        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        if not modo_sync_all:
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params={},
                mensaje="Cruce de cuentas por pagar: proceso iniciado."
            )

        try:
            from sqlalchemy import func, or_

            cuentas = SiigoCuentasPorCobrar.query.filter_by(idcliente=idcliente).all()

            compras = {
                (c.factura_proveedor, c.proveedor_identificacion): (c.idcompra, c.fecha)
                for c in SiigoCompra.query.filter_by(idcliente=idcliente).all()
            }

            matched = 0
            total = len(cuentas)
            cuentas_keys = set()

            for cuenta in cuentas:
                key = (cuenta.documento, cuenta.proveedor_identificacion)
                cuentas_keys.add(key)

                if key in compras:
                    idcompra, fecha = compras[key]
                    cuenta.idcompra = idcompra
                    cuenta.fecha = fecha

                    compra = SiigoCompra.query.filter_by(
                        idcliente=idcliente,
                        idcompra=idcompra
                    ).first()

                    if compra:
                        compra.estado = "pendiente"
                        compra.saldo = cuenta.saldo

                    matched += 1

            compras_keys = set(compras.keys())
            pagadas = compras_keys - cuentas_keys

            marcadas_pagadas = 0

            for key in pagadas:
                idcompra, _ = compras[key]

                compra = SiigoCompra.query.filter_by(
                    idcliente=idcliente,
                    idcompra=idcompra
                ).first()

                if compra:
                    compra.estado = "pagado"
                    compra.saldo = 0
                    marcadas_pagadas += 1

            ds_compras = SiigoCompra.query.filter(
                SiigoCompra.idcliente == idcliente,
                SiigoCompra.idcompra.like("DS%")
            ).all()

            ds_ajustadas = 0
            detalles_ajuste = {}

            for compra in ds_compras:
                total_pagado = db.session.query(func.sum(SiigoPagoProveedor.valor)).filter(
                    SiigoPagoProveedor.idcliente == idcliente,
                    or_(
                        func.trim(func.upper(SiigoPagoProveedor.factura_aplicada)) == func.trim(func.upper(compra.idcompra)),
                        func.trim(func.upper(SiigoPagoProveedor.factura_aplicada)) == func.trim(func.upper(compra.factura_proveedor))
                    ),
                    SiigoPagoProveedor.proveedor_identificacion == compra.proveedor_identificacion
                ).scalar() or 0

                if total_pagado > 0:
                    nuevo_saldo = max(float(compra.total or 0) - float(total_pagado), 0)
                    compra.saldo = nuevo_saldo

                    if nuevo_saldo == 0:
                        compra.estado = "pagado"
                    elif nuevo_saldo < float(compra.total or 0):
                        compra.estado = "parcial"
                    else:
                        compra.estado = "pendiente"

                    ds_ajustadas += 1

                    if len(detalles_ajuste) < 50:
                        detalles_ajuste[str(compra.idcompra)] = {
                            "factura_proveedor": str(compra.factura_proveedor or ""),
                            "total": float(compra.total or 0),
                            "pagado": float(total_pagado),
                            "saldo": float(nuevo_saldo),
                            "estado": str(compra.estado or "")
                        }

            db.session.commit()

            respuesta = {
                "mensaje": (
                    f"Cruce completado. {matched}/{total} cuentas vinculadas. "
                    f"{marcadas_pagadas} marcadas como pagadas. "
                    f"{ds_ajustadas} DS ajustadas por pagos reales."
                ),
                "cuentas_pendientes_leidas": total,
                "cuentas_vinculadas": matched,
                "compras_marcadas_pagadas": marcadas_pagadas,
                "ds_ajustadas": ds_ajustadas,
                "detalles_ds_muestra": detalles_ajuste,
                "detalles_ds_total": ds_ajustadas
            }

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="OK",
                    detalle=str(respuesta),
                    status_code=200,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify(respuesta), 200

        except Exception as e:
            db.session.rollback()
            traceback.print_exc()

            detalle = f"Error interno cruzando cuentas por pagar: {str(e)}"

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="ERROR",
                    detalle=detalle,
                    status_code=500,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify({
                "error": "Error interno cruzando cuentas por pagar",
                "detalle": str(e),
                "tipo": "cross_accounts_payable_internal_error"
            }), 500
    

    # --- Importar info de Nómina desde Archivo Excel ---
    # --- Importar info de Nómina desde Archivo Excel ---
    # --- Importar info de Nómina desde Archivo Excel ---
    @app.route("/importar/nomina-excel", methods=["POST"])
    @jwt_required()
    def importar_nomina_desde_excel():
        import pandas as pd
        from datetime import date
        import time
        import traceback

        inicio = time.time()
        endpoint_log = "/importar/nomina-excel"
        log_id = None

        claims = get_jwt()
        idcliente = claims.get("idcliente")

        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        mes = request.form.get("mes")
        anio = request.form.get("anio")

        params_log = {
            "mes": mes,
            "anio": anio,
        }

        # Crear log desde el inicio para que también queden rastros de errores de archivo/formato.
        log_id = _crear_log_sync_modulo_inicio(
            idcliente=idcliente,
            endpoint=endpoint_log,
            origen="manual_modulo",
            params=params_log,
            mensaje=(
                "Carga de nomina desde Excel: proceso iniciado"
                + (f" para periodo {mes}/{anio}." if mes and anio else ".")
            )
        )

        if "archivo" not in request.files:
            detalle = "Archivo no proporcionado."

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=400,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({"error": "Archivo no proporcionado"}), 400

        if not mes or not anio:
            detalle = "Debe indicar mes y ano de la nomina."

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=400,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({"error": "Debe indicar mes y año de la nómina"}), 400

        try:
            periodo = date(int(anio), int(mes), 1)
        except Exception:
            detalle = f"Mes o ano invalido. mes={mes}, anio={anio}"

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=400,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({"error": "Mes o año inválido"}), 400

        file = request.files["archivo"]
        filename = getattr(file, "filename", "") or ""

        try:
            df_raw = pd.read_excel(file, header=None, engine="calamine")

            # La fila 6 (índice 5) contiene encabezados según tu archivo actual
            headers = df_raw.iloc[5].tolist()
            headers = [str(h).strip().replace("\xa0", " ") if h is not None else "" for h in headers]

            # Datos desde fila 7
            df = df_raw.iloc[6:].copy()
            df.columns = headers
            df = df.dropna(how="all")
            df = df.where(pd.notnull(df), None)

            print("Encabezados detectados:", headers)
            print("Total filas a importar:", len(df))

        except Exception as e:
            traceback.print_exc()

            detalle = f"No se pudo leer el Excel de nomina. Archivo={filename}. Error={str(e)}"

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=400,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        # Resolver columnas por equivalencias
        columnas_resueltas = resolver_columnas(headers)

        faltantes = [campo for campo in REQUIRED_DB_FIELDS if campo not in columnas_resueltas]
        if faltantes:
            detalle = (
                "El archivo de nomina no contiene todas las columnas minimas requeridas. "
                f"Faltantes={faltantes}. Columnas detectadas={headers}"
            )

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=400,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({
                "error": "El archivo no contiene todas las columnas mínimas requeridas.",
                "faltantes": faltantes,
                "columnas_detectadas": headers,
                "log_id": log_id,
            }), 400

        registros_creados = 0
        errores = []
        eliminados = 0

        try:
            # Borrado previo del periodo para evitar duplicados
            eliminados = SiigoNomina.query.filter_by(
                idcliente=idcliente,
                periodo=periodo
            ).delete(synchronize_session=False)

            print(f"Registros previos eliminados para cliente={idcliente}, periodo={periodo}: {eliminados}")

            for idx, row in df.iterrows():
                try:
                    registro = SiigoNomina(
                        idcliente=idcliente,
                        periodo=periodo,
                        nombre=str(row.get(columnas_resueltas["nombre"], "") or "").strip(),
                        identificacion=str(row.get(columnas_resueltas["identificacion"], "") or "").strip(),
                        no_contrato=str(row.get(columnas_resueltas["no_contrato"], "") or "").strip(),
                        sueldo=normalizar_numero(row.get(columnas_resueltas["sueldo"])),
                        aux_transporte=normalizar_numero(row.get(columnas_resueltas["aux_transporte"])),
                        auxilio_extralegal=normalizar_numero(row.get(columnas_resueltas["auxilio_extralegal"])),
                        prima=normalizar_numero(row.get(columnas_resueltas["prima"])) if "prima" in columnas_resueltas else Decimal("0"),
                        intereses_cesantias=normalizar_numero(row.get(columnas_resueltas["intereses_cesantias"])) if "intereses_cesantias" in columnas_resueltas else Decimal("0"),
                        total_ingresos=normalizar_numero(row.get(columnas_resueltas["total_ingresos"])),
                        fondo_salud=normalizar_numero(row.get(columnas_resueltas["fondo_salud"])),
                        fondo_pension=normalizar_numero(row.get(columnas_resueltas["fondo_pension"])),
                        fondo_solidaridad=normalizar_numero(row.get(columnas_resueltas["fondo_solidaridad"])),
                        retefuente=normalizar_numero(row.get(columnas_resueltas["retefuente"])) if "retefuente" in columnas_resueltas else Decimal("0"),
                        prestamos=normalizar_numero(row.get(columnas_resueltas["prestamos"])) if "prestamos" in columnas_resueltas else Decimal("0"),
                        total_deducciones=normalizar_numero(row.get(columnas_resueltas["total_deducciones"])),
                        neto_pagar=normalizar_numero(row.get(columnas_resueltas["neto_pagar"])),
                    )
                    db.session.add(registro)
                    registros_creados += 1

                except Exception as e:
                    errores.append(f"Fila {idx + 7}: {str(e)}")

            db.session.commit()

        except Exception as e:
            db.session.rollback()
            traceback.print_exc()

            detalle = (
                f"Error al guardar nomina en BD. Periodo={periodo}. "
                f"Archivo={filename}. Error={str(e)}"
            )

            _finalizar_log_sync_modulo(
                log_id=log_id,
                idcliente=idcliente,
                endpoint=endpoint_log,
                resultado="ERROR",
                detalle=detalle,
                status_code=500,
                duracion_segundos=round(time.time() - inicio, 2),
            )

            return jsonify({
                "error": f"Error al guardar en BD: {str(e)}",
                "log_id": log_id,
            }), 500

        respuesta = {
            "mensaje": f"Nómina {periodo.strftime('%B %Y')} importada correctamente.",
            "registros_creados": registros_creados,
            "registros_reemplazados_previos": eliminados,
            "errores": errores[:10],
            "total_errores": len(errores),
            "columnas_mapeadas": columnas_resueltas,
            "periodo": periodo.isoformat(),
            "archivo": filename,
            "log_id": log_id,
        }

        # Si hubo errores parciales por fila, lo dejamos como OK porque el archivo cargó,
        # pero el detalle técnico conserva el conteo de errores.
        _finalizar_log_sync_modulo(
            log_id=log_id,
            idcliente=idcliente,
            endpoint=endpoint_log,
            resultado="OK",
            detalle=(
                f"Carga de nomina finalizada correctamente. "
                f"Periodo={periodo.isoformat()}. "
                f"Archivo={filename}. "
                f"Registros creados={registros_creados}. "
                f"Registros reemplazados previos={eliminados}. "
                f"Errores por fila={len(errores)}."
            ),
            status_code=200,
            duracion_segundos=round(time.time() - inicio, 2),
        )

        return jsonify(respuesta), 200



    # --- VALIDAR archivo de Nómina antes de importar ---
    # --- ENDPOINT de validación de Excel de Nómina ---
    @app.route("/validar/nomina-excel", methods=["POST"])
    @jwt_required()
    def validar_nomina_excel():
        import pandas as pd

        if "archivo" not in request.files:
            return jsonify({"error": "Archivo no proporcionado"}), 400

        file = request.files["archivo"]

        try:
            # Usamos calamine (más robusto con merges)
            df_raw = pd.read_excel(file, header=None, engine="calamine")

            # La fila 6 (índice 5 en 0-based) son los headers
            headers = df_raw.iloc[5].tolist()
            headers = [str(h).strip().replace("\xa0", " ") for h in headers]

            # Desde fila 7 en adelante
            df = df_raw.iloc[6:].copy()
            df.columns = headers
            df = df.dropna(how="all")

            # 🔹 Reemplazar NaN por None para que sea JSON válido
            df = df.where(pd.notnull(df), None)

            # Previsualización de las primeras filas
            preview = df.head(10).to_dict(orient="records")

        except Exception as e:
            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        return jsonify({
            "mensaje": f"Archivo leído correctamente. {len(preview)} filas en preview.",
            "preview": preview
        })



    # --- ENDPOINT: Reporte Dashboard de Nómina ---
    # --- ENDPOINT: Reporte Dashboard de Nómina (actualizado) ---
    @app.route("/reportes/nomina/dashboard", methods=["GET"])
    @jwt_required()
    def reporte_nomina_dashboard():
        from sqlalchemy.sql import text
        from datetime import datetime

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        # --- Parámetros de filtro ---
        mes = request.args.get("mes")
        anio = request.args.get("anio")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        empleado = request.args.get("empleado")

        condiciones = ["idcliente = :idcliente"]
        params = {"idcliente": idcliente}

        try:
            # Si hay rango de fechas, tiene prioridad total sobre año/mes
            if desde and hasta:
                try:
                    desde_dt = datetime.strptime(desde, "%Y-%m-%d").date()
                    hasta_dt = datetime.strptime(hasta, "%Y-%m-%d").date()
                    condiciones.append("periodo BETWEEN :desde AND :hasta")
                    params["desde"] = desde_dt
                    params["hasta"] = hasta_dt
                except ValueError:
                    return jsonify({"error": "Formato de fecha inválido. Use YYYY-MM-DD"}), 400
            else:
                if anio and anio.strip() and anio != "0":
                    condiciones.append("EXTRACT(YEAR FROM periodo) = :anio")
                    params["anio"] = int(anio)

                if mes and mes.strip() and mes != "0":
                    condiciones.append("EXTRACT(MONTH FROM periodo) = :mes")
                    params["mes"] = int(mes)

            # Filtro por empleado
            if empleado and empleado.strip():
                condiciones.append("identificacion = :empleado")
                params["empleado"] = empleado.strip()

        except Exception as e:
            print(f"[ERROR filtros nómina] {e}")
            return jsonify({"error": "Parámetros inválidos"}), 400

        where_sql = " AND ".join(condiciones)

        # --- KPIs globales ---
        sql_global = text(f"""
            SELECT
                COUNT(DISTINCT identificacion) AS empleados,
                COALESCE(SUM(sueldo), 0) AS total_sueldos,
                COALESCE(SUM(aux_transporte), 0) AS total_auxilios,
                COALESCE(SUM(auxilio_extralegal), 0) AS total_extralegal,
                COALESCE(SUM(prima), 0) AS total_prima,
                COALESCE(SUM(intereses_cesantias), 0) AS total_intereses_cesantias,
                COALESCE(SUM(total_ingresos), 0) AS total_ingresos,
                COALESCE(SUM(fondo_salud), 0) AS total_salud,
                COALESCE(SUM(fondo_pension), 0) AS total_pension,
                COALESCE(SUM(fondo_solidaridad), 0) AS total_solidaridad,
                COALESCE(SUM(retefuente), 0) AS total_retefuente,
                COALESCE(SUM(prestamos), 0) AS total_prestamos,
                COALESCE(SUM(total_deducciones), 0) AS total_deducciones,
                COALESCE(SUM(neto_pagar), 0) AS total_neto_pagar
            FROM siigo_nomina
            WHERE {where_sql}
        """)
        globales = db.session.execute(sql_global, params).mappings().first()

        # --- Totales por empleado ---
        sql_por_empleado = text(f"""
            SELECT
                nombre,
                identificacion,
                MAX(no_contrato) AS no_contrato,
                COALESCE(SUM(sueldo), 0) AS sueldo,
                COALESCE(SUM(aux_transporte), 0) AS aux_transporte,
                COALESCE(SUM(auxilio_extralegal), 0) AS auxilio_extralegal,
                COALESCE(SUM(prima), 0) AS prima,
                COALESCE(SUM(intereses_cesantias), 0) AS intereses_cesantias,
                COALESCE(SUM(total_ingresos), 0) AS total_ingresos,
                COALESCE(SUM(fondo_salud), 0) AS fondo_salud,
                COALESCE(SUM(fondo_pension), 0) AS fondo_pension,
                COALESCE(SUM(fondo_solidaridad), 0) AS fondo_solidaridad,
                COALESCE(SUM(retefuente), 0) AS retefuente,
                COALESCE(SUM(prestamos), 0) AS prestamos,
                COALESCE(SUM(total_deducciones), 0) AS total_deducciones,
                COALESCE(SUM(neto_pagar), 0) AS neto_pagar
            FROM siigo_nomina
            WHERE {where_sql}
            GROUP BY nombre, identificacion
            ORDER BY COALESCE(SUM(neto_pagar), 0) DESC
        """)
        empleados = [dict(r) for r in db.session.execute(sql_por_empleado, params).mappings().all()]

        top_empleados = empleados[:10]

        # --- Evolución mensual ---
        sql_evolucion = text(f"""
            SELECT
                TO_CHAR(periodo, 'YYYY-MM') AS periodo,
                COUNT(DISTINCT identificacion) AS empleados,
                COALESCE(SUM(total_ingresos), 0) AS total_ingresos,
                COALESCE(SUM(total_deducciones), 0) AS total_deducciones,
                COALESCE(SUM(neto_pagar), 0) AS total_neto_pagar
            FROM siigo_nomina
            WHERE {where_sql}
            GROUP BY periodo
            ORDER BY periodo
        """)
        evolucion_mensual = [
            dict(r) for r in db.session.execute(sql_evolucion, params).mappings().all()
        ]

        return jsonify({
            "globales": dict(globales) if globales else {},
            "empleados": empleados,
            "top_empleados": top_empleados,
            "evolucion_mensual": evolucion_mensual
        })


    ############ ENDPOINTS PRODUCTOS ############
    @app.route("/siigo/sync-productos", methods=["POST"])
    def siigo_sync_productos():
        idcliente = obtener_idcliente_desde_request()
        print(f"🔹 Sync productos iniciado para cliente {idcliente}")

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        # Cuando este endpoint es llamado desde /siigo/sync-all,
        # NO debe crear log individual para no duplicar el historial.
        # El log general lo crea /siigo/sync-all al final.
        es_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        endpoint_actual = "/siigo/sync-productos"
        ejecutado_en = datetime.now(timezone.utc)

        pasos_ok = 0
        pasos_error = 0
        endpoint_fallido = None
        detalle_lines = []

        def guardar_log(resultado: str, detalle: str):
            """
            Guarda historial solo cuando la ejecución es manual por módulo.
            Si viene desde sync-all, no guarda log individual para evitar duplicados.
            """
            if es_sync_all:
                return None

            try:
                log = SiigoSyncLog(
                    idcliente=idcliente,
                    fecha_programada=ejecutado_en,
                    ejecutado_en=ejecutado_en,
                    origen="manual_modulo",
                    resultado=resultado,
                    total_pasos=1,
                    pasos_ok=pasos_ok,
                    pasos_error=pasos_error,
                    endpoint_fallido=endpoint_fallido,
                    detalle=detalle,
                )
                db.session.add(log)
                db.session.commit()
                return log.id

            except Exception as log_error:
                db.session.rollback()
                raise Exception(f"Error guardando log de sincronización productos: {str(log_error)}")

        try:
            from siigo.siigo_sync_productos import sync_productos_desde_siigo

            mensaje = sync_productos_desde_siigo(idcliente)

            pasos_ok = 1
            pasos_error = 0
            detalle_lines.append(f"✅ Productos sincronizados correctamente.")
            detalle_lines.append(str(mensaje))

            detalle = "\n".join(detalle_lines)

            try:
                log_id = guardar_log("OK", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": "Los productos se procesaron, pero falló el registro del historial.",
                    "detalle": detalle,
                    "error_log": str(log_error),
                    "pasos_ok": pasos_ok,
                    "pasos_error": pasos_error,
                }), 500

            return jsonify({
                "mensaje": mensaje,
                "estado": "OK",
                "log_id": log_id,
                "pasos_ok": pasos_ok,
                "pasos_error": pasos_error,
                "detalle": detalle,
            }), 200

        except Exception as e:
            db.session.rollback()

            pasos_ok = 0
            pasos_error = 1
            endpoint_fallido = endpoint_actual

            detalle_lines.append(f"❌ Error general en sincronización de productos: {str(e)}")
            detalle = "\n".join(detalle_lines)

            try:
                log_id = guardar_log("ERROR", detalle)
            except Exception as log_error:
                return jsonify({
                    "error": str(e),
                    "detalle": detalle,
                    "error_log": str(log_error),
                }), 500

            return jsonify({
                "error": str(e),
                "detalle": detalle,
                "log_id": log_id,
            }), 500




    # Debug: traer productos en bruto desde Siigo (sin JWT)
    @app.route("/siigo/debug-productos-raw", methods=["GET"])
    def siigo_debug_productos_raw():
        import requests
        from siigo.siigo_sync_refactor import siigo_auth_json, dec_local, _headers_bearer
        from models import SiigoCredencial

        idcliente = request.args.get("idcliente", type=int)
        if not idcliente:
            return jsonify({"error": "Falta idcliente"}), 400

        # --- credenciales ---
        cred = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cred:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        access_key = dec_local(cred.client_secret)
        if not access_key:
            return jsonify({"error": "No se pudo desencriptar Access Key"}), 400

        # --- token ---
        token_data = siigo_auth_json(base_url=cred.base_url, username=cred.client_id, access_key=access_key)
        token = token_data.get("access_token")
        headers = _headers_bearer(token)
        base_url = cred.base_url.rstrip("/")

        # --- recorrer productos con paginación ---
        all_productos = []
        page = 1
        page_size = 50
        while True:
            url = f"{base_url}/v1/products?page={page}&page_size={page_size}"
            r = requests.get(url, headers=headers, timeout=60)
            if r.status_code != 200:
                return jsonify({"error": f"Error {r.status_code}", "detalle": r.text}), 500
            data = r.json() or {}
            results = data.get("results") or []
            all_productos.extend(results)

            # salir si no hay más páginas
            if len(results) < page_size:
                break
            page += 1

        return jsonify({"productos": all_productos})



    # --- Reporte: Dashboard de Vendedores ---
    # --- Reporte de Vendedores ---
    @app.route("/reportes/vendedores", methods=["GET"])
    @jwt_required()
    def reportes_vendedores():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos")
        cliente = request.args.get("cliente")

        filtros = ["f.idcliente = :idcliente", "f.estado_pago = 'pagada'"]
        params = {"idcliente": idcliente}

        if desde:
            filtros.append("f.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            filtros.append("f.fecha <= :hasta")
            params["hasta"] = hasta
        if centro_costos:
            filtros.append("f.cost_center = :centro_costos")  # 👈 nombre correcto
            params["centro_costos"] = centro_costos

        if cliente:
            filtros.append("f.cliente_nombre = :cliente")
            params["cliente"] = cliente


        filtro_sql = " AND ".join(filtros)

        # KPIs
        sql_kpis = f"""
            SELECT
                COALESCE(SUM(f.total),0) AS ventas_totales,
                COUNT(*) AS facturas,
                CASE WHEN COUNT(*) > 0 THEN ROUND(SUM(f.total)/COUNT(*),2) ELSE 0 END AS ticket_promedio
            FROM facturas_enriquecidas f
            WHERE {filtro_sql}
        """
        kpis = db.session.execute(text(sql_kpis), params).mappings().first()

        # Top 5 vendedores
        sql_top = f"""
            SELECT
                COALESCE(f.vendedor_nombre, 'Sin asignar') AS vendedor_nombre,
                SUM(f.total) AS total,
                COUNT(*) AS facturas
            FROM facturas_enriquecidas f
            WHERE {filtro_sql}
            GROUP BY COALESCE(f.vendedor_nombre, 'Sin asignar')
            ORDER BY total DESC
            LIMIT 5
        """
        top5 = db.session.execute(text(sql_top), params).mappings().all()

        # Ranking completo
        sql_rank = f"""
            SELECT
                COALESCE(f.vendedor_nombre, 'Sin asignar') AS vendedor_nombre,
                SUM(f.total) AS total,
                COUNT(*) AS facturas
            FROM facturas_enriquecidas f
            WHERE {filtro_sql}
            GROUP BY COALESCE(f.vendedor_nombre, 'Sin asignar')
            ORDER BY total DESC
        """
        ranking = db.session.execute(text(sql_rank), params).mappings().all()

        return jsonify({
            "kpis": dict(kpis) if kpis else {},
            "top5": [dict(r) for r in top5],
            "ranking": [dict(r) for r in ranking]
        })





    # --- Reporte: Dashboard de Productos ---

    # --- Reporte de Productos ---
    @app.route("/reportes/productos", methods=["GET"])
    @jwt_required()
    def reportes_productos():
        claims = get_jwt()
        idcliente = claims.get("idcliente")

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costo = request.args.get("centro_costo")  # ✅ ahora es nombre, no id
        producto_code = request.args.get("producto_code")  # ✅ filtro adicional
        ordenar_por = request.args.get("ordenar_por", "cantidad")  # cantidad | total

        filtros = ["f.idcliente = :idcliente", "f.estado_pago = 'pagada'"]
        params = {"idcliente": idcliente}

        if desde:
            filtros.append("f.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            filtros.append("f.fecha <= :hasta")
            params["hasta"] = hasta
        if centro_costo:  # ✅ usamos nombre, no id
            filtros.append("f.centro_costo_nombre = :centro_costo")
            params["centro_costo"] = centro_costo
        if producto_code:
            filtros.append("i.producto_id = :producto_code")
            params["producto_code"] = producto_code

        where = " AND ".join(filtros)
        columna_orden = "SUM(i.cantidad)" if ordenar_por == "cantidad" else "SUM(i.total_item)"

        # --- KPIs ---
        sql_kpis = text(f"""
            SELECT
                COALESCE(SUM(i.total_item),0) AS ventas_totales,
                COUNT(DISTINCT f.factura_id) AS facturas,
                CASE WHEN COUNT(DISTINCT f.factura_id) > 0
                    THEN SUM(i.total_item) / COUNT(DISTINCT f.factura_id)
                    ELSE 0 END AS ticket_promedio
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            WHERE {where}
        """)
        kpis = dict(db.session.execute(sql_kpis, params).mappings().first() or {})

        # --- Top 10 ---
        sql_top = text(f"""
            SELECT p.code, p.name as producto, SUM(i.cantidad) as cantidad, SUM(i.total_item) as total
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            JOIN siigo_productos p ON p.code = i.producto_id AND p.idcliente = f.idcliente
            WHERE {where}
            GROUP BY p.code, p.name
            ORDER BY {columna_orden} DESC
            LIMIT 10
        """)
        top10 = [dict(r) for r in db.session.execute(sql_top, params).mappings().all()]

        # --- Bottom 10 ---
        sql_bottom = text(f"""
            SELECT p.code, p.name as producto, SUM(i.cantidad) as cantidad, SUM(i.total_item) as total
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            JOIN siigo_productos p ON p.code = i.producto_id AND p.idcliente = f.idcliente
            WHERE {where}
            GROUP BY p.code, p.name
            ORDER BY {columna_orden} ASC
            LIMIT 10
        """)
        bottom10 = [dict(r) for r in db.session.execute(sql_bottom, params).mappings().all()]

        return jsonify({"kpis": kpis, "top10": top10, "bottom10": bottom10})



    # --- Catálogo de productos disponibles (para el filtro) ---
    @app.route("/catalogos/productos", methods=["GET"])
    @jwt_required()
    def catalogo_productos():
        claims = get_jwt()
        idcliente = claims.get("idcliente")

        sql = text("""
            SELECT DISTINCT p.code, p.name
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            JOIN siigo_productos p ON p.code = i.producto_id AND p.idcliente = f.idcliente
            WHERE f.idcliente = :idcliente
            ORDER BY p.code
        """)
        rows = db.session.execute(sql, {"idcliente": idcliente}).mappings().all()
        return jsonify([{"code": r["code"], "name": r["name"], "label": f"{r['code']} - {r['name']}"} for r in rows])


    # --- Detalle producto con histórico mensual ---
    @app.route("/reportes/productos/detalle", methods=["GET"])
    @jwt_required()
    def detalle_producto():
        claims = get_jwt()
        idcliente = claims.get("idcliente")

        producto_code = request.args.get("producto_code")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costo = request.args.get("centro_costo")

        if not producto_code:
            return jsonify({"error": "Falta producto_code"}), 400

        filtros = ["f.idcliente = :idcliente", "f.estado_pago = 'pagada'", "i.producto_id = :producto_code"]
        params = {"idcliente": idcliente, "producto_code": producto_code}

        if desde:
            filtros.append("f.fecha >= :desde")
            params["desde"] = desde
        if hasta:
            filtros.append("f.fecha <= :hasta")
            params["hasta"] = hasta
        if centro_costo:
            filtros.append("f.centro_costo_nombre = :centro_costo")
            params["centro_costo"] = centro_costo

        where = " AND ".join(filtros)

        # --- Totales ---
        sql_totales = text(f"""
            SELECT p.code, p.name as producto,
                SUM(i.cantidad) as cantidad,
                SUM(i.total_item) as total,
                COUNT(DISTINCT f.factura_id) as facturas
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            JOIN siigo_productos p ON p.code = i.producto_id AND p.idcliente = f.idcliente
            WHERE {where}
            GROUP BY p.code, p.name
        """)
        detalle = dict(db.session.execute(sql_totales, params).mappings().first() or {})

        # --- Histórico mensual ---
        sql_hist = text(f"""
            SELECT DATE_TRUNC('month', f.fecha) AS mes,
                SUM(i.cantidad) AS cantidad,
                SUM(i.total_item) AS total
            FROM siigo_factura_items i
            JOIN facturas_enriquecidas f ON f.factura_id = i.factura_id
            WHERE {where}
            GROUP BY DATE_TRUNC('month', f.fecha)
            ORDER BY mes
        """)
        historico = [dict(r) for r in db.session.execute(sql_hist, params).mappings().all()]

        detalle["historico"] = historico
        return jsonify(detalle)




    # Endpoint de balance inicial:
    @app.route("/siigo/test-balance-report", methods=["GET"])
    def siigo_test_balance_report():
        """
        Endpoint para generar y obtener el enlace de descarga del Balance de Prueba General desde Siigo.
        Ejemplo:
        /siigo/test-balance-report?idcliente=1&year=2025&month_start=1&month_end=12
        """
        import requests

        idcliente = request.args.get("idcliente", type=int)
        year = request.args.get("year", type=int, default=2025)
        month_start = request.args.get("month_start", type=int, default=1)
        month_end = request.args.get("month_end", type=int, default=12)
        include_tax_diff = request.args.get("include_tax_diff", type=lambda v: v.lower() == "true", default=False)

        # Validación básica
        if not idcliente:
            return jsonify({"error": "Falta idcliente (ej: /siigo/test-balance-report?idcliente=1)"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        # 🔧 Payload: dejamos account_start y account_end vacíos según la documentación
        payload = {
            "account_start": "",
            "account_end": "",
            "year": year,
            "month_start": month_start,
            "month_end": month_end,
            "includes_tax_difference": include_tax_diff,
        }

        url = f"{base_url}/v1/test-balance-report"

        try:
            r = requests.post(url, json=payload, headers=headers, timeout=120)
        except Exception as e:
            return jsonify({"error": "Error de conexión con Siigo", "detalle": str(e)}), 500

        status = r.status_code
        try:
            data = r.json()
        except Exception:
            data = {"raw": r.text}

        if status in (200, 201) and "file_url" in data:
            return jsonify({
                "mensaje": "Balance generado exitosamente",
                "status": status,
                "file_id": data.get("file_id"),
                "file_url": data.get("file_url"),
                "payload_enviado": payload
            }), 200

        return jsonify({
            "error": "Fallo al generar el balance",
            "status": status,
            "respuesta": data,
            "payload_enviado": payload
        }), status



    # --- Generar link de descarga Balance de Prueba (JWT protegido) ---
    @app.route("/siigo/balance/generar", methods=["POST"])
    @jwt_required()
    def siigo_generar_balance():
        """
        Genera el link de descarga del Balance de Prueba General desde Siigo.
        Se usa dentro del frontend del módulo Balance de Prueba.

        Body esperado (JSON o FormData):
        {
            "year": 2025,
            "month_start": 1,
            "month_end": 12,
            "includes_tax_difference": false
        }
        """
        import requests

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        # --- Leer parámetros ---
        data = request.get_json(silent=True) or request.form or {}
        year = int(data.get("year", 2025))
        month_start = int(data.get("month_start", 1))
        month_end = int(data.get("month_end", 12))
        include_tax_diff = bool(data.get("includes_tax_difference", False))

        # --- Obtener credenciales Siigo del cliente ---
        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({"error": "Auth inesperado", "detalle": str(auth_data)}), 500

        token = auth_data.get("access_token")
        if not token:
            return jsonify({"error": "No se obtuvo access_token", "detalle": auth_data}), 500

        base_url = cfg.base_url.rstrip("/")
        headers = _siigo_headers_bearer(token)

        # --- Payload de la petición ---
        payload = {
            "account_start": "",
            "account_end": "",
            "year": year,
            "month_start": month_start,
            "month_end": month_end,
            "includes_tax_difference": include_tax_diff,
        }

        url = f"{base_url}/v1/test-balance-report"

        try:
            r = requests.post(url, json=payload, headers=headers, timeout=120)
        except Exception as e:
            return jsonify({"error": "Error de conexión con Siigo", "detalle": str(e)}), 500

        status = r.status_code
        try:
            data = r.json()
        except Exception:
            data = {"raw": r.text}

        # --- Éxito ---
        if status in (200, 201) and "file_url" in data:
            return jsonify({
                "mensaje": "Balance generado exitosamente",
                "status": status,
                "file_id": data.get("file_id"),
                "file_url": data.get("file_url"),
                "payload_enviado": payload
            }), 200

        # --- Error ---
        return jsonify({
            "error": "Fallo al generar el balance",
            "status": status,
            "respuesta": data,
            "payload_enviado": payload
        }), status





    #Balance de PRUEBA
    # --- Importar Balance de Prueba desde Archivo Excel ---
    @app.route("/importar/balance-excel", methods=["POST"])
    @jwt_required()
    def importar_balance_desde_excel():
        import pandas as pd
        from datetime import datetime, timezone

        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente asociado"}), 400

        # Validar archivo
        if "archivo" not in request.files:
            return jsonify({"error": "Archivo no proporcionado"}), 400
        file = request.files["archivo"]

        # Leer parámetros de periodo
        anio = request.form.get("anio")
        mes_inicio = request.form.get("mes_inicio")
        mes_fin = request.form.get("mes_fin")

        if not anio or not mes_inicio or not mes_fin:
            return jsonify({"error": "Debe indicar año, mes de inicio y mes de fin"}), 400

        try:
            anio = int(anio)
            mes_inicio = int(mes_inicio)
            mes_fin = int(mes_fin)
        except ValueError:
            return jsonify({"error": "Mes o año inválido"}), 400

        # Leer archivo Excel con pandas y calamine
        try:
            df = pd.read_excel(file, header=4, engine="calamine")  # encabezado en fila 5
            df = df.dropna(how="all")  # eliminar filas vacías
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"No se pudo leer el Excel: {str(e)}"}), 400

        # Verificar columnas esperadas
        columnas_esperadas = [
            "Nivel",
            "Transaccional",
            "Código cuenta contable",
            "Nombre Cuenta contable",
            "Saldo Inicial",
            "Movimiento Débito",
            "Movimiento Crédito",
            "Saldo final",
        ]
        faltantes = [c for c in columnas_esperadas if c not in df.columns]
        if faltantes:
            return jsonify({"error": f"Faltan columnas esperadas: {faltantes}"}), 400

        # Eliminar registros anteriores
        try:
            BalancePrueba.query.filter_by(
                idcliente=idcliente,
                periodo_anio=anio,
                periodo_mes_inicio=mes_inicio,
                periodo_mes_fin=mes_fin
            ).delete()
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": f"Error al limpiar balances previos: {str(e)}"}), 500

        registros_creados = 0
        errores = []

        niveles_validos = {"Clase", "Grupo", "Cuenta", "Subcuenta", "Auxiliar", "Subauxiliar"}

        # Insertar cada fila
        for idx, row in df.iterrows():
            try:
                nivel = str(row.get("Nivel") or "").strip()
                if nivel not in niveles_validos:
                    continue  # ignorar filas inválidas

                registro = BalancePrueba(
                    idcliente=idcliente,
                    codigo_cuenta=str(row.get("Código cuenta contable") or "").strip(),
                    nombre_cuenta=str(row.get("Nombre Cuenta contable") or "").strip(),
                    nivel=nivel,
                    es_transaccional=str(row.get("Transaccional") or "").strip().lower() == "sí",
                    saldo_inicial=float(str(row.get("Saldo Inicial") or "0").replace(",", "").replace(" ", "")),
                    movimiento_debito=float(str(row.get("Movimiento Débito") or "0").replace(",", "").replace(" ", "")),
                    movimiento_credito=float(str(row.get("Movimiento Crédito") or "0").replace(",", "").replace(" ", "")),
                    saldo_final=float(str(row.get("Saldo final") or "0").replace(",", "").replace(" ", "")),
                    periodo_anio=anio,
                    periodo_mes_inicio=mes_inicio,
                    periodo_mes_fin=mes_fin,
                    fecha_carga=datetime.now(timezone.utc)
                )
                db.session.add(registro)
                registros_creados += 1
            except Exception as e:
                errores.append(f"Fila {idx + 5}: {e}")

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": f"Error al guardar en BD: {str(e)}"}), 500

        return jsonify({
            "mensaje": f"Balance {anio} ({mes_inicio} → {mes_fin}) importado correctamente.",
            "registros_creados": registros_creados,
            "errores_preview": errores[:5]
        }), 200


    # Endpoint que analisa el balance por clases
    # Endpoint que analiza el balance por clases contables (alineado con indicadores)
    @app.route("/reportes/balance/clases", methods=["GET"])
    @jwt_required()
    def resumen_por_clase_contable():
        idcliente = get_jwt().get("idcliente")
        anio = request.args.get("anio", type=int)
        mes_inicio = request.args.get("mes_inicio", type=int)
        mes_fin = request.args.get("mes_fin", type=int)

        if not idcliente or not anio or not mes_inicio or not mes_fin:
            return jsonify({"error": "Faltan parámetros"}), 400

        registros = db.session.query(
            BalancePrueba.codigo_cuenta,
            BalancePrueba.saldo_final
        ).filter(
            BalancePrueba.idcliente == idcliente,
            BalancePrueba.periodo_anio == anio,
            BalancePrueba.periodo_mes_inicio == mes_inicio,
            BalancePrueba.periodo_mes_fin == mes_fin,
            BalancePrueba.es_transaccional == True
        ).all()

        # Mapeo de clases por prefijo
        clases = {
            "Activo": ["1"],
            "Pasivo": ["2"],
            "Patrimonio": ["3"],
            "Ingresos": ["41", "42"],
            "Costos": ["61"],
            "Gastos": ["51", "53"]
        }

        resumen_map = {
            "Activo": 0,
            "Pasivo": 0,
            "Patrimonio": 0,
            "Ingresos": 0,
            "Costos": 0,
            "Gastos": 0
        }

        for cod, saldo in registros:
            cod_str = str(cod).strip()
            for clase, prefijos in clases.items():
                if any(cod_str.startswith(pref) for pref in prefijos):
                    resumen_map[clase] += float(saldo or 0)
                    break

        # Para coherencia visual en el gráfico
        ingresos = abs(resumen_map["Ingresos"])
        costos = abs(resumen_map["Costos"])
        gastos = abs(resumen_map["Gastos"])
        resultado_neto = ingresos - costos - gastos

        data = [
            {"clase": "Ingresos", "valor": ingresos},
            {"clase": "Costos", "valor": costos},
            {"clase": "Gastos", "valor": gastos},
            {"clase": "Resultado Neto (Utilidad o Pérdida)", "valor": resultado_neto},
            {"clase": "Activo", "valor": resumen_map["Activo"]},
            {"clase": "Pasivo", "valor": resumen_map["Pasivo"]},
            {"clase": "Patrimonio", "valor": resumen_map["Patrimonio"]},
        ]

        return jsonify({"resumen": data})





    @app.route("/reportes/balance/resumen", methods=["GET"])
    @jwt_required()
    def resumen_balance():
        idcliente = get_jwt().get("idcliente")
        anio = request.args.get("anio", type=int)
        mes_inicio = request.args.get("mes_inicio", type=int)
        mes_fin = request.args.get("mes_fin", type=int)

        if not idcliente or not anio:
            return jsonify({"error": "Faltan parámetros"}), 400

        q = db.session.query(
            BalancePrueba.nivel.label("clase"),
            db.func.sum(BalancePrueba.saldo_final).label("saldo_final")
        ).filter_by(
            idcliente=idcliente,
            periodo_anio=anio,
            periodo_mes_inicio=mes_inicio,
            periodo_mes_fin=mes_fin
        ).group_by(BalancePrueba.nivel).all()

        resumen = [
            {"clase": r.clase, "saldo_final": float(r.saldo_final or 0)}
            for r in q if (r.saldo_final or 0) != 0
        ]

        print("Resumen generado:", resumen)
        return jsonify({"resumen": resumen})


    # PAra analisis del balance por grupos contables
    @app.route("/reportes/balance/grupos", methods=["GET"])
    @jwt_required()
    def resumen_por_grupo_contable():
        idcliente = get_jwt().get("idcliente")
        anio = request.args.get("anio", type=int)
        mes_inicio = request.args.get("mes_inicio", type=int)
        mes_fin = request.args.get("mes_fin", type=int)

        if not idcliente or not anio or not mes_inicio or not mes_fin:
            return jsonify({"error": "Faltan parámetros"}), 400

         # --- 1️⃣ Obtener registros del balance de prueba ---
        registros = db.session.query(
            BalancePrueba.codigo_cuenta,
            BalancePrueba.nombre_cuenta,
            BalancePrueba.saldo_final
        ).filter(
            BalancePrueba.idcliente == idcliente,
            BalancePrueba.periodo_anio == anio,
            BalancePrueba.periodo_mes_inicio == mes_inicio,
            BalancePrueba.periodo_mes_fin == mes_fin,
            BalancePrueba.es_transaccional == True
        ).all()

        # Agrupaciones por prefijo de código
        grupos = {
            "Activo Corriente": ["11", "13"],
            "Activo No Corriente": ["15", "17"],
            "Pasivo a Corto Plazo": ["21", "22"],
            "Pasivo a Largo Plazo": ["23", "24", "25"],
            "Ingresos": ["41", "42"],
            "Costos": ["61"],
            "Gastos": ["51", "53"],
            "Patrimonio": ["3"],
            "Pasivo Total": ["2"],  # Nuevo grupo adicional si se requiere
        }


        resumen = []
        detalle = []

        for grupo, prefijos in grupos.items():
            total = 0
            cuentas = []

            for codigo, nombre, saldo in registros:
                codigo_str = str(codigo)
                if any(codigo_str.startswith(pref) for pref in prefijos):
                    total += saldo or 0
                    cuentas.append({
                        "codigo": codigo,
                        "nombre": nombre,
                        "valor": saldo or 0
                    })

            if cuentas:
                resumen.append({"grupo": grupo, "valor": total})
                detalle.append({"grupo": grupo, "cuentas": cuentas})

        def safe_val(grupo):
            """Devuelve el valor del grupo o 0 si no existe."""
            return next((item["valor"] for item in resumen if item["grupo"] == grupo), 0)

        # --- 3️⃣ Variables base coherentes con visión gerencial ---
        activo_corriente = safe_val("Activo Corriente")
        activo_no_corriente = safe_val("Activo No Corriente")
        pasivo_corto = abs(safe_val("Pasivo a Corto Plazo"))
        pasivo_largo = abs(safe_val("Pasivo a Largo Plazo"))
        ingresos = abs(safe_val("Ingresos"))  # 🔹 Ingresos siempre positivos
        costos = abs(safe_val("Costos"))      # 🔹 Costos siempre positivos
        gastos = abs(safe_val("Gastos"))      # 🔹 Gastos siempre positivos

        activo_total = activo_corriente + activo_no_corriente
        pasivo_total = pasivo_corto + pasivo_largo
        utilidad_neta = ingresos - costos - gastos

        # --- 4️⃣ Indicadores financieros consistentes ---
        indicadores = {
            "liquidez": round(activo_corriente / pasivo_corto, 2) if pasivo_corto else None,
            "apalancamiento": round(pasivo_total / activo_total, 2) if activo_total else None,
            "rentabilidad": round(utilidad_neta / ingresos, 2) if ingresos else None
        }

        # --- 5️⃣ Conclusiones automáticas coherentes ---
        conclusiones = []


        # Liquidez
        if indicadores["liquidez"] is not None:
            if indicadores["liquidez"] < 1:
                conclusiones.append("🚨 La empresa podría tener problemas de liquidez (activo corriente < pasivo corto plazo).")
            else:
                conclusiones.append("✅ Buena liquidez: el activo corriente cubre las obligaciones de corto plazo.")

        # Apalancamiento
        if indicadores["apalancamiento"] is not None:
            if indicadores["apalancamiento"] > 0.6:
                conclusiones.append("⚠ Alto nivel de endeudamiento. Evalúa reducir pasivos.")
            else:
                conclusiones.append("✅ Apalancamiento controlado.")

        # Rentabilidad
        if indicadores["rentabilidad"] is not None:
            if indicadores["rentabilidad"] < 0:
                conclusiones.append("🔻 Rentabilidad negativa: la empresa tuvo pérdidas netas.")
            elif indicadores["rentabilidad"] < 0.1:
                conclusiones.append("⚠ Rentabilidad positiva pero baja.")
            else:
                conclusiones.append("✅ Buena rentabilidad sobre ingresos.")

        # --- 6️⃣ Agregar resultado neto al resumen (antes del return)
        resumen.append({
            "grupo": "Resultado Neto (Utilidad o Pérdida)",
            "valor": utilidad_neta
        })

        # --- 7️⃣ Respuesta final
        return jsonify({
            "resumen": resumen,
            "detalle": detalle,
            "indicadores": indicadores,
            "conclusiones": conclusiones
        })





    # ================================================
    # 📊 Indicadores Financieros con Diagnóstico Ejecutivo
    # ================================================

    # Para reporte de Indicadores Financieros
    @app.route("/reportes/indicadores/financieros", methods=["GET"])
    @jwt_required()
    def indicadores_financieros():
        idcliente = get_jwt().get("idcliente")
        anio = request.args.get("anio", type=int)
        mes_inicio = request.args.get("mes_inicio", type=int)
        mes_fin = request.args.get("mes_fin", type=int)

        if not idcliente or not anio or not mes_inicio or not mes_fin:
            return jsonify({"error": "Faltan parámetros"}), 400

        registros = db.session.query(
            BalancePrueba.codigo_cuenta,
            BalancePrueba.nombre_cuenta,
            BalancePrueba.saldo_final
        ).filter(
            BalancePrueba.idcliente == idcliente,
            BalancePrueba.periodo_anio == anio,
            BalancePrueba.periodo_mes_inicio == mes_inicio,
            BalancePrueba.periodo_mes_fin == mes_fin,
            BalancePrueba.es_transaccional == True
        ).all()

        # Mapeo de grupos por prefijo
        grupos = {
            "Activo Corriente": ["11", "13"],
            "Activo No Corriente": ["15", "17"],
            "Pasivo a Corto Plazo": ["21", "22"],
            "Pasivo a Largo Plazo": ["23", "24", "25"],
            "Ingresos": ["41", "42"],
            "Costos": ["61"],
            "Gastos": ["51", "53"],
            "Patrimonio": ["3"],
        }

        # Sumas por grupo
        resumen_map = {}
        for grupo, prefijos in grupos.items():
            total = 0
            for codigo, _, saldo in registros:
                if any(str(codigo).startswith(pref) for pref in prefijos):
                    total += float(saldo or 0)
            resumen_map[grupo] = total

        # Variables base (signos coherentes para cálculo)
        activo_corriente = float(resumen_map.get("Activo Corriente", 0))
        activo_no_corriente = float(resumen_map.get("Activo No Corriente", 0))
        pasivo_corto = abs(float(resumen_map.get("Pasivo a Corto Plazo", 0)))
        pasivo_largo = abs(float(resumen_map.get("Pasivo a Largo Plazo", 0)))
        pasivo_total_real = sum(
            float(saldo or 0)
            for codigo, _, saldo in registros
            if str(codigo).startswith("2")
        )
        ingresos = abs(float(resumen_map.get("Ingresos", 0)))
        costos = abs(float(resumen_map.get("Costos", 0)))
        gastos = abs(float(resumen_map.get("Gastos", 0)))
        patrimonio = float(resumen_map.get("Patrimonio", 0))

        activo_total = activo_corriente + activo_no_corriente
        pasivo_total = abs(pasivo_total_real)
        utilidad_neta = ingresos - costos - gastos

        # Indicadores
        indicadores = {
            "liquidez": round(activo_corriente / pasivo_corto, 2) if pasivo_corto else None,
            "apalancamiento": round(pasivo_total / activo_total, 2) if activo_total else None,
            "rentabilidad": round(utilidad_neta / ingresos, 2) if ingresos else None,
            "capital_trabajo": round(activo_corriente - pasivo_corto, 2),
            "solvencia": round(activo_total / pasivo_total, 2) if pasivo_total else None,
            "autonomia": round(patrimonio / activo_total, 2) if activo_total else None,
            "porcentaje_pasivo_corto": round(pasivo_corto / pasivo_total, 2) if pasivo_total else None,
            "porcentaje_activo_no_corriente": round(activo_no_corriente / activo_total, 2) if activo_total else None,
            "cobertura_activo_pasivo": round(activo_total / pasivo_total, 2) if pasivo_total else None,
            # Solo si patrimonio > 0, de lo contrario no interpretable
            "endeudamiento_largo_plazo": round(pasivo_largo / patrimonio, 2) if patrimonio and patrimonio > 0 else None,
            # Para poder armar el resumen técnico en el front si lo requieren
            "activo_total": round(activo_total, 2),
            "pasivo_total": round(pasivo_total, 2),
            "patrimonio": round(patrimonio, 2),
            "ingresos": round(ingresos, 2),
            "costos": round(costos, 2),
            "gastos": round(gastos, 2),
            "utilidad_neta": round(utilidad_neta, 2),
        }

        # Interpretaciones dinámicas para cada indicador
        interpretaciones = {
            k: interpretar_indicador(k, v)
            for k, v in indicadores.items()
        }

        # Explicaciones breves por indicador (tooltip/tarjetas)
        explicaciones = {
            "liquidez": "Activo corriente / Pasivo a corto. >1 saludable; >2 holgado.",
            "apalancamiento": "Pasivo total / Activo total. <0.6 ideal; >0.8 alto.",
            "rentabilidad": "Utilidad neta / Ingresos. >0 indica margen neto positivo.",
            "capital_trabajo": "Activo corriente − Pasivo corto. >0 indica colchón operativo.",
            "solvencia": "Activo total / Pasivo total. >1 indica cobertura de deudas.",
            "autonomia": "Patrimonio / Activo total. >0.5 indica menor dependencia de deuda.",
            "porcentaje_pasivo_corto": "Proporción de deuda exigible pronto.",
            "porcentaje_activo_no_corriente": "Proporción de activos no líquidos.",
            "cobertura_activo_pasivo": "Cobertura de pasivos con activos.",
            "endeudamiento_largo_plazo": "Deuda estructural vs patrimonio (si patrimonio > 0).",
        }

        # Resumen técnico gerencial (valores positivos para claridad)
        resumen_financiero = [
            {"clase": "Activo total", "valor": float(activo_total), "interpretacion": "Total de activos."},
            {"clase": "Pasivo total", "valor": float(pasivo_total), "interpretacion": "Deuda total acumulada."},
            {"clase": "Patrimonio", "valor": float(patrimonio), "interpretacion": "Capital propio neto (puede ser negativo)."},
            {"clase": "Ingresos", "valor": float(ingresos), "interpretacion": "Ventas totales del período."},
            {"clase": "Costos", "valor": float(costos), "interpretacion": "Costo directo de operaciones."},
            {"clase": "Gastos", "valor": float(gastos), "interpretacion": "Gasto operativo y administrativo."},
            {"clase": "Utilidad neta", "valor": float(utilidad_neta), "interpretacion": "Resultado neto del período."},
        ]

        # Conclusiones automáticas
        conclusiones = []
        if indicadores["liquidez"] is not None:
            if indicadores["liquidez"] < 1:
                conclusiones.append("⚠ Riesgo de iliquidez: el activo corriente no cubre el pasivo a corto.")
            elif indicadores["liquidez"] > 3:
                conclusiones.append("⚠ Exceso de liquidez: posible ineficiencia en el uso del capital disponible.")
            else:
                conclusiones.append("✅ Liquidez saludable para cubrir obligaciones de corto plazo.")
        if indicadores["apalancamiento"] is not None:
            if indicadores["apalancamiento"] > 0.8:
                conclusiones.append("⚠ Apalancamiento alto: gran parte de los activos están financiados con deuda.")
            elif indicadores["apalancamiento"] > 0.6:
                conclusiones.append("• Apalancamiento moderado: monitorear endeudamiento.")
            else:
                conclusiones.append("✅ Estructura de capital sana (endeudamiento controlado).")
        if indicadores["rentabilidad"] is not None:
            if indicadores["rentabilidad"] < 0:
                conclusiones.append("🔻 Rentabilidad negativa: la empresa ha tenido pérdidas netas en el período.")
            elif indicadores["rentabilidad"] < 0.1:
                conclusiones.append("• Rentabilidad positiva pero baja: revisar eficiencia operativa.")
            else:
                conclusiones.append("✅ Buen margen neto sobre ingresos.")
        if indicadores["autonomia"] is not None and indicadores["autonomia"] < 0.3:
            conclusiones.append("⚠ Alta dependencia del financiamiento externo (autonomía < 30%).")
        if patrimonio <= 0:
            conclusiones.append("❗ Patrimonio negativo o nulo: la empresa presenta pérdida acumulada superior a su capital.")
        elif indicadores["endeudamiento_largo_plazo"] is not None and indicadores["endeudamiento_largo_plazo"] > 1:
            conclusiones.append("⚠ Endeudamiento estructural alto respecto al patrimonio.")

        return jsonify({
            "resumen_financiero": resumen_financiero,
            "indicadores": indicadores,
            "explicaciones": explicaciones,
            "interpretaciones": interpretaciones,
            "conclusiones": conclusiones
        })

    # NUEVO Reporte cruce de IVAs - MArzo 23 2026 con reporte de Auxiliar contable
    # --- Importar Movimiento Auxiliar desde Archivo Excel (VERSIÓN FINAL COMPATIBLE) ---
    @app.route("/reportes/cargar_auxiliar", methods=["POST"])
    @jwt_required()
    def importar_auxiliar_desde_excel():
        import pandas as pd
        import re
        idcliente = get_jwt().get("idcliente")
        file = request.files.get("archivo")
        if not file: return jsonify({"error": "No hay archivo"}), 400

        try:
            from models import AuxiliarContable
            df = pd.read_excel(file, header=4, engine="calamine")
            df.columns = [str(c).strip() for c in df.columns]
            
            lista_mapeada = []
            fechas_procesadas = []

            for idx, row in df.iterrows():
                cta_raw = str(row.get('Código cuenta contable') or "").strip()
                
                # REGLA DE ORO: Si la celda dice "Total" o está vacía, la saltamos
                if "Total" in cta_raw or cta_raw == "" or cta_raw.lower() == "nan":
                    continue

                f_raw = row.get('Fecha Elaboración')
                if not f_raw or str(f_raw).lower() == 'nan': continue
                
                f_dt = pd.to_datetime(f_raw, dayfirst=True)
                fechas_procesadas.append(f_dt)

                # Extraer solo el número de la cuenta (ej: 24080601)
                match = re.search(r"(\d+)", cta_raw)
                codigo_limpio = match.group(1) if match else cta_raw

                def clean_num(val):
                    if val is None or str(val).lower() == 'nan': return 0.0
                    return float(str(val).replace(",", "").replace(" ", "").replace("$", ""))

                lista_mapeada.append({
                    "idcliente": idcliente,
                    "fecha_contable": f_dt.date(),
                    "cuenta_codigo": codigo_limpio,
                    "cuenta_nombre": str(row.get('Cuenta contable') or "").strip(),
                    "debito": clean_num(row.get('Débito')),
                    "credito": clean_num(row.get('Crédito')),
                    "periodo_anio": f_dt.year,
                    "periodo_mes": f_dt.month
                })

            if lista_mapeada:
                # Borrar para reescribir
                f_min, f_max = min(fechas_procesadas), max(fechas_procesadas)
                AuxiliarContable.query.filter(AuxiliarContable.idcliente == idcliente, 
                                             AuxiliarContable.fecha_contable >= f_min, 
                                             AuxiliarContable.fecha_contable <= f_max).delete()
                db.session.bulk_insert_mappings(AuxiliarContable, lista_mapeada)
                db.session.commit()
                
            return jsonify({"detalles": {"registros_procesados": len(lista_mapeada)}}), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({"error": str(e)}), 500
        

    @app.route("/reportes/cruce_iva_v2", methods=["GET"])
    @jwt_required()
    def get_cruce_iva_v2():
        from datetime import datetime, timedelta
        from sqlalchemy import text
        
        idcliente = get_jwt().get("idcliente")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        
        # Parámetros de filtro de tasa
        inc_19 = request.args.get("inc19", "true").lower() == "true"
        inc_5 = request.args.get("inc5", "false").lower() == "true"

        # 1. Filtros dinámicos para Totales (KPIs) - Mapeo exacto según tu Excel
        f_vtas = []
        f_comps = []
        
        if inc_19:
            # Ventas 19% (Generado 01 + Devoluciones 02)
            f_vtas.append("(cuenta_codigo LIKE '24080601%' OR cuenta_codigo LIKE '24080602%')")
            # Compras 19% (Descontable 1001/02 + Servicios 1501/02)
            f_comps.extend(["cuenta_codigo LIKE '24081001%'", "cuenta_codigo LIKE '24081002%'", 
                            "cuenta_codigo LIKE '24081501%'", "cuenta_codigo LIKE '24081502%'"])
        
        if inc_5:
            # Ventas 5% (Generado 03)
            f_vtas.append("cuenta_codigo LIKE '24080603%'")
            # Compras 5% (Descontable 1003 + Servicios 1503)
            f_comps.extend(["cuenta_codigo LIKE '24081003%'", "cuenta_codigo LIKE '24081503%'"])

        sql_vtas_dinamico = " OR ".join(f_vtas) if f_vtas else "1=0"
        sql_comps_dinamico = " OR ".join(f_comps) if f_comps else "1=0"

        # 2. SQL Corregido: Calculamos NETOS reales (Crédito - Débito para ventas)
        sql = text(f"""
            SELECT 
                periodo_anio, periodo_mes,
                -- Apertura para barras del gráfico (Individuales)
                SUM(CASE WHEN (cuenta_codigo LIKE '24080601%' OR cuenta_codigo LIKE '24080602%') THEN (credito - debito) ELSE 0 END) AS v19,
                SUM(CASE WHEN cuenta_codigo LIKE '24080603%' THEN (credito - debito) ELSE 0 END) AS v5,
                SUM(CASE WHEN (cuenta_codigo LIKE '24081001%' OR cuenta_codigo LIKE '24081002%' OR cuenta_codigo LIKE '24081501%' OR cuenta_codigo LIKE '24081502%') THEN (debito - credito) ELSE 0 END) AS c19,
                SUM(CASE WHEN (cuenta_codigo LIKE '24081003%' OR cuenta_codigo LIKE '24081503%') THEN (debito - credito) ELSE 0 END) AS c5,
                
                -- Totales calculados según selectores (KPIs y Tabla)
                SUM(CASE WHEN ({sql_vtas_dinamico}) THEN (credito - debito) ELSE 0 END) AS v_total,
                SUM(CASE WHEN ({sql_comps_dinamico}) THEN (debito - credito) ELSE 0 END) AS c_total,
                
                -- ReteIVA (Cálculo Neto: Débito - Crédito)
                SUM(CASE WHEN cuenta_codigo LIKE '135517%' THEN (debito - credito) ELSE 0 END) AS rete
            FROM auxiliar_contable
            WHERE idcliente = :idc AND fecha_contable BETWEEN :d AND :h
            GROUP BY 1, 2 ORDER BY 1, 2
        """)
        
        res = db.session.execute(sql, {"idc": idcliente, "d": desde, "h": hasta}).mappings().all()
        
        series_mensuales = []
        for r in res:
            f_actual = datetime(r['periodo_anio'], r['periodo_mes'], 1)
            mes_pres = (f_actual + timedelta(days=32)).strftime("%Y-%m")

            series_mensuales.append({
                "label": f"{r['periodo_anio']}-{r['periodo_mes']:02d}",
                "iva_v19": float(r['v19'] or 0),
                "iva_v5": float(r['v5'] or 0),
                "iva_c19": float(r['c19'] or 0),
                "iva_c5": float(r['c5'] or 0),
                "iva_ventas": float(r['v_total'] or 0),
                "iva_compras": float(r['c_total'] or 0),
                "reteiva_favor": float(r['rete'] or 0),
                "saldo_iva": float(r['v_total'] or 0) - float(r['c_total'] or 0) - float(r['rete'] or 0),
                "mes_presentacion": mes_pres
            })

        def agrupar(datos, salto):
            agrupados = []
            for i in range(0, len(datos), salto):
                g = datos[i : i + salto]
                if not g: continue
                v_s = sum(x['iva_ventas'] for x in g)
                c_s = sum(x['iva_compras'] for x in g)
                r_s = sum(x['reteiva_favor'] for x in g)
                agrupados.append({
                    "label": " + ".join([x['label'] for x in g]),
                    "iva_ventas": v_s, 
                    "iva_compras": c_s, 
                    "reteiva_favor": r_s,
                    "saldo_iva": v_s - c_s - r_s,
                    "mes_presentacion": g[-1]['mes_presentacion']
                })
            return agrupados

        return jsonify({
            "series": series_mensuales,
            "kpis": {
                "iva_ventas": sum(s['iva_ventas'] for s in series_mensuales),
                "iva_compras": sum(s['iva_compras'] for s in series_mensuales),
                "reteiva_favor": sum(s['reteiva_favor'] for s in series_mensuales),
                "saldo_iva": sum(s['saldo_iva'] for s in series_mensuales)
            },
            "series_agrupadas": {
                "bimensual": agrupar(series_mensuales, 2),
                "cuatrimestral": agrupar(series_mensuales, 4)
            }
        }), 200



    #ENDPOINT PARA EL CALCULO DE REPORTE DE RETENCIONES 
    # ENDPOINT PARA EL CALCULO DE REPORTE DE RETENCIONES
    @app.route("/reportes/retenciones_v1", methods=["GET"])
    @jwt_required()
    def get_retenciones_v1():
        from sqlalchemy import text

        idcliente = get_jwt().get("idcliente")
        desde = request.args.get("desde", "2026-01-01")
        hasta = request.args.get("hasta", "2026-12-31")

        # 1) EVOLUCIÓN MENSUAL
        # - retefuente: todo 2365%
        # - reteica_conceptos: solo ICA operativo del período (236805%)
        # - 236898 NO participa en KPIs ni neteos del período
        sql_evolucion = text("""
            SELECT
                periodo_anio,
                periodo_mes,
                SUM(CASE WHEN cuenta_codigo LIKE '2365%' THEN (credito - debito) ELSE 0 END) AS retefuente,
                SUM(CASE WHEN cuenta_codigo LIKE '236805%' THEN (credito - debito) ELSE 0 END) AS reteica_conceptos
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND (
                    cuenta_codigo LIKE '2365%'
                    OR cuenta_codigo LIKE '236805%'
                )
            GROUP BY periodo_anio, periodo_mes
            ORDER BY periodo_anio, periodo_mes
        """)

        # 2) COMPOSICIÓN DETALLADA POR CUENTA REAL
        sql_composicion = text("""
            SELECT
                cuenta_codigo AS cuenta,
                cuenta_nombre AS concepto_original,
                CASE
                    WHEN cuenta_codigo LIKE '2365%' THEN 'ReteFuente'
                    WHEN cuenta_codigo LIKE '236805%' THEN 'ReteICA'
                    ELSE 'Otros'
                END AS tipo,
                SUM(credito - debito) AS valor
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND (
                    cuenta_codigo LIKE '2365%'
                    OR cuenta_codigo LIKE '236805%'
                )
            GROUP BY cuenta_codigo, cuenta_nombre
            HAVING SUM(credito - debito) <> 0
            ORDER BY ABS(SUM(credito - debito)) DESC, cuenta_codigo
        """)

        # 3) DETALLE MENSUAL SOLO DE ICA OPERATIVO
        sql_ica_detalle_mensual = text("""
            SELECT
                periodo_anio,
                periodo_mes,
                cuenta_codigo,
                cuenta_nombre,
                SUM(credito - debito) AS valor
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND cuenta_codigo LIKE '236805%'
            GROUP BY periodo_anio, periodo_mes, cuenta_codigo, cuenta_nombre
            HAVING SUM(credito - debito) <> 0
            ORDER BY periodo_anio, periodo_mes, cuenta_codigo
        """)

        # 4) DETALLE MENSUAL DE RETEFUENTE
        sql_retefuente_detalle_mensual = text("""
            SELECT
                periodo_anio,
                periodo_mes,
                cuenta_codigo,
                cuenta_nombre,
                SUM(credito - debito) AS valor
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND cuenta_codigo LIKE '2365%'
            GROUP BY periodo_anio, periodo_mes, cuenta_codigo, cuenta_nombre
            HAVING SUM(credito - debito) <> 0
            ORDER BY periodo_anio, periodo_mes, cuenta_codigo
        """)

        # 5) REFERENCIA CONTABLE OPCIONAL DE 236898
        sql_referencia_236898 = text("""
            SELECT
                SUM(credito - debito) AS valor_236898
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND cuenta_codigo = '236898'
        """)

        res_evo = db.session.execute(
            sql_evolucion,
            {"idc": idcliente, "d": desde, "h": hasta}
        ).mappings().all()

        res_comp = db.session.execute(
            sql_composicion,
            {"idc": idcliente, "d": desde, "h": hasta}
        ).mappings().all()

        res_ica_det = db.session.execute(
            sql_ica_detalle_mensual,
            {"idc": idcliente, "d": desde, "h": hasta}
        ).mappings().all()

        res_rf_det = db.session.execute(
            sql_retefuente_detalle_mensual,
            {"idc": idcliente, "d": desde, "h": hasta}
        ).mappings().all()

        ref_236898 = db.session.execute(
            sql_referencia_236898,
            {"idc": idcliente, "d": desde, "h": hasta}
        ).mappings().first()

        def normalizar_concepto(cuenta: str, nombre: str) -> str:
            nombre_l = (nombre or "").strip().lower()

            # --- ICA DETALLADO ---
            if cuenta == "23680505":
                return "ReteICA 9,66"
            if cuenta == "23680501":
                return "ReteICA 11,04"
            if cuenta == "23680507":
                return "ReteICA 8,66"

            # Detectar devoluciones ICA por nombre
            if cuenta.startswith("236805"):
                if "devol" in nombre_l and "11" in nombre_l:
                    return "Devolución ReteICA 11,04"
                if "devol" in nombre_l and ("8,66" in nombre_l or "8.66" in nombre_l or "866" in nombre_l):
                    return "Devolución ReteICA 8,66"
                if "devol" in nombre_l and ("9,66" in nombre_l or "9.66" in nombre_l or "966" in nombre_l):
                    return "Devolución ReteICA 9,66"
                if "devol" in nombre_l:
                    return f"Devolución ICA ({cuenta})"

            # --- RETEFUENTE ---
            if cuenta.startswith("2365"):
                return (nombre or cuenta).strip().title()

            return (nombre or cuenta).strip().title()

        def clasificar_visual(cuenta: str, nombre: str) -> str:
            nombre_l = (nombre or "").strip().lower()

            if cuenta.startswith("236805"):
                if "devol" in nombre_l:
                    return "ReteICA_Devolucion"
                return "ReteICA"

            if cuenta.startswith("2365"):
                return "ReteFuente"

            return "Otros"

        # --- EVOLUCIÓN ---
        evolucion = []
        total_rf = 0.0
        total_ica_conceptos = 0.0

        for r in res_evo:
            rf = float(r["retefuente"] or 0)
            ica_conceptos = float(r["reteica_conceptos"] or 0)

            total_rf += rf
            total_ica_conceptos += ica_conceptos

            evolucion.append({
                "label": f"{r['periodo_anio']}-{r['periodo_mes']:02d}",
                "retefuente": rf,
                "reteica_conceptos": ica_conceptos
            })

        # --- COMPOSICIÓN / TABLA PRINCIPAL ---
        composicion = []
        for c in res_comp:
            cuenta = str(c["cuenta"])
            concepto_original = str(c["concepto_original"] or "")
            valor = float(c["valor"] or 0)

            composicion.append({
                "cuenta": cuenta,
                "concepto": normalizar_concepto(cuenta, concepto_original),
                "tipo": clasificar_visual(cuenta, concepto_original),
                "valor": valor
            })

        # --- DETALLE ICA MENSUAL ---
        ica_detalle_mensual = []
        for r in res_ica_det:
            cuenta = str(r["cuenta_codigo"])
            nombre = str(r["cuenta_nombre"] or "")
            valor = float(r["valor"] or 0)

            ica_detalle_mensual.append({
                "label": f"{r['periodo_anio']}-{r['periodo_mes']:02d}",
                "cuenta": cuenta,
                "concepto": normalizar_concepto(cuenta, nombre),
                "tipo": clasificar_visual(cuenta, nombre),
                "valor": valor
            })

        # --- DETALLE RETEFUENTE MENSUAL ---
        retefuente_detalle_mensual = []
        for r in res_rf_det:
            cuenta = str(r["cuenta_codigo"])
            nombre = str(r["cuenta_nombre"] or "")
            valor = float(r["valor"] or 0)

            retefuente_detalle_mensual.append({
                "label": f"{r['periodo_anio']}-{r['periodo_mes']:02d}",
                "cuenta": cuenta,
                "concepto": normalizar_concepto(cuenta, nombre),
                "tipo": clasificar_visual(cuenta, nombre),
                "valor": valor
            })

        valor_236898 = float((ref_236898 or {}).get("valor_236898") or 0)

        return jsonify({
            "kpis": {
                "total_retefuente": total_rf,
                "total_reteica_conceptos": total_ica_conceptos,
                "total_general": total_rf + total_ica_conceptos
            },
            "evolucion": evolucion,
            "composicion": composicion,
            "ica_detalle_mensual": ica_detalle_mensual,
            "retefuente_detalle_mensual": retefuente_detalle_mensual,
            "referencias_contables": {
                "cuenta_236898": valor_236898,
                "nota": "La cuenta 236898 se informa solo como referencia contable y no participa en los KPIs ni en el total del período."
            }
        }), 200


    #ENDPOINT PARA EL CALCULO DEL ESTADO DE RESULTADOS (P&L)
    # ENDPOINT PARA EL CALCULO DEL ESTADO DE RESULTADOS (P&L)
    @app.route("/reportes/pnl_v1", methods=["GET"])
    @jwt_required()
    def get_pnl_v1():
        idcliente = get_jwt().get("idcliente")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        if not desde or not hasta:
            return jsonify({"error": "Debes enviar desde y hasta"}), 400

        try:
            data = construir_pnl_auxiliares(idcliente, desde, hasta)
            return jsonify(data), 200
        except Exception as e:
            return jsonify({
                "error": "No fue posible calcular el estado de resultados",
                "detalle": str(e)
            }), 500



    @app.route("/dashboard/pnl-v2", methods=["GET"])
    def get_pnl_v2():
        idcliente = request.headers.get("X-ID-CLIENTE")
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        if not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Fechas requeridas"}), 400

        # ================================
        # 1. CARGAR DATA (ajusta a tu modelo)
        # ================================
        df = obtener_auxiliar_contable(idcliente, fecha_inicio, fecha_fin)

        df["fecha"] = pd.to_datetime(df["fecha"])
        df["mes"] = df["fecha"].dt.to_period("M").astype(str)

        # ================================
        # 2. PERIODO ANTERIOR
        # ================================
        fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

        delta = fecha_fin_dt - fecha_inicio_dt

        prev_fin = fecha_inicio_dt - timedelta(days=1)
        prev_inicio = prev_fin - delta

        df_prev = obtener_auxiliar_contable(
            idcliente,
            prev_inicio.strftime("%Y-%m-%d"),
            prev_fin.strftime("%Y-%m-%d")
        )

        # ================================
        # 3. CLASIFICACIÓN CONTABLE
        # ================================
        def clasificar(cuenta):
            c = str(cuenta)
            if c.startswith("41"):
                return "ingreso"
            elif c.startswith(("6", "7")):
                return "costo"
            elif c.startswith(("51", "52")):
                return "gasto_operacional"
            elif c.startswith("42"):
                return "ingreso_no_op"
            elif c.startswith(("53", "54")):
                return "gasto_no_op"
            elif c.startswith(("5160", "5165", "5260", "5265")):
                return "dep_amort"
            return "otro"

        df["tipo"] = df["cuenta"].apply(clasificar)
        df_prev["tipo"] = df_prev["cuenta"].apply(clasificar)

        # ================================
        # 4. CALCULO KPIs
        # ================================
        def calcular(df):
            ingresos = df[df.tipo == "ingreso"]["valor"].sum()
            ingresos_no_op = df[df.tipo == "ingreso_no_op"]["valor"].sum()
            costos = df[df.tipo == "costo"]["valor"].sum()
            gastos_op = df[df.tipo == "gasto_operacional"]["valor"].sum()
            gastos_no_op = df[df.tipo == "gasto_no_op"]["valor"].sum()
            dep = df[df.tipo == "dep_amort"]["valor"].sum()

            utilidad_bruta = ingresos - costos
            utilidad_operativa = utilidad_bruta - gastos_op
            ebitda = utilidad_operativa + dep
            utilidad_neta = utilidad_operativa + ingresos_no_op - gastos_no_op

            return {
                "ingresos": ingresos + ingresos_no_op,
                "utilidad_bruta": utilidad_bruta,
                "utilidad_operativa": utilidad_operativa,
                "ebitda": ebitda,
                "utilidad_neta": utilidad_neta
            }

        actual = calcular(df)
        anterior = calcular(df_prev)

        # ================================
        # 5. VARIACIONES
        # ================================
        def variacion(act, ant):
            diff = act - ant
            pct = (diff / ant * 100) if ant != 0 else 0
            return {"actual": act, "anterior": ant, "diff": diff, "pct": pct}

        kpis = {
            k: variacion(actual[k], anterior[k])
            for k in actual.keys()
        }

        # ================================
        # 6. MATRIZ
        # ================================
        matriz = df.groupby(["cuenta", "nombre_cuenta"])["valor"].sum().reset_index()

        return jsonify({
            "kpis": kpis,
            "matriz": matriz.to_dict(orient="records")
        })



    # ENDPOINT PARA EL CALCULO DEL ANALISIS VARIACION
    # ENDPOINT PARA EL CALCULO DEL ANALISIS VARIACION
    @app.route("/reportes/analisis_variacion_v1", methods=["GET"])
    @jwt_required()
    def get_analisis_variacion_v1():
        from sqlalchemy import text

        idcliente = get_jwt().get("idcliente")
        desde = request.args.get("desde")
        hasta = request.args.get("hasta")

        if not desde or not hasta:
            return jsonify({"error": "Debes enviar desde y hasta"}), 400

        # =========================================================
        # 1) EVOLUCIÓN MENSUAL
        # =========================================================
        sql_evo = text("""
            SELECT
                periodo_anio,
                periodo_mes,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '41%' THEN (credito - debito)
                    ELSE 0
                END) AS ingresos_operacionales,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '42%' THEN (credito - debito)
                    ELSE 0
                END) AS ingresos_no_operacionales,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '6%' OR cuenta_codigo LIKE '7%' THEN (debito - credito)
                    ELSE 0
                END) AS costos_venta,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN (debito - credito)
                    ELSE 0
                END) AS gastos_operacionales,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN (debito - credito)
                    ELSE 0
                END) AS gastos_no_operacionales,

                SUM(CASE
                    WHEN cuenta_codigo LIKE '5160%'
                    OR cuenta_codigo LIKE '5165%'
                    OR cuenta_codigo LIKE '5260%'
                    OR cuenta_codigo LIKE '5265%'
                    THEN (debito - credito)
                    ELSE 0
                END) AS dep_amort
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            GROUP BY periodo_anio, periodo_mes
            ORDER BY periodo_anio, periodo_mes
        """)

        # =========================================================
        # 2) COMPOSICIÓN DETALLADA
        # =========================================================
        sql_comp = text("""
            SELECT
                periodo_anio,
                periodo_mes,
                cuenta_codigo,
                LEFT(cuenta_codigo, 4) AS cuenta_padre,
                MAX(cuenta_nombre) AS nombre_cuenta,

                CASE
                    WHEN cuenta_codigo LIKE '41%' THEN 'INGRESOS_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '42%' THEN 'INGRESOS_NO_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '6%'  OR cuenta_codigo LIKE '7%' THEN 'COSTOS_VENTA'
                    WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN 'GASTOS_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN 'GASTOS_NO_OPERACIONALES'
                    ELSE 'OTROS'
                END AS seccion,

                CASE
                    WHEN LEFT(cuenta_codigo, 1) = '4' THEN 'CREDITO_MENOS_DEBITO'
                    ELSE 'DEBITO_MENOS_CREDITO'
                END AS naturaleza,

                SUM(
                    CASE
                        WHEN LEFT(cuenta_codigo, 1) = '4' THEN (credito - debito)
                        ELSE (debito - credito)
                    END
                ) AS saldo
            FROM auxiliar_contable
            WHERE idcliente = :idc
            AND fecha_contable BETWEEN :d AND :h
            AND LEFT(cuenta_codigo, 1) IN ('4', '5', '6', '7')
            GROUP BY
                periodo_anio,
                periodo_mes,
                cuenta_codigo,
                LEFT(cuenta_codigo, 4),
                CASE
                    WHEN cuenta_codigo LIKE '41%' THEN 'INGRESOS_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '42%' THEN 'INGRESOS_NO_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '6%'  OR cuenta_codigo LIKE '7%' THEN 'COSTOS_VENTA'
                    WHEN cuenta_codigo LIKE '51%' OR cuenta_codigo LIKE '52%' THEN 'GASTOS_OPERACIONALES'
                    WHEN cuenta_codigo LIKE '53%' OR cuenta_codigo LIKE '54%' THEN 'GASTOS_NO_OPERACIONALES'
                    ELSE 'OTROS'
                END,
                CASE
                    WHEN LEFT(cuenta_codigo, 1) = '4' THEN 'CREDITO_MENOS_DEBITO'
                    ELSE 'DEBITO_MENOS_CREDITO'
                END
            HAVING SUM(
                CASE
                    WHEN LEFT(cuenta_codigo, 1) = '4' THEN (credito - debito)
                    ELSE (debito - credito)
                END
            ) <> 0
            ORDER BY periodo_anio, periodo_mes, cuenta_codigo
        """)

        res_evo = db.session.execute(sql_evo, {
            "idc": idcliente,
            "d": desde,
            "h": hasta
        }).mappings().all()

        res_comp = db.session.execute(sql_comp, {
            "idc": idcliente,
            "d": desde,
            "h": hasta
        }).mappings().all()

        evolucion = []
        totales = {
            "ingresos_operacionales": 0.0,
            "ingresos_no_operacionales": 0.0,
            "costos_venta": 0.0,
            "gastos_operacionales": 0.0,
            "gastos_no_operacionales": 0.0,
            "dep_amort": 0.0,
        }

        for r in res_evo:
            ing_op = float(r["ingresos_operacionales"] or 0)
            ing_no_op = float(r["ingresos_no_operacionales"] or 0)
            costos = float(r["costos_venta"] or 0)
            gastos_op = float(r["gastos_operacionales"] or 0)
            gastos_no_op = float(r["gastos_no_operacionales"] or 0)
            dep_amort = float(r["dep_amort"] or 0)

            ingresos_totales = ing_op + ing_no_op
            utilidad_bruta = ing_op - costos
            utilidad_operativa = utilidad_bruta - gastos_op
            ebitda = utilidad_operativa + dep_amort
            utilidad_antes_impuestos = utilidad_operativa + ing_no_op - gastos_no_op
            utilidad_neta = utilidad_antes_impuestos

            totales["ingresos_operacionales"] += ing_op
            totales["ingresos_no_operacionales"] += ing_no_op
            totales["costos_venta"] += costos
            totales["gastos_operacionales"] += gastos_op
            totales["gastos_no_operacionales"] += gastos_no_op
            totales["dep_amort"] += dep_amort

            base_margen = ingresos_totales if ingresos_totales != 0 else 0

            evolucion.append({
                "label": f"{r['periodo_anio']}-{int(r['periodo_mes']):02d}",
                "ingresos_operacionales": ing_op,
                "ingresos_no_operacionales": ing_no_op,
                "ingresos_totales": ingresos_totales,
                "costos_venta": costos,
                "gastos_operacionales": gastos_op,
                "gastos_no_operacionales": gastos_no_op,
                "utilidad_bruta": utilidad_bruta,
                "utilidad_operativa": utilidad_operativa,
                "ebitda": ebitda,
                "utilidad_antes_impuestos": utilidad_antes_impuestos,
                "utilidad_neta": utilidad_neta,
                "costos_gastos": costos + gastos_op + gastos_no_op,
                "margen_bruto": round((utilidad_bruta / base_margen) * 100, 2) if base_margen else 0,
                "margen_operativo": round((utilidad_operativa / base_margen) * 100, 2) if base_margen else 0,
                "margen_ebitda": round((ebitda / base_margen) * 100, 2) if base_margen else 0,
                "margen_neto": round((utilidad_neta / base_margen) * 100, 2) if base_margen else 0,
            })

        ingresos_operacionales = totales["ingresos_operacionales"]
        ingresos_no_operacionales = totales["ingresos_no_operacionales"]
        ingresos_totales = ingresos_operacionales + ingresos_no_operacionales
        costos_venta = totales["costos_venta"]
        gastos_operacionales = totales["gastos_operacionales"]
        gastos_no_operacionales = totales["gastos_no_operacionales"]
        dep_amort = totales["dep_amort"]

        utilidad_bruta = ingresos_operacionales - costos_venta
        utilidad_operativa = utilidad_bruta - gastos_operacionales
        ebitda = utilidad_operativa + dep_amort
        utilidad_antes_impuestos = utilidad_operativa + ingresos_no_operacionales - gastos_no_operacionales
        utilidad_neta = utilidad_antes_impuestos

        base_margen = ingresos_totales if ingresos_totales != 0 else 0

        cuentas_dict = {}

        for c in res_comp:
            cuenta_codigo = str(c["cuenta_codigo"])
            periodo = f"{c['periodo_anio']}-{int(c['periodo_mes']):02d}"

            if cuenta_codigo not in cuentas_dict:
                cuentas_dict[cuenta_codigo] = {
                    "cuenta": cuenta_codigo,
                    "cuenta_padre": str(c["cuenta_padre"]),
                    "nombre": str(c["nombre_cuenta"]).strip().title(),
                    "seccion": str(c["seccion"]),
                    "naturaleza": str(c["naturaleza"]),
                    "valores_mes": {},
                    "total": 0.0,
                }

            val = float(c["saldo"] or 0)
            cuentas_dict[cuenta_codigo]["valores_mes"][periodo] = val
            cuentas_dict[cuenta_codigo]["total"] += val

        composicion = list(cuentas_dict.values())
        composicion.sort(key=lambda x: x["cuenta"])

        return jsonify({
            "kpis": {
                "ingresos_operacionales": ingresos_operacionales,
                "ingresos_no_operacionales": ingresos_no_operacionales,
                "ingresos_totales": ingresos_totales,
                "costos_venta": costos_venta,
                "utilidad_bruta": utilidad_bruta,
                "gastos_operacionales": gastos_operacionales,
                "utilidad_operativa": utilidad_operativa,
                "ebitda": ebitda,
                "gastos_no_operacionales": gastos_no_operacionales,
                "utilidad_antes_impuestos": utilidad_antes_impuestos,
                "utilidad_neta": utilidad_neta,
                "margen_bruto": round((utilidad_bruta / base_margen) * 100, 2) if base_margen else 0,
                "margen_operativo": round((utilidad_operativa / base_margen) * 100, 2) if base_margen else 0,
                "margen_ebitda": round((ebitda / base_margen) * 100, 2) if base_margen else 0,
                "margen_neto": round((utilidad_neta / base_margen) * 100, 2) if base_margen else 0,
            },
            "evolucion": evolucion,
            "composicion": composicion
        }), 200




    # GENERAR BALANCE GENERAL A PARTIR DE LOS AUXILIARES DE CUENTAS CONTABLES

    @app.route("/reportes/balance_general/rebuild_snapshot", methods=["POST"])
    @jwt_required()
    def rebuild_balance_snapshot():
        idcliente = get_jwt().get("idcliente")
        data = request.get_json(silent=True) or {}

        fecha_corte = data.get("fecha_corte")
        comparar_con = data.get("comparar_con")

        if not fecha_corte:
            return jsonify({"error": "Debes enviar fecha_corte"}), 400

        try:
            result = regenerar_snapshots_balance(idcliente, fecha_corte, comparar_con)
            return jsonify(result), 200
        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No fue posible regenerar los snapshots del balance",
                "detalle": str(e)
            }), 500


    @app.route("/reportes/balance_general_v1", methods=["GET"])
    @jwt_required()
    def get_balance_general_v1():
        idcliente = get_jwt().get("idcliente")
        fecha_corte = request.args.get("fecha_corte")
        comparar_con = request.args.get("comparar_con")

        if not fecha_corte:
            return jsonify({"error": "Debes enviar fecha_corte"}), 400

        try:
            result = construir_balance_general(idcliente, fecha_corte, comparar_con)

            if not result.get("ok"):
                return jsonify({
                    "error": result.get("error", "No fue posible construir el balance")
                }), 404

            return jsonify(result), 200

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No fue posible consultar el balance general",
                "detalle": str(e)
            }), 500


    # ENDPOINT: Buscador inteligente de facturas por contenido
    @app.route("/reportes/facturas-buscador", methods=["GET"])
    @jwt_required()
    def get_facturas_buscador():
        from sqlalchemy import text
        from decimal import Decimal
        from datetime import datetime, timedelta
        import traceback

        try:
            idcliente = get_jwt().get("idcliente")

            q = (request.args.get("q") or "").strip().lower()
            idfactura = (request.args.get("idfactura") or "").strip()
            cliente = (request.args.get("cliente") or "").strip()
            estado_pago = (request.args.get("estado_pago") or "").strip()
            estado = (request.args.get("estado") or "").strip()

            desde = request.args.get("desde") or "2025-01-01"
            hasta = request.args.get("hasta") or datetime.now().strftime("%Y-%m-%d")

            # para incluir completo el día final
            hasta_dt = datetime.strptime(hasta, "%Y-%m-%d") + timedelta(days=1)
            hasta_sql = hasta_dt.strftime("%Y-%m-%d")

            # alias tolerantes
            aliases = {
                "zapier": ["zapier", "zappier"],
                "zappier": ["zapier", "zappier"],
            }

            terminos = aliases.get(q, [q] if q else [])

            params = {
                "idcliente": idcliente,
                "desde": desde,
                "hasta": hasta_sql,
            }

            where_main = [
                "fe.idcliente = :idcliente",
                "fe.fecha >= :desde",
                "fe.fecha < :hasta",
            ]

            if idfactura:
                where_main.append("fe.idfactura ILIKE :idfactura")
                params["idfactura"] = f"%{idfactura}%"

            if cliente:
                where_main.append("fe.cliente_nombre ILIKE :cliente")
                params["cliente"] = f"%{cliente}%"

            if estado_pago:
                where_main.append("fe.estado_pago = :estado_pago")
                params["estado_pago"] = estado_pago

            if estado:
                where_main.append("fe.estado = :estado")
                params["estado"] = estado

            # filtro de descripción
            exists_filter = ""
            descripcion_filter = ""

            if terminos:
                term_clauses_exists = []
                term_clauses_desc = []

                for i, term in enumerate(terminos):
                    key = f"term_{i}"
                    params[key] = f"%{term}%"
                    term_clauses_exists.append(f"LOWER(fi.descripcion) LIKE LOWER(:{key})")
                    term_clauses_desc.append(f"LOWER(fi2.descripcion) LIKE LOWER(:{key})")

                exists_filter = f"""
                AND EXISTS (
                    SELECT 1
                    FROM siigo_factura_items fi
                    WHERE fi.factura_id = fe.factura_id
                        AND fi.idcliente = fe.idcliente
                        AND ({' OR '.join(term_clauses_exists)})
                )
                """

                descripcion_filter = f"""
                    AND ({' OR '.join(term_clauses_desc)})
                """

            sql = text(f"""
                SELECT
                    fe.factura_id,
                    fe.idfactura,
                    fe.fecha,
                    fe.vencimiento,
                    fe.cliente_nombre,
                    fe.estado,
                    fe.estado_pago,
                    fe.subtotal,
                    fe.impuestos_total,
                    COALESCE((
                        SELECT SUM((r->>'value')::numeric)
                        FROM jsonb_array_elements(fe.retenciones) AS r
                        WHERE r->>'type' = 'ReteICA'
                    ), 0) AS reteica,
                    COALESCE((
                        SELECT SUM((r->>'value')::numeric)
                        FROM jsonb_array_elements(fe.retenciones) AS r
                        WHERE r->>'type' = 'ReteIVA'
                    ), 0) AS reteiva,
                    COALESCE((
                        SELECT SUM((r->>'value')::numeric)
                        FROM jsonb_array_elements(fe.retenciones) AS r
                        WHERE r->>'type' = 'Autorretencion'
                    ), 0) AS autorretencion,
                    COALESCE((
                        SELECT SUM((r->>'value')::numeric)
                        FROM jsonb_array_elements(fe.retenciones) AS r
                    ), 0) AS total_retenciones,
                    fe.total,
                    fe.saldo,
                    fe.public_url,
                    fe.centro_costo_nombre,
                    fe.centro_costo_codigo,
                    fe.vendedor_nombre,
                    (
                        SELECT STRING_AGG(TRIM(fi2.descripcion), ' || ' ORDER BY TRIM(fi2.descripcion))
                        FROM siigo_factura_items fi2
                        WHERE fi2.factura_id = fe.factura_id
                        AND fi2.idcliente = fe.idcliente
                        {descripcion_filter}
                    ) AS descripcion
                FROM facturas_enriquecidas fe
                WHERE {' AND '.join(where_main)}
                {exists_filter}
                ORDER BY fe.fecha DESC, fe.idfactura DESC
            """)

            rows = db.session.execute(sql, params).mappings().all()

            def norm(v):
                if isinstance(v, Decimal):
                    return float(v)
                if hasattr(v, "isoformat"):
                    return v.isoformat()
                return v

            data = [{k: norm(v) for k, v in row.items()} for row in rows]

            summary = {
                "total_registros": len(data),
                "subtotal_total": float(sum((r["subtotal"] or 0) for r in data)),
                "iva_total": float(sum((r["impuestos_total"] or 0) for r in data)),
                "reteica_total": float(sum((r["reteica"] or 0) for r in data)),
                "reteiva_total": float(sum((r["reteiva"] or 0) for r in data)),
                "autorretencion_total": float(sum((r["autorretencion"] or 0) for r in data)),
                "retenciones_total": float(sum((r["total_retenciones"] or 0) for r in data)),
                "total_facturado": float(sum((r["total"] or 0) for r in data)),
                "saldo_total": float(sum((r["saldo"] or 0) for r in data)),
            }

            return jsonify({
                "ok": True,
                "filters": {
                    "q": q,
                    "idfactura": idfactura,
                    "cliente": cliente,
                    "estado_pago": estado_pago,
                    "estado": estado,
                    "desde": desde,
                    "hasta": hasta,
                    "terminos_aplicados": terminos,
                },
                "summary": summary,
                "rows": data,
            })

        except Exception as e:
            traceback.print_exc()
            return jsonify({
                "ok": False,
                "error": str(e),
                "trace": traceback.format_exc()
            }), 500



    # ----------------------------------------------------------
    # Indicadores financieros construidos desde AuxiliarContable
    @app.route("/reportes/auxiliares/indicadores-financieros", methods=["GET"])
    @jwt_required()
    def indicadores_financieros_auxiliares():
        from datetime import date
        from calendar import monthrange

        idcliente = get_jwt().get("idcliente")
        anio = request.args.get("anio", type=int)
        mes_inicio = request.args.get("mes_inicio", type=int)
        mes_fin = request.args.get("mes_fin", type=int)

        if not idcliente or not anio or not mes_inicio or not mes_fin:
            return jsonify({"error": "Faltan parámetros"}), 400

        if mes_inicio < 1 or mes_inicio > 12 or mes_fin < 1 or mes_fin > 12 or mes_inicio > mes_fin:
            return jsonify({"error": "Rango de meses inválido"}), 400

        fecha_desde = date(anio, mes_inicio, 1)
        fecha_hasta = date(anio, mes_fin, monthrange(anio, mes_fin)[1])

        try:
            # --------------------------------------------------
            # 1) Snapshot del balance al corte final del período
            # --------------------------------------------------
            regenerar_snapshot_saldos_corte(idcliente, str(fecha_hasta))

            # --------------------------------------------------
            # 2) Balance general al corte final
            #    Ojo: esto es acumulado hasta la fecha de corte.
            # --------------------------------------------------
            balance = construir_balance_general(idcliente, str(fecha_hasta))
            if not balance.get("ok"):
                return jsonify({
                    "error": balance.get("error", "No fue posible construir el balance general")
                }), 400

            bk = balance.get("kpis", {})
            meta_balance = balance.get("meta", {})
            resumen_balance = balance.get("resumen", {})

            # --------------------------------------------------
            # 3) P&L del período seleccionado
            #    Esto sí va únicamente entre fecha_desde y fecha_hasta.
            # --------------------------------------------------
            pnl = construir_pnl_auxiliares(idcliente, str(fecha_desde), str(fecha_hasta))
            pk = pnl.get("kpis", {})
            evolucion_pnl = pnl.get("evolucion", [])

            # =========================
            # Variables base balance
            # =========================
            activo_corriente = float(bk.get("activo_corriente", 0) or 0)
            activo_no_corriente = float(bk.get("activo_no_corriente", 0) or 0)
            activo_total = float(bk.get("activos_totales", 0) or 0)

            pasivo_corto = float(bk.get("pasivo_corriente", 0) or 0)
            pasivo_largo = float(bk.get("pasivo_no_corriente", 0) or 0)
            pasivo_total = float(bk.get("pasivos_totales", 0) or 0)

            patrimonio = float(bk.get("patrimonio_total", 0) or 0)
            capital_trabajo = float(bk.get("capital_trabajo", 0) or 0)

            # =========================
            # Variables base P&L
            # =========================
            ingresos = float(pk.get("ingresos_totales", 0) or 0)
            costos = float(pk.get("costos_venta", 0) or 0)
            gastos = float(
                (pk.get("gastos_operacionales", 0) or 0) +
                (pk.get("gastos_no_operacionales", 0) or 0)
            )
            utilidad_neta = float(pk.get("utilidad_neta", 0) or 0)

            # =========================
            # Indicadores
            # =========================
            liquidez = round(activo_corriente / pasivo_corto, 2) if pasivo_corto else None
            apalancamiento = round(pasivo_total / activo_total, 2) if activo_total else None
            rentabilidad = round(utilidad_neta / ingresos, 2) if ingresos else None
            solvencia = round(activo_total / pasivo_total, 2) if pasivo_total else None
            autonomia = round(patrimonio / activo_total, 2) if activo_total else None
            porcentaje_pasivo_corto = round(pasivo_corto / pasivo_total, 2) if pasivo_total else None
            porcentaje_activo_no_corriente = round(activo_no_corriente / activo_total, 2) if activo_total else None
            cobertura_activo_pasivo = round(activo_total / pasivo_total, 2) if pasivo_total else None
            endeudamiento_largo_plazo = round(pasivo_largo / patrimonio, 2) if patrimonio and patrimonio > 0 else None

            indicadores = {
                # ratios
                "liquidez": liquidez,
                "apalancamiento": apalancamiento,
                "rentabilidad": rentabilidad,
                "capital_trabajo": round(capital_trabajo, 2),
                "solvencia": solvencia,
                "autonomia": autonomia,
                "porcentaje_pasivo_corto": porcentaje_pasivo_corto,
                "porcentaje_activo_no_corriente": porcentaje_activo_no_corriente,
                "cobertura_activo_pasivo": cobertura_activo_pasivo,
                "endeudamiento_largo_plazo": endeudamiento_largo_plazo,

                # cifras base
                "activo_total": round(activo_total, 2),
                "pasivo_total": round(pasivo_total, 2),
                "patrimonio": round(patrimonio, 2),
                "ingresos": round(ingresos, 2),
                "costos": round(costos, 2),
                "gastos": round(gastos, 2),
                "utilidad_neta": round(utilidad_neta, 2),

                # componentes adicionales útiles
                "activo_corriente": round(activo_corriente, 2),
                "activo_no_corriente": round(activo_no_corriente, 2),
                "pasivo_corto": round(pasivo_corto, 2),
                "pasivo_largo": round(pasivo_largo, 2),
            }

            # =========================
            # Interpretaciones
            # =========================
            interpretaciones = {
                k: interpretar_indicador(k, v) for k, v in indicadores.items()
            }

            explicaciones = {
                "liquidez": "Activo corriente / Pasivo corriente. Mide capacidad para cubrir obligaciones de corto plazo.",
                "apalancamiento": "Pasivo total / Activo total. Mide qué proporción de los activos está financiada con deuda.",
                "rentabilidad": "Utilidad neta / Ingresos. Mide el margen neto del período.",
                "capital_trabajo": "Activo corriente - Pasivo corriente. Mide el colchón operativo de corto plazo.",
                "solvencia": "Activo total / Pasivo total. Evalúa cobertura general de deudas con activos.",
                "autonomia": "Patrimonio / Activo total. Mide qué tanto de la empresa está financiado con recursos propios.",
                "porcentaje_pasivo_corto": "Pasivo corriente / Pasivo total. Mide concentración de deuda en el corto plazo.",
                "porcentaje_activo_no_corriente": "Activo no corriente / Activo total. Mide proporción de activos menos líquidos.",
                "cobertura_activo_pasivo": "Activo total / Pasivo total. Cobertura global de pasivos.",
                "endeudamiento_largo_plazo": "Pasivo no corriente / Patrimonio. Mide presión financiera de largo plazo frente al capital propio.",
            }

            # =========================
            # Resumen técnico
            # =========================
            resumen_financiero = [
                {
                    "clase": "Activo corriente",
                    "valor": round(activo_corriente, 2),
                    "interpretacion": "Recursos líquidos o realizables en el corto plazo."
                },
                {
                    "clase": "Activo no corriente",
                    "valor": round(activo_no_corriente, 2),
                    "interpretacion": "Activos de permanencia o recuperación a largo plazo."
                },
                {
                    "clase": "Activo total",
                    "valor": round(activo_total, 2),
                    "interpretacion": "Total de recursos controlados por la empresa al corte final."
                },
                {
                    "clase": "Pasivo corto plazo",
                    "valor": round(pasivo_corto, 2),
                    "interpretacion": "Obligaciones exigibles en el corto plazo."
                },
                {
                    "clase": "Pasivo largo plazo",
                    "valor": round(pasivo_largo, 2),
                    "interpretacion": "Obligaciones a largo plazo."
                },
                {
                    "clase": "Pasivo total",
                    "valor": round(pasivo_total, 2),
                    "interpretacion": "Total de obligaciones con terceros al corte final."
                },
                {
                    "clase": "Patrimonio",
                    "valor": round(patrimonio, 2),
                    "interpretacion": "Recursos propios o acumulados de la empresa al corte final."
                },
                {
                    "clase": "Ingresos",
                    "valor": round(ingresos, 2),
                    "interpretacion": "Ingresos del período seleccionado."
                },
                {
                    "clase": "Costos",
                    "valor": round(costos, 2),
                    "interpretacion": "Costos del período seleccionado."
                },
                {
                    "clase": "Gastos",
                    "valor": round(gastos, 2),
                    "interpretacion": "Gastos del período seleccionado."
                },
                {
                    "clase": "Utilidad neta",
                    "valor": round(utilidad_neta, 2),
                    "interpretacion": "Resultado neto del período analizado."
                },
            ]

            # =========================
            # Evolución mensual del P&L
            # =========================
            evolucion_mensual = []
            for item in evolucion_pnl:
                ingresos_mes = float(item.get("ingresos_totales", 0) or 0)
                utilidad_neta_mes = float(item.get("utilidad_neta", 0) or 0)

                evolucion_mensual.append({
                    "mes": item.get("label"),
                    "utilidad_neta": round(utilidad_neta_mes, 2),
                    "rentabilidad": round(utilidad_neta_mes / ingresos_mes, 2) if ingresos_mes else None,
                })

            # =========================
            # Diagnóstico ejecutivo corto
            # =========================
            conclusiones = []

            if liquidez is not None:
                if liquidez < 1:
                    conclusiones.append("⚠ Riesgo de iliquidez: el activo corriente no cubre el pasivo de corto plazo.")
                elif liquidez > 3:
                    conclusiones.append("⚠ Exceso de liquidez: podría existir capital ocioso o baja eficiencia en el uso de recursos.")
                else:
                    conclusiones.append("✅ La liquidez luce saludable para atender compromisos de corto plazo.")

            if apalancamiento is not None:
                if apalancamiento > 0.8:
                    conclusiones.append("⚠ El apalancamiento es alto: una parte importante de los activos está financiada con deuda.")
                elif apalancamiento > 0.6:
                    conclusiones.append("• El apalancamiento es moderado y conviene seguir monitoreándolo.")
                else:
                    conclusiones.append("✅ La estructura de endeudamiento luce controlada.")

            if rentabilidad is not None:
                if rentabilidad < 0:
                    conclusiones.append("🔻 La empresa presenta pérdida neta en el período analizado.")
                elif rentabilidad < 0.1:
                    conclusiones.append("• La empresa genera utilidad, pero con margen neto bajo.")
                else:
                    conclusiones.append("✅ La rentabilidad neta del período es positiva y saludable.")

            if autonomia is not None:
                if autonomia < 0.3:
                    conclusiones.append("⚠ La autonomía financiera es baja y existe fuerte dependencia de terceros.")
                elif autonomia < 0.5:
                    conclusiones.append("• La autonomía financiera es moderada.")
                else:
                    conclusiones.append("✅ La autonomía financiera es sólida.")

            if patrimonio <= 0:
                conclusiones.append("❗ El patrimonio es nulo o negativo, situación que requiere revisión prioritaria.")

            return jsonify({
                "resumen_financiero": resumen_financiero,
                "indicadores": indicadores,
                "explicaciones": explicaciones,
                "interpretaciones": interpretaciones,
                "conclusiones": conclusiones,
                "evolucion_mensual": evolucion_mensual,

                # NUEVO: estos dos son clave para traer la interpretación ejecutiva real del balance
                "meta_balance": meta_balance,
                "resumen_balance": resumen_balance,

                "meta": {
                    "fuente": "auxiliar_contable",
                    "anio": anio,
                    "mes_inicio": mes_inicio,
                    "mes_fin": mes_fin,
                    "fecha_desde": str(fecha_desde),
                    "fecha_hasta": str(fecha_hasta),
                    "fecha_corte_balance": str(fecha_hasta),
                    "logica_balance": "acumulado_al_corte",
                    "logica_pnl": "movimientos_del_periodo",
                    "nota": "Los indicadores de balance se calculan con saldos acumulados al corte final; la utilidad y la rentabilidad se calculan sobre el período seleccionado."
                }
            }), 200

        except Exception as e:
            current_app.logger.exception("Error en indicadores_financieros_auxiliares")
            db.session.rollback()
            return jsonify({
                "error": "No fue posible calcular indicadores financieros desde auxiliares",
                "detalle": str(e)
            }), 500


    # ----------------------------------------------------------
    # ENDPOINT: BÚSQUEDA INTELIGENTE DE FACTURAS
    # ----------------------------------------------------------
    @app.route("/reportes/busqueda-inteligente-facturas", methods=["GET"])
    @jwt_required()
    def busqueda_inteligente_facturas():
        try:
            claims = get_jwt()
            perfilid = claims.get("perfilid")
            idcliente = claims.get("idcliente")

            data = _obtener_busqueda_inteligente_facturas_data(
                idcliente=idcliente,
                perfilid=perfilid
            )

            return jsonify({
                "ok": True,
                "rows": data["rows"],
                "count": data["count"],
                "kpis": data["kpis"],
                "series": data["series"],
                "filters": data["filters"],
            })

        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 403
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500


    # ----------------------------------------------------------
    # ENDPOINT: EXPORTAR BÚSQUEDA INTELIGENTE A EXCEL
    # ----------------------------------------------------------
    @app.route("/reportes/busqueda-inteligente-facturas/export.xlsx", methods=["GET"])
    @jwt_required()
    def export_busqueda_inteligente_facturas_excel():
        from io import BytesIO
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        try:
            claims = get_jwt()
            perfilid = claims.get("perfilid")
            idcliente = claims.get("idcliente")

            data = _obtener_busqueda_inteligente_facturas_data(
                idcliente=idcliente,
                perfilid=perfilid
            )

            rows = data["rows"]
            kpis = data["kpis"]
            filtros = data["filters"]

            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            headers = [
                "Fecha",
                "Factura",
                "Cliente",
                "Centro de costo",
                "Código centro costo",
                "Estado factura",
                "Estado pago",
                "Descripción",
                "Observaciones",
                "Subtotal",
                "IVA",
                "ReteICA",
                "ReteIVA",
                "Autorretención",
                "Total retenciones",
                "Total",
                "Saldo",
                "URL",
            ]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1E3A8A")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r in rows:
                ws.append([
                    r.get("fecha"),
                    r.get("idfactura"),
                    r.get("cliente_nombre"),
                    r.get("centro_costo_nombre"),
                    r.get("centro_costo_codigo"),
                    r.get("estado"),
                    r.get("estado_pago_real") or r.get("estado_pago"),
                    r.get("descripcion"),
                    r.get("observaciones"),
                    float(r.get("subtotal") or 0),
                    float(r.get("impuestos") or 0),
                    float(r.get("reteica") or 0),
                    float(r.get("reteiva") or 0),
                    float(r.get("autorretencion") or 0),
                    float(r.get("total_retenciones") or 0),
                    float(r.get("total") or 0),
                    float(r.get("saldo") or 0),
                    r.get("public_url"),
                ])

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for col in ["J", "K", "L", "M", "N", "O", "P", "Q"]:
                for cell in ws[col]:
                    if cell.row > 1:
                        cell.number_format = '$ #,##0.00'

            widths = {
                1: 14,
                2: 18,
                3: 35,
                4: 28,
                5: 18,
                6: 18,
                7: 18,
                8: 60,
                9: 45,
                10: 16,
                11: 16,
                12: 16,
                13: 16,
                14: 18,
                15: 18,
                16: 16,
                17: 16,
                18: 28,
            }
            for i, width in widths.items():
                ws.column_dimensions[get_column_letter(i)].width = width

            rs = wb.create_sheet("Resumen")
            rs["A1"] = "Resumen búsqueda inteligente de facturas"
            rs["A1"].font = Font(bold=True, size=14)

            rs.append([])
            rs.append(["Facturas encontradas", int(data["count"])])
            rs.append(["Subtotal", float(kpis.get("subtotal") or 0)])
            rs.append(["IVA", float(kpis.get("iva") or 0)])
            rs.append(["ReteICA", float(kpis.get("reteica_total") or 0)])
            rs.append(["ReteIVA", float(kpis.get("reteiva_total") or 0)])
            rs.append(["Autorretención", float(kpis.get("autorretencion_total") or 0)])
            rs.append(["Retenciones", float(kpis.get("retenciones") or 0)])
            rs.append(["Total facturado", float(kpis.get("total_facturado") or 0)])
            rs.append(["Saldo", float(kpis.get("saldo") or 0)])

            rs.append([])
            rs.append(["Filtros aplicados", ""])
            rs.append(["Palabra clave", filtros.get("q") or ""])
            rs.append(["Factura", filtros.get("factura") or ""])
            rs.append(["Cliente", filtros.get("cliente") or ""])
            rs.append(["Centro de costo", filtros.get("cost_center") or ""])
            rs.append(["Estado pago", filtros.get("estado_pago") or ""])
            rs.append(["Estado factura", filtros.get("estado_factura") or ""])
            rs.append(["Desde", filtros.get("desde") or ""])
            rs.append(["Hasta", filtros.get("hasta") or ""])

            for row in range(3, 11):
                rs[f"B{row}"].number_format = '$ #,##0.00'

            rs.column_dimensions["A"].width = 28
            rs.column_dimensions["B"].width = 28

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return send_file(
                output,
                as_attachment=True,
                download_name="busqueda_inteligente_facturas.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except ValueError as ve:
            return jsonify({"ok": False, "error": str(ve)}), 403
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500


    # =========================================================
    # DASHBOARD / RESUMEN EJECUTIVO INTELIGENTE
    # =========================================================




    # =========================================================
    # ENDPOINT METADATA DASHBOARD
    # =========================================================
    @app.route("/dashboard/resumen-ejecutivo/metadata", methods=["GET"])
    @jwt_required()
    def dashboard_resumen_ejecutivo_metadata():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        try:
            corte = _resolver_corte_confiable_auxiliar(idcliente)

            if not corte["ultima_fecha_auxiliar"]:
                return jsonify({
                    "ultima_fecha_auxiliar": None,
                    "fecha_corte_confiable": None,
                    "desde_sugerido": None,
                    "hasta_sugerido": None,
                    "mes_actual_parcial": False,
                    "modo_periodo": "sin_datos",
                    "mensaje_contexto": "No hay información disponible en auxiliar contable."
                }), 200

            mensaje_contexto = None
            if corte["mes_actual_parcial"]:
                mensaje_contexto = (
                    "Se detectó que el mes más reciente en auxiliares está parcial. "
                    "El dashboard usará por defecto el año corrido cerrado hasta el último mes completo."
                )
            else:
                mensaje_contexto = (
                    "El dashboard usará por defecto el año corrido cerrado hasta el último corte mensual disponible."
                )

            return jsonify({
                "ultima_fecha_auxiliar": corte["ultima_fecha_auxiliar"].strftime("%Y-%m-%d"),
                "fecha_corte_confiable": corte["fecha_corte_confiable"].strftime("%Y-%m-%d"),
                "desde_sugerido": corte["desde_ytd"].strftime("%Y-%m-%d"),
                "hasta_sugerido": corte["hasta_ytd"].strftime("%Y-%m-%d"),
                "mes_actual_parcial": corte["mes_actual_parcial"],
                "modo_periodo": corte["modo_periodo"],
                "mensaje_contexto": mensaje_contexto
            }), 200

        except Exception as e:
            return jsonify({
                "error": "No fue posible obtener metadata del dashboard",
                "detalle": str(e)
            }), 500

    # =========================================================
    # ENDPOINT RESUMEN EJECUTIVO
    # =========================================================
    @app.route("/dashboard/resumen-ejecutivo", methods=["GET"])
    @jwt_required()
    def dashboard_resumen_ejecutivo():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        desde = request.args.get("desde")
        hasta = request.args.get("hasta")
        centro_costos = request.args.get("centro_costos", type=int)

        try:
            corte = _resolver_corte_confiable_auxiliar(idcliente)
            config = _obtener_config_dashboard(idcliente)

            if not corte["ultima_fecha_auxiliar"]:
                return jsonify({
                    "error": "No hay auxiliar contable cargado para construir el dashboard"
                }), 400

            # Si el frontend no manda fechas, usar por defecto YTD cerrado
            if not desde or not hasta:
                fecha_desde = corte["desde_ytd"]
                fecha_hasta = corte["hasta_ytd"]
                desde = fecha_desde.strftime("%Y-%m-%d")
                hasta = fecha_hasta.strftime("%Y-%m-%d")
                rango_auto = True
            else:
                try:
                    fecha_desde = datetime.strptime(desde, "%Y-%m-%d").date()
                    fecha_hasta = datetime.strptime(hasta, "%Y-%m-%d").date()
                    rango_auto = False
                except Exception:
                    return jsonify({"error": "Formato de fecha inválido. Usa YYYY-MM-DD"}), 400

            # Ajuste automático al corte confiable
            fecha_hasta_ajustada = fecha_hasta
            ajuste_por_corte = False

            if fecha_hasta > corte["fecha_corte_confiable"]:
                fecha_hasta_ajustada = corte["fecha_corte_confiable"]
                ajuste_por_corte = True

                if fecha_desde > fecha_hasta_ajustada:
                    fecha_desde = corte["desde_ytd"]

                desde = fecha_desde.strftime("%Y-%m-%d")
                hasta = fecha_hasta_ajustada.strftime("%Y-%m-%d")
            else:
                hasta = fecha_hasta.strftime("%Y-%m-%d")
                desde = fecha_desde.strftime("%Y-%m-%d")

            # =========================================================
            # 1. P&L ACTUAL
            # =========================================================
            pnl_actual = construir_pnl_auxiliares(idcliente, desde, hasta)
            kpis_actual = pnl_actual.get("kpis", {})
            evolucion_actual = pnl_actual.get("evolucion", [])
            composicion_actual = pnl_actual.get("composicion", [])

            hay_datos_auxiliar_actual = bool(evolucion_actual) or bool(composicion_actual)

            mensaje_contexto = None
            if not hay_datos_auxiliar_actual and corte["ultima_fecha_auxiliar"]:
                mensaje_contexto = (
                    f"No hay información de auxiliar contable para el período seleccionado. "
                    f"Última fecha disponible: {corte['ultima_fecha_auxiliar'].strftime('%Y-%m-%d')}."
                )
            elif ajuste_por_corte:
                mensaje_contexto = (
                    f"El rango solicitado excedía el corte confiable del auxiliar. "
                    f"Se ajustó automáticamente hasta {corte['fecha_corte_confiable'].strftime('%Y-%m-%d')}."
                )
            elif corte["mes_actual_parcial"]:
                mensaje_contexto = (
                    f"El mes más reciente cargado en auxiliares está parcial. "
                    f"El análisis ejecutivo usa corte confiable hasta {corte['fecha_corte_confiable'].strftime('%Y-%m-%d')}."
                )

            # =========================================================
            # 2. PERIODO ANTERIOR
            # =========================================================
            delta = fecha_hasta_ajustada - fecha_desde
            prev_hasta = fecha_desde - timedelta(days=1)
            prev_desde = prev_hasta - delta

            pnl_anterior = construir_pnl_auxiliares(
                idcliente,
                prev_desde.strftime("%Y-%m-%d"),
                prev_hasta.strftime("%Y-%m-%d"),
            )
            kpis_anterior = pnl_anterior.get("kpis", {})

            # =========================================================
            # 3. TENDENCIA ÚLTIMOS N MESES
            # =========================================================
            meses_grafica = int(config.get("meses_grafica") or 6) if config else 6
            meses_grafica = max(3, min(meses_grafica, 24))

            hasta_nm = fecha_hasta_ajustada
            desde_nm = _first_day_of_month(_shift_months(fecha_hasta_ajustada, -(meses_grafica - 1)))

            pnl_nm = construir_pnl_auxiliares(
                idcliente,
                desde_nm.strftime("%Y-%m-%d"),
                hasta_nm.strftime("%Y-%m-%d"),
            )
            evolucion_nm = pnl_nm.get("evolucion", [])

            # =========================================================
            # 4. EFICIENCIA OPERATIVA
            # =========================================================
            ventas_actual = _safe_float(kpis_actual.get("ingresos_operacionales", 0))
            ventas_anterior = _safe_float(kpis_anterior.get("ingresos_operacionales", 0))

            ebitda_actual = _safe_float(kpis_actual.get("ebitda", 0))
            ebitda_anterior = _safe_float(kpis_anterior.get("ebitda", 0))

            eficiencia_actual = (ebitda_actual / ventas_actual * 100) if ventas_actual else 0
            eficiencia_anterior = (ebitda_anterior / ventas_anterior * 100) if ventas_anterior else 0

            eficiencias_nm = []
            serie_mensual = []

            for row in evolucion_nm:
                ventas_mes = _safe_float(row.get("ingresos_operacionales", 0))
                ebitda_mes = _safe_float(row.get("ebitda", 0))
                gastos_op_mes = _safe_float(row.get("gastos_operacionales", 0))
                dep_mes = _safe_float(row.get("dep_amort", 0))
                eficiencia_mes = (ebitda_mes / ventas_mes * 100) if ventas_mes else 0

                eficiencias_nm.append(eficiencia_mes)

                serie_mensual.append({
                    "label": row.get("label"),
                    "ventas": _round2(ventas_mes),
                    "ebitda": _round2(ebitda_mes),
                    "eficiencia_operativa": _round2(eficiencia_mes),
                    "gastos_operacionales": _round2(gastos_op_mes),
                    "dep_amort": _round2(dep_mes),
                })

            promedio_nm = _round2(sum(eficiencias_nm) / len(eficiencias_nm)) if eficiencias_nm else 0
            meta_eficiencia = _resolver_meta_eficiencia(config)

            # =========================================================
            # 5. CAJA DISPONIBLE + RUNWAY PARAMETRIZADOS
            # =========================================================
            caja_info = _calcular_caja_disponible_parametrizada(idcliente, hasta, config)
            runway_info = _calcular_cash_runway_parametrizado(
                idcliente=idcliente,
                fecha_hasta=fecha_hasta_ajustada,
                config=config,
                caja_info=caja_info
            )

            # =========================================================
            # 6. TOP GASTOS
            # =========================================================
            top_gastos_limite = int(config.get("top_gastos") or 5) if config else 5
            top_gastos = _calcular_top_gastos(composicion_actual, limite=top_gastos_limite)

            # =========================================================
            # 7. TOP CLIENTES / PROVEEDORES
            # =========================================================
            params = {"idcliente": idcliente}
            condiciones_ing = ["idcliente = :idcliente"]
            condiciones_egr = ["idcliente = :idcliente"]

            if fecha_desde:
                condiciones_ing.append("fecha >= :desde")
                condiciones_egr.append("fecha >= :desde")
                params["desde"] = fecha_desde

            if fecha_hasta_ajustada:
                condiciones_ing.append("fecha <= :hasta")
                condiciones_egr.append("fecha <= :hasta")
                params["hasta"] = fecha_hasta_ajustada

            if centro_costos:
                condiciones_ing.append("cost_center = :centro_costos")
                condiciones_egr.append("cost_center = :centro_costos")
                params["centro_costos"] = centro_costos

            where_ing = " AND ".join(condiciones_ing)
            where_egr = " AND ".join(condiciones_egr)

            limit_clientes = int(config.get("top_clientes") or 5) if config else 5
            limit_proveedores = int(config.get("top_proveedores") or 5) if config else 5

            sql_top_clientes = text(f"""
                SELECT
                    cliente_nombre AS nombre,
                    COALESCE(SUM(total), 0) AS total
                FROM facturas_enriquecidas
                WHERE {where_ing}
                GROUP BY cliente_nombre
                ORDER BY total DESC
                LIMIT {limit_clientes}
            """)

            sql_top_proveedores = text(f"""
                SELECT
                    COALESCE(proveedor_nombre, 'Sin proveedor') AS nombre,
                    COALESCE(SUM(total), 0) AS total
                FROM siigo_compras
                WHERE {where_egr}
                GROUP BY COALESCE(proveedor_nombre, 'Sin proveedor')
                ORDER BY total DESC
                LIMIT {limit_proveedores}
            """)

            top_clientes = [
                {"nombre": str(r["nombre"]), "total": _round2(r["total"])}
                for r in db.session.execute(sql_top_clientes, params).mappings().all()
            ]

            top_proveedores = [
                {"nombre": str(r["nombre"]), "total": _round2(r["total"])}
                for r in db.session.execute(sql_top_proveedores, params).mappings().all()
            ]

            # =========================================================
            # 8. EXPLICACIONES + ACCIONES + ALERTAS
            # =========================================================
            runway_ref = _safe_float(runway_info["actual"]) if runway_info.get("actual") is not None else 0

            explicaciones, acciones, alertas = _construir_explicaciones_y_acciones(
                eficiencia_actual=eficiencia_actual,
                eficiencia_anterior=eficiencia_anterior,
                promedio_6m=promedio_nm,
                runway_meses=runway_ref,
                top_gastos=top_gastos,
                ventas_actual=ventas_actual,
                ventas_anterior=ventas_anterior,
                ebitda_actual=ebitda_actual,
                ebitda_anterior=ebitda_anterior,
            )

            # Si runway no está parametrizado, no forzar alerta engañosa
            if runway_info.get("requiere_parametrizacion"):
                alertas = [{
                    "nivel": "media",
                    "titulo": "Cash runway pendiente",
                    "descripcion": runway_info.get("mensaje") or "El indicador requiere parametrización."
                }]

            # =========================================================
            # 9. RESPUESTA FINAL
            # =========================================================
            return jsonify({
                "periodo": {
                    "desde": desde,
                    "hasta": hasta,
                    "anterior_desde": prev_desde.strftime("%Y-%m-%d"),
                    "anterior_hasta": prev_hasta.strftime("%Y-%m-%d"),
                    "rango_auto": rango_auto,
                    "ajuste_por_corte": ajuste_por_corte,
                },
                "metadata": {
                    "hay_datos_auxiliar_actual": hay_datos_auxiliar_actual,
                    "ultima_fecha_auxiliar": (
                        corte["ultima_fecha_auxiliar"].strftime("%Y-%m-%d")
                        if corte["ultima_fecha_auxiliar"] else None
                    ),
                    "fecha_corte_confiable": (
                        corte["fecha_corte_confiable"].strftime("%Y-%m-%d")
                        if corte["fecha_corte_confiable"] else None
                    ),
                    "mes_actual_parcial": corte["mes_actual_parcial"],
                    "modo_periodo": corte["modo_periodo"],
                    "mensaje_contexto": mensaje_contexto,
                    "config_dashboard": {
                        "existe_config": bool(config),
                        "mostrar_caja": bool(config.get("mostrar_caja")) if config else False,
                        "mostrar_runway": bool(config.get("mostrar_runway")) if config else False,
                        "modo_caja": config.get("modo_caja") if config else "sin_configurar",
                        "modo_runway": config.get("modo_runway") if config else "sin_configurar",
                        "meses_grafica": meses_grafica,
                        "indicador_estrella": config.get("indicador_estrella") if config else "eficiencia_operativa",
                    }
                },
                "kpis": {
                    "ventas_netas": _variacion(
                        kpis_actual.get("ingresos_operacionales", 0),
                        kpis_anterior.get("ingresos_operacionales", 0)
                    ),
                    "ebitda": _variacion(
                        kpis_actual.get("ebitda", 0),
                        kpis_anterior.get("ebitda", 0)
                    ),
                    "utilidad_operativa": _variacion(
                        kpis_actual.get("utilidad_operativa", 0),
                        kpis_anterior.get("utilidad_operativa", 0)
                    ),
                    "eficiencia_operativa": {
                        "actual": _round2(eficiencia_actual),
                        "anterior": _round2(eficiencia_anterior),
                        "diff": _round2(eficiencia_actual - eficiencia_anterior),
                        "promedio_6m": _round2(promedio_nm),
                        "meta": _round2(meta_eficiencia),
                    },
                    "caja_disponible": caja_info,
                    "cash_runway": runway_info
                },
                "series": {
                    "mensual": serie_mensual
                },
                "top_gastos": top_gastos,
                "top_clientes": top_clientes,
                "top_proveedores": top_proveedores,
                "explicaciones": explicaciones,
                "acciones": acciones,
                "alertas": alertas
            }), 200

        except Exception as e:
            return jsonify({
                "error": "No fue posible construir el resumen ejecutivo",
                "detalle": str(e)
            }), 500



    # =========================================================
    # ENDPOINT CONFIG DASHBOARD RESUMEN - GET
    # =========================================================
    @app.route("/dashboard/resumen-config", methods=["GET"])
    @jwt_required()
    def dashboard_resumen_config_get():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        try:
            sql = text("""
                SELECT
                    id,
                    idcliente,
                    activo,
                    mostrar_caja,
                    mostrar_runway,
                    modo_caja,
                    cuentas_incluidas,
                    cuentas_excluidas,
                    modo_runway,
                    meses_promedio_runway,
                    meta_eficiencia_operativa,
                    meta_ebitda,
                    meta_margen_ebitda,
                    meses_grafica,
                    top_clientes,
                    top_proveedores,
                    top_gastos,
                    indicador_estrella,
                    modo_periodo_default
                FROM dashboard_resumen_config
                WHERE idcliente = :idc
                LIMIT 1
            """)

            row = db.session.execute(sql, {"idc": idcliente}).mappings().first()
            return jsonify(_serializar_dashboard_resumen_config(row)), 200

        except Exception as e:
            return jsonify({
                "error": "No fue posible obtener la configuración del dashboard",
                "detalle": str(e)
            }), 500


    # =========================================================
    # ENDPOINT CONFIG DASHBOARD RESUMEN - PUT
    # =========================================================
    @app.route("/dashboard/resumen-config", methods=["PUT"])
    @jwt_required()
    def dashboard_resumen_config_put():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        data = request.get_json(silent=True) or {}

        try:
            activo = bool(data.get("activo", True))
            mostrar_caja = bool(data.get("mostrar_caja", False))
            mostrar_runway = bool(data.get("mostrar_runway", False))

            modo_caja = str(data.get("modo_caja") or "sin_configurar").strip()
            modo_runway = str(data.get("modo_runway") or "sin_configurar").strip()
            indicador_estrella = str(data.get("indicador_estrella") or "eficiencia_operativa").strip()
            modo_periodo_default = str(data.get("modo_periodo_default") or "ytd_cerrado").strip()

            cuentas_incluidas = _limpiar_cuentas_config(data.get("cuentas_incluidas"))
            cuentas_excluidas = _limpiar_cuentas_config(data.get("cuentas_excluidas"))

            meses_promedio_runway = int(data.get("meses_promedio_runway") or 3)
            meta_eficiencia_operativa = float(data.get("meta_eficiencia_operativa") or 20)
            meta_ebitda = data.get("meta_ebitda")
            meta_margen_ebitda = data.get("meta_margen_ebitda")
            meses_grafica = int(data.get("meses_grafica") or 6)
            top_clientes = int(data.get("top_clientes") or 5)
            top_proveedores = int(data.get("top_proveedores") or 5)
            top_gastos = int(data.get("top_gastos") or 5)

            # Normalizar nullables
            meta_ebitda = float(meta_ebitda) if meta_ebitda not in (None, "", "null") else None
            meta_margen_ebitda = float(meta_margen_ebitda) if meta_margen_ebitda not in (None, "", "null") else None

            # Validaciones controladas
            modos_caja_validos = {"sin_configurar", "inclusion", "exclusion"}
            modos_runway_validos = {"sin_configurar", "burn_operativo", "egresos_promedio", "personalizado"}
            indicadores_validos = {
                "eficiencia_operativa",
                "ebitda",
                "ventas_netas",
                "utilidad_operativa",
                "caja_disponible",
                "cash_runway",
            }
            modos_periodo_validos = {"ytd_cerrado", "manual", "ultimo_mes_cerrado"}

            if modo_caja not in modos_caja_validos:
                return jsonify({"error": "modo_caja inválido"}), 400

            if modo_runway not in modos_runway_validos:
                return jsonify({"error": "modo_runway inválido"}), 400

            if indicador_estrella not in indicadores_validos:
                return jsonify({"error": "indicador_estrella inválido"}), 400

            if modo_periodo_default not in modos_periodo_validos:
                return jsonify({"error": "modo_periodo_default inválido"}), 400

            if not (1 <= meses_promedio_runway <= 12):
                return jsonify({"error": "meses_promedio_runway debe estar entre 1 y 12"}), 400

            if not (3 <= meses_grafica <= 24):
                return jsonify({"error": "meses_grafica debe estar entre 3 y 24"}), 400

            if not (3 <= top_clientes <= 20):
                return jsonify({"error": "top_clientes debe estar entre 3 y 20"}), 400

            if not (3 <= top_proveedores <= 20):
                return jsonify({"error": "top_proveedores debe estar entre 3 y 20"}), 400

            if not (3 <= top_gastos <= 20):
                return jsonify({"error": "top_gastos debe estar entre 3 y 20"}), 400

            # Upsert manual
            sql_check = text("""
                SELECT id
                FROM dashboard_resumen_config
                WHERE idcliente = :idc
                LIMIT 1
            """)
            existing = db.session.execute(sql_check, {"idc": idcliente}).mappings().first()

            if existing:
                sql_update = text("""
                    UPDATE dashboard_resumen_config
                    SET
                        activo = :activo,
                        mostrar_caja = :mostrar_caja,
                        mostrar_runway = :mostrar_runway,
                        modo_caja = :modo_caja,
                        cuentas_incluidas = CAST(:cuentas_incluidas AS JSONB),
                        cuentas_excluidas = CAST(:cuentas_excluidas AS JSONB),
                        modo_runway = :modo_runway,
                        meses_promedio_runway = :meses_promedio_runway,
                        meta_eficiencia_operativa = :meta_eficiencia_operativa,
                        meta_ebitda = :meta_ebitda,
                        meta_margen_ebitda = :meta_margen_ebitda,
                        meses_grafica = :meses_grafica,
                        top_clientes = :top_clientes,
                        top_proveedores = :top_proveedores,
                        top_gastos = :top_gastos,
                        indicador_estrella = :indicador_estrella,
                        modo_periodo_default = :modo_periodo_default,
                        actualizado_en = NOW()
                    WHERE idcliente = :idc
                """)

                db.session.execute(sql_update, {
                    "idc": idcliente,
                    "activo": activo,
                    "mostrar_caja": mostrar_caja,
                    "mostrar_runway": mostrar_runway,
                    "modo_caja": modo_caja,
                    "cuentas_incluidas": json.dumps(cuentas_incluidas),
                    "cuentas_excluidas": json.dumps(cuentas_excluidas),
                    "modo_runway": modo_runway,
                    "meses_promedio_runway": meses_promedio_runway,
                    "meta_eficiencia_operativa": meta_eficiencia_operativa,
                    "meta_ebitda": meta_ebitda,
                    "meta_margen_ebitda": meta_margen_ebitda,
                    "meses_grafica": meses_grafica,
                    "top_clientes": top_clientes,
                    "top_proveedores": top_proveedores,
                    "top_gastos": top_gastos,
                    "indicador_estrella": indicador_estrella,
                    "modo_periodo_default": modo_periodo_default,
                })
            else:
                sql_insert = text("""
                    INSERT INTO dashboard_resumen_config (
                        idcliente,
                        activo,
                        mostrar_caja,
                        mostrar_runway,
                        modo_caja,
                        cuentas_incluidas,
                        cuentas_excluidas,
                        modo_runway,
                        meses_promedio_runway,
                        meta_eficiencia_operativa,
                        meta_ebitda,
                        meta_margen_ebitda,
                        meses_grafica,
                        top_clientes,
                        top_proveedores,
                        top_gastos,
                        indicador_estrella,
                        modo_periodo_default,
                        creado_en,
                        actualizado_en
                    )
                    VALUES (
                        :idc,
                        :activo,
                        :mostrar_caja,
                        :mostrar_runway,
                        :modo_caja,
                        CAST(:cuentas_incluidas AS JSONB),
                        CAST(:cuentas_excluidas AS JSONB),
                        :modo_runway,
                        :meses_promedio_runway,
                        :meta_eficiencia_operativa,
                        :meta_ebitda,
                        :meta_margen_ebitda,
                        :meses_grafica,
                        :top_clientes,
                        :top_proveedores,
                        :top_gastos,
                        :indicador_estrella,
                        :modo_periodo_default,
                        NOW(),
                        NOW()
                    )
                """)

                db.session.execute(sql_insert, {
                    "idc": idcliente,
                    "activo": activo,
                    "mostrar_caja": mostrar_caja,
                    "mostrar_runway": mostrar_runway,
                    "modo_caja": modo_caja,
                    "cuentas_incluidas": json.dumps(cuentas_incluidas),
                    "cuentas_excluidas": json.dumps(cuentas_excluidas),
                    "modo_runway": modo_runway,
                    "meses_promedio_runway": meses_promedio_runway,
                    "meta_eficiencia_operativa": meta_eficiencia_operativa,
                    "meta_ebitda": meta_ebitda,
                    "meta_margen_ebitda": meta_margen_ebitda,
                    "meses_grafica": meses_grafica,
                    "top_clientes": top_clientes,
                    "top_proveedores": top_proveedores,
                    "top_gastos": top_gastos,
                    "indicador_estrella": indicador_estrella,
                    "modo_periodo_default": modo_periodo_default,
                })

            db.session.commit()

            sql_get = text("""
                SELECT
                    id,
                    idcliente,
                    activo,
                    mostrar_caja,
                    mostrar_runway,
                    modo_caja,
                    cuentas_incluidas,
                    cuentas_excluidas,
                    modo_runway,
                    meses_promedio_runway,
                    meta_eficiencia_operativa,
                    meta_ebitda,
                    meta_margen_ebitda,
                    meses_grafica,
                    top_clientes,
                    top_proveedores,
                    top_gastos,
                    indicador_estrella,
                    modo_periodo_default
                FROM dashboard_resumen_config
                WHERE idcliente = :idc
                LIMIT 1
            """)
            row = db.session.execute(sql_get, {"idc": idcliente}).mappings().first()

            return jsonify(_serializar_dashboard_resumen_config(row)), 200

        except Exception as e:
            db.session.rollback()
            return jsonify({
                "error": "No fue posible guardar la configuración del dashboard",
                "detalle": str(e)
            }), 500



    # =========================================================
    # ENDPOINT BUSCAR CUENTAS AUXILIAR PARA CONFIGURACIÓN
    # =========================================================
    @app.route("/dashboard/resumen-config/buscar-cuentas", methods=["GET"])
    @jwt_required()
    def dashboard_resumen_config_buscar_cuentas():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        if not idcliente:
            return jsonify({"error": "Token sin cliente"}), 403

        q = request.args.get("q", "").strip()
        limite = request.args.get("limite", default=20, type=int)

        try:
            cuentas = _buscar_cuentas_auxiliar_para_config(
                idcliente=idcliente,
                q=q,
                limite=limite,
            )

            return jsonify({
                "q": q,
                "total": len(cuentas),
                "items": cuentas
            }), 200

        except Exception as e:
            return jsonify({
                "error": "No fue posible buscar cuentas contables",
                "detalle": str(e)
            }), 500



    # Endpoint para llenar tabla de stagin compras
    @app.route("/siigo/sync-documentos-soporte-staging", methods=["POST"])
    def siigo_sync_documentos_soporte_staging():
        """
        Sincroniza documentos soporte desde Siigo API hacia tabla staging.

        No toca siigo_compras.
        No toca siigo_compras_items.
        No afecta reportes actuales.

        Si se ejecuta manualmente, registra historial.
        Si viene desde sync-all, no registra log individual porque sync-all ya registra toda la ejecución.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        batch = request.args.get("batch", default=50, type=int)
        max_pages = request.args.get("max_pages", default=None, type=int)

        endpoint_log = "/siigo/sync-documentos-soporte-staging"
        log_id = None
        inicio = time.time()

        params_log = {
            "batch": batch,
            "max_pages": max_pages,
        }

        if not modo_sync_all:
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje="Documento Soporte API staging: proceso iniciado."
            )

        try:
            resultado = sync_documentos_soporte_staging_desde_siigo(
                idcliente=idcliente,
                batch_size=batch if batch else 50,
                max_pages=max_pages,
            )

            status = 500 if isinstance(resultado, dict) and resultado.get("error") else 200

            if not modo_sync_all:
                if status >= 400:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="ERROR",
                        detalle=f"Error actualizando staging de Documento Soporte: {resultado}",
                        status_code=status,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )
                else:
                    _finalizar_log_sync_modulo(
                        log_id=log_id,
                        idcliente=idcliente,
                        endpoint=endpoint_log,
                        resultado="OK",
                        detalle=f"Staging de Documento Soporte actualizado correctamente: {resultado}",
                        status_code=200,
                        duracion_segundos=round(time.time() - inicio, 2),
                    )

            return jsonify(resultado), status

        except Exception as e:
            db.session.rollback()
            traceback.print_exc()

            detalle = f"Error sincronizando documentos soporte staging: {str(e)}"

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="ERROR",
                    detalle=detalle,
                    status_code=500,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify({
                "error": "Error sincronizando documentos soporte staging",
                "detalle": str(e),
                "tipo": "documentos_soporte_staging_error",
            }), 500



    @app.route("/siigo/insert-documentos-soporte-desde-staging", methods=["POST"])
    def siigo_insert_documentos_soporte_desde_staging():
        """
        Inserta en siigo_compras únicamente DS nuevos desde staging.

        No actualiza documentos existentes.
        No inserta Failed, Draft, Rejected, Sent.
        No usa balance API como saldo definitivo.

        dry_run = true  -> Simulación
        dry_run = false -> Inserción real

        Si se ejecuta manualmente, registra historial.
        Si viene desde sync-all, no registra log individual porque sync-all ya registra toda la ejecución.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        modo_sync_all = request.headers.get("X-SYNC-ALL") == "1"

        payload = request.get_json(silent=True) or {}

        fecha_desde = (
            request.args.get("fecha_desde")
            or payload.get("fecha_desde")
        )

        dry_run = (
            request.args.get("dry_run", "").lower() in ("1", "true", "yes")
            or bool(payload.get("dry_run"))
        )

        max_registros = request.args.get("max_registros", default=None, type=int)

        endpoint_log = "/siigo/insert-documentos-soporte-desde-staging"
        log_id = None
        inicio = time.time()

        nombre_proceso = (
            "Simulación de inserción de Documento Soporte"
            if dry_run
            else "Inserción de Documento Soporte en compras"
        )

        params_log = {
            "dry_run": dry_run,
            "fecha_desde": fecha_desde,
            "max_registros": max_registros,
        }

        if not modo_sync_all:
            log_id = _crear_log_sync_modulo_inicio(
                idcliente=idcliente,
                endpoint=endpoint_log,
                origen="manual_modulo",
                params=params_log,
                mensaje=(
                    f"{nombre_proceso}: proceso iniciado"
                    + (f" con fecha desde {fecha_desde}." if fecha_desde else " sin límite de fecha.")
                )
            )

        try:
            resultado = insertar_documentos_soporte_desde_staging(
                idcliente=idcliente,
                fecha_desde=fecha_desde,
                dry_run=dry_run,
                max_registros=max_registros,
            )

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="OK",
                    detalle=f"{nombre_proceso} finalizada correctamente: {resultado}",
                    status_code=200,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify(resultado), 200

        except Exception as e:
            db.session.rollback()
            traceback.print_exc()

            detalle = f"Error en {nombre_proceso}: {str(e)}"

            if not modo_sync_all:
                _finalizar_log_sync_modulo(
                    log_id=log_id,
                    idcliente=idcliente,
                    endpoint=endpoint_log,
                    resultado="ERROR",
                    detalle=detalle,
                    status_code=500,
                    duracion_segundos=round(time.time() - inicio, 2),
                )

            return jsonify({
                "error": "Error insertando documentos soporte desde staging",
                "detalle": str(e),
                "tipo": "documentos_soporte_insert_error",
                "dry_run": dry_run,
                "fecha_desde": fecha_desde,
            }), 500



    #Endpoints para consultar historial ed sincronizaciones
    @app.route("/config/siigo-sync-history", methods=["GET"])
    @jwt_required()
    def get_siigo_sync_history():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)
        limit = request.args.get("limit", default=10, type=int)

        if perfilid == 0 and q_idcliente:
            idcliente = q_idcliente

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        if limit < 1:
            limit = 10
        if limit > 50:
            limit = 50

        cliente = Cliente.query.get(idcliente)
        tz_str = cliente.timezone if cliente and cliente.timezone else "America/Bogota"

        logs = (
            SiigoSyncLog.query
            .filter_by(idcliente=idcliente)
            .order_by(SiigoSyncLog.ejecutado_en.desc().nullslast(), SiigoSyncLog.creado_en.desc())
            .limit(limit)
            .all()
        )

        data = []

        for log in logs:
            resumen = _resumir_detalle_sync(log.detalle or "")

            data.append({
                "id": log.id,
                "idcliente": log.idcliente,
                "origen": log.origen or "desconocido",
                "resultado": log.resultado,
                "ejecutado_en": log.ejecutado_en.isoformat() if log.ejecutado_en else None,
                "creado_en": log.creado_en.isoformat() if log.creado_en else None,
                "total_pasos": log.total_pasos if log.total_pasos is not None else resumen["total_pasos"],
                "pasos_ok": log.pasos_ok if log.pasos_ok is not None else resumen["pasos_ok"],
                "pasos_error": log.pasos_error if log.pasos_error is not None else resumen["pasos_error"],
                "endpoint_fallido": log.endpoint_fallido or resumen["endpoint_fallido"],
                "detalle": log.detalle or "",
            })

        return jsonify({
            "timezone": tz_str,
            "items": data,
        }), 200



    # Endpoints de reporte de consultas cotizaciones
    # ---- ENDPOINT DEBUG consulta cotizaciones
    @app.route("/siigo/debug-cotizaciones", methods=["GET"])
    def siigo_debug_cotizaciones():
        """
        Endpoint temporal SOLO LECTURA.
        Consulta cotizaciones desde Siigo API y devuelve el JSON crudo.

        No guarda.
        No actualiza.
        No borra.
        No toca tablas de InsightFlow.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        page_size = request.args.get("page_size", default=10, type=int)
        page = request.args.get("page", default=1, type=int)

        # Opcionales para explorar qué filtros acepta realmente Siigo.
        # Si Siigo no los soporta, devolverá el error de la API y lo veremos en pantalla.
        created_start = request.args.get("created_start")
        created_end = request.args.get("created_end")
        date_start = request.args.get("date_start")
        date_end = request.args.get("date_end")

        params = {
            "page_size": page_size,
            "page": page,
        }

        if created_start:
            params["created_start"] = created_start
        if created_end:
            params["created_end"] = created_end
        if date_start:
            params["date_start"] = date_start
        if date_end:
            params["date_end"] = date_end

        query_string = urlencode(params)
        url = f"{base_url}/v1/quotations?{query_string}"

        try:
            r = requests.get(url, headers=headers, timeout=90)

            try:
                data = r.json()
            except ValueError:
                return jsonify({
                    "error": "Siigo respondió algo que no es JSON",
                    "status_code": r.status_code,
                    "url": url,
                    "text": r.text
                }), 500

            results = []

            if isinstance(data, dict):
                if isinstance(data.get("results"), list):
                    results = data.get("results")
                elif isinstance(data.get("data"), dict) and isinstance(data["data"].get("results"), list):
                    results = data["data"].get("results")

            nombres_detectados = []
            for item in results:
                if isinstance(item, dict):
                    nombre = item.get("name") or item.get("number") or item.get("id")
                    if nombre:
                        nombres_detectados.append(str(nombre))

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Consulta de cotizaciones realizada sin guardar información en InsightFlow.",
                "status_code": r.status_code,
                "url": url,
                "cantidad_results": len(results),
                "nombres_detectados": nombres_detectados,
                "data": data
            }), r.status_code

        except Exception as e:
            return jsonify({
                "error": "Error consultando cotizaciones en Siigo",
                "detalle": str(e)
            }), 500


    @app.route("/siigo/debug-cotizaciones/<string:cotizacion_id>", methods=["GET"])
    def siigo_debug_cotizacion_detalle(cotizacion_id):
        """
        Endpoint temporal SOLO LECTURA.
        Consulta una cotización específica por ID Siigo.

        No guarda.
        No actualiza.
        No borra.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        url = f"{base_url}/v1/quotations/{cotizacion_id}"

        try:
            r = requests.get(url, headers=headers, timeout=90)

            try:
                data = r.json()
            except ValueError:
                return jsonify({
                    "error": "Siigo respondió algo que no es JSON",
                    "status_code": r.status_code,
                    "url": url,
                    "text": r.text
                }), 500

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Consulta de detalle de cotización realizada sin guardar información en InsightFlow.",
                "status_code": r.status_code,
                "url": url,
                "data": data
            }), r.status_code

        except Exception as e:
            return jsonify({
                "error": "Error consultando detalle de cotización en Siigo",
                "detalle": str(e)
            }), 500


    @app.route("/siigo/debug-cotizaciones-buscar", methods=["GET"])
    def siigo_debug_cotizaciones_buscar():
        """
        Endpoint temporal SOLO LECTURA.
        Busca una cotización por name, recorriendo páginas del endpoint /v1/quotations.

        Útil para encontrar el ID Siigo de una cotización visible como C-1-481.
        """
        idcliente = obtener_idcliente_desde_request()
        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        name = (request.args.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Debes enviar ?name=C-1-481"}), 400

        cfg = SiigoCredencial.query.filter_by(idcliente=idcliente).first()
        if not cfg:
            return jsonify({"error": "Credenciales Siigo no configuradas"}), 400

        auth_data = _siigo_auth_json_for_client(cfg)
        if not isinstance(auth_data, dict):
            return jsonify({
                "error": "Respuesta inesperada del auth de Siigo",
                "detalle": str(auth_data)
            }), 400

        token = auth_data.get("access_token")
        if not token:
            return jsonify({
                "error": "No se obtuvo access_token",
                "detalle": auth_data
            }), 400

        headers = _siigo_headers_bearer(token)
        base_url = cfg.base_url.rstrip("/")

        page_size = request.args.get("page_size", default=50, type=int)
        max_pages = request.args.get("max_pages", default=20, type=int)

        encontrados = []
        paginas_consultadas = 0
        ultimo_status_code = None
        ultima_url = None

        try:
            for page in range(1, max_pages + 1):
                url = f"{base_url}/v1/quotations?page_size={page_size}&page={page}"
                ultima_url = url

                r = requests.get(url, headers=headers, timeout=90)
                ultimo_status_code = r.status_code
                paginas_consultadas += 1

                try:
                    data = r.json()
                except ValueError:
                    return jsonify({
                        "error": "Siigo respondió algo que no es JSON",
                        "status_code": r.status_code,
                        "url": url,
                        "text": r.text
                    }), 500

                if not r.ok:
                    return jsonify({
                        "error": "Siigo respondió error consultando cotizaciones",
                        "status_code": r.status_code,
                        "url": url,
                        "data": data
                    }), r.status_code

                results = []
                if isinstance(data, dict):
                    if isinstance(data.get("results"), list):
                        results = data.get("results")
                    elif isinstance(data.get("data"), dict) and isinstance(data["data"].get("results"), list):
                        results = data["data"].get("results")

                for item in results:
                    if not isinstance(item, dict):
                        continue

                    item_name = str(item.get("name") or "").strip()
                    if item_name.upper() == name.upper():
                        encontrados.append(item)

                if encontrados:
                    break

                if not results:
                    break

            return jsonify({
                "modo": "solo_lectura",
                "mensaje": "Búsqueda de cotización realizada sin guardar información en InsightFlow.",
                "cotizacion_buscada": name,
                "encontrada": len(encontrados) > 0,
                "cantidad_encontrada": len(encontrados),
                "paginas_consultadas": paginas_consultadas,
                "ultimo_status_code": ultimo_status_code,
                "ultima_url": ultima_url,
                "resultados": encontrados
            }), 200

        except Exception as e:
            return jsonify({
                "error": "Error buscando cotización en Siigo",
                "detalle": str(e)
            }), 500




# Hasta aqui endpoints de reportes
#_____________________________________________________________________________________________________________________________________

#----------------------------------------------------------------------------------------------------------------------------------
# No tocar de qui para abajo 
 
    # NO TOCAR DE AQEUI PARA ABAJO
    # --- Registrar rutas de permisos ---
    from permisos_routes import register_permisos_routes
    register_permisos_routes(app)


    @app.before_request
    def verificar_permisos_global():
        # 🔓 Preflight CORS
        if request.method == "OPTIONS":
            return jsonify({"status": "ok"}), 200

        # 🔓 Rutas públicas
        if request.path.startswith("/auth") or request.path == "/":
            return

        # ⚠️ Rutas internas / técnicas exentas
        rutas_exentas = [
            "/siigo/sync-catalogos",
            "/siigo/sync-customers",
            "/siigo/sync-proveedores",
            "/siigo/sync-productos",
            "/siigo/sync-facturas",
            "/siigo/sync-notas-credito",
            "/siigo/sync-compras",
                # ✅ Documento Soporte API usado internamente por Sync-all
           "/siigo/sync-documentos-soporte-staging",
            "/siigo/insert-documentos-soporte-desde-staging",
            "/siigo/sync-accounts-payable",
            "/siigo/cross-accounts-payable",
            "/siigo/sync-all",
            "/config/siigo-sync-status",
            "/ping",
        ]

        for ruta in rutas_exentas:
            if ruta in request.path:
                return

        # ✅ Verificar JWT válido
        try:
            verify_jwt_in_request(optional=False)
        except Exception as e:
            return jsonify({"error": f"Token inválido o faltante: {str(e)}"}), 401

        claims = get_jwt()

        # 👑 SuperAdmin entra sin restricciones
        if _is_superadmin(claims):
            return

        idperfil = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        if not idperfil or not idcliente:
            return jsonify({
                "error": "Token inválido: falta perfil o cliente",
                "ruta": request.path
            }), 403

        print(
            f"[PERMISOS] Ruta: {request.path}, "
            f"Método: {request.method}, "
            f"Perfil: {idperfil}, Cliente: {idcliente}"
        )

        codigo = None

        # ======================================================
        # Clientes
        # ======================================================
        if request.path.startswith("/clientes"):
            codigo = "ver_clientes"

        # ======================================================
        # Administración global SuperAdmin
        # Nunca debe ser usada por usuarios cliente.
        # ======================================================
        elif request.path.startswith("/admin"):
            codigo = "admin_panel"

        # ======================================================
        # Perfiles de cliente
        # ======================================================
        elif request.path.startswith("/perfiles"):
            if request.method == "GET":
                codigo = "ver_perfiles"
            else:
                codigo = "editar_perfiles"

        # ======================================================
        # Usuarios de cliente
        # ======================================================
        elif request.path.startswith("/usuarios"):
            if request.method == "GET":
                codigo = "ver_usuarios"
            else:
                codigo = "editar_usuarios"

        # ======================================================
        # Permisos / mis permisos / permisos por perfil
        # Estos endpoints tienen validación propia en permisos_routes.py
        # y son necesarios para construir menú/sidebar.
        # ======================================================
        elif request.path.startswith("/api/mis_permisos"):
            return

        elif request.path.startswith("/api/perfiles"):
            return

        elif request.path.startswith("/api/usuarios"):
            return

        elif request.path.startswith("/api/permisos"):
            return

        # ======================================================
        # Notificaciones
        # De momento se permite si el usuario tiene dashboard.
        # ======================================================
        elif request.path.startswith("/api/notificaciones"):
            codigo = "ver_dashboard"

        # ======================================================
        # Siigo
        # ======================================================
        elif request.path.startswith("/siigo"):
            codigo = "ver_siigo"

        # ======================================================
        # Dashboard / configuración
        # ======================================================
        elif request.path.startswith("/dashboard"):
            if "resumen-config/buscar-cuentas" in request.path:
                codigo = "ver_configuraciones_varias"
            elif "resumen-config" in request.path or "configuraciones_varias" in request.path:
                codigo = "ver_configuraciones_varias"
            elif "resumen-ejecutivo" in request.path:
                codigo = "ver_resumen_ejecutivo"
            else:
                codigo = "ver_dashboard"

        # ======================================================
        # Reportes
        # ======================================================
        elif request.path.startswith("/reportes"):
            # Normaliza rutas:
            # /reportes/facturas_enriquecidas -> /reportes/facturas-enriquecidas
            # /reportes/analisis_variacion_v1 -> /reportes/analisis-variacion-v1
            path_norm = request.path.lower().replace("_", "-")

            # ------------------------------------------
            # Cargue de Auxiliar Contable
            # Usado por páginas del módulo financiero:
            # Cruce de IVA, Retenciones, P&L, Variación,
            # Balance General e Indicadores Auxiliares.
            # ------------------------------------------
            if (
                "cargar-auxiliar" in path_norm
                or "cargar-auxiliares" in path_norm
                or "upload-auxiliar" in path_norm
                or "importar-auxiliar" in path_norm
            ):
                permisos_financieros_auxiliar = [
                    "ver_reporte_cruceivas",
                    "ver_reporte_retenciones",
                    "ver_reporte_estado_resultados",
                    "ver_reporte_analisis_variacion",
                    "ver_reporte_balance_general",
                    "ver_reporte_indicadores_auxiliares",
                ]

                tiene_permiso_paquete = any(
                    cliente_tiene_permiso_en_paquete(idcliente, permiso)
                    for permiso in permisos_financieros_auxiliar
                )

                tiene_permiso_perfil = any(
                    _perfil_tiene_permiso(idperfil, idcliente, permiso)
                    for permiso in permisos_financieros_auxiliar
                )

                if not tiene_permiso_paquete:
                    return jsonify({
                        "error": "Acceso denegado: el paquete contratado no incluye cargue de auxiliar contable.",
                        "ruta": request.path,
                        "motivo": "cargue_auxiliar_no_incluido_en_paquete"
                    }), 403

                if not tiene_permiso_perfil:
                    return jsonify({
                        "error": "Acceso denegado: el perfil no tiene permisos financieros para cargar auxiliar contable.",
                        "ruta": request.path,
                        "motivo": "cargue_auxiliar_no_asignado_al_perfil"
                    }), 403

                return

            # ------------------------------------------
            # Buscador Inteligente de Facturas
            # Ejemplo:
            # /reportes/busqueda-inteligente-facturas
            # /reportes/financiero/buscador-facturas
            # ------------------------------------------
            if (
                "busqueda-inteligente-facturas" in path_norm
                or "buscador-facturas" in path_norm
                or "facturas-buscador" in path_norm
                or "busqueda-facturas" in path_norm
                or "facturas-inteligente" in path_norm
            ):
                codigo = "ver_reporte_buscador_facturas"
                

            # ------------------------------------------
            # Cuentas por Cobrar / Cartera / Aging
            # Ejemplo:
            # /reportes/cuentas-por-cobrar
            # /reportes/financiero/cxc
            # ------------------------------------------
            elif (
                "cuentas-por-cobrar" in path_norm
                or "cuentasporcobrar" in path_norm
                or "cxc" in path_norm
                or "cartera" in path_norm
                or "aging" in path_norm
            ):
                codigo = "ver_reporte_cxc"

            # ------------------------------------------
            # Ventas por Vendedor
            # Ejemplo:
            # /reportes/vendedores
            # ------------------------------------------
            elif (
                "vendedores" in path_norm
                or "vendedor" in path_norm
                or "ventas-vendedor" in path_norm
                or "sales-by-vendor" in path_norm
            ):
                codigo = "ver_reporte_vendedores"

            # ------------------------------------------
            # Ventas por Producto
            # Ejemplo:
            # /reportes/productos
            # ------------------------------------------
            elif (
                "productos" in path_norm
                or "producto" in path_norm
                or "ventas-producto" in path_norm
                or "sales-by-product" in path_norm
            ):
                codigo = "ver_reporte_productos"

            # ------------------------------------------
            # Ingresos por Ventas y endpoints auxiliares de ventas
            # Ejemplos usados por la página de ventas:
            # /reportes/facturas-enriquecidas
            # /reportes/facturas-por-cliente
            # /reportes/facturas-por-estado
            # /reportes/facturas-detalle-mes
            # /reportes/financiero/ventas
            # ------------------------------------------
            elif (
                "ventas" in path_norm
                or "ingresos" in path_norm
                or "facturas-enriquecidas" in path_norm
                or "facturas-por-cliente" in path_norm
                or "facturas-por-estado" in path_norm
                or "facturas-detalle-mes" in path_norm
                or "facturacion-ventas" in path_norm
                or "sales" in path_norm
            ):
                codigo = "ver_reporte_ventas"

            # ------------------------------------------
            # Facturación Clientes
            # Se deja después de ventas para que facturas-por-cliente
            # siga perteneciendo al reporte de ventas.
            # ------------------------------------------
            elif (
                "facturas-cliente" in path_norm
                or "facturas-clientes" in path_norm
                or "facturacion-clientes" in path_norm
                or "analisis-clientes" in path_norm
                or "reporte-clientes" in path_norm
            ):
                codigo = "ver_reporte_clientes"

            # ------------------------------------------
            # Compras / Gastos / Egresos
            # ------------------------------------------
            elif (
                "compras-gastos" in path_norm
                or "compras-y-gastos" in path_norm
                or "compras-gastos" in path_norm
                or "egresos" in path_norm
                or "gastos" in path_norm
            ):
                codigo = "ver_reporte_compras_gastos"

            # ------------------------------------------
            # Nómina
            # ------------------------------------------
            elif "nomina" in path_norm or "nómina" in path_norm:
                codigo = "ver_reporte_nomina"

            # ------------------------------------------
            # Compras a Proveedores
            # ------------------------------------------
            elif (
                "proveedores" in path_norm
                or "proveedor" in path_norm
                or "compras-proveedores" in path_norm
            ):
                codigo = "ver_reporte_proveedores"

            # ------------------------------------------
            # Financiero Consolidado
            # ------------------------------------------
            elif "consolidado" in path_norm:
                codigo = "ver_reporte_consolidado"

            # ------------------------------------------
            # Cruce de IVAs
            # ------------------------------------------
            elif (
                "cruce-iva" in path_norm
                or "cruce-ivas" in path_norm
                or "cruceiva" in path_norm
            ):
                codigo = "ver_reporte_cruceivas"

            # ------------------------------------------
            # Retenciones
            # ------------------------------------------
            elif "retenciones" in path_norm or "retencion" in path_norm:
                codigo = "ver_reporte_retenciones"

            # ------------------------------------------
            # Estado de Resultados / P&L
            # ------------------------------------------
            elif (
                "pnl-v1" in path_norm
                or "pnl" in path_norm
                or "estado-resultados" in path_norm
                or "estado-de-resultados" in path_norm
                or "estadoresultados" in path_norm
            ):
                codigo = "ver_reporte_estado_resultados"

            # ------------------------------------------
            # Análisis de Variación
            # ------------------------------------------
            elif (
                "analisis-variacion" in path_norm
                or "analisis-variacion-v1" in path_norm
                or "variacion" in path_norm
            ):
                codigo = "ver_reporte_analisis_variacion"

            # ------------------------------------------
            # Indicadores financieros auxiliares
            # ------------------------------------------
            elif (
                "auxiliares/indicadores-financieros" in path_norm
                or "indicadores-financieros-auxiliares" in path_norm
                or "indicadores-auxiliares" in path_norm
            ):
                codigo = "ver_reporte_indicadores_auxiliares"

            # ------------------------------------------
            # Indicadores financieros antiguos
            # ------------------------------------------
            elif "indicadores" in path_norm:
                codigo = "ver_reporte_indicadores"

            # ------------------------------------------
            # Balance General
            # ------------------------------------------
            elif (
                "balance-general" in path_norm
                or "balancegeneral" in path_norm
            ):
                codigo = "ver_reporte_balance_general"

            # ------------------------------------------
            # Balance antiguo / balance prueba
            # ------------------------------------------
            elif "balance" in path_norm:
                codigo = "ver_reporte_balance"

            # ------------------------------------------
            # Ruta no mapeada
            # No usamos ver_reportes para evitar permisos demasiado amplios.
            # ------------------------------------------
            else:
                return jsonify({
                    "error": "Ruta de reporte no mapeada en control de permisos.",
                    "ruta": request.path,
                    "motivo": "reporte_sin_mapeo",
                    "recomendacion": (
                        "Agrega esta ruta al mapeo de permisos en before_request "
                        "para asociarla con un permiso específico."
                    )
                }), 403
            
        # Si no hay permiso requerido, no aplica control global
        if not codigo:
            return

        # 🔒 1. Verificar que el paquete contratado incluya el permiso
        if not cliente_tiene_permiso_en_paquete(idcliente, codigo):
            return jsonify({
                "error": f"Acceso denegado: el paquete contratado no incluye el permiso '{codigo}'",
                "permiso": codigo,
                "ruta": request.path,
                "motivo": "permiso_no_incluido_en_paquete"
            }), 403

        # 🔒 2. Verificar que el perfil del usuario tenga el permiso asignado
        if not _perfil_tiene_permiso(idperfil, idcliente, codigo):
            return jsonify({
                "error": f"Acceso denegado: falta permiso '{codigo}'",
                "permiso": codigo,
                "ruta": request.path,
                "motivo": "permiso_no_asignado_al_perfil"
            }), 403


    @app.route("/config/sync", methods=["GET", "OPTIONS"])
    @jwt_required()
    @cross_origin()
    def get_sync_config():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)

        if perfilid == 0:
            if q_idcliente:
                idcliente = q_idcliente
            elif not idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        config = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        if not config:
            return jsonify({
                "idcliente": idcliente,
                "hora_ejecucion": None,
                "frecuencia_dias": 1,
                "activo": False,
                "ultimo_ejecutado": None,
                "ultimo_auto_ejecutado": None,
                "resultado_ultima_sync": None,
                "detalle_ultima_sync": None,
                "created_at": None,

                # Fecha global oficial para documentos transaccionales Siigo
                "sync_fecha_desde": None,

                # Compatibilidad con implementación anterior de Documento Soporte
                "ds_fecha_desde": None,
            }), 200

        return jsonify(config.as_dict()), 200


    @app.route("/config/sync", methods=["POST", "OPTIONS"])
    @jwt_required()
    @cross_origin()
    def save_sync_config():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)

        if perfilid == 0:
            if q_idcliente:
                idcliente = q_idcliente
            elif not idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        data = request.get_json() or {}

        hora_ejecucion = _parse_time_hh_mm(data.get("hora_ejecucion") or "02:00")
        frecuencia_dias_raw = data.get("frecuencia_dias", 1)
        activo_raw = data.get("activo", True)

        # Fecha global oficial:
        # aplica a facturas de venta, notas crédito, compras y Documento Soporte API.
        # ds_fecha_desde queda como compatibilidad, pero se sincroniza con la global.
        sync_fecha_desde = _parse_date_yyyy_mm_dd(
            data.get("sync_fecha_desde") or data.get("ds_fecha_desde")
        )

        try:
            frecuencia_dias = int(frecuencia_dias_raw or 1)
            if frecuencia_dias < 1:
                frecuencia_dias = 1
        except Exception:
            frecuencia_dias = 1

        config = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        if not config:
            config = SiigoSyncConfig(
                idcliente=idcliente,
                hora_ejecucion=hora_ejecucion,
                frecuencia_dias=frecuencia_dias,
                activo=bool(activo_raw),

                # Fecha global oficial
                sync_fecha_desde=sync_fecha_desde,

                # Compatibilidad Documento Soporte
                ds_fecha_desde=sync_fecha_desde,
            )
        else:
            config.hora_ejecucion = hora_ejecucion
            config.frecuencia_dias = frecuencia_dias
            config.activo = bool(activo_raw)

            # Fecha global oficial
            config.sync_fecha_desde = sync_fecha_desde

            # Compatibilidad Documento Soporte
            config.ds_fecha_desde = sync_fecha_desde

        db.session.add(config)
        db.session.commit()

        return jsonify({
            "mensaje": "Configuración guardada",
            "config": config.as_dict()
        }), 200


    @app.route("/config/siigo-sync-status", methods=["GET", "OPTIONS"])
    @jwt_required()
    @cross_origin()
    def config_siigo_sync_status():
        claims = get_jwt()
        perfilid = claims.get("perfilid")
        idcliente = claims.get("idcliente")

        q_idcliente = request.args.get("idcliente", type=int)

        if perfilid == 0:
            if q_idcliente:
                idcliente = q_idcliente
            elif not idcliente:
                return jsonify({"error": "Falta idcliente"}), 400
        else:
            if not idcliente:
                return jsonify({"error": "No autorizado"}), 403

        cliente = Cliente.query.get(idcliente)
        tz_str = cliente.timezone if cliente and cliente.timezone else "America/Bogota"

        config = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        if not config:
            return jsonify({
                "pendientes": 0,
                "ultimo_ejec": None,
                "ultimo_auto_ejec": None,
                "resultado": None,
                "detalle": "",
                "hora_ejecucion": None,
                "frecuencia_dias": 1,
                "activo": False,
                "timezone": tz_str,

                # Fecha global oficial para documentos transaccionales Siigo
                "sync_fecha_desde": None,

                # Compatibilidad con implementación previa de Documento Soporte
                "ds_fecha_desde": None,
            }), 200

        # Última ejecución general: manual o cron
        if config.ultimo_ejecutado:
            dt_local = utc_to_local(config.ultimo_ejecutado, tz_str)
            ultimo_ejec = dt_local.isoformat()
        else:
            ultimo_ejec = None

        # Última ejecución automática: solo cron
        if config.ultimo_auto_ejecutado:
            dt_auto_local = utc_to_local(config.ultimo_auto_ejecutado, tz_str)
            ultimo_auto_ejec = dt_auto_local.isoformat()
        else:
            ultimo_auto_ejec = None

        return jsonify({
            "pendientes": 0,
            "ultimo_ejec": ultimo_ejec,
            "ultimo_auto_ejec": ultimo_auto_ejec,
            "resultado": config.resultado_ultima_sync,
            "detalle": config.detalle_ultima_sync or "",
            "hora_ejecucion": config.hora_ejecucion.strftime("%H:%M") if config.hora_ejecucion else None,
            "frecuencia_dias": config.frecuencia_dias,
            "activo": config.activo,
            "timezone": tz_str,

            # Fecha global oficial
            "sync_fecha_desde": config.sync_fecha_desde.isoformat() if config.sync_fecha_desde else None,

            # Compatibilidad Documento Soporte
            "ds_fecha_desde": config.ds_fecha_desde.isoformat() if config.ds_fecha_desde else None,
        }), 200



    # Enpoint para llevar a cabo la ejecucion del boton de sincronizacion de todo en siigo
    # Endpoint para sincronizar todo (invoca internamente otros endpoints)
    # Endpoint para sincronizar todo (invoca internamente otros endpoints)
    @app.route("/siigo/sync-all", methods=["POST"])
    def siigo_sync_all():
        idcliente = obtener_idcliente_desde_request()
        print(f"🔹 Sync-all iniciado para cliente {idcliente}")

        if not idcliente:
            return jsonify({"error": "Cliente no autorizado"}), 403

        data = request.get_json(silent=True) or {}
        origen = data.get("origen", "cron")
        es_cron = origen == "cron"

        log_parts = []
        overall_status = "OK"

        # 🕒 Obtener zona horaria del cliente
        cliente = Cliente.query.get_or_404(idcliente)
        tz_str = cliente.timezone or "America/Bogota"
        print(f"🌎 Zona horaria detectada para cliente {idcliente}: {tz_str}")
        print(f"🧭 Origen sync-all: {origen}")

        # 🧩 Configuración actual del cliente
        config_actual = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        # ==========================================================
        # Fecha global de datos Siigo
        # ==========================================================
        # Aplica a documentos transaccionales:
        # - Facturas de venta
        # - Notas crédito
        # - Compras
        # - Documento Soporte API al insertar desde staging
        #
        # No aplica a:
        # - Catálogos
        # - Clientes
        # - Proveedores
        # - Productos
        # - Accounts payable
        # - Cross accounts payable
        #
        # Prioridad:
        # 1. sync_fecha_desde enviada por frontend
        # 2. ds_fecha_desde enviada por compatibilidad
        # 3. config.sync_fecha_desde
        # 4. config.ds_fecha_desde por compatibilidad
        # 5. Sin fecha
        # ==========================================================
        sync_fecha_desde = None

        if data.get("sync_fecha_desde"):
            sync_fecha_desde = data.get("sync_fecha_desde")
        elif data.get("ds_fecha_desde"):
            sync_fecha_desde = data.get("ds_fecha_desde")
        elif config_actual and getattr(config_actual, "sync_fecha_desde", None):
            sync_fecha_desde = config_actual.sync_fecha_desde.isoformat()
        elif config_actual and getattr(config_actual, "ds_fecha_desde", None):
            sync_fecha_desde = config_actual.ds_fecha_desde.isoformat()

        # Validar formato. Si viene inválida, la dejamos sin límite.
        if sync_fecha_desde:
            fecha_validada = _parse_date_yyyy_mm_dd(sync_fecha_desde)
            sync_fecha_desde = fecha_validada.isoformat() if fecha_validada else None

        print(
            f"📅 Fecha global de datos Siigo para cliente {idcliente}: "
            f"{sync_fecha_desde or 'SIN LÍMITE'}"
        )

        # Parámetros por módulo
        facturas_params = {}
        facturas_deep_params = {
            "deep": 1,
            "batch": 100,
            "only_missing": 1,
        }
        notas_credito_params = {}
        compras_params = {}
        ds_insert_params = {
            "dry_run": 0
        }

        if sync_fecha_desde:
            facturas_params["since"] = sync_fecha_desde
            facturas_deep_params["since"] = sync_fecha_desde
            notas_credito_params["since"] = sync_fecha_desde
            compras_params["since"] = sync_fecha_desde
            ds_insert_params["fecha_desde"] = sync_fecha_desde

        # 🔁 Secuencia de endpoints Siigo a ejecutar
        sequence = [
            ("/siigo/sync-catalogos", {}),
            ("/siigo/sync-customers", {}),
            ("/siigo/sync-proveedores", {}),
            ("/siigo/sync-productos", {}),

            # Documentos transaccionales con fecha global
            ("/siigo/sync-facturas", facturas_params),
            ("/siigo/sync-facturas", facturas_deep_params),
            ("/siigo/sync-notas-credito", notas_credito_params),
            ("/siigo/sync-compras", compras_params),

            # Documento Soporte API
            # Staging trae información desde API, pero la inserción a compras
            # sí respeta la fecha global mediante fecha_desde.
            ("/siigo/sync-documentos-soporte-staging", {"batch": 50}),
            ("/siigo/insert-documentos-soporte-desde-staging", ds_insert_params),

            # Cuentas por pagar y cruce NO llevan fecha global.
            ("/siigo/sync-accounts-payable", {}),
            ("/siigo/cross-accounts-payable", {}),
        ]

        print("🚀 === INICIO SECUENCIA SYNC-ALL ===")

        with app.test_client() as client:
            for ep, params in sequence:
                try:
                    print(f"➡️  Ejecutando {ep} ...")
                    inicio = time.time()

                    max_reintentos_429 = 3
                    intento_429 = 0

                    while True:
                        resp = client.post(
                            ep,
                            headers={
                                "X-ID-CLIENTE": str(idcliente),
                                "X-SYNC-ALL": "1"
                            },
                            query_string=params
                        )

                        status = resp.status_code
                        body = resp.get_data(as_text=True)

                        if status != 429:
                            break

                        intento_429 += 1

                        if intento_429 > max_reintentos_429:
                            print(f"❌ {ep} mantuvo 429 después de {max_reintentos_429} reintentos.")
                            break

                        espera = _extraer_espera_rate_limit(body, default=6)

                        print(
                            f"⏳ {ep} respondió 429. "
                            f"Reintento {intento_429}/{max_reintentos_429} en {espera}s..."
                        )

                        time.sleep(espera)

                    dur = round(time.time() - inicio, 1)
                    print(f"✅ {ep} completado en {dur}s -> {status}")

                    log_parts.append(f"{ep} {params} -> {status}: {body}")

                    # 📊 Guardar métrica individual del endpoint
                    try:
                        resumen = body[:300] if body else None
                        metric = SiigoSyncMetric(
                            idcliente=idcliente,
                            endpoint=ep,
                            duracion_segundos=dur,
                            status_code=status,
                            resultado="OK" if status < 400 else "ERROR",
                            detalle_resumen=resumen
                        )
                        db.session.add(metric)
                        db.session.commit()
                    except Exception as e:
                        print(f"⚠️  Error guardando métrica: {e}")

                    if status >= 400:
                        overall_status = "ERROR"
                        break

                    # Pausa defensiva para evitar rate limit de Siigo entre módulos
                    if ep.startswith("/siigo/"):
                        time.sleep(2)

                except Exception as e:
                    overall_status = "ERROR"
                    log_parts.append(f"{ep} excepción: {str(e)}")
                    break

        detalle = "\n".join(log_parts)

        # ✅ Usar pytz localmente
        import pytz
        tz_obj = pytz.timezone(tz_str)
        now_local = datetime.now(tz_obj)

        print(f"🕒 Fecha/hora local: {now_local.isoformat()}")
        print(f"🕐 Offset local detectado: {now_local.utcoffset()}")
        print(f"📦 Guardando resultado sync-all para cliente {idcliente}")

        # 🧩 Actualizar configuración o crearla
        config = SiigoSyncConfig.query.filter_by(idcliente=idcliente).first()

        if config:
            # IMPORTANTE:
            # Sync-all manual o por cron NO debe cambiar hora_ejecucion.
            # hora_ejecucion solo se cambia desde /config/sync.

            # Si el frontend manda fecha global explícitamente, actualizamos configuración.
            # Esto permite que el botón manual respete lo que el usuario acaba de configurar.
            if "sync_fecha_desde" in data or "ds_fecha_desde" in data:
                fecha_global = _parse_date_yyyy_mm_dd(
                    data.get("sync_fecha_desde") or data.get("ds_fecha_desde")
                )
                config.sync_fecha_desde = fecha_global
                config.ds_fecha_desde = fecha_global

            # Última ejecución general: manual o cron.
            config.ultimo_ejecutado = now_local

            # Última ejecución automática: SOLO cron.
            # Esto evita que una ejecución manual bloquee la ejecución automática del día siguiente.
            if es_cron:
                config.ultimo_auto_ejecutado = now_local

            config.resultado_ultima_sync = overall_status
            config.detalle_ultima_sync = detalle[:10000]
            db.session.add(config)

        else:
            nueva_fecha_global = None
            if data.get("sync_fecha_desde") or data.get("ds_fecha_desde"):
                nueva_fecha_global = _parse_date_yyyy_mm_dd(
                    data.get("sync_fecha_desde") or data.get("ds_fecha_desde")
                )

            config = SiigoSyncConfig(
                idcliente=idcliente,
                hora_ejecucion=_parse_time_hh_mm("02:00"),
                frecuencia_dias=1,
                activo=True,
                ultimo_ejecutado=now_local,
                ultimo_auto_ejecutado=now_local if es_cron else None,
                resultado_ultima_sync=overall_status,
                detalle_ultima_sync=detalle[:10000],
                sync_fecha_desde=nueva_fecha_global,
                ds_fecha_desde=nueva_fecha_global,
            )
            db.session.add(config)

        # 🧾 Registrar log histórico
        resumen_sync = _resumir_detalle_sync(detalle)

        logrec = SiigoSyncLog(
            idcliente=idcliente,
            fecha_programada=now_local,
            ejecutado_en=now_local,
            resultado=overall_status,
            detalle=detalle[:10000],
            origen=origen,
            total_pasos=resumen_sync["total_pasos"],
            pasos_ok=resumen_sync["pasos_ok"],
            pasos_error=resumen_sync["pasos_error"],
            endpoint_fallido=resumen_sync["endpoint_fallido"],
        )
        db.session.add(logrec)
        db.session.commit()

        # 🟢 Crear notificación para administradores del cliente
        try:
            titulo = "Sincronización automática completada" if es_cron else "Sincronización manual completada"

            ep_fallido = None
            for line in reversed(log_parts):
                if ("->" in line and "ERROR" in line) or ("excepción" in line):
                    ep_fallido = line.split(" ")[0]
                    break

            if overall_status == "OK":
                if es_cron:
                    mensaje = (
                        f"✅ La sincronización automática de Siigo finalizó correctamente "
                        f"el {now_local.strftime('%d/%m/%Y %H:%M')} ({tz_str})."
                    )
                else:
                    mensaje = (
                        f"✅ La sincronización manual de Siigo finalizó correctamente "
                        f"el {now_local.strftime('%d/%m/%Y %H:%M')} ({tz_str})."
                    )
                nivel = "success"
            else:
                if es_cron:
                    mensaje = (
                        f"❌ La sincronización automática de Siigo falló en "
                        f"{ep_fallido or 'uno de los módulos'} el "
                        f"{now_local.strftime('%d/%m/%Y %H:%M')} ({tz_str}). "
                        f"Revisa los reportes de integración para más detalles."
                    )
                else:
                    mensaje = (
                        f"❌ La sincronización manual de Siigo falló en "
                        f"{ep_fallido or 'uno de los módulos'} el "
                        f"{now_local.strftime('%d/%m/%Y %H:%M')} ({tz_str}). "
                        f"Revisa los reportes de integración para más detalles."
                    )
                nivel = "error"

            notif = SystemNotification(
                idcliente=idcliente,
                tipo="SYNC_RESULT",
                titulo=titulo,
                mensaje=mensaje,
                nivel=nivel,
                leido=False
            )
            db.session.add(notif)
            db.session.commit()
            print(f"📢 Notificación creada para cliente {idcliente}: {nivel}")

        except Exception as e:
            print(f"⚠️ Error creando notificación: {e}")

        print("✅ Registro en BD completado.\n")

        return jsonify({
            "status": overall_status,
            "detalle": detalle,
            "sync_fecha_desde": sync_fecha_desde,
            "ds_fecha_desde": sync_fecha_desde,
            "origen": origen,
            "ultimo_auto_ejecutado": (
                config.ultimo_auto_ejecutado.isoformat()
                if config and config.ultimo_auto_ejecutado else None
            )
        })

# Este endpoint no esta en suso pero se deja para futura posible verificacion o uso
    # --- CRON: Verificador automático de sincronización Siigo (cada 4 horas) ---
    @app.route("/cron/siigo-verifier", methods=["GET"])
    def cron_siigo_verifier():
        """
        Verifica en siigo_sync_config qué clientes deben ejecutar sync-all
        en las próximas 4 horas (según su zona horaria y frecuencia).
        Ejecuta si están dentro del rango.
        """
        from datetime import datetime, timedelta
        import pytz

        print("⏰ CRON Siigo-verifier iniciado...")

        ahora_utc = datetime.utcnow().replace(tzinfo=pytz.utc)
        rango_horas = 4  # 🔹 Cada 4 horas
        print(f"🕓 Hora actual UTC: {ahora_utc.isoformat()}")

        # Buscar configuraciones activas
        configs = SiigoSyncConfig.query.filter_by(activo=True).all()
        if not configs:
            print("⚠️ No hay configuraciones activas en siigo_sync_config.")
            return jsonify({"mensaje": "Sin configuraciones activas"})

        ejecutados = []
        skip_por_frecuencia = []
        fuera_de_rango = []

        for cfg in configs:
            cliente = Cliente.query.get(cfg.idcliente)
            if not cliente:
                continue

            tz_str = cliente.timezone or "America/Bogota"
            tz_obj = pytz.timezone(tz_str)
            ahora_local = ahora_utc.astimezone(tz_obj)

            hora_prog = cfg.hora_ejecucion
            fecha_prog_local = datetime.combine(ahora_local.date(), hora_prog)
            fecha_prog_local = tz_obj.localize(fecha_prog_local)

            # Ajuste si ya pasó la hora programada hoy → usar la de mañana
            if fecha_prog_local < ahora_local:
                fecha_prog_local += timedelta(days=1)

            # Verificar frecuencia
            if cfg.ultimo_ejecutado:
                dias_desde_ultima = (ahora_utc - cfg.ultimo_ejecutado).days
                if dias_desde_ultima < cfg.frecuencia_dias:
                    skip_por_frecuencia.append(cliente.idcliente)
                    continue

            # Si está dentro de las próximas 4 horas, ejecutar
            diff_horas = (fecha_prog_local - ahora_local).total_seconds() / 3600
            if 0 <= diff_horas <= rango_horas:
                print(f"🚀 Ejecutando sync-all para cliente {cliente.idcliente} ({tz_str})")
                try:
                    with app.test_client() as client:
                        resp = client.post(
                            "/siigo/sync-all",
                            headers={"X-ID-CLIENTE": str(cliente.idcliente)},
                            json={"origen": "cron"}
                        )
                        print(f"✅ Cliente {cliente.idcliente} → {resp.status_code}")
                        ejecutados.append(cliente.idcliente)
                except Exception as e:
                    print(f"❌ Error en cliente {cliente.idcliente}: {e}")
            else:
                fuera_de_rango.append(cliente.idcliente)

        print(f"🟢 Ejecutados: {ejecutados}")
        print(f"⏸ Omitidos por frecuencia: {skip_por_frecuencia}")
        print(f"⏰ Fuera de rango horario: {fuera_de_rango}")

        return jsonify({
            "hora_utc": ahora_utc.isoformat(),
            "ejecutados": ejecutados,
            "omitidos_por_frecuencia": skip_por_frecuencia,
            "fuera_de_rango": fuera_de_rango,
            "total_activos": len(configs)
        })



    @app.route("/api/notificaciones", methods=["GET"])
    @jwt_required()
    def get_notificaciones():
        claims = get_jwt()
        idcliente = claims.get("idcliente")
        perfilid = claims.get("perfilid")

        if not idcliente:
            return jsonify({"error": "No autorizado"}), 403

        # Solo administradores del cliente o superadmin (perfilid == 0)
        if perfilid not in [0, 1]:
            return jsonify([])

        notifs = (
            SystemNotification.query
            .filter_by(idcliente=idcliente, leido=False)
            .order_by(SystemNotification.creado_en.desc())
            .limit(5)
            .all()
        )

        return jsonify([
            {
                "id": n.id,
                "titulo": n.titulo,
                "mensaje": n.mensaje,
                "nivel": n.nivel,
                "fecha": n.creado_en.isoformat()
            }
            for n in notifs
        ])



    @app.route("/api/notificaciones/marcar-leida/<int:notif_id>", methods=["POST"])
    @jwt_required()
    def marcar_notificacion_leida(notif_id):
        idcliente = get_jwt().get("idcliente")
        notif = SystemNotification.query.filter_by(id=notif_id, idcliente=idcliente).first()
        if notif:
            notif.leido = True
            db.session.commit()
        return jsonify({"ok": True})




    # Para verificar la conexion del BAckend
    @app.route("/ping")
    def ping():
        return {"message": "pong"}, 200



    return app

app = create_app()  # 👈 ESTA LÍNEA ES CLAVE PARA RAILWAY (Gunicorn la necesita)


if __name__ == "__main__":
    app.run(debug=True)

